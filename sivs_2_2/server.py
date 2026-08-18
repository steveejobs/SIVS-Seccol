#!/usr/bin/env python3
"""SIVS — servidor local, API, persistência SQLite, backup e documentos técnicos."""

from __future__ import annotations

import argparse
import base64
import binascii
import collections
import contextlib
import csv
import hashlib
import hmac
import html
import http.client
import io
import json
import math
import mimetypes
import os
import re
import secrets
import shutil
import ssl
import sqlite3
import tempfile
import threading
import time
import traceback
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from http import HTTPStatus
from http.cookies import SimpleCookie
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse
from pypdf import PdfReader
from pypdf.errors import PyPdfError


BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
DEFAULT_DB = BASE_DIR / "data" / "sivs.db"


def load_local_env() -> None:
    """Carrega somente variáveis simples do .env local, sem substituir o ambiente."""
    env_path = BASE_DIR.parent / ".env"
    try:
        lines = env_path.read_text(encoding="utf-8").splitlines()
    except OSError:
        return
    for line in lines:
        text = line.strip()
        if not text or text.startswith("#") or "=" not in text:
            continue
        key, value = text.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key.replace("_", "").isalnum():
            os.environ.setdefault(key, value)


load_local_env()


def bounded_env_int(name, default, minimum, maximum):
    try:
        return max(minimum, min(int(os.environ.get(name, default)), maximum))
    except (TypeError, ValueError):
        return default


SESSION_SECONDS = 12 * 60 * 60
SESSION_IDLE_SECONDS = 60 * 60
SESSION_ACTIVE_SECONDS = 5 * 60
TELEMETRY_RETENTION_DAYS = bounded_env_int(
    "SIVS_TELEMETRY_RETENTION_DAYS", 180, 30, 3650
)
PBKDF2_ITERATIONS = 310_000
MAX_BODY = 16 * 1024 * 1024
MAX_IMPORT_BODY = 128 * 1024 * 1024
MAX_ATTACHMENT = 10 * 1024 * 1024
MAX_TENDER_DOCUMENT = 20 * 1024 * 1024
MAX_RECORD_PAYLOAD = 1024 * 1024
MAX_FISCAL_CERTIFICATE = 2 * 1024 * 1024
PARTNER_LOOKUP_TIMEOUT = 5
PARTNER_LOOKUP_CACHE_SECONDS = 15 * 60
VERSION = "2.2.0"


class BusinessKeyConflict(ValueError):
    """Identificador operacional duplicado dentro da empresa ativa."""


class InventoryWorkflowConflict(ValueError):
    """Transição incompatível com reservas ou baixas de estoque ativas."""


def mountinfo_has_path(contents: str, expected: str) -> bool:
    for line in contents.splitlines():
        fields = line.split()
        if len(fields) < 5:
            continue
        mount_path = re.sub(
            r"\\([0-7]{3})",
            lambda match: chr(int(match.group(1), 8)),
            fields[4],
        )
        if mount_path == expected:
            return True
    return False


def database_directory_is_mount(path: Path) -> bool:
    """Detecta inclusive bind mounts Linux, que Path.is_mount pode nao reconhecer."""
    resolved = str(path.expanduser().resolve())
    mountinfo = Path("/proc/self/mountinfo")
    if mountinfo.exists():
        try:
            return mountinfo_has_path(mountinfo.read_text(encoding="utf-8"), resolved)
        except OSError:
            pass
    return path.is_mount()


def require_persistent_database(path: Path) -> bool:
    """Interrompe a producao se o SQLite apontar para um diretorio nao montado."""
    required = os.environ.get("SIVS_REQUIRE_PERSISTENT_DB") == "1"
    if not required:
        return False
    database_dir = path.expanduser().resolve().parent
    if not database_dir.exists():
        raise RuntimeError(
            f"Diretorio persistente do SQLite nao existe: {database_dir}"
        )
    if not database_directory_is_mount(database_dir):
        raise RuntimeError(
            "Persistencia obrigatoria ausente: monte um volume no diretorio "
            f"{database_dir} antes de iniciar o SIVS"
        )
    return True


def database_readonly_connection(path: Path) -> sqlite3.Connection:
    """Abre um SQLite existente sem permitir criacao ou gravacao acidental."""
    uri = f"file:{path.expanduser().resolve().as_posix()}?mode=ro"
    return sqlite3.connect(uri, uri=True, timeout=20)


def validate_persistent_database_state(path: Path) -> dict:
    """Recusa em producao uma base ausente, vazia, corrompida ou nao configurada."""
    database_path = path.expanduser().resolve()
    allow_empty = os.environ.get("SIVS_ALLOW_EMPTY_DB_INITIALIZATION") == "1"
    if not database_path.exists() or database_path.stat().st_size == 0:
        if allow_empty:
            return {"bootstrap": True, "configured": False, "users": 0}
        raise RuntimeError(
            "Banco persistente ausente ou vazio. O SIVS recusou criar uma base nova "
            "durante o deploy. Para a primeira instalacao apenas, defina temporariamente "
            "SIVS_ALLOW_EMPTY_DB_INITIALIZATION=1 e remova a variavel apos criar o administrador."
        )
    try:
        connection = database_readonly_connection(database_path)
        try:
            integrity = connection.execute("PRAGMA quick_check").fetchone()
            if not integrity or integrity[0] != "ok":
                raise RuntimeError(f"PRAGMA quick_check retornou: {integrity!r}")
            tables = {
                row[0] for row in connection.execute(
                    "SELECT name FROM sqlite_master WHERE type='table'"
                )
            }
            if not {"users", "setup_state"}.issubset(tables):
                raise RuntimeError("schema essencial ausente")
            users = int(connection.execute("SELECT COUNT(*) FROM users").fetchone()[0])
            setup = connection.execute(
                "SELECT configured FROM setup_state WHERE id=1"
            ).fetchone()
            configured = bool(setup and setup[0])
        finally:
            connection.close()
    except (OSError, sqlite3.Error, RuntimeError) as exc:
        raise RuntimeError(
            f"Banco persistente invalido ou corrompido em {database_path}: {exc}"
        ) from exc
    if not configured or users < 1:
        if allow_empty:
            return {"bootstrap": True, "configured": configured, "users": users}
        raise RuntimeError(
            "Banco persistente sem configuracao administrativa valida. O deploy foi "
            "interrompido para impedir que uma base zerada substitua a base operacional."
        )
    return {"bootstrap": False, "configured": True, "users": users}


def create_prestart_database_backup(path: Path, retention: int | None = None) -> Path:
    """Cria snapshot SQLite consistente no volume antes de migracoes e inicializacao."""
    database_path = path.expanduser().resolve()
    if not database_path.exists() or database_path.stat().st_size == 0:
        raise RuntimeError("Nao ha banco persistente para o snapshot pre-start")
    keep = retention if retention is not None else bounded_env_int(
        "SIVS_PRESTART_BACKUP_RETENTION", 7, 2, 30
    )
    backup_dir = database_path.parent / "prestart-backups"
    backup_dir.mkdir(mode=0o700, parents=True, exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S.%fZ")
    destination_path = backup_dir / f"sivs-prestart-{stamp}.sqlite3"
    source = database_readonly_connection(database_path)
    destination = sqlite3.connect(destination_path)
    try:
        source.backup(destination)
        integrity = destination.execute("PRAGMA quick_check").fetchone()
        if not integrity or integrity[0] != "ok":
            raise RuntimeError(f"snapshot pre-start invalido: {integrity!r}")
    except Exception:
        destination.close()
        source.close()
        destination_path.unlink(missing_ok=True)
        raise
    else:
        destination.close()
        source.close()
    try:
        os.chmod(destination_path, 0o600)
    except OSError:
        pass
    snapshots = sorted(
        backup_dir.glob("sivs-prestart-*.sqlite3"),
        key=lambda candidate: candidate.stat().st_mtime,
        reverse=True,
    )
    for obsolete in snapshots[keep:]:
        obsolete.unlink(missing_ok=True)
    return destination_path


PNCP_MAX_REQUESTS_PER_SEARCH = 9
PNCP_TEXT_QUERIES_PER_SEARCH = 8
PNCP_TEXT_RESULTS_PER_QUERY = 50

DEFAULT_TENDER_SEARCH_QUERIES = [
    "cabine de segurança biológica",
    "capela de exaustão",
    "fluxo laminar",
    "filtro HEPA",
    "sala limpa",
    "contador de partículas",
    "qualificação de HVAC",
    "biodescontaminação",
]

DEFAULT_TENDER_KEYWORDS = [
    "controle de contaminação ambiental", "certificação de área limpa",
    "certificação de sala limpa", "qualificação de área limpa", "qualificação de sala limpa",
    "área limpa", "sala limpa", "ambiente controlado", "área classificada",
    "cabine de segurança biológica", "cabine de segurança microbiológica",
    "certificação de cabine", "manutenção de cabine de segurança biológica",
    "fluxo unidirecional", "módulo de fluxo unidirecional", "fluxo laminar",
    "capela de fluxo laminar", "manutenção de fluxo laminar", "capela de exaustão",
    "manutenção de capela de exaustão", "filtro HEPA", "filtros HEPA", "filtro ULPA", "filtros ULPA",
    "integridade de filtro HEPA", "teste de integridade HEPA", "teste PAO",
    "fotômetro de aerossol", "gerador de aerossol", "contagem de partículas",
    "contador de partículas", "partículas em suspensão", "qualificação de HVAC",
    "validação de HVAC", "balanceamento de insuflamento e exaustão",
    "pressão diferencial entre salas", "teste de recuperação", "trocas de ar por hora",
    "vazão de ar insuflado", "biodescontaminação", "descontaminação por VHP",
    "peróxido de hidrogênio vaporizado", "unidade de ventilação e descontaminação",
    "unidade de filtragem refrigerada", "rack isolador", "isolador farmacêutico",
    "difusor motorizado", "projeto de área limpa", "projeto de centro cirúrgico",
    "certificação de centro cirúrgico", "qualificação ISO 5", "certificação ISO 5", "ISO 14644",
    "velocidade e uniformidade do fluxo de ar", "ensaio de inflow",
    "perda de carga dos filtros", "detecção de vazamento HEPA", "ensaio de fumaça",
    "visualização do fluxo de ar", "avaliação de alarmes da CSB",
    "ensaio de intensidade de iluminação", "ensaio de vibração", "ensaio de ruído",
    "substituição de filtro HEPA", "troca de filtros HEPA", "reparo de filtro HEPA", "estanqueidade de filtro HEPA",
    "balanceamento de insuflamento e retorno", "radiação de lâmpada germicida",
    "saturação de filtro HEPA", "pressão diferencial do filtro", "selos de vedação",
    "componentes eletromecânicos", "certificação de fluxo laminar",
    "certificação de capela", "reforma de cabine", "reforma de capela"
]

# Termos que aumentam a aderência, mas não disparam oportunidades isoladamente.
SECCOL_CONTEXT_TERMS = [
    "reprodução assistida", "fertilização humana", "fertilização animal", "oncologia",
    "centro cirúrgico", "análises clínicas", "indústria farmacêutica", "nutrição parenteral",
    "nutrição enteral", "indústria de injetáveis", "indústria alimentícia", "cosmético",
    "indústria química", "indústria agropecuária", "biotecnologia", "instituto de pesquisa",
    "universidade", "laboratório P3", "laboratório NB3", "hospital", "life science",
    "contador de partículas", "fotômetro", "gerador de aerossol", "balometer", "luxímetro",
    "decibelímetro", "termo anemômetro", "manômetro digital", "termo higrômetro",
    "radiômetro", "ISO 21501-4", "ANVISA", "SBCC", "certificação", "qualificação",
    "manutenção", "reforma", "ensaio", "serviço técnico"
]

SOURCE_CATALOG = [
    ("pncp", "PNCP — Portal Nacional de Contratações Públicas", "https://pncp.gov.br/app/editais", "Nacional", "API automática", "Oficial"),
    ("comprasgov", "Compras.gov.br — Governo Federal", "https://www.gov.br/compras/pt-br", "Nacional/Federal", "API automática de contingência", "Oficial"),
    ("dou", "Diário Oficial da União — Imprensa Nacional", "https://www.in.gov.br/consulta", "Nacional", "Consulta manual/alerta oficial", "Oficial"),
    ("transparencia", "Portal da Transparência — Licitações", "https://portaldatransparencia.gov.br/licitacoes", "Federal", "Consulta manual — API planejada", "Oficial"),
    ("licitacoese", "Licitações-e — Banco do Brasil", "https://www.licitacoes-e.com.br/", "Nacional", "Consulta manual/autenticada", "Plataforma"),
    ("bll", "BLL Compras", "https://bll.org.br/editais/", "Nacional", "Consulta manual — sem scraping", "Plataforma"),
    ("pcp", "Portal de Compras Públicas", "https://www.portaldecompraspublicas.com.br/processos", "Nacional", "Consulta manual — automação externa vedada", "Plataforma"),
    ("licitanet", "Licitanet", "https://licitanet.com.br/", "Nacional", "Consulta manual", "Plataforma"),
    ("bnc", "Bolsa Nacional de Compras — BNC", "https://bnc.org.br/", "Nacional", "Consulta manual", "Plataforma"),
    ("sp", "Portal de Compras do Estado de São Paulo", "https://compras.sp.gov.br/", "SP", "Consulta manual", "Estadual"),
    ("mg", "Portal de Compras de Minas Gerais", "https://compras.mg.gov.br/", "MG", "Consulta manual", "Estadual"),
    ("rj", "Compras do Estado do Rio de Janeiro", "https://www.compras.rj.gov.br/", "RJ", "Consulta manual", "Estadual"),
    ("es", "Portal de Compras do Espírito Santo", "https://compras.es.gov.br/", "ES", "Consulta manual", "Estadual"),
    ("rs", "Compras Eletrônicas do Rio Grande do Sul", "https://www.compras.rs.gov.br/", "RS", "Consulta manual", "Estadual"),
    ("sc", "Portal de Compras de Santa Catarina", "https://compras.sc.gov.br/", "SC", "Consulta manual", "Estadual"),
    ("pr", "Compras Paraná", "https://www.comprasparana.pr.gov.br/", "PR", "Consulta manual", "Estadual"),
    ("ba", "ComprasNet Bahia", "https://www.comprasnet.ba.gov.br/", "BA", "Consulta manual", "Estadual"),
    ("ce", "Portal de Compras do Ceará", "https://www.portalcompras.ce.gov.br/", "CE", "Consulta manual", "Estadual"),
    ("pe", "Portal de Compras de Pernambuco", "https://www.peintegrado.pe.gov.br/", "PE", "Consulta manual/autenticada", "Estadual"),
    ("pb", "Central de Compras da Paraíba", "https://centraldecompras.pb.gov.br/", "PB", "Consulta manual", "Estadual"),
    ("rn", "Portal de Compras do Rio Grande do Norte", "https://portaldecompras.rn.gov.br/", "RN", "Consulta manual", "Estadual"),
    ("al", "Portal de Compras de Alagoas", "https://compras.al.gov.br/", "AL", "Consulta manual", "Estadual"),
    ("se", "ComprasNet Sergipe", "https://www.comprasnet.se.gov.br/", "SE", "Consulta manual", "Estadual"),
    ("go", "SISLOG — Compras de Goiás", "https://sislog.go.gov.br/", "GO", "Consulta manual", "Estadual"),
    ("to", "Central de Compras do Tocantins", "https://centraldecompras.to.gov.br/", "TO", "Consulta manual prioritária", "Estadual"),
    ("mt", "Portal de Aquisições de Mato Grosso", "https://aquisicoes.seplag.mt.gov.br/", "MT", "Consulta manual", "Estadual"),
    ("ms", "Portal de Compras de Mato Grosso do Sul", "https://www.compras.ms.gov.br/", "MS", "Consulta manual", "Estadual"),
    ("df", "Portal de Compras do Distrito Federal", "https://www.compras.df.gov.br/", "DF", "Consulta manual", "Distrital"),
    ("pa", "Portal de Compras do Pará", "https://www.compraspara.pa.gov.br/", "PA", "Consulta manual prioritária", "Estadual"),
    ("am", "e-Compras Amazonas", "https://www.e-compras.am.gov.br/publico/", "AM", "Consulta manual", "Estadual"),
    ("ro", "SUPEL — Licitações de Rondônia", "https://rondonia.ro.gov.br/supel/", "RO", "Consulta manual", "Estadual"),
    ("ac", "Portal de Licitações do Acre", "https://www.ac.gov.br/", "AC", "Consulta manual", "Estadual"),
    ("rr", "Portal de Compras de Roraima", "https://www.gov.br/compras/pt-br", "RR", "Cobertura PNCP/Compras.gov", "Estadual"),
    ("ap", "Central de Licitações do Amapá", "https://compras.portal.ap.gov.br/", "AP", "Consulta manual", "Estadual"),
    ("ma", "Portal de Compras do Maranhão", "https://www.compras.ma.gov.br/", "MA", "Consulta manual", "Estadual"),
    ("pi", "Portal de Compras do Piauí", "https://sistemas.tce.pi.gov.br/licitacoesweb/", "PI", "Consulta manual", "Estadual"),
    ("cnes", "CNES — Estabelecimentos de Saúde", "https://cnes.datasus.gov.br/", "Nacional", "Prospecção privada — não é edital", "Mercado privado"),
    ("anahp", "ANAHP — Hospitais Privados", "https://www.anahp.com.br/", "Nacional", "Prospecção privada — não é edital", "Mercado privado"),
]

# Catálogo curado em 15/08/2026. Para normas comerciais, a instalação inclui
# uma ficha autoral de referência e o link oficial; a íntegra licenciada deve
# ser anexada pela empresa que detém a licença de uso.
NORM_CATALOG = [
    {
        "key": "iso-14644-1-2015", "code": "ISO 14644-1:2015", "organization": "ISO",
        "edition": "2ª edição · 2015", "status": "Publicada — em revisão sistemática",
        "scope": "Classificação da limpeza do ar pela concentração de partículas em salas e zonas limpas.",
        "application": "Classificação ISO de áreas limpas e dispositivos separativos.",
        "tests": "Contagem de partículas; definição de classe de limpeza; plano de amostragem.",
        "url": "https://www.iso.org/standard/53394.html", "license": "Comercial/licenciada",
    },
    {
        "key": "iso-14644-2-2015", "code": "ISO 14644-2:2015", "organization": "ISO",
        "edition": "2ª edição · 2015", "status": "Publicada — em revisão sistemática",
        "scope": "Requisitos mínimos para o plano de monitoramento do desempenho de salas limpas.",
        "application": "Definição de frequência, parâmetros e evidências do monitoramento continuado.",
        "tests": "Monitoramento de partículas e parâmetros que afetam a concentração no ar.",
        "url": "https://www.iso.org/standard/53393.html", "license": "Comercial/licenciada",
    },
    {
        "key": "iso-14644-3-2019", "code": "ISO 14644-3:2019", "organization": "ISO",
        "edition": "2ª edição · 2019", "status": "Publicada",
        "scope": "Métodos de ensaio para caracterizar o desempenho de salas e zonas limpas.",
        "application": "Base técnica principal dos ensaios de qualificação e certificação de áreas limpas.",
        "tests": "Fluxo e velocidade do ar; diferencial de pressão; integridade HEPA; visualização; recuperação; partículas.",
        "url": "https://www.iso.org/standard/60598.html", "license": "Comercial/licenciada",
    },
    {
        "key": "iso-14644-4-2022", "code": "ISO 14644-4:2022", "organization": "ISO",
        "edition": "2ª edição · 2022", "status": "Publicada",
        "scope": "Projeto, construção e partida de salas limpas e ambientes controlados associados.",
        "application": "Projetos, reformas, comissionamento e critérios de aceitação de áreas limpas.",
        "tests": "Requisitos do usuário; projeto; construção; comissionamento; entrega documental.",
        "url": "https://www.iso.org/standard/72379.html", "license": "Comercial/licenciada",
    },
    {
        "key": "iso-14644-5-2025", "code": "ISO 14644-5:2025", "organization": "ISO",
        "edition": "2ª edição · 2025", "status": "Publicada",
        "scope": "Programa de controle operacional para salas limpas em funcionamento.",
        "application": "Operação, limpeza, manutenção, pessoal, materiais e acompanhamento de desempenho.",
        "tests": "Controles operacionais; procedimentos; treinamento; limpeza; manutenção e monitoramento.",
        "url": "https://www.iso.org/standard/88599.html", "license": "Comercial/licenciada",
    },
    {
        "key": "iso-14644-7-2004", "code": "ISO 14644-7:2004", "organization": "ISO",
        "edition": "1ª edição · 2004", "status": "Publicada — revisão em desenvolvimento",
        "scope": "Requisitos mínimos para dispositivos separativos, como cabines, isoladores e ambientes de luvas.",
        "application": "Cabines de segurança, fluxos unidirecionais, isoladores e equipamentos de contenção.",
        "tests": "Desempenho do dispositivo; contenção; fluxo de ar; integridade de barreiras e filtros.",
        "url": "https://www.iso.org/standard/38264.html", "license": "Comercial/licenciada",
    },
    {
        "key": "iso-iec-17025-2017", "code": "ISO/IEC 17025:2017", "organization": "ISO/IEC",
        "edition": "3ª edição · 2017", "status": "Publicada — confirmada em 2023",
        "scope": "Requisitos gerais para competência, imparcialidade e operação consistente de laboratórios.",
        "application": "Governança dos ensaios e calibrações, rastreabilidade, validade de resultados e laudos.",
        "tests": "Competência; métodos; equipamentos; rastreabilidade; incerteza; validade; relato de resultados.",
        "url": "https://www.iso.org/standard/66912.html", "license": "Comercial/licenciada",
    },
    {
        "key": "nsf-ansi-49-2022", "code": "NSF/ANSI 49-2022", "organization": "NSF/ANSI",
        "edition": "Edição 2022", "status": "Publicada",
        "scope": "Desempenho, construção e certificação em campo de cabines de segurança biológica Classe II.",
        "application": "Certificação de cabines de segurança biológica Classe II.",
        "tests": "Velocidade de inflow/downflow; integridade HEPA; fumaça; alarmes; iluminação; ruído e vibração.",
        "url": "https://www.nsf.org/lab-testing/biosafety-cabinetry/biosafety-cabinet-certification",
        "license": "Comercial/licenciada",
    },
    {
        "key": "anvisa-rdc-50-2002", "code": "ANVISA RDC 50/2002", "organization": "ANVISA",
        "edition": "RDC nº 50 · 2002", "status": "Referência regulatória — verificar alterações aplicáveis",
        "scope": "Regulamento técnico para planejamento, programação, elaboração e avaliação de projetos físicos de estabelecimentos de saúde.",
        "application": "Projetos e avaliações de ambientes assistenciais e áreas de apoio em saúde.",
        "tests": "Requisitos de projeto físico, fluxos, instalações e ambientes de estabelecimentos assistenciais.",
        "url": "https://bvsms.saude.gov.br/bvs/saudelegis/anvisa/2002/res0050_21_02_2002.html",
        "license": "Acesso público",
    },
    {
        "key": "anvisa-rdc-67-2007", "code": "ANVISA RDC 67/2007", "organization": "ANVISA",
        "edition": "RDC nº 67 · 2007", "status": "Vigente com alterações — confirmar escopo aplicável",
        "scope": "Boas práticas de manipulação de preparações magistrais e oficinais para uso humano.",
        "application": "Ambientes e equipamentos de farmácias de manipulação, conforme o processo e o risco.",
        "tests": "Controles ambientais, instalações, equipamentos, limpeza e qualificação aplicáveis.",
        "url": "https://www.gov.br/anvisa/pt-br/setorregulado/regularizacao/farmacias-e-drogarias/boas-praticas-farmaceuticas",
        "license": "Acesso público",
    },
    {
        "key": "anvisa-rdc-658-2022", "code": "ANVISA RDC 658/2022", "organization": "ANVISA",
        "edition": "RDC nº 658 · 2022", "status": "Vigente",
        "scope": "Diretrizes gerais de Boas Práticas de Fabricação de medicamentos.",
        "application": "Avaliação de sistemas, instalações e ambientes de fabricação farmacêutica.",
        "tests": "Sistema da qualidade; instalações; equipamentos; documentação; produção; controle e validação.",
        "url": "https://www.gov.br/anvisa/pt-br/assuntos/noticias-anvisa/2022/revisaco-entenda-como-ficaram-as-normas-da-area-de-fiscalizacao",
        "license": "Acesso público",
    },
    {
        "key": "anvisa-in-138-2022", "code": "ANVISA IN 138/2022", "organization": "ANVISA",
        "edition": "IN nº 138 · 2022", "status": "Vigente",
        "scope": "Diretrizes complementares sobre qualificação e validação na fabricação de medicamentos.",
        "application": "Planos, protocolos, execução e documentação de qualificação e validação.",
        "tests": "URS; qualificação de projeto, instalação, operação e desempenho; validação e controle de mudanças.",
        "url": "https://www.gov.br/anvisa/pt-br/assuntos/noticias-anvisa/2022/revisaco-entenda-como-ficaram-as-normas-da-area-de-fiscalizacao",
        "license": "Acesso público",
    },
    {
        "key": "iso-21501-4-2018-amd1-2023", "code": "ISO 21501-4:2018 + Amd 1:2023",
        "organization": "ISO", "edition": "2ª edição · 2018 + Emenda 1 · 2023",
        "status": "Publicada — revisão em desenvolvimento",
        "scope": "Calibração e verificação de contadores ópticos de partículas em suspensão no ar para espaços limpos.",
        "application": "Controle dos contadores usados na classificação e no monitoramento de partículas.",
        "tests": "Eficiência de detecção; resolução; falso contagem; vazão; resposta; calibração e verificação do LSAPC.",
        "url": "https://www.iso.org/standard/58073.html", "license": "Comercial/licenciada",
    },
    {
        "key": "iest-rp-cc006-4", "code": "IEST-RP-CC006.4", "organization": "IEST",
        "edition": "Revisão .4", "status": "Publicada",
        "scope": "Métodos para caracterização do desempenho de salas e zonas limpas em diferentes fases operacionais.",
        "application": "Especificação e ensaios complementares de conformidade operacional de áreas limpas.",
        "tests": "Desempenho de salas limpas; fases de ocupação; critérios e documentação de ensaios.",
        "url": "https://www.iest.org/Standards-RPs/Recommended-Practices/IEST-RP-CC006",
        "license": "Comercial/licenciada",
    },
    {
        "key": "iest-rp-cc019-1", "code": "IEST-RP-CC019.1", "organization": "IEST",
        "edition": "Revisão .1 · 2006", "status": "Publicada",
        "scope": "Qualificações recomendadas para organizações e profissionais que testam e certificam salas limpas e dispositivos de ar limpo.",
        "application": "Matriz de competência, escolaridade, treinamento e experiência da equipe de certificação.",
        "tests": "Categorias profissionais; requisitos de competência; treinamento; experiência e evidência de qualificação.",
        "url": "https://www.iest.org/Standards-RPs/Recommended-Practices/IEST-RP-CC019",
        "license": "Comercial/licenciada",
    },
    {
        "key": "iest-rp-cc034-5", "code": "IEST-RP-CC034.5", "organization": "IEST",
        "edition": "Revisão .5", "status": "Publicada",
        "scope": "Definições, equipamentos e procedimentos para ensaios de vazamento em filtros HEPA e ULPA.",
        "application": "Integridade de filtros na fábrica, antes da instalação e instalados em áreas/equipamentos.",
        "tests": "Geração e medição de aerossol; varredura; vazamento local; aceitação e registro de filtros HEPA/ULPA.",
        "url": "https://www.iest.org/Standards-RPs/Recommended-Practices/IEST-RP-CC034",
        "license": "Comercial/licenciada",
    },
    {
        "key": "ashrae-110-2016-ra2025", "code": "ANSI/ASHRAE 110-2016 (RA 2025)",
        "organization": "ASHRAE", "edition": "Edição 2016 · reafirmada em 2025", "status": "Publicada",
        "scope": "Métodos quantitativos e qualitativos para avaliar a contenção de capelas laboratoriais.",
        "application": "Ensaios de capelas de exaustão convencionais, bypass, ar auxiliar e VAV.",
        "tests": "Velocidade de face; visualização; contenção com gás traçador; condição como fabricada, instalada ou usada.",
        "url": "https://www.ashrae.org/technical-resources/standards-and-guidelines/titles-purposes-and-scopes",
        "license": "Comercial/licenciada",
    },
    {
        "key": "ashrae-111-2024", "code": "ANSI/ASHRAE 111-2024", "organization": "ASHRAE",
        "edition": "Edição 2024 · errata de 05/09/2025", "status": "Publicada",
        "scope": "Medição, ensaio, ajuste, balanceamento, avaliação e relato de desempenho de sistemas HVAC prediais.",
        "application": "Balanceamento, vazões, trocas de ar, pressurização e avaliação de contaminação cruzada.",
        "tests": "Vazão; balanceamento; taxa de renovação; pressão; ventilação externa; validação e relatório dos dados.",
        "url": "https://www.ashrae.org/technical-resources/standards-and-guidelines/titles-purposes-and-scopes",
        "license": "Comercial/licenciada",
    },
]

# Portfólio operacional confirmado pela direção em 15/08/2026: tudo o que é
# apresentado no site oficial integra a produção/fornecimento da SECCOL ou o
# seu patrimônio técnico. Os grupos abaixo mantêm essas naturezas separadas.
SECCOL_PRODUCT_CATALOG = [
    {
        "key": "produto-area-limpa", "code": "SEC-PRO-001", "title": "Área Limpa / Sala Limpa",
        "family": "Ambientes controlados", "kind": "Produção, projeto e fornecimento SECCOL",
        "description": "Solução de ambiente controlado para reduzir introdução, geração e retenção de contaminantes.",
        "source": "https://www.seccol.com.br/area-limpa.html",
        "norms": ["iso-14644-1-2015", "iso-14644-2-2015", "iso-14644-3-2019",
                  "iso-14644-4-2022", "iso-14644-5-2025", "ashrae-111-2024",
                  "anvisa-rdc-50-2002"],
    },
    {
        "key": "produto-cabine-seguranca-biologica", "code": "SEC-PRO-002",
        "title": "Cabine de Segurança Biológica", "family": "Contenção biológica",
        "kind": "Produção e fornecimento SECCOL",
        "description": "Equipamento de contenção primária para proteção do operador, produto e ambiente.",
        "source": "https://www.seccol.com.br/quem.html",
        "norms": ["nsf-ansi-49-2022", "iso-14644-3-2019", "iso-14644-7-2004",
                  "iest-rp-cc034-5"],
    },
    {
        "key": "produto-capela-exaustao", "code": "SEC-PRO-003", "title": "Capela de Exaustão",
        "family": "Proteção laboratorial", "kind": "Produção e fornecimento SECCOL",
        "description": "Barreira de proteção laboratorial com exaustão para contenção de emissões do processo.",
        "source": "https://www.seccol.com.br/quem.html",
        "norms": ["ashrae-110-2016-ra2025", "ashrae-111-2024", "iso-iec-17025-2017"],
    },
    {
        "key": "produto-fluxo-unidirecional", "code": "SEC-PRO-004",
        "title": "Equipamento de Fluxo Unidirecional (Laminar)", "family": "Ar limpo localizado",
        "kind": "Produção e fornecimento SECCOL",
        "description": "Equipamento para manipulação protegida de materiais biológicos ou estéreis.",
        "source": "https://www.seccol.com.br/quem.html",
        "norms": ["iso-14644-1-2015", "iso-14644-3-2019", "iso-14644-7-2004",
                  "iest-rp-cc034-5"],
    },
    {
        "key": "produto-unidade-descontaminacao", "code": "SEC-PRO-005",
        "title": "Unidade de Descontaminação / Ventilação", "family": "Tratamento de ar",
        "kind": "Produção e fornecimento SECCOL",
        "description": "Unidade para ventilação controlada e suporte a processos de descontaminação.",
        "source": "https://www.seccol.com.br/quem.html",
        "norms": ["iso-14644-3-2019", "iso-14644-5-2025", "ashrae-111-2024"],
    },
    {
        "key": "produto-filtro-hepa-ulpa", "code": "SEC-PRO-006", "title": "Filtro Absoluto HEPA / ULPA",
        "family": "Filtração absoluta", "kind": "Componente fornecido pela SECCOL",
        "description": "Elemento filtrante de reposição para sistemas e equipamentos de ar limpo.",
        "source": "https://www.seccol.com.br/teste-equipamento.html",
        "norms": ["iso-14644-3-2019", "iest-rp-cc034-5"],
    },
    {
        "key": "produto-motor-eletrico", "code": "SEC-PRO-007", "title": "Motor Elétrico de Reposição",
        "family": "Componentes eletromecânicos", "kind": "Componente fornecido pela SECCOL",
        "description": "Motor de reposição para manutenção e recuperação de equipamentos atendidos.",
        "source": "https://www.seccol.com.br/teste-equipamento.html", "norms": [],
    },
]

SECCOL_INSTRUMENT_CATALOG = [
    ("instrumento-contador-particulas", "SEC-INS-001", "Contador de Partículas", "Contagem e classificação de partículas no ar", ["iso-21501-4-2018-amd1-2023", "iso-14644-1-2015", "iso-14644-3-2019"]),
    ("instrumento-fotometro-pao", "SEC-INS-002", "Fotômetro e Gerador de Aerossol (PAO)", "Integridade de sistemas de filtragem HEPA/ULPA", ["iso-14644-3-2019", "iest-rp-cc034-5"]),
    ("instrumento-balometer", "SEC-INS-003", "Balometer", "Vazão, pressão, temperatura, umidade e velocidade do ar", ["iso-14644-3-2019", "ashrae-111-2024"]),
    ("instrumento-luximetro", "SEC-INS-004", "Luxímetro", "Medição de iluminância", ["iso-14644-3-2019", "iso-iec-17025-2017"]),
    ("instrumento-decibelimetro", "SEC-INS-005", "Decibelímetro", "Medição de nível de pressão sonora", ["iso-14644-3-2019", "iso-iec-17025-2017"]),
    ("instrumento-termoanemometro", "SEC-INS-006", "Termoanemômetro de Fio Quente", "Medição de velocidade em baixos fluxos de ar", ["iso-14644-3-2019", "ashrae-111-2024"]),
    ("instrumento-manometro", "SEC-INS-007", "Manômetro Digital", "Pressão diferencial de ambientes e saturação de filtros", ["iso-14644-3-2019", "ashrae-111-2024"]),
    ("instrumento-alicate-amperimetro", "SEC-INS-008", "Alicate Amperímetro", "Medição elétrica sem interrupção do circuito", ["iso-iec-17025-2017"]),
    ("instrumento-ampola-fumaca", "SEC-INS-009", "Ampola de Fumaça", "Visualização de movimentação e sentido do fluxo de ar", ["iso-14644-3-2019"]),
    ("instrumento-termohigrometro", "SEC-INS-010", "Termohigrômetro", "Medição de temperatura e umidade relativa", ["iso-14644-3-2019", "iso-iec-17025-2017"]),
    ("instrumento-radiometro-uvc", "SEC-INS-011", "Radiômetro UVC", "Medição da emissão de fontes ultravioleta germicidas", ["iso-iec-17025-2017"]),
    ("instrumento-vhp", "SEC-INS-012", "VHP — Vapor de Peróxido de Hidrogênio", "Biodescontaminação de ambientes, superfícies e equipamentos", ["iso-14644-5-2025"]),
]

SECCOL_SERVICE_CATALOG = [
    ("servico-manutencao-equipamento", "SEC-SRV-001", "Manutenção de Equipamentos", "Manutenção", "Prevenção e correção para reduzir paradas operacionais", ["iso-14644-5-2025"]),
    ("servico-reforma-equipamento", "SEC-SRV-002", "Reforma de Equipamentos", "Reforma", "Recuperação funcional, estrutural e eletromecânica", ["iso-14644-5-2025"]),
    ("servico-certificacao-equipamento", "SEC-SRV-003", "Certificação de Equipamentos", "Certificação", "Ensaios, avaliação e emissão de certificado técnico", ["iso-14644-3-2019", "iso-iec-17025-2017"]),
    ("servico-certificacao-area-limpa", "SEC-SRV-004", "Certificação de Área Limpa", "Certificação", "Classificação, qualificação e relatório técnico do ambiente", ["iso-14644-1-2015", "iso-14644-2-2015", "iso-14644-3-2019"]),
    ("servico-projeto-area-limpa", "SEC-SRV-005", "Projeto e Execução de Área Limpa", "Engenharia", "Levantamento, fluxos, projeto, execução e entrega controlada", ["iso-14644-4-2022", "anvisa-rdc-50-2002"]),
    ("servico-projeto-centro-cirurgico", "SEC-SRV-006", "Projeto e Execução de Centro Cirúrgico", "Engenharia", "Projeto técnico e execução de ambiente cirúrgico", ["anvisa-rdc-50-2002", "ashrae-111-2024"]),
    ("servico-monitoramento-descontaminacao", "SEC-SRV-007", "Monitoramento de Unidade de Descontaminação / Ventilação", "Monitoramento", "Acompanhamento de desempenho segundo normas aplicáveis", ["iso-14644-2-2015", "iso-14644-5-2025", "ashrae-111-2024"]),
    ("ensaio-velocidade-uniformidade", "SEC-SRV-008", "Velocidade e Uniformidade do Fluxo de Ar", "Ensaio de ar", "Medição e avaliação da uniformidade do fluxo", ["iso-14644-3-2019", "ashrae-111-2024"]),
    ("ensaio-inflow", "SEC-SRV-009", "Velocidade do Fluxo de Ar de Face (Inflow)", "Ensaio de contenção", "Medição da velocidade de entrada em cabine", ["nsf-ansi-49-2022"]),
    ("ensaio-perda-carga-filtro", "SEC-SRV-010", "Perda de Carga do Sistema HEPA / ULPA", "Ensaio de filtragem", "Diferença de pressão do sistema de filtragem", ["iso-14644-3-2019", "iest-rp-cc034-5"]),
    ("ensaio-integridade-hepa-pao", "SEC-SRV-011", "Integridade e Estanqueidade HEPA / ULPA (PAO)", "Ensaio de filtragem", "Detecção e localização de vazamentos no sistema instalado", ["iso-14644-3-2019", "iest-rp-cc034-5"]),
    ("ensaio-visualizacao-fumaca", "SEC-SRV-012", "Visualização do Fluxo de Ar (Fumaça)", "Ensaio de ar", "Avaliação visual do sentido e do padrão de fluxo", ["iso-14644-3-2019", "nsf-ansi-49-2022"]),
    ("ensaio-contagem-particulas", "SEC-SRV-013", "Contagem de Partículas em Suspensão", "Ensaio de partículas", "Contagem eletrônica e classificação de limpeza", ["iso-14644-1-2015", "iso-21501-4-2018-amd1-2023"]),
    ("ensaio-alarmes-csb", "SEC-SRV-014", "Avaliação dos Alarmes da CSB", "Ensaio de segurança", "Verificação funcional dos alarmes da cabine", ["nsf-ansi-49-2022"]),
    ("ensaio-iluminacao", "SEC-SRV-015", "Intensidade de Iluminação", "Ensaio ambiental", "Medição de iluminância em equipamento ou ambiente", ["iso-14644-3-2019", "iso-iec-17025-2017"]),
    ("ensaio-vibracao", "SEC-SRV-016", "Ensaio de Vibração", "Ensaio ambiental", "Medição e avaliação de vibração do equipamento", ["nsf-ansi-49-2022", "iso-iec-17025-2017"]),
    ("ensaio-ruido", "SEC-SRV-017", "Ensaio de Ruído", "Ensaio ambiental", "Medição de nível de pressão sonora", ["nsf-ansi-49-2022", "iso-iec-17025-2017"]),
    ("servico-filtros-inspecao-substituicao", "SEC-SRV-018", "Inspeção e Substituição de Filtros", "Manutenção", "Inspeção ou troca de filtros grossos e absolutos", ["iso-14644-5-2025", "iest-rp-cc034-5"]),
    ("servico-balanceamento", "SEC-SRV-019", "Balanceamento de Insuflamento, Exaustão e Retorno", "TAB/HVAC", "Ajuste e balanceamento dos sistemas de ar", ["ashrae-111-2024", "iso-14644-3-2019"]),
    ("servico-limpeza-interna", "SEC-SRV-020", "Limpeza Interna de Equipamento", "Manutenção", "Limpeza técnica da parte interna do equipamento", ["iso-14644-5-2025"]),
    ("ensaio-radiacao-uvc", "SEC-SRV-021", "Eficiência de Radiação UVC", "Ensaio de radiação", "Medição da eficiência de lâmpadas germicidas", ["iso-iec-17025-2017"]),
    ("ensaio-saturacao-filtro", "SEC-SRV-022", "Saturação de Filtros por Pressão Diferencial", "Ensaio de filtragem", "Medição da pressão diferencial para avaliar saturação", ["iso-14644-3-2019"]),
    ("ensaio-eletrico-motor", "SEC-SRV-023", "Tensão e Corrente Elétrica do Motor", "Ensaio elétrico", "Medição elétrica do conjunto motriz", ["iso-iec-17025-2017"]),
    ("servico-reparo-filtro", "SEC-SRV-024", "Reparo de Filtro HEPA", "Manutenção", "Reparo do meio filtrante ou da estrutura, quando tecnicamente admissível", ["iest-rp-cc034-5"]),
    ("servico-manometro-vedacao", "SEC-SRV-025", "Revisão de Manômetro e Selos de Vedação", "Manutenção", "Revisão do indicador de pressão e das vedações", ["iso-14644-5-2025"]),
    ("servico-componentes-eletromecanicos", "SEC-SRV-026", "Verificação de Componentes Eletromecânicos", "Manutenção", "Inspeção funcional dos componentes do equipamento", ["iso-14644-5-2025"]),
    ("ensaio-vazao-trocas-ar", "SEC-SRV-027", "Vazão e Número de Trocas de Ar", "Ensaio de área limpa", "Cálculo de vazão insuflada e renovações por hora", ["iso-14644-3-2019", "ashrae-111-2024"]),
    ("ensaio-recuperacao", "SEC-SRV-028", "Teste de Recuperação", "Ensaio de área limpa", "Determinação da capacidade de recuperação da limpeza do ambiente", ["iso-14644-3-2019"]),
    ("ensaio-pressao-entre-salas", "SEC-SRV-029", "Pressão Diferencial entre Salas", "Ensaio de área limpa", "Medição do diferencial e verificação da cascata de pressão", ["iso-14644-3-2019", "ashrae-111-2024"]),
]

NORMATIVE_REQUIRED_MODULES = {"certificados", "laudos_tecnicos", "estudos_tecnicos"}

MODULES = {
    # Administrativo
    "arquivos": "Arquivos administrativos",
    "clientes_fornecedores": "Clientes e fornecedores",
    "clientes": "Clientes",
    "fornecedores": "Fornecedores",
    "contatos": "Contatos",
    "importacoes_xml": "Importação XML NF-e",
    "solicitacoes_compra": "Solicitações de compra",
    "pedidos_compra": "Pedidos de compra",
    "ramais": "Ramais",
    # Comercial e inteligência
    "crm": "CRM",
    "propostas": "Propostas",
    "contratos": "Contratos",
    "licitacoes": "Licitações",
    "editais": "Busca de editais",
    "fontes": "Fontes de busca",
    "concorrentes": "Concorrentes",
    # Operação técnica
    "equipamentos": "Equipamentos",
    "chamados": "Chamados",
    "agendamentos": "Agendamentos",
    "ordens_servico": "Ordens de Serviço",
    "servicos": "Serviços técnicos",
    "calibracoes": "Calibrações",
    "certificados": "Certificados",
    "padroes": "Padrões metrológicos",
    "planilhas_calibracao": "Planilhas de calibração",
    "laudos_tecnicos": "Laudos técnicos",
    "estudos_tecnicos": "Estudos técnicos",
    # Qualidade e pessoas
    "qualidade": "Qualidade",
    "normas_tecnicas": "Normas técnicas",
    "documentos_qualidade": "Documentos controlados",
    "reclamacoes": "Reclamações",
    "nao_conformidades": "Trabalhos não conformes",
    "colaboradores": "Colaboradores",
    "treinamentos": "Treinamentos e competências",
    # Ativos e frota
    "frota": "Frota",
    "manutencao_frota": "Controle veicular",
    # Vendas, estoque, financeiro e fiscal
    "produtos": "Produtos",
    "catalogo_servicos": "Catálogo de serviços e ensaios",
    "instrumentos_seccol": "Instrumentos técnicos SECCOL",
    "estoque": "Estoque e lotes",
    "vendas": "Vendas",
    "fiscal": "Fiscal",
    "contas_pagar": "Contas a pagar",
    "contas_receber": "Contas a receber",
    "boletos": "Boletos e remessas",
    "financeiro": "Financeiro",
    "caixa": "Caixa",
    "controladoria": "Controladoria",
    # Gestão
    "produtividade": "Produtividade",
    "metas": "Metas",
}

ROLE_MODULES = {
    "admin": set(MODULES),
    "manager": set(MODULES),
    "operator": set(MODULES) - {
        "documentos_qualidade", "fiscal", "normas_tecnicas",
        "certificados", "laudos_tecnicos", "estudos_tecnicos",
        "controladoria",
    },
    "viewer": set(),
    "technician": {"equipamentos", "chamados", "agendamentos", "ordens_servico", "servicos",
                   "calibracoes", "certificados", "padroes", "laudos_tecnicos", "estudos_tecnicos",
                   "catalogo_servicos", "instrumentos_seccol", "frota", "manutencao_frota"},
    "quality": {"arquivos", "equipamentos", "calibracoes", "certificados", "padroes",
                "planilhas_calibracao", "laudos_tecnicos", "estudos_tecnicos", "normas_tecnicas",
                "qualidade", "documentos_qualidade", "reclamacoes",
                "nao_conformidades", "colaboradores", "treinamentos", "catalogo_servicos",
                "instrumentos_seccol"},
    "fiscal": {"clientes", "fornecedores", "contatos", "importacoes_xml", "pedidos_compra",
               "produtos", "catalogo_servicos", "estoque", "vendas", "fiscal", "contas_pagar", "contas_receber",
               "boletos", "financeiro", "caixa"},
    "approver": {"solicitacoes_compra", "pedidos_compra", "propostas", "contratos", "financeiro"},
}

# Leitura e exportação são permissões independentes da escrita. A coluna
# company_memberships.permissions pode ampliar ou restringir estes conjuntos
# com as chaves read/write/export/deny_read/deny_write/deny_export.
ROLE_READ_MODULES = {
    "admin": set(MODULES),
    "manager": set(MODULES),
    "operator": set(ROLE_MODULES["operator"]),
    "viewer": set(MODULES),
    "technician": set(ROLE_MODULES["technician"]) | {
        "clientes", "contatos", "normas_tecnicas", "documentos_qualidade",
    },
    "quality": set(ROLE_MODULES["quality"]) | {
        "clientes", "fornecedores", "chamados", "agendamentos", "ordens_servico", "servicos",
    },
    "fiscal": set(ROLE_MODULES["fiscal"]) | {
        "contratos", "solicitacoes_compra", "controladoria",
    },
    "approver": set(ROLE_MODULES["approver"]) | {
        "clientes", "fornecedores", "licitacoes", "contas_pagar", "contas_receber", "vendas",
    },
}

ROLE_EXPORT_MODULES = {
    "admin": set(MODULES),
    "manager": set(MODULES),
    "operator": set(),
    "viewer": set(),
    "technician": set(),
    "quality": set(),
    "fiscal": set(),
    "approver": set(),
}

PARTY_MODULE = "clientes_fornecedores"
PARTY_PHYSICAL_MODULES = ("clientes", "fornecedores")

# Estoque usa um ledger dedicado. Quantidades são persistidas como micros para
# evitar deriva de ponto flutuante; a API converte de/para unidades decimais.
INVENTORY_QUANTITY_SCALE = 1_000_000
INVENTORY_MOVEMENT_TYPES = {
    "PURCHASE_IN", "SALE_OUT", "SERVICE_ORDER_OUT", "RESERVE",
    "RELEASE_RESERVATION", "TRANSFER_IN", "TRANSFER_OUT", "RETURN_IN",
    "RETURN_OUT", "ADJUSTMENT_IN", "ADJUSTMENT_OUT",
}

# Códigos oficiais de UF usados no protocolo nacional da NF-e. Endpoints e
# schemas permanecem versionados/configuráveis: não são regras tributárias e
# nunca determinam alíquota, CST, CFOP ou qualquer cálculo fiscal.
UF_CODES = {
    "RO": "11", "AC": "12", "AM": "13", "RR": "14", "PA": "15", "AP": "16",
    "TO": "17", "MA": "21", "PI": "22", "CE": "23", "RN": "24", "PB": "25",
    "PE": "26", "AL": "27", "SE": "28", "BA": "29", "MG": "31", "ES": "32",
    "RJ": "33", "SP": "35", "PR": "41", "SC": "42", "RS": "43", "MS": "50",
    "MT": "51", "GO": "52", "DF": "53",
}
SEFAZ_OFFICIAL_REFERENCE = "https://www.nfe.fazenda.gov.br/portal/webservices.aspx"
SEFAZ_SCHEMA_REFERENCE = (
    "https://www.nfe.fazenda.gov.br/portal/listaConteudo.aspx?"
    "tipoConteudo=BMPFMBoln3w="
)
SEFAZ_GO_ENDPOINTS = {
    "HOMOLOGATION": {
        "status": "https://homolog.sefaz.go.gov.br/nfe/services/NFeStatusServico4",
        "authorization": "https://homolog.sefaz.go.gov.br/nfe/services/NFeAutorizacao4",
        "authorization_return": "https://homolog.sefaz.go.gov.br/nfe/services/NFeRetAutorizacao4",
        "protocol": "https://homolog.sefaz.go.gov.br/nfe/services/NFeConsultaProtocolo4",
        "events": "https://homolog.sefaz.go.gov.br/nfe/services/NFeRecepcaoEvento4",
        "invalidation": "https://homolog.sefaz.go.gov.br/nfe/services/NFeInutilizacao4",
    },
    "PRODUCTION": {
        "status": "https://nfe.sefaz.go.gov.br/nfe/services/NFeStatusServico4",
        "authorization": "https://nfe.sefaz.go.gov.br/nfe/services/NFeAutorizacao4",
        "authorization_return": "https://nfe.sefaz.go.gov.br/nfe/services/NFeRetAutorizacao4",
        "protocol": "https://nfe.sefaz.go.gov.br/nfe/services/NFeConsultaProtocolo4",
        "events": "https://nfe.sefaz.go.gov.br/nfe/services/NFeRecepcaoEvento4",
        "invalidation": "https://nfe.sefaz.go.gov.br/nfe/services/NFeInutilizacao4",
    },
}

# Catálogo de autorização funcional. As permissões de leitura, escrita e
# exportação continuam sendo a primeira barreira; estas ações refinam o que a
# pessoa pode executar dentro de cada módulo. Associações antigas sem a chave
# ``actions`` preservam o comportamento anterior derivado do perfil-base.
VALUE_SENSITIVE_MODULES = {
    "importacoes_xml", "solicitacoes_compra", "pedidos_compra", "crm",
    "propostas", "contratos", "licitacoes", "editais", "chamados", "ordens_servico",
    "servicos", "calibracoes", "reclamacoes", "nao_conformidades", "frota",
    "manutencao_frota", "produtos", "catalogo_servicos", "estoque", "vendas",
    "fiscal", "contas_pagar", "contas_receber", "boletos", "financeiro", "caixa",
}
SENSITIVE_PAYLOAD_FIELDS = {
    "produtos": {"preco_venda"},
}
READ_ONLY_MODULES = {"controladoria"}
VALUE_DEPENDENT_ACTIONS = {
    "create", "update", "manage_items", "bill_sales", "settle_financial",
    "receive_stock", "register_fiscal", "convert_tender", "export_accounting",
}

MODULE_ACTION_LABELS = {
    "create": "Criar cadastros",
    "update": "Editar cadastros",
    "delete": "Enviar para a lixeira",
    "restore": "Restaurar itens da lixeira",
    "view_values": "Visualizar preços e valores",
    "transition": "Alterar etapa do fluxo",
    "manage_items": "Incluir e alterar itens",
    "manage_attachments": "Anexar evidências e documentos",
    "request_approval": "Solicitar aprovação",
    "decide_approval": "Decidir aprovações",
    "partner_control": "Aprovar ou bloquear parceiro",
    "import_xml": "Importar XML fiscal",
    "manage_warehouses": "Criar depósitos",
    "move_stock": "Registrar movimentos",
    "adjust_stock": "Ajustar estoque",
    "transfer_stock": "Transferir entre depósitos",
    "reserve_stock": "Reservar estoque",
    "release_stock": "Liberar reservas",
    "fulfill_stock": "Baixar estoque",
    "receive_stock": "Receber compras no estoque",
    "bill_sales": "Marcar venda como faturada",
    "settle_financial": "Baixar pagamento ou recebimento",
    "cancel_financial": "Cancelar título financeiro",
    "register_fiscal": "Registrar documento fiscal local",
    "manage_fiscal_config": "Configurar integração fiscal",
    "manage_fiscal_certificate": "Gerenciar certificado digital A1",
    "check_sefaz_status": "Consultar disponibilidade da SEFAZ",
    "export_accounting": "Gerar pacote para a contabilidade",
    "issue_report": "Emitir documento técnico",
    "search_tenders": "Executar pesquisa de editais",
    "manage_tender_schedules": "Gerenciar planos de pesquisa",
    "triage_tenders": "Analisar e classificar editais",
    "convert_tender": "Converter edital em licitação",
    "view_billing": "Consultar faturamento",
    "view_cashflow": "Consultar fluxo de caixa",
    "view_inventory_value": "Consultar valor do estoque",
    "view_overdue": "Consultar títulos vencidos",
}


def build_module_actions():
    actions = {}
    for module in MODULES:
        current = [] if module in READ_ONLY_MODULES else [
            "create", "update", "delete", "restore", "manage_attachments",
        ]
        if module in VALUE_SENSITIVE_MODULES:
            current.append("view_values")
        if module in MODULE_STATUS_TRANSITIONS:
            current.append("transition")
        actions[module] = current
    for module in ITEM_DOCUMENT_MODULES:
        actions[module].append("manage_items")
    return actions
INVENTORY_IN_TYPES = {"PURCHASE_IN", "TRANSFER_IN", "RETURN_IN", "ADJUSTMENT_IN"}
INVENTORY_OUT_TYPES = {
    "SALE_OUT", "SERVICE_ORDER_OUT", "TRANSFER_OUT", "RETURN_OUT", "ADJUSTMENT_OUT",
}
ITEM_DOCUMENT_MODULES = {
    "propostas", "vendas", "solicitacoes_compra", "pedidos_compra", "ordens_servico",
}
RESERVABLE_ITEM_MODULES = {"vendas", "ordens_servico"}
BUSINESS_UNIQUE_FIELDS = {
    "produtos": "codigo", "catalogo_servicos": "codigo",
    "instrumentos_seccol": "codigo", "solicitacoes_compra": "numero",
    "pedidos_compra": "numero", "propostas": "numero", "vendas": "documento",
    "ordens_servico": "numero", "certificados": "numero",
    "laudos_tecnicos": "numero", "estudos_tecnicos": "numero",
    "documentos_qualidade": "codigo",
}

# Campos de cadastros operacionais que apontam para registros mestres. O nome
# continua no payload para relatórios legados, mas ``<campo>_id`` e
# record_relationships são a fonte relacional controlada pelo servidor.
RECORD_REFERENCE_RULES = {
    "cliente": {"modules": PARTY_PHYSICAL_MODULES, "party_role": "C", "relation": "Cliente"},
    "fornecedor": {"modules": PARTY_PHYSICAL_MODULES, "party_role": "F", "relation": "Fornecedor"},
    "cliente_fornecedor": {"modules": PARTY_PHYSICAL_MODULES, "party_role": "A", "relation": "Parceiro"},
    "destinatario": {"modules": PARTY_PHYSICAL_MODULES, "party_role": "A", "relation": "Destinatário"},
    "parceiro": {"modules": PARTY_PHYSICAL_MODULES, "party_role": "A", "relation": "Parceiro"},
    "equipamento": {"modules": ("equipamentos",), "relation": "Equipamento"},
    "os": {"modules": ("ordens_servico",), "relation": "Ordem de Serviço"},
    "solicitacao": {"modules": ("solicitacoes_compra",), "relation": "Solicitação de origem"},
    "produto": {"modules": ("produtos",), "relation": "Produto"},
    "colaborador": {"modules": ("colaboradores",), "relation": "Colaborador"},
    "certificado": {"modules": ("certificados",), "relation": "Certificado"},
    "norma": {"modules": ("normas_tecnicas",), "relation": "Norma técnica"},
    "placa": {"modules": ("frota",), "relation": "Veículo"},
}

DEFAULT_STATUSES = {
    "Ativo", "Em andamento", "Pendente", "A revisar", "Aprovado", "Pago",
    "Concluído", "Cancelado",
}

MODULE_STATUSES = {
    "crm": {"Novo lead", "Contato realizado", "Qualificado", "Proposta", "Negociação", "Ganho", "Perdido"},
    "propostas": {"Rascunho", "Enviada", "Em negociação", "Aprovada", "Recusada"},
    "licitacoes": {"Captação", "Análise", "Documentação", "Proposta enviada", "Disputa", "Habilitação", "Homologada", "Perdida"},
    "chamados": {"Aberto", "Em atendimento", "Aguardando cliente", "Concluído", "Cancelado"},
    "ordens_servico": {"Aberta", "Agendada", "Em execução", "Pausada", "Aguardando aprovação", "Concluída", "Cancelada"},
    "solicitacoes_compra": {"Rascunho", "Pendente de aprovação", "Aprovada", "Rejeitada", "Convertida em pedido"},
    "pedidos_compra": {"Rascunho", "Emitido", "Aguardando fornecedor", "Recebido parcial", "Recebido", "Cancelado"},
    "vendas": {"Rascunho", "Confirmado", "Separação", "Faturado", "Concluído", "Cancelado"},
    "contas_pagar": {"Em aberto", "Parcial", "Pago", "Vencido", "Cancelado"},
    "contas_receber": {"Em aberto", "Parcial", "Recebido", "Vencido", "Cancelado"},
    "certificados": {"Rascunho", "Em revisão", "Aguardando aprovação", "Aprovado", "Publicado", "Obsoleto"},
    "laudos_tecnicos": {"Rascunho", "Em revisão", "Aguardando aprovação", "Aprovado", "Emitido", "Obsoleto"},
    "estudos_tecnicos": {"Rascunho", "Em revisão", "Aguardando aprovação", "Aprovado", "Emitido", "Obsoleto"},
    "normas_tecnicas": {"Publicada", "Publicada — em revisão sistemática", "Publicada — revisão em desenvolvimento", "Vigente", "Obsoleta"},
    "documentos_qualidade": {"Rascunho", "Em revisão", "Aguardando aprovação", "Vigente", "Obsoleto"},
    "fiscal": {"Rascunho", "Registrado localmente", "Aguardando processamento fiscal", "Autorizado", "Rejeitado", "Cancelado"},
    "importacoes_xml": {"Importada", "Validada", "Rejeitada"},
}

# Estado inicial determinístico de cada fluxo especializado. Conjuntos são
# apropriados para validar, mas não podem definir o primeiro estado porque sua
# ordem não é um contrato estável entre execuções do Python.
MODULE_INITIAL_STATUSES = {
    "crm": "Novo lead",
    "propostas": "Rascunho",
    "licitacoes": "Captação",
    "chamados": "Aberto",
    "ordens_servico": "Aberta",
    "solicitacoes_compra": "Rascunho",
    "pedidos_compra": "Rascunho",
    "vendas": "Rascunho",
    "contas_pagar": "Em aberto",
    "contas_receber": "Em aberto",
    "certificados": "Rascunho",
    "laudos_tecnicos": "Rascunho",
    "estudos_tecnicos": "Rascunho",
    "normas_tecnicas": "Publicada",
    "documentos_qualidade": "Rascunho",
    "fiscal": "Rascunho",
    "importacoes_xml": "Importada",
}

MODULE_STATUS_TRANSITIONS = {
    "propostas": {
        "Rascunho": {"Enviada", "Recusada"},
        "Enviada": {"Rascunho", "Em negociação", "Aprovada", "Recusada"},
        "Em negociação": {"Enviada", "Aprovada", "Recusada"},
        "Aprovada": set(),
        "Recusada": {"Rascunho"},
    },
    "solicitacoes_compra": {
        "Rascunho": {"Pendente de aprovação"},
        "Pendente de aprovação": {"Aprovada", "Rejeitada"},
        "Aprovada": {"Convertida em pedido"},
        "Rejeitada": {"Rascunho"},
        "Convertida em pedido": set(),
    },
    "pedidos_compra": {
        "Rascunho": {"Emitido", "Cancelado"},
        "Emitido": {"Aguardando fornecedor", "Recebido parcial", "Recebido", "Cancelado"},
        "Aguardando fornecedor": {"Recebido parcial", "Recebido", "Cancelado"},
        "Recebido parcial": {"Recebido"},
        "Recebido": set(),
        "Cancelado": set(),
    },
    "vendas": {
        "Rascunho": {"Confirmado", "Cancelado"},
        "Confirmado": {"Separação", "Cancelado"},
        "Separação": {"Faturado", "Cancelado"},
        "Faturado": {"Concluído"},
        "Concluído": set(),
        "Cancelado": set(),
    },
    "ordens_servico": {
        "Aberta": {"Agendada", "Em execução", "Cancelada"},
        "Agendada": {"Em execução", "Cancelada"},
        "Em execução": {"Pausada", "Aguardando aprovação", "Concluída"},
        "Pausada": {"Em execução", "Cancelada"},
        "Aguardando aprovação": {"Em execução", "Concluída"},
        "Concluída": set(),
        "Cancelada": set(),
    },
    "contas_pagar": {
        "Em aberto": {"Parcial", "Pago", "Vencido", "Cancelado"},
        "Parcial": {"Pago", "Vencido", "Cancelado"},
        "Vencido": {"Parcial", "Pago", "Cancelado"},
        "Pago": set(),
        "Cancelado": set(),
    },
    "contas_receber": {
        "Em aberto": {"Parcial", "Recebido", "Vencido", "Cancelado"},
        "Parcial": {"Recebido", "Vencido", "Cancelado"},
        "Vencido": {"Parcial", "Recebido", "Cancelado"},
        "Recebido": set(),
        "Cancelado": set(),
    },
}

MODULE_ACTIONS = build_module_actions()
for module in set(MODULES) - READ_ONLY_MODULES:
    MODULE_ACTIONS[module].extend(["request_approval", "decide_approval"])
for module in {"clientes", "fornecedores"}:
    MODULE_ACTIONS[module].append("partner_control")
MODULE_ACTIONS["importacoes_xml"].append("import_xml")
MODULE_ACTIONS["estoque"].extend([
    "manage_warehouses", "move_stock", "adjust_stock", "transfer_stock",
    "reserve_stock", "release_stock",
])
for module in {"vendas", "ordens_servico"}:
    MODULE_ACTIONS[module].extend(["reserve_stock", "release_stock", "fulfill_stock"])
MODULE_ACTIONS["vendas"].append("bill_sales")
MODULE_ACTIONS["pedidos_compra"].append("receive_stock")
for module in {"contas_pagar", "contas_receber"}:
    MODULE_ACTIONS[module].extend(["settle_financial", "cancel_financial"])
MODULE_ACTIONS["fiscal"].extend([
    "register_fiscal", "manage_fiscal_config", "manage_fiscal_certificate",
    "check_sefaz_status", "export_accounting",
])
MODULE_ACTIONS["editais"].extend([
    "search_tenders", "manage_tender_schedules", "triage_tenders", "convert_tender",
])
for module in {"certificados", "laudos_tecnicos", "estudos_tecnicos"}:
    MODULE_ACTIONS[module].append("issue_report")
MODULE_ACTIONS["controladoria"] = [
    "view_billing", "view_cashflow", "view_inventory_value", "view_overdue",
]
MODULE_ACTIONS = {
    module: tuple(dict.fromkeys(actions)) for module, actions in MODULE_ACTIONS.items()
}

ACCESS_CATEGORIES = (
    ("administrativo", "Administrativo", (
        "arquivos", "clientes", "fornecedores", "contatos", "importacoes_xml", "ramais",
    )),
    ("compras", "Compras", ("solicitacoes_compra", "pedidos_compra")),
    ("comercial", "Comercial e vendas", (
        "crm", "propostas", "contratos", "vendas", "licitacoes", "editais", "fontes",
        "concorrentes",
    )),
    ("servico", "Serviço técnico", (
        "equipamentos", "chamados", "agendamentos", "ordens_servico", "servicos",
        "calibracoes", "certificados", "padroes", "planilhas_calibracao",
        "laudos_tecnicos", "estudos_tecnicos",
    )),
    ("qualidade", "Qualidade e pessoas", (
        "qualidade", "normas_tecnicas", "documentos_qualidade", "reclamacoes",
        "nao_conformidades", "colaboradores", "treinamentos",
    )),
    ("ativos", "Ativos e catálogo", (
        "frota", "manutencao_frota", "produtos", "catalogo_servicos",
        "instrumentos_seccol", "estoque",
    )),
    ("financeiro", "Financeiro, fiscal e controladoria", (
        "fiscal", "contas_pagar", "contas_receber", "boletos", "financeiro", "caixa",
        "controladoria",
    )),
    ("gestao", "Gestão", ("produtividade", "metas")),
)

# Contrato de obrigatoriedade espelhado dos 46 formulários especializados.
REQUIRED_PAYLOAD_FIELDS = {
    "arquivos": ("identificador", "categoria"),
    "clientes": ("tipo_pessoa", "documento", "razao_social"),
    "fornecedores": ("tipo_pessoa", "documento", "razao_social", "avaliacao"),
    "contatos": ("cliente_fornecedor", "tipo_contato", "cargo"),
    "importacoes_xml": ("chave", "numero", "fornecedor", "data_emissao"),
    "solicitacoes_compra": ("numero", "solicitante", "prioridade", "justificativa"),
    "pedidos_compra": ("numero", "fornecedor", "condicao_pagamento", "centro_custo"),
    "ramais": ("nome_ramal", "ramal", "setor"),
    "crm": ("etapa", "origem", "proximo_passo", "probabilidade"),
    "propostas": ("numero", "cliente", "validade", "etapa", "local_execucao"),
    "contratos": ("numero", "cliente", "gestor", "inicio", "fim"),
    "licitacoes": ("orgao", "edital", "portal", "modalidade", "data_abertura", "etapa"),
    "concorrentes": ("cnpj", "especialidade", "regiao", "fonte"),
    "instrumentos_seccol": ("codigo", "tipo", "fabricante", "modelo", "numero_serie", "proxima_calibracao"),
    "equipamentos": ("cliente", "tipo", "fabricante", "modelo", "numero_serie", "localizacao"),
    "chamados": ("cliente", "solicitante", "tipo", "prioridade"),
    "agendamentos": ("cliente", "tecnico", "data", "hora", "local", "tipo_servico"),
    "ordens_servico": ("numero", "cliente", "tecnico", "tipo_os", "local_execucao"),
    "servicos": ("cliente", "equipamento", "tecnico", "tipo_servico"),
    "calibracoes": ("os", "equipamento", "tecnico", "data_calibracao", "regra_decisao"),
    "certificados": ("numero", "os", "equipamento", "data_emissao", "revisao", "aprovador"),
    "laudos_tecnicos": ("numero", "os", "cliente", "local_avaliado", "responsavel_tecnico", "data_emissao", "metodo", "regra_decisao", "conclusao"),
    "estudos_tecnicos": ("numero", "cliente", "objeto", "responsavel_tecnico", "data_emissao", "premissas", "metodologia", "recomendacoes"),
    "padroes": ("codigo", "tipo", "fabricante", "numero_serie", "faixa_medicao", "proxima_calibracao", "rastreabilidade"),
    "planilhas_calibracao": ("codigo", "grandeza", "versao", "criterio_aceitacao"),
    "normas_tecnicas": ("codigo", "organismo", "edicao", "escopo_resumido", "aplicabilidade_seccol", "referencia_oficial", "licenciamento", "verificado_em"),
    "qualidade": ("tipo", "norma", "responsavel_qualidade", "acao_corretiva"),
    "documentos_qualidade": ("codigo", "tipo", "revisao", "elaborador", "aprovador", "data_vigencia"),
    "reclamacoes": ("cliente", "canal", "procedente", "causa", "tratativa"),
    "nao_conformidades": ("origem", "requisito", "causa_raiz", "correcao", "acao_corretiva"),
    "colaboradores": ("cpf", "cargo", "setor", "email"),
    "treinamentos": ("colaborador", "competencia", "data", "validade", "resultado"),
    "frota": ("placa", "veiculo", "renavam", "chassi", "responsavel_veiculo"),
    "manutencao_frota": ("placa", "tipo", "quilometragem", "oficina", "data_servico"),
    "produtos": ("codigo", "familia", "tipo_item", "descricao", "ncm", "unidade", "preco_venda"),
    "catalogo_servicos": ("codigo", "categoria", "tipo_servico", "descricao", "fonte_oficial", "verificado_em"),
    "estoque": ("produto", "lote", "quantidade", "localizacao", "movimento"),
    "vendas": ("cliente", "documento", "vendedor", "forma_pagamento", "condicao_pagamento"),
    "fiscal": ("tipo_nota", "numero", "serie", "chave", "destinatario", "cfop", "finalidade"),
    "contas_pagar": ("fornecedor", "documento", "parcela", "categoria", "centro_custo"),
    "contas_receber": ("cliente", "documento", "parcela", "categoria", "centro_custo"),
    "boletos": ("cliente", "nosso_numero", "banco", "conta", "vencimento_original"),
    "financeiro": ("tipo_lancamento", "categoria", "documento", "conta", "centro_custo"),
    "caixa": ("tipo_movimento", "categoria", "conta", "operador", "forma_pagamento"),
    "produtividade": ("colaborador", "periodo", "indicador", "resultado"),
    "metas": ("responsavel_meta", "indicador", "periodo", "meta"),
}

DATE_FIELDS = {
    "importacoes_xml": {"data_emissao"}, "propostas": {"validade"},
    "contratos": {"inicio", "fim"}, "licitacoes": {"data_abertura"},
    "instrumentos_seccol": {"proxima_calibracao"}, "equipamentos": {"proxima_calibracao"},
    "agendamentos": {"data"}, "calibracoes": {"data_calibracao", "proxima_calibracao"},
    "certificados": {"data_emissao"}, "laudos_tecnicos": {"data_emissao"},
    "estudos_tecnicos": {"data_emissao"}, "padroes": {"proxima_calibracao"},
    "normas_tecnicas": {"verificado_em"}, "documentos_qualidade": {"data_vigencia"},
    "treinamentos": {"data", "validade"}, "frota": {"seguro_vencimento"},
    "manutencao_frota": {"data_servico"}, "catalogo_servicos": {"verificado_em"},
    "estoque": {"validade"}, "boletos": {"vencimento_original"},
}

DATETIME_FIELDS = {"ordens_servico": {"inicio", "fim"}}
TIME_FIELDS = {"agendamentos": {"hora"}}
NUMBER_FIELDS = {
    "crm": {"probabilidade"}, "ordens_servico": {"tempo_minutos"},
    "frota": {"quilometragem"}, "manutencao_frota": {"quilometragem", "proxima_km"},
    "produtos": {"preco_venda"}, "estoque": {"quantidade"},
    "produtividade": {"resultado", "horas"}, "metas": {"meta", "realizado"},
}

EMAIL_FIELDS = {"clientes": {"email"}, "fornecedores": {"email"}, "contatos": {"email"}, "colaboradores": {"email"}}
URL_FIELDS = {
    "licitacoes": {"portal"}, "concorrentes": {"fonte"},
    "instrumentos_seccol": {"fonte_oficial"}, "produtos": {"fonte_oficial"},
    "catalogo_servicos": {"fonte_oficial"}, "normas_tecnicas": {"referencia_oficial"},
}


def _blank(value) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _valid_cpf(value: str) -> bool:
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) != 11 or len(set(digits)) == 1:
        return False
    for size in (9, 10):
        total = sum(int(digits[index]) * (size + 1 - index) for index in range(size))
        check = (total * 10 % 11) % 10
        if check != int(digits[size]):
            return False
    return True


def _valid_cnpj(value: str) -> bool:
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) != 14 or len(set(digits)) == 1:
        return False
    for weights, position in (((5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2), 12),
                              ((6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2), 13)):
        remainder = sum(int(digits[index]) * weight for index, weight in enumerate(weights)) % 11
        check = 0 if remainder < 2 else 11 - remainder
        if check != int(digits[position]):
            return False
    return True


def _validate_document(value: str, label="CPF/CNPJ") -> None:
    digits = re.sub(r"\D", "", str(value or ""))
    valid = _valid_cpf(digits) if len(digits) == 11 else _valid_cnpj(digits) if len(digits) == 14 else False
    if not valid:
        raise ValueError(f"{label} inválido")


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def json_dumps(value) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"), allow_nan=False)


def _reject_json_constant(value):
    raise ValueError(f"Constante numérica JSON inválida: {value}")


def json_loads_strict(value):
    return json.loads(value, parse_constant=_reject_json_constant)


class Database:
    def __init__(self, path: Path):
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.local = threading.local()
        self.initialize()
        self.secure_files()

    def secure_files(self) -> None:
        """Aplica menor privilégio aos arquivos SQLite em plataformas POSIX."""
        for candidate in (self.path, Path(str(self.path) + "-wal"), Path(str(self.path) + "-shm")):
            if candidate.exists():
                try:
                    os.chmod(candidate, 0o600)
                except OSError:
                    pass

    def connection(self) -> sqlite3.Connection:
        connection = getattr(self.local, "connection", None)
        if connection is None:
            connection = sqlite3.connect(self.path, timeout=20)
            connection.row_factory = sqlite3.Row
            connection.execute("PRAGMA foreign_keys = ON")
            connection.execute("PRAGMA journal_mode = WAL")
            connection.execute("PRAGMA busy_timeout = 20000")
            connection.execute("PRAGMA synchronous = NORMAL")
            self.local.connection = connection
            self.secure_files()
        return connection

    def close_thread_connection(self) -> None:
        """Fecha a conexao pertencente a thread atual."""
        connection = getattr(self.local, "connection", None)
        if connection is None:
            return
        try:
            if connection.in_transaction:
                connection.rollback()
            connection.close()
        finally:
            del self.local.connection

    @contextlib.contextmanager
    def transaction(self, immediate=False):
        """Unidade de trabalho com suporte seguro a chamadas aninhadas."""
        db = self.connection()
        depth = int(getattr(self.local, "transaction_depth", 0))
        savepoint = f"sivs_sp_{depth}"
        if depth == 0:
            if db.in_transaction:
                db.commit()
            db.execute("BEGIN IMMEDIATE" if immediate else "BEGIN")
        else:
            db.execute(f"SAVEPOINT {savepoint}")
        self.local.transaction_depth = depth + 1
        try:
            yield db
            self.local.transaction_depth = depth
            if depth == 0:
                db.commit()
            else:
                db.execute(f"RELEASE SAVEPOINT {savepoint}")
        except Exception:
            self.local.transaction_depth = depth
            if depth == 0:
                db.rollback()
            else:
                db.execute(f"ROLLBACK TO SAVEPOINT {savepoint}")
                db.execute(f"RELEASE SAVEPOINT {savepoint}")
            raise

    def commit_if_outer(self) -> None:
        if (int(getattr(self.local, "transaction_depth", 0)) == 0 and
                not bool(getattr(self.local, "manual_transaction", False))):
            self.connection().commit()

    def begin_manual_transaction(self, immediate=True) -> None:
        """Inicia unidade longa usada por importadores sem permitir commits intermediários."""
        if int(getattr(self.local, "transaction_depth", 0)) or getattr(
                self.local, "manual_transaction", False):
            raise RuntimeError("Já existe uma transação ativa")
        db = self.connection()
        if db.in_transaction:
            db.commit()
        db.execute("BEGIN IMMEDIATE" if immediate else "BEGIN")
        self.local.manual_transaction = True

    def finish_manual_transaction(self, commit=True) -> None:
        if not getattr(self.local, "manual_transaction", False):
            return
        db = self.connection()
        try:
            db.commit() if commit else db.rollback()
        finally:
            self.local.manual_transaction = False

    def abort_manual_transaction(self) -> None:
        self.finish_manual_transaction(commit=False)

    def initialize(self) -> None:
        db = self.connection()
        db.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                email TEXT NOT NULL UNIQUE COLLATE NOCASE,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'admin',
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS sessions (
                token_hash TEXT PRIMARY KEY,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                csrf_token TEXT NOT NULL,
                expires_at INTEGER NOT NULL,
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                module TEXT NOT NULL,
                title TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'Ativo',
                amount REAL,
                due_date TEXT,
                payload TEXT NOT NULL DEFAULT '{}',
                created_by INTEGER REFERENCES users(id),
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_records_module ON records(module);
            CREATE INDEX IF NOT EXISTS idx_records_status ON records(status);
            CREATE INDEX IF NOT EXISTS idx_records_due_date ON records(due_date);
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER REFERENCES users(id),
                action TEXT NOT NULL,
                entity_type TEXT NOT NULL,
                entity_id TEXT,
                detail TEXT,
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS record_versions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                record_id INTEGER NOT NULL,
                snapshot TEXT NOT NULL,
                changed_by INTEGER REFERENCES users(id),
                created_at TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_record_versions_record ON record_versions(record_id);
            CREATE TABLE IF NOT EXISTS tender_searches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                keywords TEXT NOT NULL,
                uf TEXT,
                days INTEGER NOT NULL,
                sources_searched TEXT NOT NULL,
                found_count INTEGER NOT NULL DEFAULT 0,
                new_count INTEGER NOT NULL DEFAULT 0,
                error_detail TEXT,
                created_by INTEGER REFERENCES users(id),
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS tender_results (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_key TEXT NOT NULL,
                external_id TEXT NOT NULL,
                title TEXT NOT NULL,
                object_text TEXT NOT NULL,
                agency TEXT,
                uf TEXT,
                municipality TEXT,
                modality TEXT,
                estimated_value REAL,
                published_at TEXT,
                deadline TEXT,
                source_url TEXT,
                matched_terms TEXT NOT NULL DEFAULT '[]',
                relevance_score INTEGER NOT NULL DEFAULT 0,
                status TEXT NOT NULL DEFAULT 'Novo',
                raw_json TEXT NOT NULL DEFAULT '{}',
                converted_record_id INTEGER REFERENCES records(id),
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(source_key, external_id)
            );
            CREATE INDEX IF NOT EXISTS idx_tender_results_status ON tender_results(status);
            CREATE INDEX IF NOT EXISTS idx_tender_results_deadline ON tender_results(deadline);
            CREATE TABLE IF NOT EXISTS tender_details (
                tender_result_id INTEGER PRIMARY KEY REFERENCES tender_results(id) ON DELETE CASCADE,
                company_id INTEGER NOT NULL REFERENCES companies(id),
                official_data TEXT NOT NULL DEFAULT '{}',
                items_json TEXT NOT NULL DEFAULT '[]',
                documents_json TEXT NOT NULL DEFAULT '[]',
                value_source TEXT NOT NULL DEFAULT 'unavailable',
                analysis_json TEXT NOT NULL DEFAULT '{}',
                refreshed_at TEXT NOT NULL,
                refresh_error TEXT
            );
            CREATE INDEX IF NOT EXISTS idx_tender_details_company ON tender_details(company_id,refreshed_at);
            CREATE TABLE IF NOT EXISTS subjects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                normalized_name TEXT NOT NULL UNIQUE,
                status TEXT NOT NULL DEFAULT 'Ativo',
                created_by INTEGER REFERENCES users(id),
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS record_relationships (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                from_record_id INTEGER NOT NULL REFERENCES records(id) ON DELETE CASCADE,
                to_record_id INTEGER NOT NULL REFERENCES records(id) ON DELETE CASCADE,
                relationship_type TEXT NOT NULL,
                created_by INTEGER REFERENCES users(id),
                created_at TEXT NOT NULL,
                CHECK(from_record_id != to_record_id),
                UNIQUE(from_record_id,to_record_id,relationship_type)
            );
            CREATE INDEX IF NOT EXISTS idx_relationships_from ON record_relationships(from_record_id);
            CREATE INDEX IF NOT EXISTS idx_relationships_to ON record_relationships(to_record_id);
            """
        )
        db.executescript(
            """
            CREATE TABLE IF NOT EXISTS companies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                cnpj TEXT,
                phone TEXT,
                email TEXT,
                address TEXT,
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS company_memberships (
                company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                role TEXT NOT NULL DEFAULT 'operator',
                permissions TEXT NOT NULL DEFAULT '{}',
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                PRIMARY KEY(company_id,user_id)
            );
            CREATE TABLE IF NOT EXISTS company_settings (
                company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
                key TEXT NOT NULL,
                value TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                PRIMARY KEY(company_id,key)
            );
            CREATE TABLE IF NOT EXISTS record_subjects (
                record_id INTEGER NOT NULL REFERENCES records(id) ON DELETE CASCADE,
                subject_id INTEGER NOT NULL REFERENCES subjects(id) ON DELETE CASCADE,
                relationship_type TEXT NOT NULL DEFAULT 'Relacionado a',
                is_primary INTEGER NOT NULL DEFAULT 0,
                created_by INTEGER REFERENCES users(id),
                created_at TEXT NOT NULL,
                PRIMARY KEY(record_id,subject_id,relationship_type)
            );
            CREATE INDEX IF NOT EXISTS idx_record_subjects_subject ON record_subjects(subject_id);
            CREATE TABLE IF NOT EXISTS attachments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
                record_id INTEGER REFERENCES records(id) ON DELETE CASCADE,
                filename TEXT NOT NULL,
                mime_type TEXT NOT NULL DEFAULT 'application/octet-stream',
                content BLOB NOT NULL,
                size INTEGER NOT NULL,
                category TEXT,
                version TEXT,
                uploaded_by INTEGER REFERENCES users(id),
                created_at TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_attachments_record ON attachments(record_id);
            CREATE TABLE IF NOT EXISTS approvals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
                record_id INTEGER NOT NULL REFERENCES records(id) ON DELETE CASCADE,
                approval_type TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'Pendente',
                requested_to INTEGER REFERENCES users(id),
                decided_by INTEGER REFERENCES users(id),
                comment TEXT,
                requested_at TEXT NOT NULL,
                decided_at TEXT
            );
            CREATE INDEX IF NOT EXISTS idx_approvals_record ON approvals(record_id);
            CREATE TABLE IF NOT EXISTS notifications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
                user_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
                title TEXT NOT NULL,
                message TEXT,
                record_id INTEGER REFERENCES records(id) ON DELETE CASCADE,
                level TEXT NOT NULL DEFAULT 'info',
                read_at TEXT,
                created_at TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_notifications_user ON notifications(company_id,user_id,read_at);
            CREATE TABLE IF NOT EXISTS fiscal_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
                record_id INTEGER NOT NULL REFERENCES records(id) ON DELETE CASCADE,
                event_type TEXT NOT NULL,
                status TEXT NOT NULL,
                protocol TEXT,
                response_detail TEXT,
                created_by INTEGER REFERENCES users(id),
                created_at TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_fiscal_events_record ON fiscal_events(record_id);
            CREATE TABLE IF NOT EXISTS search_schedules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
                name TEXT NOT NULL,
                keywords TEXT NOT NULL,
                uf TEXT,
                days INTEGER NOT NULL DEFAULT 7,
                frequency TEXT NOT NULL DEFAULT 'manual',
                active INTEGER NOT NULL DEFAULT 1,
                last_run_at TEXT,
                next_run_at TEXT,
                created_by INTEGER REFERENCES users(id),
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS tender_jobs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
                schedule_id INTEGER REFERENCES search_schedules(id) ON DELETE SET NULL,
                status TEXT NOT NULL DEFAULT 'queued',
                request_json TEXT NOT NULL,
                progress INTEGER NOT NULL DEFAULT 0,
                stage TEXT NOT NULL DEFAULT 'Pesquisa enfileirada',
                result_json TEXT,
                error_detail TEXT,
                created_by INTEGER REFERENCES users(id),
                created_at TEXT NOT NULL,
                started_at TEXT,
                heartbeat_at TEXT,
                finished_at TEXT
            );
            CREATE INDEX IF NOT EXISTS idx_tender_jobs_company
              ON tender_jobs(company_id,created_at DESC);
            CREATE TABLE IF NOT EXISTS schema_migrations (
                version INTEGER PRIMARY KEY,
                name TEXT NOT NULL,
                applied_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS setup_state (
                id INTEGER PRIMARY KEY CHECK(id=1),
                configured INTEGER NOT NULL DEFAULT 0,
                configured_at TEXT
            );
            CREATE TABLE IF NOT EXISTS system_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER REFERENCES companies(id) ON DELETE CASCADE,
                user_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
                severity TEXT NOT NULL,
                category TEXT NOT NULL,
                event_type TEXT NOT NULL,
                message TEXT NOT NULL,
                detail TEXT,
                request_id TEXT,
                path TEXT,
                method TEXT,
                client_ip TEXT,
                user_agent TEXT,
                resolved_at TEXT,
                resolved_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                created_at TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_system_events_company_created
              ON system_events(company_id,created_at DESC);
            CREATE INDEX IF NOT EXISTS idx_system_events_open_severity
              ON system_events(company_id,resolved_at,severity,created_at DESC);
            """
        )

        # Domínios estruturais do ERP. Os cadastros mestres existentes em
        # records continuam canônicos nesta etapa; estoque e fiscal ganham
        # tabelas próprias porque exigem invariantes que um payload genérico
        # não consegue garantir com segurança.
        db.executescript(
            """
            CREATE TABLE IF NOT EXISTS holdings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                active INTEGER NOT NULL DEFAULT 1 CHECK(active IN (0,1)),
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS branches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
                code TEXT NOT NULL,
                name TEXT NOT NULL,
                cnpj TEXT,
                address TEXT,
                active INTEGER NOT NULL DEFAULT 1 CHECK(active IN (0,1)),
                is_headquarters INTEGER NOT NULL DEFAULT 0 CHECK(is_headquarters IN (0,1)),
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(company_id,code)
            );
            CREATE INDEX IF NOT EXISTS idx_branches_company_active
              ON branches(company_id,active,name);

            CREATE TABLE IF NOT EXISTS warehouses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
                branch_id INTEGER NOT NULL REFERENCES branches(id),
                code TEXT NOT NULL,
                name TEXT NOT NULL,
                location TEXT,
                active INTEGER NOT NULL DEFAULT 1 CHECK(active IN (0,1)),
                created_by INTEGER REFERENCES users(id),
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(company_id,code)
            );
            CREATE INDEX IF NOT EXISTS idx_warehouses_company_active
              ON warehouses(company_id,active,name);

            CREATE TABLE IF NOT EXISTS inventory_balances (
                company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
                warehouse_id INTEGER NOT NULL REFERENCES warehouses(id),
                product_record_id INTEGER NOT NULL REFERENCES records(id),
                lot_key TEXT NOT NULL DEFAULT '',
                physical_quantity_micros INTEGER NOT NULL DEFAULT 0
                  CHECK(physical_quantity_micros >= 0),
                reserved_quantity_micros INTEGER NOT NULL DEFAULT 0
                  CHECK(reserved_quantity_micros >= 0),
                inventory_value_cents INTEGER NOT NULL DEFAULT 0
                  CHECK(inventory_value_cents >= 0),
                revision INTEGER NOT NULL DEFAULT 1 CHECK(revision >= 1),
                updated_at TEXT NOT NULL,
                PRIMARY KEY(company_id,warehouse_id,product_record_id,lot_key),
                CHECK(reserved_quantity_micros <= physical_quantity_micros)
            );
            CREATE INDEX IF NOT EXISTS idx_inventory_balances_product
              ON inventory_balances(company_id,product_record_id,warehouse_id);

            CREATE TABLE IF NOT EXISTS inventory_reservations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
                warehouse_id INTEGER NOT NULL REFERENCES warehouses(id),
                product_record_id INTEGER NOT NULL REFERENCES records(id),
                lot_key TEXT NOT NULL DEFAULT '',
                quantity_micros INTEGER NOT NULL CHECK(quantity_micros > 0),
                status TEXT NOT NULL DEFAULT 'ACTIVE'
                  CHECK(status IN ('ACTIVE','RELEASED','FULFILLED')),
                origin_type TEXT NOT NULL,
                origin_id TEXT NOT NULL,
                reference TEXT,
                expires_at TEXT,
                created_by INTEGER REFERENCES users(id),
                released_by INTEGER REFERENCES users(id),
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_inventory_reservations_company_status
              ON inventory_reservations(company_id,status,expires_at);
            CREATE UNIQUE INDEX IF NOT EXISTS idx_inventory_reservations_one_active_origin
              ON inventory_reservations(
                company_id,warehouse_id,product_record_id,lot_key,origin_type,origin_id
              ) WHERE status='ACTIVE';

            CREATE TABLE IF NOT EXISTS inventory_movements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
                warehouse_id INTEGER NOT NULL REFERENCES warehouses(id),
                counterpart_warehouse_id INTEGER REFERENCES warehouses(id),
                product_record_id INTEGER NOT NULL REFERENCES records(id),
                lot_key TEXT NOT NULL DEFAULT '',
                movement_type TEXT NOT NULL CHECK(movement_type IN (
                  'PURCHASE_IN','SALE_OUT','SERVICE_ORDER_OUT','RESERVE',
                  'RELEASE_RESERVATION','TRANSFER_IN','TRANSFER_OUT','RETURN_IN',
                  'RETURN_OUT','ADJUSTMENT_IN','ADJUSTMENT_OUT'
                )),
                quantity_micros INTEGER NOT NULL CHECK(quantity_micros > 0),
                physical_delta_micros INTEGER NOT NULL DEFAULT 0,
                reserved_delta_micros INTEGER NOT NULL DEFAULT 0,
                unit_cost_cents INTEGER,
                value_delta_cents INTEGER NOT NULL DEFAULT 0,
                balance_value_cents INTEGER NOT NULL DEFAULT 0,
                origin_type TEXT NOT NULL,
                origin_id TEXT NOT NULL,
                reference TEXT,
                reason TEXT,
                reservation_id INTEGER REFERENCES inventory_reservations(id),
                created_by INTEGER REFERENCES users(id),
                created_at TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_inventory_movements_history
              ON inventory_movements(company_id,product_record_id,created_at DESC,id DESC);
            CREATE INDEX IF NOT EXISTS idx_inventory_movements_warehouse
              ON inventory_movements(company_id,warehouse_id,created_at DESC,id DESC);

            CREATE TABLE IF NOT EXISTS document_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
                record_id INTEGER NOT NULL REFERENCES records(id) ON DELETE CASCADE,
                item_kind TEXT NOT NULL CHECK(item_kind IN ('PRODUCT','SERVICE')),
                catalog_record_id INTEGER NOT NULL REFERENCES records(id),
                description TEXT NOT NULL,
                quantity_micros INTEGER NOT NULL CHECK(quantity_micros > 0),
                unit_price_cents INTEGER NOT NULL CHECK(unit_price_cents >= 0),
                discount_cents INTEGER NOT NULL DEFAULT 0 CHECK(discount_cents >= 0),
                total_cents INTEGER NOT NULL CHECK(total_cents >= 0),
                warehouse_id INTEGER REFERENCES warehouses(id),
                lot_key TEXT NOT NULL DEFAULT '',
                reservation_id INTEGER REFERENCES inventory_reservations(id),
                notes TEXT,
                sort_order INTEGER NOT NULL DEFAULT 0,
                revision INTEGER NOT NULL DEFAULT 1 CHECK(revision >= 1),
                created_by INTEGER REFERENCES users(id),
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_document_items_record
              ON document_items(company_id,record_id,sort_order,id);
            CREATE INDEX IF NOT EXISTS idx_document_items_catalog
              ON document_items(company_id,catalog_record_id);
            CREATE UNIQUE INDEX IF NOT EXISTS idx_document_items_reservation
              ON document_items(reservation_id) WHERE reservation_id IS NOT NULL;

            CREATE TRIGGER IF NOT EXISTS trg_document_item_scope_insert
            BEFORE INSERT ON document_items FOR EACH ROW
            WHEN COALESCE((SELECT company_id FROM records WHERE id=NEW.record_id
                           AND module IN ('propostas','vendas','solicitacoes_compra',
                                          'pedidos_compra','ordens_servico')),-1) != NEW.company_id
              OR COALESCE((SELECT company_id FROM records WHERE id=NEW.catalog_record_id
                           AND module=CASE NEW.item_kind WHEN 'PRODUCT' THEN 'produtos'
                                     ELSE 'catalogo_servicos' END),-1) != NEW.company_id
              OR (NEW.warehouse_id IS NOT NULL AND
                  COALESCE((SELECT company_id FROM warehouses WHERE id=NEW.warehouse_id),-1)
                    != NEW.company_id)
            BEGIN
              SELECT RAISE(ABORT, 'Item fora da empresa, documento ou catálogo incompatível');
            END;
            CREATE TRIGGER IF NOT EXISTS trg_document_item_scope_update
            BEFORE UPDATE OF company_id,record_id,item_kind,catalog_record_id,warehouse_id
            ON document_items FOR EACH ROW
            WHEN COALESCE((SELECT company_id FROM records WHERE id=NEW.record_id
                           AND module IN ('propostas','vendas','solicitacoes_compra',
                                          'pedidos_compra','ordens_servico')),-1) != NEW.company_id
              OR COALESCE((SELECT company_id FROM records WHERE id=NEW.catalog_record_id
                           AND module=CASE NEW.item_kind WHEN 'PRODUCT' THEN 'produtos'
                                     ELSE 'catalogo_servicos' END),-1) != NEW.company_id
              OR (NEW.warehouse_id IS NOT NULL AND
                  COALESCE((SELECT company_id FROM warehouses WHERE id=NEW.warehouse_id),-1)
                    != NEW.company_id)
            BEGIN
              SELECT RAISE(ABORT, 'Item fora da empresa, documento ou catálogo incompatível');
            END;

            CREATE TRIGGER IF NOT EXISTS trg_inventory_movement_immutable_update
            BEFORE UPDATE ON inventory_movements BEGIN
              SELECT RAISE(ABORT, 'Movimento de estoque é imutável');
            END;
            CREATE TRIGGER IF NOT EXISTS trg_inventory_movement_immutable_delete
            BEFORE DELETE ON inventory_movements BEGIN
              SELECT RAISE(ABORT, 'Movimento de estoque é imutável');
            END;
            CREATE TRIGGER IF NOT EXISTS trg_warehouse_same_company_insert
            BEFORE INSERT ON warehouses FOR EACH ROW
            WHEN COALESCE((SELECT company_id FROM branches WHERE id=NEW.branch_id),-1) != NEW.company_id
            BEGIN
              SELECT RAISE(ABORT, 'Depósito e unidade devem pertencer à mesma empresa');
            END;
            CREATE TRIGGER IF NOT EXISTS trg_warehouse_same_company_update
            BEFORE UPDATE OF company_id,branch_id ON warehouses FOR EACH ROW
            WHEN COALESCE((SELECT company_id FROM branches WHERE id=NEW.branch_id),-1) != NEW.company_id
            BEGIN
              SELECT RAISE(ABORT, 'Depósito e unidade devem pertencer à mesma empresa');
            END;
            CREATE TRIGGER IF NOT EXISTS trg_inventory_balance_scope_insert
            BEFORE INSERT ON inventory_balances FOR EACH ROW
            WHEN COALESCE((SELECT company_id FROM warehouses WHERE id=NEW.warehouse_id),-1) != NEW.company_id
              OR COALESCE((SELECT company_id FROM records WHERE id=NEW.product_record_id AND module='produtos'),-1) != NEW.company_id
            BEGIN
              SELECT RAISE(ABORT, 'Saldo de estoque fora da empresa ou produto inválido');
            END;
            CREATE TRIGGER IF NOT EXISTS trg_inventory_balance_scope_update
            BEFORE UPDATE OF company_id,warehouse_id,product_record_id ON inventory_balances FOR EACH ROW
            WHEN COALESCE((SELECT company_id FROM warehouses WHERE id=NEW.warehouse_id),-1) != NEW.company_id
              OR COALESCE((SELECT company_id FROM records WHERE id=NEW.product_record_id AND module='produtos'),-1) != NEW.company_id
            BEGIN
              SELECT RAISE(ABORT, 'Saldo de estoque fora da empresa ou produto inválido');
            END;
            CREATE TRIGGER IF NOT EXISTS trg_inventory_reservation_scope_insert
            BEFORE INSERT ON inventory_reservations FOR EACH ROW
            WHEN COALESCE((SELECT company_id FROM warehouses WHERE id=NEW.warehouse_id),-1) != NEW.company_id
              OR COALESCE((SELECT company_id FROM records WHERE id=NEW.product_record_id AND module='produtos'),-1) != NEW.company_id
            BEGIN
              SELECT RAISE(ABORT, 'Reserva fora da empresa ou produto inválido');
            END;
            CREATE TRIGGER IF NOT EXISTS trg_inventory_reservation_scope_update
            BEFORE UPDATE OF company_id,warehouse_id,product_record_id ON inventory_reservations FOR EACH ROW
            WHEN COALESCE((SELECT company_id FROM warehouses WHERE id=NEW.warehouse_id),-1) != NEW.company_id
              OR COALESCE((SELECT company_id FROM records WHERE id=NEW.product_record_id AND module='produtos'),-1) != NEW.company_id
            BEGIN
              SELECT RAISE(ABORT, 'Reserva fora da empresa ou produto inválido');
            END;
            CREATE TRIGGER IF NOT EXISTS trg_inventory_movement_scope_insert
            BEFORE INSERT ON inventory_movements FOR EACH ROW
            WHEN COALESCE((SELECT company_id FROM warehouses WHERE id=NEW.warehouse_id),-1) != NEW.company_id
              OR COALESCE((SELECT company_id FROM records WHERE id=NEW.product_record_id AND module='produtos'),-1) != NEW.company_id
              OR (NEW.counterpart_warehouse_id IS NOT NULL AND
                  COALESCE((SELECT company_id FROM warehouses WHERE id=NEW.counterpart_warehouse_id),-1) != NEW.company_id)
            BEGIN
              SELECT RAISE(ABORT, 'Movimento de estoque fora da empresa ou produto inválido');
            END;

            CREATE TABLE IF NOT EXISTS fiscal_schema_versions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                document_type TEXT NOT NULL,
                version TEXT NOT NULL,
                environment TEXT NOT NULL CHECK(environment IN ('HOMOLOGATION','PRODUCTION','BOTH')),
                schema_reference TEXT NOT NULL,
                valid_from TEXT,
                valid_to TEXT,
                active INTEGER NOT NULL DEFAULT 1 CHECK(active IN (0,1)),
                created_at TEXT NOT NULL,
                UNIQUE(document_type,version,environment)
            );
            CREATE TABLE IF NOT EXISTS fiscal_operations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
                code TEXT NOT NULL,
                name TEXT NOT NULL,
                direction TEXT NOT NULL CHECK(direction IN ('IN','OUT','BOTH')),
                parameters_json TEXT NOT NULL DEFAULT '{}',
                version INTEGER NOT NULL DEFAULT 1,
                valid_from TEXT,
                valid_to TEXT,
                active INTEGER NOT NULL DEFAULT 1 CHECK(active IN (0,1)),
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(company_id,code,version)
            );
            CREATE TABLE IF NOT EXISTS tax_profiles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
                name TEXT NOT NULL,
                tax_regime TEXT,
                parameters_json TEXT NOT NULL DEFAULT '{}',
                version INTEGER NOT NULL DEFAULT 1,
                valid_from TEXT,
                valid_to TEXT,
                active INTEGER NOT NULL DEFAULT 1 CHECK(active IN (0,1)),
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(company_id,name,version)
            );
            CREATE TABLE IF NOT EXISTS company_fiscal_profiles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
                branch_id INTEGER REFERENCES branches(id),
                tax_profile_id INTEGER NOT NULL REFERENCES tax_profiles(id),
                parameters_json TEXT NOT NULL DEFAULT '{}',
                valid_from TEXT,
                valid_to TEXT,
                active INTEGER NOT NULL DEFAULT 1 CHECK(active IN (0,1)),
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS product_fiscal_profiles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
                product_record_id INTEGER NOT NULL REFERENCES records(id),
                tax_profile_id INTEGER REFERENCES tax_profiles(id),
                ncm TEXT,
                cest TEXT,
                merchandise_origin TEXT,
                cclass_trib TEXT,
                parameters_json TEXT NOT NULL DEFAULT '{}',
                version INTEGER NOT NULL DEFAULT 1,
                valid_from TEXT,
                valid_to TEXT,
                active INTEGER NOT NULL DEFAULT 1 CHECK(active IN (0,1)),
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(company_id,product_record_id,version)
            );
            CREATE TABLE IF NOT EXISTS tax_rules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
                fiscal_operation_id INTEGER REFERENCES fiscal_operations(id),
                tax_profile_id INTEGER REFERENCES tax_profiles(id),
                tax_code TEXT NOT NULL,
                priority INTEGER NOT NULL DEFAULT 100,
                conditions_json TEXT NOT NULL DEFAULT '{}',
                result_json TEXT NOT NULL DEFAULT '{}',
                version INTEGER NOT NULL DEFAULT 1,
                valid_from TEXT,
                valid_to TEXT,
                active INTEGER NOT NULL DEFAULT 1 CHECK(active IN (0,1)),
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS fiscal_documents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
                branch_id INTEGER NOT NULL REFERENCES branches(id),
                record_id INTEGER REFERENCES records(id),
                fiscal_operation_id INTEGER REFERENCES fiscal_operations(id),
                tax_profile_id INTEGER REFERENCES tax_profiles(id),
                fiscal_schema_version_id INTEGER REFERENCES fiscal_schema_versions(id),
                document_type TEXT NOT NULL,
                environment TEXT NOT NULL CHECK(environment IN ('HOMOLOGATION','PRODUCTION')),
                status TEXT NOT NULL DEFAULT 'DRAFT',
                access_key TEXT,
                protocol TEXT,
                totals_json TEXT NOT NULL DEFAULT '{}',
                payload_json TEXT NOT NULL DEFAULT '{}',
                revision INTEGER NOT NULL DEFAULT 1,
                created_by INTEGER REFERENCES users(id),
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_fiscal_documents_company_status
              ON fiscal_documents(company_id,status,created_at DESC);
            CREATE TABLE IF NOT EXISTS fiscal_document_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
                fiscal_document_id INTEGER NOT NULL REFERENCES fiscal_documents(id) ON DELETE CASCADE,
                line_number INTEGER NOT NULL,
                product_record_id INTEGER REFERENCES records(id),
                service_record_id INTEGER REFERENCES records(id),
                quantity_micros INTEGER NOT NULL DEFAULT 0,
                unit_value_micros INTEGER NOT NULL DEFAULT 0,
                fiscal_classification_json TEXT NOT NULL DEFAULT '{}',
                calculation_json TEXT NOT NULL DEFAULT '{}',
                UNIQUE(fiscal_document_id,line_number)
            );
            CREATE TABLE IF NOT EXISTS fiscal_certificates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
                branch_id INTEGER REFERENCES branches(id),
                certificate_type TEXT NOT NULL,
                subject_name TEXT,
                fingerprint_sha256 TEXT NOT NULL,
                encrypted_content BLOB,
                valid_from TEXT,
                valid_to TEXT,
                status TEXT NOT NULL DEFAULT 'INACTIVE',
                created_by INTEGER REFERENCES users(id),
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(company_id,fingerprint_sha256)
            );
            CREATE TABLE IF NOT EXISTS xml_documents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
                fiscal_document_id INTEGER REFERENCES fiscal_documents(id) ON DELETE CASCADE,
                fiscal_event_id INTEGER REFERENCES fiscal_events(id) ON DELETE CASCADE,
                fiscal_schema_version_id INTEGER REFERENCES fiscal_schema_versions(id),
                document_role TEXT NOT NULL,
                sha256 TEXT NOT NULL,
                content BLOB NOT NULL,
                created_at TEXT NOT NULL,
                UNIQUE(company_id,sha256,document_role)
            );
            CREATE TABLE IF NOT EXISTS sefaz_configurations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
                branch_id INTEGER NOT NULL REFERENCES branches(id) ON DELETE CASCADE,
                environment TEXT NOT NULL DEFAULT 'HOMOLOGATION'
                  CHECK(environment IN ('HOMOLOGATION','PRODUCTION')),
                uf TEXT NOT NULL,
                state_code TEXT NOT NULL,
                service_version TEXT NOT NULL DEFAULT '4.00',
                status_service_url TEXT NOT NULL,
                authorization_service_url TEXT,
                authorization_return_url TEXT,
                protocol_service_url TEXT,
                event_service_url TEXT,
                invalidation_service_url TEXT,
                source_url TEXT NOT NULL,
                source_verified_at TEXT NOT NULL,
                enabled INTEGER NOT NULL DEFAULT 0 CHECK(enabled IN (0,1)),
                last_status_code TEXT,
                last_status_reason TEXT,
                last_checked_at TEXT,
                created_by INTEGER REFERENCES users(id),
                updated_by INTEGER REFERENCES users(id),
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(company_id,branch_id,environment)
            );
            CREATE INDEX IF NOT EXISTS idx_sefaz_config_company_environment
              ON sefaz_configurations(company_id,environment,enabled);
            CREATE TABLE IF NOT EXISTS accounting_exports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
                period TEXT NOT NULL,
                format_version TEXT NOT NULL,
                sha256 TEXT NOT NULL,
                file_size INTEGER NOT NULL CHECK(file_size >= 0),
                totals_json TEXT NOT NULL DEFAULT '{}',
                generated_by INTEGER REFERENCES users(id),
                created_at TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_accounting_exports_company_period
              ON accounting_exports(company_id,period,created_at DESC);

            CREATE TRIGGER IF NOT EXISTS trg_company_fiscal_profile_scope_insert
            BEFORE INSERT ON company_fiscal_profiles FOR EACH ROW
            WHEN (NEW.branch_id IS NOT NULL AND
                  COALESCE((SELECT company_id FROM branches WHERE id=NEW.branch_id),-1) != NEW.company_id)
              OR COALESCE((SELECT company_id FROM tax_profiles WHERE id=NEW.tax_profile_id),-1) != NEW.company_id
            BEGIN
              SELECT RAISE(ABORT, 'Perfil fiscal fora da empresa');
            END;
            CREATE TRIGGER IF NOT EXISTS trg_product_fiscal_profile_scope_insert
            BEFORE INSERT ON product_fiscal_profiles FOR EACH ROW
            WHEN COALESCE((SELECT company_id FROM records WHERE id=NEW.product_record_id AND module='produtos'),-1) != NEW.company_id
              OR (NEW.tax_profile_id IS NOT NULL AND
                  COALESCE((SELECT company_id FROM tax_profiles WHERE id=NEW.tax_profile_id),-1) != NEW.company_id)
            BEGIN
              SELECT RAISE(ABORT, 'Perfil fiscal de produto fora da empresa');
            END;
            CREATE TRIGGER IF NOT EXISTS trg_tax_rule_scope_insert
            BEFORE INSERT ON tax_rules FOR EACH ROW
            WHEN (NEW.fiscal_operation_id IS NOT NULL AND
                  COALESCE((SELECT company_id FROM fiscal_operations WHERE id=NEW.fiscal_operation_id),-1) != NEW.company_id)
              OR (NEW.tax_profile_id IS NOT NULL AND
                  COALESCE((SELECT company_id FROM tax_profiles WHERE id=NEW.tax_profile_id),-1) != NEW.company_id)
            BEGIN
              SELECT RAISE(ABORT, 'Regra tributária fora da empresa');
            END;
            CREATE TRIGGER IF NOT EXISTS trg_fiscal_document_scope_insert
            BEFORE INSERT ON fiscal_documents FOR EACH ROW
            WHEN COALESCE((SELECT company_id FROM branches WHERE id=NEW.branch_id),-1) != NEW.company_id
              OR (NEW.record_id IS NOT NULL AND
                  COALESCE((SELECT company_id FROM records WHERE id=NEW.record_id),-1) != NEW.company_id)
              OR (NEW.fiscal_operation_id IS NOT NULL AND
                  COALESCE((SELECT company_id FROM fiscal_operations WHERE id=NEW.fiscal_operation_id),-1) != NEW.company_id)
              OR (NEW.tax_profile_id IS NOT NULL AND
                  COALESCE((SELECT company_id FROM tax_profiles WHERE id=NEW.tax_profile_id),-1) != NEW.company_id)
            BEGIN
              SELECT RAISE(ABORT, 'Documento fiscal fora da empresa');
            END;
            CREATE TRIGGER IF NOT EXISTS trg_fiscal_document_item_scope_insert
            BEFORE INSERT ON fiscal_document_items FOR EACH ROW
            WHEN COALESCE((SELECT company_id FROM fiscal_documents WHERE id=NEW.fiscal_document_id),-1) != NEW.company_id
              OR (NEW.product_record_id IS NOT NULL AND
                  COALESCE((SELECT company_id FROM records WHERE id=NEW.product_record_id AND module='produtos'),-1) != NEW.company_id)
              OR (NEW.service_record_id IS NOT NULL AND
                  COALESCE((SELECT company_id FROM records WHERE id=NEW.service_record_id AND module='catalogo_servicos'),-1) != NEW.company_id)
            BEGIN
              SELECT RAISE(ABORT, 'Item fiscal fora da empresa');
            END;
            CREATE TRIGGER IF NOT EXISTS trg_fiscal_certificate_scope_insert
            BEFORE INSERT ON fiscal_certificates FOR EACH ROW
            WHEN NEW.branch_id IS NOT NULL AND
                 COALESCE((SELECT company_id FROM branches WHERE id=NEW.branch_id),-1) != NEW.company_id
            BEGIN
              SELECT RAISE(ABORT, 'Certificado fiscal fora da empresa');
            END;
            CREATE TRIGGER IF NOT EXISTS trg_fiscal_certificate_scope_update
            BEFORE UPDATE OF company_id,branch_id ON fiscal_certificates FOR EACH ROW
            WHEN NEW.branch_id IS NOT NULL AND
                 COALESCE((SELECT company_id FROM branches WHERE id=NEW.branch_id),-1) != NEW.company_id
            BEGIN
              SELECT RAISE(ABORT, 'Certificado fiscal fora da empresa');
            END;
            CREATE TRIGGER IF NOT EXISTS trg_sefaz_configuration_scope_insert
            BEFORE INSERT ON sefaz_configurations FOR EACH ROW
            WHEN COALESCE((SELECT company_id FROM branches WHERE id=NEW.branch_id),-1) != NEW.company_id
            BEGIN
              SELECT RAISE(ABORT, 'Configuração SEFAZ fora da empresa');
            END;
            CREATE TRIGGER IF NOT EXISTS trg_sefaz_configuration_scope_update
            BEFORE UPDATE OF company_id,branch_id ON sefaz_configurations FOR EACH ROW
            WHEN COALESCE((SELECT company_id FROM branches WHERE id=NEW.branch_id),-1) != NEW.company_id
            BEGIN
              SELECT RAISE(ABORT, 'Configuração SEFAZ fora da empresa');
            END;
            CREATE TRIGGER IF NOT EXISTS trg_xml_document_scope_insert
            BEFORE INSERT ON xml_documents FOR EACH ROW
            WHEN (NEW.fiscal_document_id IS NOT NULL AND
                  COALESCE((SELECT company_id FROM fiscal_documents WHERE id=NEW.fiscal_document_id),-1) != NEW.company_id)
              OR (NEW.fiscal_event_id IS NOT NULL AND
                  COALESCE((SELECT company_id FROM fiscal_events WHERE id=NEW.fiscal_event_id),-1) != NEW.company_id)
            BEGIN
              SELECT RAISE(ABORT, 'XML fiscal fora da empresa');
            END;
            """
        )

        def ensure_column(table, name, definition):
            columns = {row["name"] for row in db.execute(f"PRAGMA table_info({table})").fetchall()}
            if name not in columns:
                db.execute(f"ALTER TABLE {table} ADD COLUMN {name} {definition}")

        ensure_column("records", "deleted_at", "TEXT")
        ensure_column("records", "subject_id", "INTEGER REFERENCES subjects(id)")
        ensure_column("records", "company_id", "INTEGER REFERENCES companies(id)")
        ensure_column("records", "revision", "INTEGER NOT NULL DEFAULT 1")
        ensure_column("companies", "holding_id", "INTEGER REFERENCES holdings(id)")
        ensure_column("companies", "legal_name", "TEXT")
        ensure_column("companies", "state_registration", "TEXT")
        ensure_column("companies", "municipal_registration", "TEXT")
        ensure_column("companies", "uf", "TEXT")
        ensure_column("companies", "municipality_code", "TEXT")
        ensure_column("companies", "tax_regime", "TEXT")
        ensure_column("branches", "state_registration", "TEXT")
        ensure_column("branches", "municipal_registration", "TEXT")
        ensure_column("branches", "uf", "TEXT")
        ensure_column("branches", "municipality_code", "TEXT")
        ensure_column("fiscal_certificates", "serial_number", "TEXT")
        ensure_column("fiscal_certificates", "issuer_name", "TEXT")
        ensure_column("fiscal_certificates", "key_algorithm", "TEXT")
        ensure_column("fiscal_certificates", "last_used_at", "TEXT")
        ensure_column("subjects", "company_id", "INTEGER REFERENCES companies(id)")
        ensure_column("sessions", "company_id", "INTEGER REFERENCES companies(id)")
        ensure_column("sessions", "public_id", "TEXT")
        ensure_column("sessions", "last_activity_at", "INTEGER")
        ensure_column("sessions", "ip_address", "TEXT")
        ensure_column("sessions", "user_agent", "TEXT")
        ensure_column("audit_log", "company_id", "INTEGER REFERENCES companies(id)")
        ensure_column("tender_searches", "company_id", "INTEGER REFERENCES companies(id)")
        ensure_column("tender_results", "company_id", "INTEGER REFERENCES companies(id)")
        ensure_column("tender_results", "relevance_feedback", "TEXT")
        ensure_column("tender_results", "feedback_reason", "TEXT")
        ensure_column("tender_results", "feedback_at", "TEXT")
        ensure_column("tender_results", "feedback_by", "INTEGER REFERENCES users(id)")
        ensure_column("tender_details", "analysis_json", "TEXT NOT NULL DEFAULT '{}'")
        ensure_column("record_versions", "company_id", "INTEGER REFERENCES companies(id)")
        ensure_column("approvals", "requested_by", "INTEGER REFERENCES users(id)")
        ensure_column("approvals", "record_revision", "INTEGER NOT NULL DEFAULT 1")
        ensure_column("approvals", "request_comment", "TEXT")
        ensure_column("approvals", "decision_comment", "TEXT")
        ensure_column("attachments", "sha256", "TEXT")
        ensure_column("attachments", "license_confirmed", "INTEGER NOT NULL DEFAULT 0")
        ensure_column("inventory_balances", "inventory_value_cents", "INTEGER NOT NULL DEFAULT 0")
        ensure_column("inventory_movements", "unit_cost_cents", "INTEGER")
        ensure_column("inventory_movements", "value_delta_cents", "INTEGER NOT NULL DEFAULT 0")
        ensure_column("inventory_movements", "balance_value_cents", "INTEGER NOT NULL DEFAULT 0")
        for row in db.execute(
                "SELECT token_hash FROM sessions WHERE public_id IS NULL OR public_id=''"
        ).fetchall():
            db.execute(
                "UPDATE sessions SET public_id=? WHERE token_hash=?",
                (secrets.token_hex(12), row["token_hash"]),
            )
        db.execute(
            "UPDATE sessions SET last_activity_at=COALESCE(last_activity_at,expires_at-?)",
            (SESSION_SECONDS,),
        )
        db.execute(
            "CREATE UNIQUE INDEX IF NOT EXISTS idx_sessions_public_id ON sessions(public_id)"
        )
        telemetry_cutoff = (
            datetime.now(timezone.utc) - timedelta(days=TELEMETRY_RETENTION_DAYS)
        ).isoformat(timespec="seconds")
        db.execute(
            "DELETE FROM system_events WHERE resolved_at IS NOT NULL AND created_at<?",
            (telemetry_cutoff,),
        )

        # Índices alinhados às consultas reais: todos os registros são sempre
        # filtrados pela empresa ativa e, em seguida, por módulo/situação/prazo.
        # Os índices antigos por coluna isolada permanecem por compatibilidade.
        db.executescript(
            """
            CREATE INDEX IF NOT EXISTS idx_records_company_module_active_updated
              ON records(company_id,module,deleted_at,updated_at DESC);
            CREATE INDEX IF NOT EXISTS idx_records_company_status_active_due
              ON records(company_id,status,deleted_at,due_date);
            CREATE INDEX IF NOT EXISTS idx_records_company_subject_active
              ON records(company_id,subject_id,deleted_at);
            CREATE UNIQUE INDEX IF NOT EXISTS idx_records_company_party_document_active
              ON records(
                company_id,
                replace(replace(replace(replace(
                  CAST(json_extract(payload,'$.documento') AS TEXT),
                  '.',''),'-',''),'/',''),' ','')
              )
              WHERE deleted_at IS NULL
                AND module IN ('clientes','fornecedores')
                AND length(replace(replace(replace(replace(
                  CAST(json_extract(payload,'$.documento') AS TEXT),
                  '.',''),'-',''),'/',''),' ','')) IN (11,14);
            CREATE INDEX IF NOT EXISTS idx_subjects_company_status_name
              ON subjects(company_id,status,normalized_name);
            CREATE INDEX IF NOT EXISTS idx_record_subjects_record
              ON record_subjects(record_id,subject_id);

            -- A aplicação já valida esse vínculo; as travas no banco impedem
            -- que importadores ou evoluções futuras criem relação entre empresas.
            CREATE TRIGGER IF NOT EXISTS trg_relationship_same_company_insert
            BEFORE INSERT ON record_relationships
            FOR EACH ROW
            WHEN COALESCE((SELECT company_id FROM records WHERE id=NEW.from_record_id), -1)
                 != COALESCE((SELECT company_id FROM records WHERE id=NEW.to_record_id), -1)
            BEGIN
              SELECT RAISE(ABORT, 'Relacionamentos devem permanecer na mesma empresa');
            END;

            CREATE TRIGGER IF NOT EXISTS trg_record_subject_same_company_insert
            BEFORE INSERT ON record_subjects
            FOR EACH ROW
            WHEN COALESCE((SELECT company_id FROM records WHERE id=NEW.record_id), -1)
                 != COALESCE((SELECT company_id FROM subjects WHERE id=NEW.subject_id), -1)
            BEGIN
              SELECT RAISE(ABORT, 'Assunto deve pertencer à mesma empresa do registro');
            END;
            """
        )

        now = utc_now()
        default_holding = db.execute("SELECT id FROM holdings ORDER BY id LIMIT 1").fetchone()
        if default_holding:
            default_holding_id = default_holding["id"]
        else:
            default_holding_id = db.execute(
                "INSERT INTO holdings(name,created_at,updated_at) VALUES(?,?,?)",
                ("Holding principal", now, now),
            ).lastrowid
        default_company = db.execute("SELECT id FROM companies ORDER BY id LIMIT 1").fetchone()
        if not default_company:
            legacy = db.execute("SELECT value FROM settings WHERE key='company'").fetchone()
            legacy_company = json.loads(legacy["value"] or "{}") if legacy else {}
            cursor = db.execute(
                """INSERT INTO companies
                   (name,cnpj,phone,address,created_at,updated_at,holding_id)
                   VALUES(?,?,?,?,?,?,?)""",
                (legacy_company.get("name") or "SECCOL", legacy_company.get("cnpj"),
                 legacy_company.get("phone"), legacy_company.get("address"), now, now,
                 default_holding_id),
            )
            default_company_id = cursor.lastrowid
        else:
            default_company_id = default_company["id"]
        db.execute(
            "UPDATE companies SET holding_id=? WHERE holding_id IS NULL",
            (default_holding_id,),
        )
        for company in db.execute("SELECT id,name,cnpj FROM companies").fetchall():
            self.ensure_company_structure(
                company["id"], company["name"], company["cnpj"], now=now,
            )

        for table in ("records", "subjects", "sessions", "audit_log", "tender_searches",
                      "tender_results", "record_versions"):
            db.execute(f"UPDATE {table} SET company_id=? WHERE company_id IS NULL", (default_company_id,))

        # A chave normalizada inclui a empresa para permitir assuntos iguais em empresas distintas.
        subject_rows = db.execute("SELECT id,company_id,normalized_name FROM subjects").fetchall()
        for subject in subject_rows:
            prefix = f'{subject["company_id"]}:'
            if not str(subject["normalized_name"] or "").startswith(prefix):
                db.execute("UPDATE subjects SET normalized_name=? WHERE id=?",
                           (prefix + str(subject["normalized_name"] or ""), subject["id"]))

        users = db.execute("SELECT id,role,created_at,updated_at FROM users").fetchall()
        for user in users:
            db.execute(
                """INSERT OR IGNORE INTO company_memberships
                   (company_id,user_id,role,permissions,active,created_at,updated_at)
                   VALUES(?,?,?,'{}',1,?,?)""",
                (default_company_id, user["id"], user["role"], user["created_at"], user["updated_at"]),
            )
        db.execute("CREATE INDEX IF NOT EXISTS idx_records_company_module ON records(company_id,module)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_subjects_company ON subjects(company_id)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_tender_results_company ON tender_results(company_id,status)")
        db.execute(
            """UPDATE approvals SET requested_by=COALESCE(requested_by,
                   (SELECT created_by FROM records WHERE records.id=approvals.record_id)),
                   request_comment=COALESCE(request_comment,comment),
                   record_revision=COALESCE(record_revision,1)"""
        )
        db.execute(
            """UPDATE approvals SET status='Cancelada por migração'
               WHERE status='Pendente' AND EXISTS (
                 SELECT 1 FROM approvals newer
                 WHERE newer.company_id=approvals.company_id
                   AND newer.record_id=approvals.record_id
                   AND newer.approval_type=approvals.approval_type
                   AND newer.status='Pendente' AND newer.id>approvals.id
               )"""
        )
        db.execute(
            """CREATE UNIQUE INDEX IF NOT EXISTS idx_approvals_one_pending
               ON approvals(company_id,record_id,approval_type) WHERE status='Pendente'"""
        )
        db.execute(
            """UPDATE tender_jobs SET status='failed',finished_at=?,
               error_detail='Execução interrompida por reinicialização do servidor.'
               WHERE status IN ('queued','running')""",
            (utc_now(),),
        )
        db.execute(
            """CREATE UNIQUE INDEX IF NOT EXISTS idx_tender_jobs_one_active
               ON tender_jobs(company_id) WHERE status IN ('queued','running')"""
        )
        db.execute("INSERT OR IGNORE INTO setup_state(id,configured) VALUES(1,0)")
        if db.execute("SELECT COUNT(*) FROM users").fetchone()[0]:
            db.execute(
                "UPDATE setup_state SET configured=1,configured_at=COALESCE(configured_at,?) WHERE id=1",
                (utc_now(),),
            )
        db.execute(
            """INSERT OR IGNORE INTO schema_migrations(version,name,applied_at)
               VALUES(220,'hardening-v2.2',?)""", (utc_now(),)
        )
        db.execute(
            """INSERT OR IGNORE INTO schema_migrations(version,name,applied_at)
               VALUES(221,'relational-master-record-links',?)""", (utc_now(),)
        )
        db.execute(
            """INSERT OR IGNORE INTO schema_migrations(version,name,applied_at)
               VALUES(222,'operational-control-center',?)""", (utc_now(),)
        )
        db.execute(
            """INSERT OR IGNORE INTO schema_migrations(version,name,applied_at)
               VALUES(223,'erp-multicompany-inventory-fiscal-foundation',?)""", (utc_now(),)
        )
        db.execute(
            """INSERT OR IGNORE INTO schema_migrations(version,name,applied_at)
               VALUES(224,'commercial-service-purchase-document-items',?)""", (utc_now(),)
        )
        db.execute(
            """INSERT OR IGNORE INTO schema_migrations(version,name,applied_at)
               VALUES(225,'tender-keywords-and-quality-feedback',?)""", (utc_now(),)
        )
        db.execute(
            """INSERT OR IGNORE INTO schema_migrations(version,name,applied_at)
               VALUES(226,'functional-access-costed-inventory-controllership',?)""", (utc_now(),)
        )
        db.execute(
            """INSERT OR IGNORE INTO schema_migrations(version,name,applied_at)
               VALUES(227,'sefaz-readiness-a1-vault-accounting-export',?)""", (utc_now(),)
        )
        db.commit()
        self.seed_sources(default_company_id)
        self.seed_norms(default_company_id)
        self.seed_seccol_portfolio(default_company_id)
        self.migrate_subjects()
        self.migrate_record_references()
        missing_hashes = self.connection().execute(
            "SELECT id,content FROM attachments WHERE sha256 IS NULL OR sha256=''"
        ).fetchall()
        for attachment in missing_hashes:
            self.connection().execute(
                "UPDATE attachments SET sha256=? WHERE id=?",
                (hashlib.sha256(attachment["content"]).hexdigest(), attachment["id"]),
            )
        self.commit_if_outer()

    def ensure_company_structure(self, company_id, company_name, company_cnpj=None, now=None):
        """Garante a hierarquia mínima Company -> Branch -> Warehouse de forma idempotente."""
        db = self.connection()
        now = now or utc_now()
        branch = db.execute(
            "SELECT id FROM branches WHERE company_id=? ORDER BY is_headquarters DESC,id LIMIT 1",
            (company_id,),
        ).fetchone()
        if branch:
            branch_id = branch["id"]
        else:
            branch_id = db.execute(
                """INSERT INTO branches
                   (company_id,code,name,cnpj,active,is_headquarters,created_at,updated_at)
                   VALUES(?,'MATRIZ',?,?,1,1,?,?)""",
                (company_id, f"Matriz — {company_name}", company_cnpj, now, now),
            ).lastrowid
        warehouse = db.execute(
            "SELECT id FROM warehouses WHERE company_id=? ORDER BY id LIMIT 1",
            (company_id,),
        ).fetchone()
        if not warehouse:
            db.execute(
                """INSERT INTO warehouses
                   (company_id,branch_id,code,name,location,active,created_at,updated_at)
                   VALUES(?,?,'PRINCIPAL','Depósito principal','Matriz',1,?,?)""",
                (company_id, branch_id, now, now),
            )
        self.commit_if_outer()

    @staticmethod
    def normalize_subject(value):
        text = unicodedata.normalize("NFD", str(value or "").strip().lower())
        return "".join(char for char in text if unicodedata.category(char) != "Mn")

    def ensure_subject(self, name, user_id=None, company_id=None):
        name = str(name or "").strip()[:180]
        company_id = int(company_id or 1)
        normalized = self.normalize_subject(name)
        if not normalized:
            return None
        normalized_key = f"{company_id}:{normalized}"
        now = utc_now()
        self.connection().execute(
            """INSERT OR IGNORE INTO subjects
               (name,normalized_name,status,created_by,created_at,updated_at,company_id)
               VALUES(?,?,'Ativo',?,?,?,?)""",
            (name, normalized_key, user_id, now, now, company_id))
        row = self.connection().execute(
            "SELECT id FROM subjects WHERE company_id=? AND normalized_name=?",
            (company_id, normalized_key)).fetchone()
        return row["id"] if row else None

    def sync_relationships(self, record_id, payload, user_id=None, company_id=None):
        db = self.connection()
        record = db.execute("SELECT company_id FROM records WHERE id=?", (record_id,)).fetchone()
        company_id = int(company_id or (record["company_id"] if record else 1) or 1)
        subject_id = self.ensure_subject(payload.get("assunto"), user_id, company_id)
        db.execute("UPDATE records SET subject_id=? WHERE id=?", (subject_id, record_id))
        db.execute("DELETE FROM record_subjects WHERE record_id=?", (record_id,))
        if subject_id:
            db.execute(
                """INSERT OR IGNORE INTO record_subjects
                   (record_id,subject_id,relationship_type,is_primary,created_by,created_at)
                   VALUES(?,?,'Assunto principal',1,?,?)""",
                (record_id, subject_id, user_id, utc_now()))
        additional_subjects = payload.get("assuntos_adicionais") or []
        if isinstance(additional_subjects, str):
            additional_subjects = [item.strip() for item in additional_subjects.split(",") if item.strip()]
        if isinstance(additional_subjects, list):
            for additional_name in additional_subjects[:30]:
                additional_id = self.ensure_subject(additional_name, user_id, company_id)
                if additional_id and additional_id != subject_id:
                    db.execute(
                        """INSERT OR IGNORE INTO record_subjects
                           (record_id,subject_id,relationship_type,is_primary,created_by,created_at)
                           VALUES(?,?,'Relacionado a',0,?,?)""",
                        (record_id, additional_id, user_id, utc_now()))
        requested = []
        primary = str(payload.get("registro_relacionado") or "").strip()
        if primary:
            requested.append({"record": primary, "type": payload.get("tipo_relacao") or "Relacionado a"})
        if isinstance(payload.get("relacionamentos"), list):
            requested.extend(payload["relacionamentos"][:50])
        db.execute("DELETE FROM record_relationships WHERE from_record_id=?", (record_id,))
        for relation in requested:
            if not isinstance(relation, dict):
                continue
            reference = str(relation.get("record") or relation.get("registro") or "")
            match = re.fullmatch(r"([a-z_]+):(\d+)", reference)
            if not match or match.group(1) not in MODULES:
                raise ValueError("Referência de relacionamento inválida")
            target_module, target_id = match.group(1), int(match.group(2))
            if target_id == record_id:
                raise ValueError("Um registro não pode ser relacionado a ele próprio")
            exists = db.execute(
                """SELECT id FROM records
                   WHERE id=? AND module=? AND company_id=? AND deleted_at IS NULL""",
                (target_id, target_module, company_id)).fetchone()
            if not exists:
                raise ValueError(
                    "O registro relacionado não existe, pertence a outro módulo/empresa ou foi excluído"
                )
            relation_type = str(relation.get("type") or relation.get("tipo") or "Relacionado a").strip()[:80]
            db.execute(
                """INSERT OR IGNORE INTO record_relationships
                   (from_record_id,to_record_id,relationship_type,created_by,created_at) VALUES(?,?,?,?,?)""",
                (record_id, target_id, relation_type, user_id, utc_now()))
    def migrate_subjects(self):
        db = self.connection()
        rows = db.execute("SELECT id,payload,created_by,company_id FROM records WHERE subject_id IS NULL").fetchall()
        for row in rows:
            try:
                payload = json.loads(row["payload"] or "{}")
                if payload.get("assunto"):
                    self.sync_relationships(row["id"], payload, row["created_by"], row["company_id"])
            except (ValueError, TypeError, json.JSONDecodeError):
                continue

    def migrate_record_references(self):
        """Recupera vínculos antigos quando o nome identifica um único cadastro mestre."""
        db = self.connection()
        rows = db.execute(
            """SELECT id,module,payload,company_id,created_by FROM records
               WHERE deleted_at IS NULL AND module NOT IN ('clientes','fornecedores')"""
        ).fetchall()
        for row in rows:
            try:
                payload = json.loads(row["payload"] or "{}")
            except (ValueError, TypeError, json.JSONDecodeError):
                continue
            changed = False
            new_relations = []
            for field, rule in RECORD_REFERENCE_RULES.items():
                id_key = f"{field}_id"
                display = str(payload.get(field) or "").strip()
                if payload.get(id_key) or not display:
                    continue
                placeholders = ",".join("?" for _ in rule["modules"])
                candidates = db.execute(
                    f"""SELECT id,module,title,payload FROM records
                        WHERE company_id=? AND deleted_at IS NULL
                          AND module IN ({placeholders}) AND title=? COLLATE NOCASE LIMIT 3""",
                    (row["company_id"], *rule["modules"], display),
                ).fetchall()
                expected_role = rule.get("party_role")
                if expected_role:
                    compatible = []
                    for candidate in candidates:
                        candidate_payload = json.loads(candidate["payload"] or "{}")
                        fallback = "F" if candidate["module"] == "fornecedores" else "C"
                        role = str(candidate_payload.get("tipo_cadastro") or fallback).strip()
                        role = {"Cliente (C)": "C", "Fornecedor (F)": "F",
                                "Cliente e fornecedor (A)": "A", "C e F": "A"}.get(role, role)
                        if expected_role == "A" or role == "A" or role == expected_role:
                            compatible.append(candidate)
                    candidates = compatible
                if len(candidates) != 1:
                    continue
                target = candidates[0]
                payload[id_key] = target["id"]
                payload[field] = target["title"]
                new_relations.append((target["id"], rule["relation"]))
                changed = True
            if not changed:
                continue
            db.execute("UPDATE records SET payload=? WHERE id=?", (json_dumps(payload), row["id"]))
            for target_id, relation_type in new_relations:
                db.execute(
                    """INSERT OR IGNORE INTO record_relationships
                       (from_record_id,to_record_id,relationship_type,created_by,created_at)
                       VALUES(?,?,?,?,?)""",
                    (row["id"], target_id, relation_type, row["created_by"], utc_now()),
                )
        self.commit_if_outer()

    def seed_sources(self, company_id=None):
        db = self.connection()
        now = utc_now()
        company_ids = [company_id] if company_id else [row["id"] for row in db.execute(
            "SELECT id FROM companies WHERE active=1").fetchall()]
        for current_company_id in company_ids:
            for key, name, url, coverage, collection_mode, category in SOURCE_CATALOG:
                exists = db.execute(
                    """SELECT id,title,payload FROM records WHERE company_id=? AND module='fontes'
                       AND json_extract(payload,'$.source_key')=?""",
                    (current_company_id, key),
                ).fetchone()
                if exists:
                    try:
                        payload = json.loads(exists["payload"] or "{}")
                    except json.JSONDecodeError:
                        payload = {}
                    payload.update({
                        "source_key": key, "url": url, "abrangencia": coverage,
                        "modo_coleta": collection_mode, "categoria": category,
                        "verificado_em": "2026-08-15",
                        "assunto": f"Fonte de busca — {name}",
                    })
                    if not payload.get("palavra_chave"):
                        payload["palavra_chave"] = ", ".join(DEFAULT_TENDER_KEYWORDS)
                    encoded_payload = json_dumps(payload)
                    if exists["title"] != name or exists["payload"] != encoded_payload:
                        db.execute(
                            "UPDATE records SET title=?,payload=?,updated_at=? WHERE id=?",
                            (name, encoded_payload, now, exists["id"]),
                        )
                    self.sync_relationships(exists["id"], payload, None, current_company_id)
                    continue
                payload = {
                    "source_key": key, "url": url, "abrangencia": coverage,
                    "modo_coleta": collection_mode, "categoria": category,
                    "palavra_chave": ", ".join(DEFAULT_TENDER_KEYWORDS),
                    "verificado_em": "2026-08-15", "ativo": True,
                    "ultima_execucao": None, "ultimo_sucesso": None,
                    "assunto": f"Fonte de busca — {name}",
                }
                cursor = db.execute(
                    """INSERT INTO records
                       (module,title,status,payload,company_id,created_at,updated_at)
                       VALUES('fontes',?,'Ativo',?,?,?,?)""",
                    (name, json_dumps(payload), current_company_id, now, now),
                )
                self.sync_relationships(cursor.lastrowid, payload, None, current_company_id)
        self.commit_if_outer()

    def seed_norms(self, company_id=None):
        """Instala a base normativa sem redistribuir textos integrais protegidos."""
        db = self.connection()
        now = utc_now()
        company_ids = [company_id] if company_id else [row["id"] for row in db.execute(
            "SELECT id FROM companies WHERE active=1").fetchall()]
        for current_company_id in company_ids:
            for norm in NORM_CATALOG:
                row = db.execute(
                    """SELECT id FROM records WHERE company_id=? AND module='normas_tecnicas'
                       AND json_extract(payload,'$.norm_key')=? AND deleted_at IS NULL""",
                    (current_company_id, norm["key"]),
                ).fetchone()
                if row:
                    record_id = row["id"]
                else:
                    payload = {
                        "norm_key": norm["key"], "codigo": norm["code"],
                        "organismo": norm["organization"], "edicao": norm["edition"],
                        "escopo_resumido": norm["scope"],
                        "aplicabilidade_seccol": norm["application"],
                        "ensaios_base": norm["tests"], "referencia_oficial": norm["url"],
                        "licenciamento": norm["license"],
                        "documento_status": (
                            "Ficha de referência anexada; adicionar a cópia integral licenciada da SECCOL"
                            if norm["license"] == "Comercial/licenciada"
                            else "Ficha de referência anexada; texto oficial disponível no endereço público"
                        ),
                        "verificado_em": "2026-08-15",
                        "assunto": f'Base normativa — {norm["code"]}',
                        "notes": "Confirmar edição, emendas, escopo contratual e requisitos regulatórios antes da emissão.",
                    }
                    cursor = db.execute(
                        """INSERT INTO records
                           (module,title,status,payload,company_id,created_at,updated_at)
                           VALUES('normas_tecnicas',?,?,?,?,?,?)""",
                        (norm["code"], norm["status"], json_dumps(payload),
                         current_company_id, now, now),
                    )
                    record_id = cursor.lastrowid
                    self.sync_relationships(record_id, payload, None, current_company_id)
                attached = db.execute(
                    """SELECT id FROM attachments WHERE company_id=? AND record_id=?
                       AND category='Ficha de referência normativa' LIMIT 1""",
                    (current_company_id, record_id),
                ).fetchone()
                if attached:
                    continue
                warning = (
                    "Esta ficha NÃO substitui a norma integral. A cópia integral é comercial/licenciada "
                    "e deve ser anexada somente se a SECCOL possuir licença válida."
                    if norm["license"] == "Comercial/licenciada"
                    else "Esta ficha NÃO substitui a leitura do ato oficial e de suas alterações posteriores."
                )
                reference_sheet = (
                    f"FICHA DE REFERÊNCIA NORMATIVA — SIVS SECCOL 2.2\n\n"
                    f"Código: {norm['code']}\nOrganismo: {norm['organization']}\n"
                    f"Edição: {norm['edition']}\nSituação cadastrada: {norm['status']}\n"
                    f"Verificação do catálogo: 15/08/2026\n\nESCOPO RESUMIDO\n{norm['scope']}\n\n"
                    f"APLICABILIDADE SECCOL\n{norm['application']}\n\n"
                    f"ENSAIOS/CONTROLES RELACIONADOS\n{norm['tests']}\n\n"
                    f"REFERÊNCIA OFICIAL\n{norm['url']}\n\nCONTROLE DOCUMENTAL\n{warning}\n"
                    "Antes de assinar laudo, estudo ou certificado, o responsável técnico deve verificar "
                    "a edição vigente, emendas, erratas, requisitos do contrato, método aprovado, evidências, "
                    "incerteza quando aplicável e regra de decisão.\n"
                ).encode("utf-8")
                db.execute(
                    """INSERT INTO attachments
                       (company_id,record_id,filename,mime_type,content,size,category,version,
                        created_at,sha256,license_confirmed)
                       VALUES(?,?,?,?,?,?,?,?,?,?,0)""",
                    (current_company_id, record_id, f"ficha-{norm['key']}.txt", "text/plain; charset=utf-8",
                     reference_sheet, len(reference_sheet), "Ficha de referência normativa",
                     norm["edition"], now, hashlib.sha256(reference_sheet).hexdigest()),
                )
        self.commit_if_outer()

    def seed_seccol_portfolio(self, company_id=None):
        """Instala portfólio, instrumentos próprios e ensaios sem misturar execução operacional."""
        db = self.connection()
        now = utc_now()
        company_ids = [company_id] if company_id else [row["id"] for row in db.execute(
            "SELECT id FROM companies WHERE active=1").fetchall()]
        norm_codes = {item["key"]: item["code"] for item in NORM_CATALOG}

        for current_company_id in company_ids:
            norm_rows = db.execute(
                """SELECT id,json_extract(payload,'$.norm_key') norm_key FROM records
                   WHERE company_id=? AND module='normas_tecnicas' AND deleted_at IS NULL""",
                (current_company_id,),
            ).fetchall()
            norm_ids = {row["norm_key"]: row["id"] for row in norm_rows if row["norm_key"]}

            catalog_items = []
            for item in SECCOL_PRODUCT_CATALOG:
                catalog_items.append(("produtos", {
                    "catalog_key": item["key"], "codigo": item["code"], "title": item["title"],
                    "familia": item["family"], "tipo_item": item["kind"], "descricao": item["description"],
                    "unidade": "UN", "origem_operacional": "Produção/fornecimento SECCOL",
                    "fonte_oficial": item["source"], "norm_keys": item["norms"],
                }))
            for key, code, title, use, norms in SECCOL_INSTRUMENT_CATALOG:
                catalog_items.append(("instrumentos_seccol", {
                    "catalog_key": key, "codigo": code, "title": title,
                    "tipo": "Instrumento técnico próprio", "propriedade": "SECCOL",
                    "fabricante": "Conforme patrimônio e certificado", "uso_tecnico": use,
                    "controle_metrologico": "Controlar identificação, série, certificado e validade",
                    "fonte_oficial": "https://www.seccol.com.br/equipamentos.html", "norm_keys": norms,
                }))
            for key, code, title, category, description, norms in SECCOL_SERVICE_CATALOG:
                if key.startswith(("servico-projeto", "servico-monitoramento")):
                    source = "https://www.seccol.com.br/quem.html"
                elif key in {"servico-certificacao-area-limpa", "ensaio-vazao-trocas-ar",
                             "ensaio-recuperacao", "ensaio-pressao-entre-salas"}:
                    source = "https://www.seccol.com.br/area-limpa.html"
                else:
                    source = "https://www.seccol.com.br/teste-equipamento.html"
                catalog_items.append(("catalogo_servicos", {
                    "catalog_key": key, "codigo": code, "title": title, "categoria": category,
                    "tipo_servico": title, "descricao": description,
                    "origem_operacional": "Serviço/ensaio executado pela SECCOL",
                    "fonte_oficial": source, "norm_keys": norms,
                }))

            for module, item in catalog_items:
                row = db.execute(
                    """SELECT id,payload,deleted_at FROM records WHERE company_id=? AND module=?
                       AND json_extract(payload,'$.catalog_key')=? ORDER BY id LIMIT 1""",
                    (current_company_id, module, item["catalog_key"]),
                ).fetchone()
                if row and row["deleted_at"]:
                    continue
                norm_names = [norm_codes[key] for key in item["norm_keys"] if key in norm_codes]
                authoritative = {
                    key: value for key, value in item.items() if key not in {"title", "norm_keys"}
                }
                authoritative.update({
                    "catalogo_seccol": True,
                    "classificacao_catalogo": (
                        "Produto/solução" if module == "produtos"
                        else "Instrumento técnico próprio" if module == "instrumentos_seccol"
                        else "Serviço/ensaio"
                    ),
                    "normas_aplicaveis": norm_names,
                    "verificado_em": "2026-08-15",
                    "assunto": f'Portfólio SECCOL — {item["title"]}',
                    "notes": (
                        "Item confirmado pela direção como integrante da produção, fornecimento ou patrimônio "
                        "técnico da SECCOL. Confirmar configuração, modelo, série, escopo e normas aplicáveis "
                        "antes de proposta, fabricação, ensaio, laudo ou certificado."
                    ),
                })
                if row:
                    try:
                        payload = json.loads(row["payload"] or "{}")
                    except json.JSONDecodeError:
                        payload = {}
                    for key, value in authoritative.items():
                        if key in {"fonte_oficial", "verificado_em", "normas_aplicaveis",
                                   "catalogo_seccol", "classificacao_catalogo"} or not payload.get(key):
                            payload[key] = value
                    record_id = row["id"]
                    db.execute("UPDATE records SET payload=?,updated_at=? WHERE id=?",
                               (json_dumps(payload), now, record_id))
                else:
                    payload = authoritative
                    cursor = db.execute(
                        """INSERT INTO records
                           (module,title,status,payload,company_id,created_at,updated_at)
                           VALUES(?,?, 'Ativo',?,?,?,?)""",
                        (module, item["title"], json_dumps(payload), current_company_id, now, now),
                    )
                    record_id = cursor.lastrowid

                subject_id = self.ensure_subject(payload["assunto"], None, current_company_id)
                db.execute("UPDATE records SET subject_id=? WHERE id=?", (subject_id, record_id))
                if subject_id:
                    db.execute(
                        """INSERT OR IGNORE INTO record_subjects
                           (record_id,subject_id,relationship_type,is_primary,created_at)
                           VALUES(?,?,'Assunto principal',1,?)""",
                        (record_id, subject_id, now),
                    )
                for norm_key in item["norm_keys"]:
                    norm_id = norm_ids.get(norm_key)
                    if norm_id:
                        db.execute(
                            """INSERT OR IGNORE INTO record_relationships
                               (from_record_id,to_record_id,relationship_type,created_at)
                               VALUES(?,?,'Fundamentado em',?)""",
                            (record_id, norm_id, now),
                        )

                attached = db.execute(
                    """SELECT id FROM attachments WHERE company_id=? AND record_id=?
                       AND category='Ficha de portfólio SECCOL' LIMIT 1""",
                    (current_company_id, record_id),
                ).fetchone()
                if not attached:
                    reference_sheet = (
                        f"FICHA DE PORTFÓLIO SECCOL — SIVS 2.2\n\n"
                        f"Código: {item.get('codigo', '')}\nItem: {item['title']}\n"
                        f"Classificação: {authoritative['classificacao_catalogo']}\n"
                        f"Descrição/uso: {item.get('descricao') or item.get('uso_tecnico', '')}\n"
                        f"Origem operacional: {item.get('origem_operacional') or item.get('propriedade', 'SECCOL')}\n"
                        f"Normas inicialmente relacionadas: {', '.join(norm_names) or 'Definir por aplicação'}\n"
                        f"Fonte oficial: {item['fonte_oficial']}\nVerificação: 15/08/2026\n\n"
                        "CONTROLE\nA matriz normativa é inicial. O responsável técnico deve confirmar modelo, "
                        "configuração, escopo contratado, método, requisitos regulatórios e edição vigente.\n"
                    ).encode("utf-8")
                    db.execute(
                        """INSERT INTO attachments
                           (company_id,record_id,filename,mime_type,content,size,category,version,
                            created_at,sha256,license_confirmed)
                           VALUES(?,?,?,?,?,?,?,?,?,?,0)""",
                        (current_company_id, record_id, f'ficha-{item["catalog_key"]}.txt',
                         "text/plain; charset=utf-8", reference_sheet, len(reference_sheet),
                         "Ficha de portfólio SECCOL", "Catálogo 2026-08-15", now,
                         hashlib.sha256(reference_sheet).hexdigest()),
                    )
        self.commit_if_outer()

    def validate_normative_base(self, module, payload, company_id):
        if module not in NORMATIVE_REQUIRED_MODULES:
            return
        requested = []
        primary = str(payload.get("registro_relacionado") or "").strip()
        if primary:
            requested.append(primary)
        for relation in payload.get("relacionamentos") or []:
            if isinstance(relation, dict):
                requested.append(str(relation.get("record") or relation.get("registro") or ""))
        target_ids = []
        for reference in requested:
            try:
                target_ids.append(int(reference.rsplit(":", 1)[-1]))
            except (ValueError, TypeError):
                continue
        if target_ids:
            placeholders = ",".join("?" for _ in target_ids)
            exists = self.connection().execute(
                f"""SELECT 1 FROM records WHERE company_id=? AND module='normas_tecnicas'
                    AND deleted_at IS NULL AND status NOT IN ('Obsoleta','Cancelada')
                    AND id IN ({placeholders}) LIMIT 1""",
                (company_id, *target_ids),
            ).fetchone()
            if exists:
                return
        raise ValueError(
            "Base normativa obrigatória: relacione ao menos uma norma técnica vigente antes de salvar."
        )

    def scalar(self, sql: str, params=()):
        row = self.connection().execute(sql, params).fetchone()
        return row[0] if row else None

    def execute(self, sql: str, params=()) -> sqlite3.Cursor:
        db = self.connection()
        cursor = db.execute(sql, params)
        self.commit_if_outer()
        return cursor

    def audit(self, user_id, action, entity_type, entity_id=None, detail=None, company_id=None):
        if company_id is None and user_id is not None:
            membership = self.connection().execute(
                "SELECT company_id FROM company_memberships WHERE user_id=? AND active=1 ORDER BY company_id LIMIT 1",
                (user_id,)).fetchone()
            company_id = membership["company_id"] if membership else None
        self.execute(
            """INSERT INTO audit_log
               (user_id,action,entity_type,entity_id,detail,created_at,company_id)
               VALUES(?,?,?,?,?,?,?)""",
            (user_id, action, entity_type, str(entity_id) if entity_id is not None else None,
             json_dumps(detail) if detail is not None else None, utc_now(), company_id),
        )

    @staticmethod
    def _log_text(value, limit):
        text = str(value or "").replace("\r", " ").replace("\n", " ").strip()
        return text[:limit]

    def system_event(self, severity, category, event_type, message, *, company_id=None,
                     user_id=None, detail=None, request_id=None, path=None, method=None,
                     client_ip=None, user_agent=None):
        """Registra telemetria sanitizada sem segredos nem conteúdo de formulários."""
        safe_detail = None
        if detail is not None:
            safe_detail = json_dumps(detail)[:8000]
        self.execute(
            """INSERT INTO system_events
               (company_id,user_id,severity,category,event_type,message,detail,request_id,
                path,method,client_ip,user_agent,created_at)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                company_id, user_id, self._log_text(severity, 16),
                self._log_text(category, 40), self._log_text(event_type, 80),
                self._log_text(message, 500), safe_detail,
                self._log_text(request_id, 40) or None, self._log_text(path, 300) or None,
                self._log_text(method, 12) or None, self._log_text(client_ip, 80) or None,
                self._log_text(user_agent, 300) or None, utc_now(),
            ),
        )


def password_hash(password: str) -> str:
    salt = secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode(), salt, PBKDF2_ITERATIONS)
    return f"pbkdf2_sha256${PBKDF2_ITERATIONS}${base64.b64encode(salt).decode()}${base64.b64encode(digest).decode()}"


def password_verify(password: str, encoded: str) -> bool:
    try:
        algorithm, iterations, salt_text, digest_text = encoded.split("$", 3)
        if algorithm != "pbkdf2_sha256":
            return False
        salt = base64.b64decode(salt_text)
        expected = base64.b64decode(digest_text)
        actual = hashlib.pbkdf2_hmac("sha256", password.encode(), salt, int(iterations))
        return hmac.compare_digest(actual, expected)
    except (ValueError, TypeError):
        return False


class SIVSHandler(BaseHTTPRequestHandler):
    server_version = f"SIVS/{VERSION}"

    def version_string(self):
        return self.server_version

    def send_response(self, code, message=None):
        self._response_status = int(code)
        return super().send_response(code, message)

    @property
    def db(self) -> Database:
        return self.server.db  # type: ignore[attr-defined]

    def log_message(self, format, *args):
        print(f"[{self.log_date_time_string()}] {format % args}")

    def security_headers(self, frame_policy="DENY"):
        self.send_header("Cache-Control", "no-store")
        self.send_header("Pragma", "no-cache")
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("X-Frame-Options", frame_policy)
        self.send_header("Referrer-Policy", "no-referrer")
        self.send_header("Permissions-Policy", "camera=(), microphone=(), geolocation=(), payment=()")
        self.send_header("Cross-Origin-Opener-Policy", "same-origin")

    def client_ip(self):
        if os.environ.get("SIVS_TRUST_PROXY") == "1":
            forwarded = self.headers.get("X-Forwarded-For", "").split(",", 1)[0].strip()
            if forwarded:
                return forwarded[:80]
        return str(self.client_address[0])[:80]

    def allow_request(self, bucket, limit, window_seconds):
        allowed, retry_after = self.server.rate_limit(  # type: ignore[attr-defined]
            bucket, self.client_ip(), limit, window_seconds
        )
        if not allowed:
            self.send_json(
                {
                    "ok": False,
                    "error": "rate_limited",
                    "message": "Muitas tentativas. Aguarde antes de tentar novamente.",
                },
                429,
                headers={"Retry-After": str(retry_after)},
            )
        return allowed

    def session_cookie(self, value="", max_age=0):
        secure = os.environ.get("SIVS_SECURE_COOKIE") == "1"
        if os.environ.get("SIVS_TRUST_PROXY") == "1":
            secure = secure or self.headers.get("X-Forwarded-Proto", "").lower() == "https"
        attributes = [
            f"sivs_session={value}", "Path=/", "HttpOnly", "SameSite=Strict",
            f"Max-Age={int(max_age)}",
        ]
        if secure:
            attributes.append("Secure")
        return "; ".join(attributes)

    def _safe_dispatch(self, callback):
        self._response_started = False
        self._response_status = 500
        self._request_id = secrets.token_hex(8)
        started = time.perf_counter()
        try:
            return callback()
        except (BrokenPipeError, ConnectionResetError):
            return None
        except Exception:
            self.db.abort_manual_transaction()
            print(f"[ERRO {self._request_id}] Falha não tratada em {self.command} {self.path}")
            traceback.print_exc()
            session = getattr(self, "_request_session", None)
            try:
                self.db.system_event(
                    "error", "application", "unhandled_exception",
                    "Falha não tratada durante a requisição.",
                    company_id=session["company_id"] if session else None,
                    user_id=session["id"] if session else None,
                    detail={"exception": traceback.format_exc(limit=8)[:6000]},
                    request_id=self._request_id, path=self.route()[0], method=self.command,
                    client_ip=self.client_ip(), user_agent=self.headers.get("User-Agent", ""),
                )
            except Exception:
                traceback.print_exc()
            if not self._response_started:
                return self.error_json(
                    "Não foi possível concluir a operação. Informe o código de referência ao suporte.",
                    500, "internal_error", request_id=self._request_id,
                )
            return None
        finally:
            duration_ms = (time.perf_counter() - started) * 1000
            try:
                self.server.record_request(  # type: ignore[attr-defined]
                    self.command, self.route()[0], self._response_status, duration_ms,
                )
            except Exception:
                pass

    def send_json(self, data, status=200, headers=None):
        try:
            body = json_dumps(data).encode("utf-8")
        except (ValueError, TypeError):
            status = 500
            body = json_dumps({
                "ok": False, "error": "invalid_server_data",
                "message": "O servidor recusou dados não compatíveis com JSON estrito.",
                "requestId": getattr(self, "_request_id", None),
            }).encode("utf-8")
        self._response_started = True
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.security_headers()
        for key, value in (headers or {}).items():
            self.send_header(key, value)
        self.end_headers()
        self.wfile.write(body)

    def error_json(self, message, status=400, code="bad_request", request_id=None):
        payload = {"ok": False, "error": code, "message": message}
        if request_id:
            payload["requestId"] = request_id
        self.send_json(payload, status)

    def parse_json(self, max_bytes=MAX_BODY):
        try:
            length = int(self.headers.get("Content-Length", "0"))
        except ValueError:
            raise ValueError("Tamanho inválido")
        if length < 0:
            raise ValueError("Tamanho inválido")
        if length > max_bytes:
            raise ValueError(f"Conteúdo excede o limite de {max_bytes // (1024 * 1024)} MB")
        content_type = self.headers.get("Content-Type", "").split(";", 1)[0].strip().lower()
        if length and content_type != "application/json":
            raise ValueError("Envie o corpo como application/json")
        raw = self.rfile.read(length)
        if not raw:
            return {}
        try:
            value = json_loads_strict(raw)
        except (json.JSONDecodeError, UnicodeDecodeError, ValueError) as exc:
            raise ValueError(f"JSON inválido: {exc}") from None
        if not isinstance(value, dict):
            raise ValueError("O corpo deve ser um objeto JSON")
        return value

    def cookies(self):
        cookie = SimpleCookie()
        cookie.load(self.headers.get("Cookie", ""))
        return cookie

    def session(self):
        cookie = self.cookies().get("sivs_session")
        if not cookie:
            return None
        token_hash = hashlib.sha256(cookie.value.encode()).hexdigest()
        row = self.db.connection().execute(
            """SELECT s.token_hash,s.public_id,s.csrf_token,s.expires_at,s.last_activity_at,
                      s.company_id,s.ip_address,s.user_agent,
                      u.id,u.name,u.email,u.active,cm.role,cm.permissions,c.name company_name
               FROM sessions s
               JOIN users u ON u.id=s.user_id
               JOIN company_memberships cm
                 ON cm.user_id=u.id AND cm.company_id=s.company_id AND cm.active=1
               JOIN companies c ON c.id=s.company_id AND c.active=1
               WHERE s.token_hash=?""",
            (token_hash,),
        ).fetchone()
        now = int(time.time())
        idle_expired = bool(row and row["last_activity_at"] and
                            row["last_activity_at"] < now - SESSION_IDLE_SECONDS)
        if (not row or not row["active"] or not row["company_id"] or
                row["expires_at"] < now or idle_expired):
            if row:
                self.db.execute("DELETE FROM sessions WHERE token_hash=?", (token_hash,))
            return None
        if not row["last_activity_at"] or row["last_activity_at"] < now - 60:
            self.db.execute(
                "UPDATE sessions SET last_activity_at=? WHERE token_hash=?",
                (now, token_hash),
            )
        return row

    def require_auth(self, csrf=False):
        session = self.session()
        if not session:
            self.error_json("Sessão ausente ou expirada", 401, "unauthorized")
            return None
        self._request_session = session
        if csrf and not hmac.compare_digest(self.headers.get("X-CSRF-Token", ""), session["csrf_token"]):
            self.error_json("Token de segurança inválido", 403, "csrf_invalid")
            return None
        return session

    def require_admin(self, session):
        if session["role"] != "admin":
            self.error_json("Esta operação exige perfil de administrador", 403, "forbidden")
            return False
        return True

    @staticmethod
    def permission_spec(session):
        try:
            raw = session["permissions"] if "permissions" in session.keys() else "{}"
            value = json_loads_strict(raw or "{}")
            return value if isinstance(value, dict) else {}
        except (ValueError, TypeError, json.JSONDecodeError):
            return {}

    def allowed_modules(self, session, action):
        role = str(session["role"])
        base = {
            "read": ROLE_READ_MODULES,
            "write": ROLE_MODULES,
            "export": ROLE_EXPORT_MODULES,
        }.get(action, {}).get(role, set())
        allowed = set(base)
        permissions = self.permission_spec(session)
        additions = permissions.get(action, [])
        denials = permissions.get(f"deny_{action}", [])
        if isinstance(additions, list):
            allowed.update(module for module in additions if module in MODULES)
        if isinstance(denials, list):
            allowed.difference_update(module for module in denials if module in MODULES)
        # Clientes e fornecedores aparecem em uma única aba, mas continuam
        # usando os módulos físicos legados para preservar referências e
        # permissões existentes.
        allowed.discard(PARTY_MODULE)
        if any(module in allowed for module in PARTY_PHYSICAL_MODULES):
            allowed.add(PARTY_MODULE)
        return allowed

    def require_module_read(self, session, module):
        if module not in self.allowed_modules(session, "read"):
            self.error_json("Seu perfil não possui permissão para consultar este módulo", 403, "forbidden")
            return False
        return True

    def require_module_write(self, session, module):
        if module not in self.allowed_modules(session, "write"):
            self.error_json("Seu perfil não possui permissão para alterar este módulo", 403, "forbidden")
            return False
        return True

    def require_module_export(self, session, module):
        if module not in self.allowed_modules(session, "export"):
            self.error_json("Seu perfil não possui permissão para exportar este módulo", 403, "forbidden")
            return False
        return True

    @staticmethod
    def operation_is_read_only(action):
        return (action == "view_values" or action.startswith("view_")
                or action == "decide_approval")

    @classmethod
    def operation_defaults(cls, module, readable, writable, role=None):
        available = MODULE_ACTIONS.get(module, ())
        selected = set()
        if module in readable:
            selected.update(action for action in available if cls.operation_is_read_only(action))
            if role not in {"admin", "manager", "approver"}:
                selected.discard("decide_approval")
        if module in writable:
            selected.update(action for action in available if not cls.operation_is_read_only(action))
        return selected

    def allowed_operations(self, session, module):
        if module not in MODULE_ACTIONS:
            return set()
        if module == PARTY_MODULE:
            combined = set()
            for physical_module in PARTY_PHYSICAL_MODULES:
                combined.update(self.allowed_operations(session, physical_module))
            return combined.intersection(MODULE_ACTIONS[PARTY_MODULE])
        readable = self.allowed_modules(session, "read")
        writable = self.allowed_modules(session, "write")
        allowed = self.operation_defaults(module, readable, writable, str(session["role"]))
        overrides = self.permission_spec(session).get("actions")
        if isinstance(overrides, dict) and module in overrides:
            values = overrides.get(module)
            allowed = set(values) if isinstance(values, list) else set()
        valid = set(MODULE_ACTIONS[module])
        allowed.intersection_update(valid)
        if module not in readable:
            return set()
        if module not in writable:
            allowed = {action for action in allowed if self.operation_is_read_only(action)}
        if module in VALUE_SENSITIVE_MODULES and "view_values" not in allowed:
            allowed.difference_update(VALUE_DEPENDENT_ACTIONS)
        return allowed

    def require_operation(self, session, module, action):
        if action not in self.allowed_operations(session, module):
            label = MODULE_ACTION_LABELS.get(action, action.replace("_", " "))
            self.error_json(
                f"Seu acesso não permite: {label.lower()} neste módulo",
                403, "operation_forbidden",
            )
            return False
        return True

    def capabilities(self, session):
        role = str(session["role"])
        capabilities = {
            "settings": role == "admin",
            "users": role == "admin",
            "companies_create": role == "admin",
            "audit": role in {"admin", "manager"},
            "trash": role in {"admin", "manager"},
            "approvals": role in {"admin", "manager", "approver"} or bool(
                self.allowed_modules(session, "write")
            ),
            "full_backup": role == "admin",
            "import": role == "admin",
            "control_center": role == "admin",
        }
        custom = self.permission_spec(session).get("capabilities")
        if isinstance(custom, dict):
            for key, value in custom.items():
                if key in {"audit", "trash", "approvals"} and isinstance(value, bool):
                    capabilities[key] = value
        return capabilities

    def effective_permission_spec(self, role, desired, desired_capabilities=None,
                                  desired_actions=None):
        if not isinstance(desired, dict):
            raise ValueError("Permissões efetivas devem ser um objeto")
        selected = {}
        for action in ("read", "write", "export"):
            values = desired.get(action, [])
            if not isinstance(values, list) or len(values) > len(MODULES):
                raise ValueError(f"Permissões de {action} inválidas")
            if any(module not in MODULES for module in values):
                raise ValueError(f"Permissões de {action} contêm módulo inválido")
            selected[action] = set(values) - {PARTY_MODULE}
        if not selected["write"].issubset(selected["read"]):
            raise ValueError("Todo módulo editável também deve permitir leitura")
        if not selected["export"].issubset(selected["read"]):
            raise ValueError("Todo módulo exportável também deve permitir leitura")
        bases = {
            "read": set(ROLE_READ_MODULES.get(role, set())) - {PARTY_MODULE},
            "write": set(ROLE_MODULES.get(role, set())) - {PARTY_MODULE},
            "export": set(ROLE_EXPORT_MODULES.get(role, set())) - {PARTY_MODULE},
        }
        spec = {}
        for action in ("read", "write", "export"):
            additions = sorted(selected[action] - bases[action])
            denials = sorted(bases[action] - selected[action])
            if additions:
                spec[action] = additions
            if denials:
                spec[f"deny_{action}"] = denials
        if desired_capabilities is not None:
            if not isinstance(desired_capabilities, dict):
                raise ValueError("Capacidades efetivas devem ser um objeto")
            custom = {}
            for key, value in desired_capabilities.items():
                if key not in {"audit", "trash", "approvals"} or not isinstance(value, bool):
                    raise ValueError("Capacidade personalizada inválida")
                custom[key] = value
            if custom:
                spec["capabilities"] = custom
        if desired_actions is not None:
            if not isinstance(desired_actions, dict) or any(
                    module not in MODULE_ACTIONS for module in desired_actions):
                raise ValueError("Funções efetivas devem ser organizadas por módulo")
            custom_actions = {}
            for module in MODULE_ACTIONS:
                values = [] if module == PARTY_MODULE else desired_actions.get(module, [])
                if not isinstance(values, list) or len(values) > len(MODULE_ACTIONS[module]):
                    raise ValueError(f"Funções inválidas em {MODULES[module]}")
                if any(action not in MODULE_ACTIONS[module] for action in values):
                    raise ValueError(f"Função desconhecida em {MODULES[module]}")
                selected_actions = set(values)
                if (module in VALUE_SENSITIVE_MODULES
                        and module in selected["export"]
                        and "view_values" not in selected_actions):
                    raise ValueError(
                        f"{MODULES[module]} exige visualização de valores para exportar"
                    )
                for action in selected_actions:
                    if module not in selected["read"]:
                        raise ValueError(f"{MODULES[module]} exige permissão de consulta")
                    if not self.operation_is_read_only(action) and module not in selected["write"]:
                        raise ValueError(f"{MODULES[module]} exige permissão de edição")
                    if (module in VALUE_SENSITIVE_MODULES
                            and action in VALUE_DEPENDENT_ACTIONS
                            and "view_values" not in selected_actions):
                        raise ValueError(
                            f"{MODULES[module]} exige visualização de valores para criar ou editar"
                        )
                defaults = self.operation_defaults(
                    module, selected["read"], selected["write"], role,
                )
                if selected_actions != defaults:
                    custom_actions[module] = sorted(selected_actions)
            if custom_actions:
                spec["actions"] = custom_actions
        return spec

    def access_control_catalog(self):
        categories = []
        for key, label, modules in ACCESS_CATEGORIES:
            categories.append({
                "key": key,
                "label": label,
                "modules": [{
                    "key": module,
                    "label": MODULES[module],
                    "readOnly": module in READ_ONLY_MODULES,
                    "actions": [{"key": action, "label": MODULE_ACTION_LABELS[action]}
                                for action in MODULE_ACTIONS[module]],
                } for module in modules],
            })
        role_defaults = {}
        for role in ROLE_MODULES:
            readable = set(ROLE_READ_MODULES.get(role, set())) - {PARTY_MODULE}
            writable = set(ROLE_MODULES.get(role, set())) - {PARTY_MODULE}
            role_defaults[role] = {
                "permissions": {
                    "read": sorted(readable),
                    "write": sorted(writable),
                    "export": sorted(set(ROLE_EXPORT_MODULES.get(role, set())) - {PARTY_MODULE}),
                },
                "actions": {
                    module: sorted(self.operation_defaults(module, readable, writable, role))
                    for module in MODULE_ACTIONS if module != PARTY_MODULE
                },
                "capabilities": {
                    "audit": role in {"admin", "manager"},
                    "trash": role in {"admin", "manager"},
                    "approvals": role in {"admin", "manager", "approver"} or bool(writable),
                },
            }
        return {"categories": categories, "roleDefaults": role_defaults}

    def route(self):
        parsed = urlparse(self.path)
        return parsed.path.rstrip("/") or "/", parse_qs(parsed.query)

    def do_GET(self):
        return self._safe_dispatch(lambda: self._do_get())

    def _do_get(self):
        path, query = self.route()
        if path.startswith("/api/"):
            return self.api_get(path, query)
        return self.static_get(path)

    def do_POST(self):
        return self._safe_dispatch(lambda: self.api_write("POST", self.route()[0]))

    def do_PUT(self):
        return self._safe_dispatch(lambda: self.api_write("PUT", self.route()[0]))

    def do_DELETE(self):
        return self._safe_dispatch(lambda: self.api_write("DELETE", self.route()[0]))

    def api_get(self, path, query):
        if path == "/api/status":
            configured = bool(self.db.scalar("SELECT configured FROM setup_state WHERE id=1"))
            return self.send_json({"ok": True, "configured": configured, "version": VERSION})
        if path == "/api/me":
            session = self.require_auth()
            if not session:
                return
            return self.send_json({"ok": True, "user": self.user_json(session),
                                   "csrfToken": session["csrf_token"],
                                   "capabilities": self.capabilities(session)})
        session = self.require_auth()
        if not session:
            return
        company_id = session["company_id"]
        if path == "/api/control-center":
            return self.control_center_get(session, query)
        if path == "/api/companies":
            return self.companies_get(session)
        if path == "/api/branches":
            rows = self.db.connection().execute(
                """SELECT id,company_id,code,name,cnpj,address,active,is_headquarters
                   FROM branches WHERE company_id=? ORDER BY is_headquarters DESC,name""",
                (company_id,),
            ).fetchall()
            return self.send_json({"ok": True, "items": [dict(row) for row in rows]})
        if path == "/api/inventory":
            return self.inventory_get(session, query)
        if path == "/api/management/overview":
            return self.management_overview(session)
        if path == "/api/partners/options":
            readable = [module for module in PARTY_PHYSICAL_MODULES
                        if module in self.allowed_modules(session, "read")]
            if not readable:
                return self.send_json({"ok": True, "items": [], "counts": {"C": 0, "F": 0, "A": 0}})
            placeholders = ",".join("?" for _ in readable)
            rows = self.db.connection().execute(
                f"""SELECT id,module,title,status,
                            COALESCE(json_extract(payload,'$.tipo_cadastro'),
                              CASE WHEN module='fornecedores' THEN 'F' ELSE 'C' END) party_type,
                            json_extract(payload,'$.codigo_cadastro') code,
                            json_extract(payload,'$.documento') document
                     FROM records
                     WHERE company_id=? AND deleted_at IS NULL AND module IN ({placeholders})
                     ORDER BY title COLLATE NOCASE,id LIMIT 5000""",
                (company_id, *readable),
            ).fetchall()
            items = [dict(row) for row in rows]
            counts = {role: sum(1 for item in items if item["party_type"] == role)
                      for role in ("C", "F", "A")}
            return self.send_json({"ok": True, "items": items, "counts": counts})
        if path == "/api/notifications":
            rows = self.db.connection().execute(
                """SELECT * FROM notifications
                   WHERE company_id=? AND (user_id IS NULL OR user_id=?)
                   ORDER BY read_at IS NULL DESC,id DESC LIMIT 100""",
                (company_id, session["id"])).fetchall()
            return self.send_json({"ok": True, "items": [dict(row) for row in rows]})
        if path == "/api/approvals":
            readable = sorted(self.allowed_modules(session, "read"))
            if not readable:
                return self.send_json({"ok": True, "items": []})
            status = (query.get("status") or ["Pendente"])[0].strip()
            placeholders = ",".join("?" for _ in readable)
            sql = """SELECT a.*,r.module,r.title,u0.name requested_by_name,
                            u1.name requested_to_name,u2.name decided_by_name
                     FROM approvals a JOIN records r ON r.id=a.record_id
                     LEFT JOIN users u0 ON u0.id=a.requested_by
                     LEFT JOIN users u1 ON u1.id=a.requested_to
                     LEFT JOIN users u2 ON u2.id=a.decided_by
                     WHERE a.company_id=? AND r.deleted_at IS NULL
                     AND (r.module IN (""" + placeholders + ") OR a.requested_to=?)"
            params = [company_id, *readable, session["id"]]
            if session["role"] not in {"admin", "manager", "approver"}:
                sql += " AND (a.requested_by=? OR a.requested_to=?)"
                params.extend([session["id"], session["id"]])
            if status:
                sql += " AND a.status=?"
                params.append(status)
            sql += " ORDER BY CASE a.status WHEN 'Pendente' THEN 0 ELSE 1 END,a.id DESC LIMIT 300"
            rows = self.db.connection().execute(sql, params).fetchall()
            items = []
            for row in rows:
                item = dict(row)
                item["can_decide"] = self.approval_can_decide(session, item)
                items.append(item)
            return self.send_json({"ok": True, "items": items})
        if path == "/api/fiscal/readiness":
            return self.fiscal_readiness(session)
        if path == "/api/accounting/export":
            return self.accounting_export(query, session)
        if path == "/api/fiscal/events":
            if not self.require_module_read(session, "fiscal"):
                return
            record_id = (query.get("record_id") or [""])[0]
            sql = """SELECT f.*,r.title,u.name user_name FROM fiscal_events f
                     JOIN records r ON r.id=f.record_id LEFT JOIN users u ON u.id=f.created_by
                     WHERE f.company_id=?"""
            params = [company_id]
            if str(record_id).isdigit():
                sql += " AND f.record_id=?"
                params.append(int(record_id))
            sql += " ORDER BY f.id DESC LIMIT 300"
            rows = self.db.connection().execute(sql, params).fetchall()
            return self.send_json({"ok": True, "items": [dict(row) for row in rows]})
        if path.startswith("/api/attachments/") and path.rsplit("/", 1)[-1].isdigit():
            return self.attachment_download(int(path.rsplit("/", 1)[-1]), session)
        if (path.startswith("/api/reports/") and path.endswith("/preview") and
                path.split("/")[3].isdigit()):
            return self.technical_report_preview(int(path.split("/")[3]), session)
        if path == "/api/modules":
            readable = self.allowed_modules(session, "read")
            return self.send_json({
                "ok": True,
                "modules": {key: value for key, value in MODULES.items() if key in readable},
                "readableModules": sorted(readable),
                "writableModules": sorted(self.allowed_modules(session, "write")),
                "exportableModules": sorted(self.allowed_modules(session, "export")),
                "actionPermissions": {
                    module: sorted(self.allowed_operations(session, module))
                    for module in readable if module in MODULE_ACTIONS
                },
                "capabilities": self.capabilities(session),
            })
        if path == "/api/dashboard":
            return self.dashboard(session)
        if path == "/api/search":
            return self.global_search(query, session)
        if path == "/api/partner-lookup":
            return self.partner_lookup(query, session)
        if path == "/api/assistant/capabilities":
            readable = sorted(self.allowed_modules(session, "read"))
            return self.send_json({
                "ok": True,
                "aiConfigured": bool(os.environ.get("OPENROUTER_API_KEY")),
                "readableModules": readable,
            })
        if path == "/api/settings":
            if not self.require_admin(session):
                return
            company = self.db.connection().execute(
                """SELECT c.id,c.name,c.cnpj,c.phone,c.email,c.address,c.active,
                          c.holding_id,h.name holding_name
                   FROM companies c LEFT JOIN holdings h ON h.id=c.holding_id
                   WHERE c.id=?""", (company_id,)).fetchone()
            rows = self.db.connection().execute(
                "SELECT key,value FROM company_settings WHERE company_id=?", (company_id,)).fetchall()
            settings = {row["key"]: json.loads(row["value"]) for row in rows}
            settings["company"] = dict(company) if company else {}
            settings["branches"] = [dict(row) for row in self.db.connection().execute(
                """SELECT id,code,name,cnpj,address,active,is_headquarters
                   FROM branches WHERE company_id=? ORDER BY is_headquarters DESC,name""",
                (company_id,),
            ).fetchall()]
            return self.send_json({"ok": True, "settings": settings})
        if path == "/api/audit":
            if not self.capabilities(session)["audit"]:
                return self.error_json("Seu perfil não consulta a trilha de auditoria", 403, "forbidden")
            rows = self.db.connection().execute(
                """SELECT a.*,u.name user_name FROM audit_log a LEFT JOIN users u ON u.id=a.user_id
                   WHERE a.company_id=? ORDER BY a.id DESC LIMIT 100""", (company_id,)
            ).fetchall()
            return self.send_json({"ok": True, "items": [dict(row) for row in rows]})
        if path == "/api/users":
            if not self.require_admin(session):
                return
            rows = self.db.connection().execute(
                """SELECT u.id,u.name,u.email,cm.role,cm.active,u.created_at,u.updated_at,cm.permissions
                   FROM company_memberships cm JOIN users u ON u.id=cm.user_id
                   WHERE cm.company_id=? ORDER BY u.name""", (company_id,)
            ).fetchall()
            items = []
            for row in rows:
                item = dict(row)
                item["permissions"] = self.permission_spec(row)
                item["effective_permissions"] = {
                    action: sorted(module for module in self.allowed_modules(row, action)
                                   if module in MODULES and module != PARTY_MODULE)
                    for action in ("read", "write", "export")
                }
                item["effective_capabilities"] = {
                    key: self.capabilities(row)[key]
                    for key in ("audit", "trash", "approvals")
                }
                item["effective_actions"] = {
                    module: sorted(self.allowed_operations(row, module))
                    for module in MODULE_ACTIONS if module != PARTY_MODULE
                }
                items.append(item)
            return self.send_json({
                "ok": True, "items": items,
                "accessControl": self.access_control_catalog(),
            })
        if path == "/api/trash":
            if not self.capabilities(session)["trash"]:
                return self.error_json("Seu perfil não consulta a lixeira", 403, "forbidden")
            readable = sorted(self.allowed_modules(session, "read"))
            if not readable:
                return self.send_json({"ok": True, "items": []})
            placeholders = ",".join("?" for _ in readable)
            rows = self.db.connection().execute(
                f"""SELECT * FROM records WHERE company_id=? AND deleted_at IS NOT NULL
                   AND module IN ({placeholders}) ORDER BY deleted_at DESC LIMIT 200""",
                (company_id, *readable),
            ).fetchall()
            self.db.audit(session["id"], "read", "trash", detail={"count": len(rows)}, company_id=company_id)
            return self.send_json({"ok": True, "items": self.records_json(rows, session)})
        if path == "/api/tenders/sources":
            # O catálogo sustenta o painel de editais. Exigir também acesso ao
            # módulo administrativo "fontes" quebrava a tela para perfis
            # comerciais com permissão explícita apenas em editais.
            if not self.require_module_read(session, "editais"):
                return
            rows = self.db.connection().execute(
                """SELECT * FROM records WHERE company_id=? AND module='fontes'
                   AND deleted_at IS NULL ORDER BY title""", (company_id,)
            ).fetchall()
            return self.send_json({"ok": True, "items": self.records_json(rows),
                                   "defaultKeywords": DEFAULT_TENDER_KEYWORDS})
        if path == "/api/tenders/results":
            if not self.require_module_read(session, "editais"):
                return
            return self.tender_results_get(query, session)
        if path == "/api/competitors/insights":
            return self.competitor_insights(session)
        if path.startswith("/api/tenders/results/"):
            if not self.require_module_read(session, "editais"):
                return
            return self.tender_result_get(path, session)
        if path.startswith("/api/tenders/jobs/") and path.rsplit("/", 1)[-1].isdigit():
            if not self.require_module_read(session, "editais"):
                return
            return self.tender_job_get(int(path.rsplit("/", 1)[-1]), session)
        if path == "/api/tenders/history":
            if not self.require_module_read(session, "editais"):
                return
            rows = self.db.connection().execute(
                "SELECT * FROM tender_searches WHERE company_id=? ORDER BY id DESC LIMIT 50", (company_id,)
            ).fetchall()
            return self.send_json({"ok": True, "items": [dict(row) for row in rows]})
        if path == "/api/tenders/schedules":
            if not self.require_module_read(session, "editais"):
                return
            rows = self.db.connection().execute(
                "SELECT * FROM search_schedules WHERE company_id=? ORDER BY active DESC,name", (company_id,)
            ).fetchall()
            return self.send_json({"ok": True, "items": [dict(row) for row in rows]})
        if path == "/api/relations/options":
            readable = sorted(self.allowed_modules(session, "read") - {"fontes"})
            if not readable:
                return self.send_json({"ok": True, "items": []})
            placeholders = ",".join("?" for _ in readable)
            rows = self.db.connection().execute(
                f"""SELECT id,module,title,status,
                            CASE WHEN module IN ('clientes','fornecedores')
                                 THEN COALESCE(json_extract(payload,'$.tipo_cadastro'),
                                      CASE WHEN module='fornecedores' THEN 'F' ELSE 'C' END)
                                 ELSE NULL END party_type,
                            CASE WHEN module IN ('clientes','fornecedores')
                                 THEN json_extract(payload,'$.codigo_cadastro') ELSE NULL END code,
                            CASE WHEN module IN ('clientes','fornecedores')
                                 THEN json_extract(payload,'$.documento') ELSE NULL END document
                     FROM records
                   WHERE company_id=? AND deleted_at IS NULL AND module IN ({placeholders})
                   ORDER BY CASE WHEN module='normas_tecnicas' THEN 0 ELSE 1 END,
                            updated_at DESC LIMIT 3000""", (company_id, *readable)
            ).fetchall()
            return self.send_json({"ok": True, "items": [dict(row) for row in rows]})
        if path == "/api/subjects":
            readable = sorted(self.allowed_modules(session, "read"))
            if not readable:
                return self.send_json({"ok": True, "items": []})
            placeholders = ",".join("?" for _ in readable)
            search = (query.get("q") or [""])[0].strip()
            sql = f"""SELECT s.id,s.name,s.status,s.created_at,s.updated_at,COUNT(DISTINCT r.id) record_count,
                      MAX(r.updated_at) last_activity
                      FROM subjects s
                      LEFT JOIN record_subjects rs ON rs.subject_id=s.id
                      LEFT JOIN records r ON r.id=rs.record_id AND r.company_id=? AND r.deleted_at IS NULL
                       AND r.module IN ({placeholders})
                      WHERE s.company_id=?"""
            params = [company_id, *readable, company_id]
            if search:
                sql += " AND s.name LIKE ?"
                params.append(f"%{search}%")
            sql += " GROUP BY s.id ORDER BY COALESCE(MAX(r.updated_at),s.updated_at) DESC LIMIT 500"
            rows = self.db.connection().execute(sql, params).fetchall()
            return self.send_json({"ok": True, "items": [dict(row) for row in rows]})
        if path.startswith("/api/subjects/") and path.rsplit("/", 1)[-1].isdigit():
            subject_id = int(path.rsplit("/", 1)[-1])
            subject = self.db.connection().execute(
                "SELECT * FROM subjects WHERE id=? AND company_id=?", (subject_id, company_id)).fetchone()
            if not subject:
                return self.error_json("Assunto não encontrado", 404)
            readable = sorted(self.allowed_modules(session, "read"))
            if not readable:
                rows = []
            else:
                placeholders = ",".join("?" for _ in readable)
                rows = self.db.connection().execute(
                    f"""SELECT DISTINCT r.* FROM records r JOIN record_subjects rs ON rs.record_id=r.id
                       WHERE rs.subject_id=? AND r.company_id=? AND r.deleted_at IS NULL
                       AND r.module IN ({placeholders}) ORDER BY r.updated_at DESC""",
                    (subject_id, company_id, *readable)).fetchall()
            return self.send_json({"ok": True, "subject": dict(subject),
                                   "records": self.records_json(rows, session)})
        if path == "/api/backup":
            return self.error_json(
                "Use POST com uma senha de proteção para gerar o backup criptografado",
                405, "method_not_allowed",
            )
        if path == "/api/export":
            return self.export_data(query, session)
        item_match = re.fullmatch(r"/api/records/(\d+)/items", path)
        if item_match:
            return self.record_items_get(int(item_match.group(1)), session)
        if path.startswith("/api/records"):
            return self.records_get(path, query, session)
        return self.error_json("Rota não encontrada", 404, "not_found")

    def api_write(self, method, path):
        if method == "POST" and path == "/api/setup":
            return self.initial_setup()
        if method == "POST" and path == "/api/login":
            return self.login()
        session = self.require_auth(csrf=True)
        if not session:
            return
        read_only_allowed = {
            "/api/logout", "/api/company/switch", "/api/notifications/read",
            "/api/assistant/query", "/api/telemetry/client-error",
            "/api/tenders/keywords/import",
        }
        if session["role"] == "viewer" and path not in read_only_allowed:
            return self.error_json("Perfil de consulta não pode alterar dados", 403, "read_only")
        if method == "POST" and path == "/api/logout":
            token = self.cookies().get("sivs_session")
            if token:
                self.db.audit(session["id"], "logout", "session", company_id=session["company_id"])
                self.db.execute("DELETE FROM sessions WHERE token_hash=?", (hashlib.sha256(token.value.encode()).hexdigest(),))
            headers = {"Set-Cookie": self.session_cookie()}
            return self.send_json({"ok": True}, headers=headers)
        if method == "POST" and path == "/api/company/switch":
            return self.company_switch(session)
        if method == "POST" and path == "/api/companies":
            if not self.require_admin(session):
                return
            return self.company_create(session)
        if method == "POST" and path == "/api/branches":
            if not self.require_admin(session):
                return
            return self.branch_create(session)
        if method == "POST" and path == "/api/inventory/warehouses":
            return self.inventory_warehouse_create(session)
        if method == "POST" and path == "/api/inventory/movements":
            return self.inventory_movement_create(session)
        if method == "POST" and path == "/api/inventory/reservations":
            return self.inventory_reservation_create(session)
        if (method == "POST" and path.startswith("/api/inventory/reservations/")
                and path.endswith("/release")):
            return self.inventory_reservation_release(path, session)
        reserve_match = re.fullmatch(
            r"/api/records/(\d+)/(reserve-items|release-items|fulfill-items)", path,
        )
        if method == "POST" and reserve_match:
            return self.record_items_reservation(
                int(reserve_match.group(1)), reserve_match.group(2), session,
            )
        receive_match = re.fullmatch(r"/api/records/(\d+)/receive-items", path)
        if method == "POST" and receive_match:
            return self.record_items_receive(int(receive_match.group(1)), session)
        item_collection = re.fullmatch(r"/api/records/(\d+)/items", path)
        if method == "POST" and item_collection:
            return self.record_item_create(int(item_collection.group(1)), session)
        item_member = re.fullmatch(r"/api/records/(\d+)/items/(\d+)", path)
        if method in {"PUT", "DELETE"} and item_member:
            return self.record_item_write(
                method, int(item_member.group(1)), int(item_member.group(2)), session,
            )
        if method == "POST" and path == "/api/notifications/read":
            return self.notifications_read(session)
        if method == "POST" and path == "/api/assistant/query":
            return self.assistant_query(session)
        if method == "POST" and path == "/api/telemetry/client-error":
            return self.client_error_report(session)
        if method == "DELETE" and path.startswith("/api/control-center/sessions/"):
            return self.control_center_session_delete(path, session)
        if method == "POST" and path.startswith("/api/control-center/events/") and path.endswith("/resolve"):
            return self.control_center_event_resolve(path, session)
        if method == "PUT" and path == "/api/settings":
            if not self.require_admin(session):
                return
            return self.settings_update(session)
        if method == "PUT" and path == "/api/fiscal/configuration":
            return self.fiscal_configuration_update(session)
        if method == "POST" and path == "/api/fiscal/certificate":
            return self.fiscal_certificate_upload(session)
        if method == "DELETE" and re.fullmatch(r"/api/fiscal/certificate/\d+", path):
            return self.fiscal_certificate_delete(path, session)
        if method == "POST" and path == "/api/fiscal/sefaz/status":
            return self.fiscal_sefaz_status(session)
        if method == "POST" and path == "/api/backup":
            if not self.capabilities(session)["full_backup"]:
                return self.error_json("O backup de desastre exige administrador", 403, "forbidden")
            return self.database_backup(session)
        if method == "POST" and path == "/api/users":
            return self.user_create(session)
        if (method == "POST" and path.startswith("/api/users/")
                and path.endswith("/password")):
            return self.user_password_reset(path, session)
        if method == "PUT" and path.startswith("/api/users/"):
            return self.user_update(path, session)
        if method == "DELETE" and (path == "/api/trash" or path.startswith("/api/trash/")):
            return self.trash_purge(path, session)
        if method == "POST" and path.startswith("/api/restore/"):
            return self.record_restore(path, session)
        if method == "POST" and path == "/api/tenders/search":
            if not self.require_operation(session, "editais", "search_tenders"):
                return
            return self.tender_search(session)
        if method == "POST" and path == "/api/tenders/keywords/import":
            if not self.require_module_read(session, "editais"):
                return
            return self.tender_keywords_import(session)
        if method == "POST" and path == "/api/tenders/schedules":
            if not self.require_operation(session, "editais", "manage_tender_schedules"):
                return
            return self.search_schedule_save(session)
        if method == "PUT" and path.startswith("/api/tenders/results/"):
            if not self.require_operation(session, "editais", "triage_tenders"):
                return
            return self.tender_result_update(path, session)
        if method == "POST" and path.startswith("/api/tenders/results/") and path.endswith("/refresh"):
            if not self.require_operation(session, "editais", "triage_tenders"):
                return
            return self.tender_result_refresh(path, session)
        if method == "POST" and path.startswith("/api/tenders/results/") and path.endswith("/analyze"):
            if not self.require_operation(session, "editais", "triage_tenders"):
                return
            return self.tender_result_analyze(path, session)
        if method == "POST" and path.startswith("/api/tenders/convert/"):
            if (not self.require_operation(session, "editais", "convert_tender") or
                    not self.require_operation(session, "licitacoes", "create")):
                return
            return self.tender_convert(path, session)
        if method == "POST" and path == "/api/xml/import":
            if not self.require_module_write(session, "importacoes_xml"):
                return
            return self.xml_import(session)
        if method == "POST" and path.startswith("/api/records/") and path.endswith("/attachments"):
            return self.attachment_upload(path, session)
        if method == "POST" and path.startswith("/api/records/") and path.endswith("/approval"):
            return self.approval_create(path, session)
        if method == "POST" and path.startswith("/api/approvals/"):
            return self.approval_decide(path, session)
        if (method == "POST" and path.startswith("/api/reports/") and path.endswith("/issue") and
                path.split("/")[3].isdigit()):
            return self.technical_report_issue(int(path.split("/")[3]), session)
        if method == "POST" and path.startswith("/api/fiscal/"):
            if not self.require_module_write(session, "fiscal"):
                return
            return self.fiscal_action(path, session)
        if method == "POST" and path.startswith("/api/subjects/"):
            if session["role"] not in {"admin", "manager"}:
                return self.error_json(
                    "Somente administradores e gestores podem reorganizar assuntos", 403, "forbidden")
            return self.subject_action(path, session)
        if path.startswith("/api/records"):
            return self.records_write(method, path, session)
        if method == "POST" and path == "/api/import":
            if not self.require_admin(session):
                return
            return self.import_data(session)
        return self.error_json("Rota não encontrada", 404, "not_found")

    def user_json(self, row):
        keys = set(row.keys())
        company_id = row["company_id"] if "company_id" in keys else None
        memberships = self.db.connection().execute(
            """SELECT c.id,c.name,c.cnpj,c.holding_id holdingId,h.name holdingName,cm.role
               FROM company_memberships cm JOIN companies c ON c.id=cm.company_id
               LEFT JOIN holdings h ON h.id=c.holding_id
               WHERE cm.user_id=? AND cm.active=1 AND c.active=1 ORDER BY c.name""",
            (row["id"],)).fetchall()
        current = next((item for item in memberships if item["id"] == company_id), None)
        return {
            "id": row["id"], "name": row["name"], "email": row["email"], "role": row["role"],
            "companyId": company_id, "companyName": current["name"] if current else None,
            "companies": [dict(item) for item in memberships],
        }

    def initial_setup(self):
        if not self.allow_request("setup", 8, 10 * 60):
            return
        try:
            data = self.parse_json()
        except ValueError as exc:
            return self.error_json(str(exc))
        name = str(data.get("name", "")).strip()
        email = str(data.get("email", "")).strip().lower()
        password = str(data.get("password", ""))
        company = str(data.get("company", "SIVS")).strip() or "SIVS"
        if len(name) < 2 or not re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", email) or len(password) < 10:
            return self.error_json("Informe nome, e-mail válido e senha com pelo menos 10 caracteres")
        now = utc_now()
        try:
            with self.db.transaction(immediate=True):
                setup = self.db.connection().execute(
                    "SELECT configured FROM setup_state WHERE id=1"
                ).fetchone()
                if (setup and setup["configured"]) or self.db.scalar("SELECT COUNT(*) FROM users"):
                    return self.error_json(
                        "O administrador inicial já foi criado", 409, "already_configured"
                    )
                default_company = self.db.connection().execute(
                    "SELECT id FROM companies ORDER BY id LIMIT 1"
                ).fetchone()
                if not default_company:
                    raise sqlite3.IntegrityError("empresa base ausente")
                company_id = default_company["id"]
                self.db.execute(
                    "UPDATE companies SET name=?,updated_at=? WHERE id=?",
                    (company, now, company_id),
                )
                self.db.execute(
                    """UPDATE holdings SET name=?,updated_at=? WHERE id=(
                         SELECT holding_id FROM companies WHERE id=?
                       ) AND name='Holding principal'""",
                    (f"Holding {company}", now, company_id),
                )
                self.db.execute(
                    """UPDATE branches SET name=?,updated_at=?
                       WHERE company_id=? AND is_headquarters=1""",
                    (f"Matriz — {company}", now, company_id),
                )
                cursor = self.db.execute(
                    """INSERT INTO users(name,email,password_hash,role,created_at,updated_at)
                       VALUES(?,?,?,?,?,?)""",
                    (name, email, password_hash(password), "admin", now, now),
                )
                user_id = cursor.lastrowid
                self.db.execute(
                    """INSERT INTO company_memberships
                       (company_id,user_id,role,permissions,active,created_at,updated_at)
                       VALUES(?,?,'admin','{}',1,?,?)""",
                    (company_id, user_id, now, now),
                )
                updated = self.db.execute(
                    """UPDATE setup_state SET configured=1,configured_at=?
                       WHERE id=1 AND configured=0""",
                    (now,),
                )
                if updated.rowcount != 1:
                    raise sqlite3.IntegrityError("configuração concorrente")
                self.db.seed_sources(company_id)
                self.db.seed_norms(company_id)
                self.db.seed_seccol_portfolio(company_id)
                self.db.audit(
                    user_id, "setup", "system", detail={"company": company},
                    company_id=company_id,
                )
        except sqlite3.IntegrityError:
            return self.error_json(
                "O administrador inicial já foi criado", 409, "already_configured"
            )
        return self.create_session(user_id, company_id)

    def login(self):
        if not self.allow_request("login", 12, 5 * 60):
            return
        try:
            data = self.parse_json()
        except ValueError as exc:
            return self.error_json(str(exc))
        email = str(data.get("email", "")).strip().lower()
        password = str(data.get("password", ""))
        row = self.db.connection().execute("SELECT * FROM users WHERE email=? AND active=1", (email,)).fetchone()
        if not row or not password_verify(password, row["password_hash"]):
            company_id = None
            if row:
                membership = self.db.connection().execute(
                    "SELECT company_id FROM company_memberships WHERE user_id=? AND active=1 ORDER BY company_id LIMIT 1",
                    (row["id"],),
                ).fetchone()
                company_id = membership["company_id"] if membership else None
            self.db.system_event(
                "warning", "security", "login_failed", "Tentativa de acesso rejeitada.",
                company_id=company_id, user_id=row["id"] if row else None,
                path="/api/login", method="POST", client_ip=self.client_ip(),
                user_agent=self.headers.get("User-Agent", ""),
            )
            time.sleep(0.2)
            return self.error_json("E-mail ou senha incorretos", 401, "invalid_credentials")
        requested_company = data.get("company_id")
        membership = None
        if requested_company:
            try:
                requested_company = int(requested_company)
            except (ValueError, TypeError):
                return self.error_json("Empresa inválida")
            membership = self.db.connection().execute(
                """SELECT cm.company_id FROM company_memberships cm JOIN companies c ON c.id=cm.company_id
                   WHERE cm.user_id=? AND cm.company_id=? AND cm.active=1 AND c.active=1""",
                (row["id"], requested_company)).fetchone()
        if not membership:
            membership = self.db.connection().execute(
                """SELECT cm.company_id FROM company_memberships cm JOIN companies c ON c.id=cm.company_id
                   WHERE cm.user_id=? AND cm.active=1 AND c.active=1 ORDER BY cm.company_id LIMIT 1""",
                (row["id"],)).fetchone()
        if not membership:
            return self.error_json("Usuário sem empresa ativa vinculada", 403, "no_company")
        self.db.audit(row["id"], "login", "session", company_id=membership["company_id"])
        return self.create_session(row["id"], membership["company_id"])

    def create_session(self, user_id, company_id=None):
        if company_id is None:
            membership = self.db.connection().execute(
                "SELECT company_id FROM company_memberships WHERE user_id=? AND active=1 ORDER BY company_id LIMIT 1",
                (user_id,)).fetchone()
            if not membership:
                return self.error_json("Usuário sem empresa vinculada", 403, "no_company")
            company_id = membership["company_id"]
        raw_token = secrets.token_urlsafe(36)
        token_hash = hashlib.sha256(raw_token.encode()).hexdigest()
        csrf_token = secrets.token_urlsafe(24)
        public_id = secrets.token_hex(12)
        expires = int(time.time()) + SESSION_SECONDS
        now_epoch = int(time.time())
        self.db.execute("DELETE FROM sessions WHERE expires_at < ?", (int(time.time()),))
        self.db.execute(
            """INSERT INTO sessions
               (token_hash,user_id,csrf_token,expires_at,created_at,company_id,public_id,
                last_activity_at,ip_address,user_agent)
               VALUES(?,?,?,?,?,?,?,?,?,?)""",
            (
                token_hash, user_id, csrf_token, expires, utc_now(), company_id, public_id,
                now_epoch, self.client_ip(), str(self.headers.get("User-Agent", ""))[:300],
            ),
        )
        row = self.db.connection().execute(
            """SELECT u.id,u.name,u.email,u.active,s.company_id,cm.role,c.name company_name
               FROM users u JOIN sessions s ON s.user_id=u.id
               JOIN company_memberships cm ON cm.user_id=u.id AND cm.company_id=s.company_id
               JOIN companies c ON c.id=s.company_id
               WHERE u.id=? AND s.token_hash=?""", (user_id, token_hash)).fetchone()
        cookie = self.session_cookie(raw_token, SESSION_SECONDS)
        return self.send_json({"ok": True, "user": self.user_json(row), "csrfToken": csrf_token},
                              headers={"Set-Cookie": cookie})

    def companies_get(self, session):
        rows = self.db.connection().execute(
            """SELECT c.id,c.name,c.cnpj,c.phone,c.email,c.address,c.active,
                      c.holding_id,h.name holding_name,cm.role
               FROM company_memberships cm JOIN companies c ON c.id=cm.company_id
               LEFT JOIN holdings h ON h.id=c.holding_id
               WHERE cm.user_id=? AND cm.active=1 AND c.active=1 ORDER BY c.name""",
            (session["id"],)).fetchall()
        branches = self.db.connection().execute(
            """SELECT id,company_id,code,name,cnpj,address,active,is_headquarters
               FROM branches WHERE company_id=? ORDER BY is_headquarters DESC,name""",
            (session["company_id"],),
        ).fetchall()
        return self.send_json({"ok": True, "currentCompanyId": session["company_id"],
                               "items": [dict(row) for row in rows],
                               "branches": [dict(row) for row in branches]})

    def company_create(self, session):
        try:
            data = self.parse_json()
        except ValueError as exc:
            return self.error_json(str(exc))
        name = str(data.get("name", "")).strip()
        if len(name) < 2:
            return self.error_json("Informe o nome da empresa")
        cnpj = str(data.get("cnpj") or "").strip() or None
        email = str(data.get("email") or "").strip().lower() or None
        if cnpj and not _valid_cnpj(cnpj):
            return self.error_json("CNPJ da empresa inválido")
        if email and not re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", email):
            return self.error_json("E-mail da empresa inválido")
        now = utc_now()
        with self.db.transaction(immediate=True):
            holding_id = self.db.scalar(
                "SELECT holding_id FROM companies WHERE id=?", (session["company_id"],)
            )
            cursor = self.db.execute(
                """INSERT INTO companies
                   (name,cnpj,phone,email,address,created_at,updated_at,holding_id)
                   VALUES(?,?,?,?,?,?,?,?)""",
                (name, cnpj, str(data.get("phone") or "").strip() or None, email,
                 str(data.get("address") or "").strip() or None, now, now, holding_id),
            )
            company_id = cursor.lastrowid
            self.db.ensure_company_structure(company_id, name, cnpj, now=now)
            self.db.execute(
                """INSERT INTO company_memberships
                   (company_id,user_id,role,permissions,active,created_at,updated_at)
                   VALUES(?,?,'admin','{}',1,?,?)""", (company_id, session["id"], now, now))
            self.db.seed_sources(company_id)
            self.db.seed_norms(company_id)
            self.db.seed_seccol_portfolio(company_id)
            self.db.audit(session["id"], "create", "company", company_id,
                          {"name": name}, company_id=company_id)
        return self.send_json({"ok": True, "id": company_id}, 201)

    def company_switch(self, session):
        try:
            data = self.parse_json()
            company_id = int(data.get("company_id"))
        except (ValueError, TypeError):
            return self.error_json("Empresa inválida")
        membership = self.db.connection().execute(
            """SELECT 1 FROM company_memberships cm JOIN companies c ON c.id=cm.company_id
               WHERE cm.user_id=? AND cm.company_id=? AND cm.active=1 AND c.active=1""",
            (session["id"], company_id)).fetchone()
        if not membership:
            return self.error_json("Usuário não possui acesso a esta empresa", 403, "forbidden")
        self.db.execute("UPDATE sessions SET company_id=? WHERE token_hash=?", (company_id, session["token_hash"]))
        self.db.audit(session["id"], "switch", "company", company_id, company_id=company_id)
        return self.send_json({"ok": True, "companyId": company_id})

    def branch_create(self, session):
        try:
            data = self.parse_json(max_bytes=64 * 1024)
        except ValueError as exc:
            return self.error_json(str(exc))
        code = re.sub(r"\s+", "-", str(data.get("code") or "").strip().upper())
        name = str(data.get("name") or "").strip()
        cnpj = re.sub(r"\D", "", str(data.get("cnpj") or "")) or None
        address = str(data.get("address") or "").strip() or None
        if not re.fullmatch(r"[A-Z0-9][A-Z0-9_-]{0,39}", code):
            return self.error_json("Código da unidade inválido")
        if len(name) < 2 or len(name) > 160:
            return self.error_json("Informe o nome da unidade")
        if cnpj and not _valid_cnpj(cnpj):
            return self.error_json("CNPJ da unidade inválido")
        now = utc_now()
        try:
            with self.db.transaction(immediate=True):
                cursor = self.db.execute(
                    """INSERT INTO branches
                       (company_id,code,name,cnpj,address,active,is_headquarters,created_at,updated_at)
                       VALUES(?,?,?,?,?,1,0,?,?)""",
                    (session["company_id"], code, name, cnpj, address, now, now),
                )
                branch_id = cursor.lastrowid
                self.db.audit(
                    session["id"], "create", "branch", branch_id,
                    {"code": code, "name": name}, company_id=session["company_id"],
                )
        except sqlite3.IntegrityError:
            return self.error_json(
                "Já existe uma unidade com este código na empresa ativa", 409, "duplicate_branch",
            )
        return self.send_json({"ok": True, "id": branch_id}, 201)

    @staticmethod
    def inventory_micros(value, label="Quantidade"):
        try:
            number = Decimal(str(value))
        except (InvalidOperation, ValueError, TypeError):
            raise ValueError(f"{label}: informe um número válido") from None
        if not number.is_finite() or number <= 0 or number > Decimal("1000000000"):
            raise ValueError(f"{label}: informe um valor positivo de até 1 bilhão")
        quantum = Decimal(1) / INVENTORY_QUANTITY_SCALE
        normalized = number.quantize(quantum, rounding=ROUND_HALF_UP)
        if normalized != number:
            raise ValueError(f"{label}: use no máximo seis casas decimais")
        return int(normalized * INVENTORY_QUANTITY_SCALE)

    @staticmethod
    def inventory_units(micros):
        quantity = Decimal(int(micros or 0)) / INVENTORY_QUANTITY_SCALE
        return int(quantity) if quantity == quantity.to_integral_value() else float(quantity)

    @staticmethod
    def inventory_origin(data):
        origin_type = str(data.get("originType") or "").strip().upper()
        origin_id = str(data.get("originId") or "").strip()
        if not re.fullmatch(r"[A-Z][A-Z0-9_]{1,39}", origin_type):
            raise ValueError("Informe uma origem operacional válida")
        if not origin_id or len(origin_id) > 120:
            raise ValueError("Informe o identificador da origem")
        return origin_type, origin_id

    @staticmethod
    def inventory_lot(value):
        lot = str(value or "").strip()
        if len(lot) > 120:
            raise ValueError("Lote excede 120 caracteres")
        return lot

    def inventory_scope(self, company_id, warehouse_id, product_id, *, active=True):
        warehouse = self.db.connection().execute(
            """SELECT w.id,w.name,w.branch_id FROM warehouses w
               JOIN branches b ON b.id=w.branch_id AND b.company_id=w.company_id
               WHERE w.id=? AND w.company_id=? AND (?=0 OR (w.active=1 AND b.active=1))""",
            (warehouse_id, company_id, 1 if active else 0),
        ).fetchone()
        product = self.db.connection().execute(
            """SELECT id,title,payload FROM records
               WHERE id=? AND company_id=? AND module='produtos' AND deleted_at IS NULL""",
            (product_id, company_id),
        ).fetchone()
        if not warehouse:
            raise ValueError("Depósito não existe ou não está ativo na empresa atual")
        if not product:
            raise ValueError("Produto não existe ou não está ativo na empresa atual")
        return warehouse, product

    def inventory_balance(self, company_id, warehouse_id, product_id, lot_key, now):
        db = self.db.connection()
        db.execute(
            """INSERT OR IGNORE INTO inventory_balances
               (company_id,warehouse_id,product_record_id,lot_key,
                physical_quantity_micros,reserved_quantity_micros,revision,updated_at)
               VALUES(?,?,?,?,0,0,1,?)""",
            (company_id, warehouse_id, product_id, lot_key, now),
        )
        return db.execute(
            """SELECT * FROM inventory_balances
               WHERE company_id=? AND warehouse_id=? AND product_record_id=? AND lot_key=?""",
            (company_id, warehouse_id, product_id, lot_key),
        ).fetchone()

    def inventory_log_movement(self, *, company_id, warehouse_id, product_id, lot_key,
                               movement_type, quantity_micros, physical_delta_micros,
                               reserved_delta_micros, origin_type, origin_id, reference,
                               reason, reservation_id, created_by, counterpart_id=None,
                               created_at=None, unit_cost_cents=None,
                               value_delta_cents=0, balance_value_cents=0):
        return self.db.connection().execute(
            """INSERT INTO inventory_movements
               (company_id,warehouse_id,counterpart_warehouse_id,product_record_id,lot_key,
                movement_type,quantity_micros,physical_delta_micros,reserved_delta_micros,
                unit_cost_cents,value_delta_cents,balance_value_cents,origin_type,origin_id,
                reference,reason,reservation_id,created_by,created_at)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (company_id, warehouse_id, counterpart_id, product_id, lot_key,
             movement_type, quantity_micros, physical_delta_micros, reserved_delta_micros,
             unit_cost_cents, value_delta_cents, balance_value_cents, origin_type, origin_id,
             reference, reason, reservation_id, created_by, created_at or utc_now()),
        ).lastrowid

    @staticmethod
    def inventory_proportional_value(value_cents, quantity_micros, physical_micros):
        if physical_micros <= 0 or quantity_micros <= 0 or value_cents <= 0:
            return 0
        if quantity_micros >= physical_micros:
            return int(value_cents)
        return int((Decimal(value_cents) * Decimal(quantity_micros) /
                    Decimal(physical_micros)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))

    @staticmethod
    def inventory_average_cost(value_cents, physical_micros):
        if value_cents <= 0 or physical_micros <= 0:
            return None
        return int((Decimal(value_cents) * INVENTORY_QUANTITY_SCALE /
                    Decimal(physical_micros)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))

    def inventory_get(self, session, query):
        if not self.require_module_read(session, "estoque"):
            return
        company_id = session["company_id"]
        db = self.db.connection()
        warehouses = [dict(row) for row in db.execute(
            """SELECT w.id,w.code,w.name,w.location,w.active,w.branch_id,b.name branch_name
               FROM warehouses w JOIN branches b ON b.id=w.branch_id
               WHERE w.company_id=? ORDER BY w.active DESC,w.name""",
            (company_id,),
        ).fetchall()]
        branches = [dict(row) for row in db.execute(
            """SELECT id,code,name,active,is_headquarters FROM branches
               WHERE company_id=? ORDER BY is_headquarters DESC,name""",
            (company_id,),
        ).fetchall()]
        products = [dict(row) for row in db.execute(
            """SELECT id,title,json_extract(payload,'$.codigo') code,
                      COALESCE(json_extract(payload,'$.unidade'),'UN') unit
               FROM records WHERE company_id=? AND module='produtos' AND deleted_at IS NULL
               ORDER BY title COLLATE NOCASE""",
            (company_id,),
        ).fetchall()]
        balance_rows = db.execute(
            """SELECT b.*,w.name warehouse_name,r.title product_name,
                      json_extract(r.payload,'$.codigo') product_code,
                      COALESCE(json_extract(r.payload,'$.unidade'),'UN') unit
               FROM inventory_balances b
               JOIN warehouses w ON w.id=b.warehouse_id AND w.company_id=b.company_id
               JOIN records r ON r.id=b.product_record_id AND r.company_id=b.company_id
               WHERE b.company_id=?
               ORDER BY r.title COLLATE NOCASE,w.name,b.lot_key""",
            (company_id,),
        ).fetchall()
        balances = []
        show_values = "view_values" in self.allowed_operations(session, "estoque")
        total_inventory_value_cents = 0
        total_reserved_value_cents = 0
        unvalued_balances = 0
        for row in balance_rows:
            item = dict(row)
            physical = item.pop("physical_quantity_micros")
            reserved = item.pop("reserved_quantity_micros")
            inventory_value = int(item.pop("inventory_value_cents") or 0)
            reserved_value = self.inventory_proportional_value(
                inventory_value, reserved, physical,
            )
            average_cost = self.inventory_average_cost(inventory_value, physical)
            if physical > 0 and inventory_value <= 0:
                unvalued_balances += 1
            total_inventory_value_cents += inventory_value
            total_reserved_value_cents += reserved_value
            item.update(
                physicalQuantity=self.inventory_units(physical),
                reservedQuantity=self.inventory_units(reserved),
                availableQuantity=self.inventory_units(physical - reserved),
                lot=item.pop("lot_key"),
                inventoryValueCents=inventory_value if show_values else None,
                reservedValueCents=reserved_value if show_values else None,
                availableValueCents=(inventory_value - reserved_value) if show_values else None,
                averageUnitCostCents=average_cost if show_values else None,
                valuesRestricted=not show_values,
            )
            balances.append(item)
        movement_rows = db.execute(
            """SELECT m.*,w.name warehouse_name,cw.name counterpart_warehouse_name,
                      r.title product_name,json_extract(r.payload,'$.codigo') product_code,
                      u.name created_by_name
               FROM inventory_movements m
               JOIN warehouses w ON w.id=m.warehouse_id AND w.company_id=m.company_id
               LEFT JOIN warehouses cw ON cw.id=m.counterpart_warehouse_id
               JOIN records r ON r.id=m.product_record_id AND r.company_id=m.company_id
               LEFT JOIN users u ON u.id=m.created_by
               WHERE m.company_id=? ORDER BY m.id DESC LIMIT 300""",
            (company_id,),
        ).fetchall()
        movements = []
        for row in movement_rows:
            item = dict(row)
            item["quantity"] = self.inventory_units(item.pop("quantity_micros"))
            item["physicalDelta"] = self.inventory_units(item.pop("physical_delta_micros"))
            item["reservedDelta"] = self.inventory_units(item.pop("reserved_delta_micros"))
            unit_cost = item.pop("unit_cost_cents")
            value_delta = item.pop("value_delta_cents")
            balance_value = item.pop("balance_value_cents")
            item["unitCostCents"] = unit_cost if show_values else None
            item["valueDeltaCents"] = value_delta if show_values else None
            item["balanceValueCents"] = balance_value if show_values else None
            item["valuesRestricted"] = not show_values
            item["lot"] = item.pop("lot_key")
            movements.append(item)
        reservation_rows = db.execute(
            """SELECT q.*,w.name warehouse_name,r.title product_name,
                      json_extract(r.payload,'$.codigo') product_code,u.name created_by_name
               FROM inventory_reservations q
               JOIN warehouses w ON w.id=q.warehouse_id AND w.company_id=q.company_id
               JOIN records r ON r.id=q.product_record_id AND r.company_id=q.company_id
               LEFT JOIN users u ON u.id=q.created_by
               WHERE q.company_id=?
               ORDER BY CASE q.status WHEN 'ACTIVE' THEN 0 ELSE 1 END,q.id DESC LIMIT 300""",
            (company_id,),
        ).fetchall()
        reservations = []
        for row in reservation_rows:
            item = dict(row)
            item["quantity"] = self.inventory_units(item.pop("quantity_micros"))
            item["lot"] = item.pop("lot_key")
            reservations.append(item)
        legacy_count = self.db.scalar(
            """SELECT COUNT(*) FROM records
               WHERE company_id=? AND module='estoque' AND deleted_at IS NULL""",
            (company_id,),
        )
        return self.send_json({
            "ok": True, "warehouses": warehouses, "branches": branches,
            "products": products, "balances": balances, "movements": movements,
            "reservations": reservations, "movementTypes": sorted(INVENTORY_MOVEMENT_TYPES),
            "legacyRecordCount": int(legacy_count or 0),
            "valueVisible": show_values,
            "valuation": {
                "inventoryValueCents": total_inventory_value_cents if show_values else None,
                "reservedValueCents": total_reserved_value_cents if show_values else None,
                "availableValueCents": (
                    total_inventory_value_cents - total_reserved_value_cents
                ) if show_values else None,
                "unvaluedBalances": unvalued_balances if show_values else None,
            },
        })

    def inventory_warehouse_create(self, session):
        if not self.require_operation(session, "estoque", "manage_warehouses"):
            return
        try:
            data = self.parse_json(max_bytes=64 * 1024)
            branch_id = int(data.get("branchId"))
        except ValueError as exc:
            return self.error_json(str(exc))
        except (TypeError, OverflowError):
            return self.error_json("Selecione uma unidade válida")
        code = re.sub(r"\s+", "-", str(data.get("code") or "").strip().upper())
        name = str(data.get("name") or "").strip()
        location = str(data.get("location") or "").strip() or None
        if not re.fullmatch(r"[A-Z0-9][A-Z0-9_-]{0,39}", code):
            return self.error_json("Código do depósito inválido")
        if len(name) < 2 or len(name) > 160:
            return self.error_json("Informe o nome do depósito")
        branch = self.db.connection().execute(
            "SELECT id FROM branches WHERE id=? AND company_id=? AND active=1",
            (branch_id, session["company_id"]),
        ).fetchone()
        if not branch:
            return self.error_json("Unidade não existe ou não está ativa nesta empresa")
        now = utc_now()
        try:
            with self.db.transaction(immediate=True):
                warehouse_id = self.db.execute(
                    """INSERT INTO warehouses
                       (company_id,branch_id,code,name,location,active,created_by,created_at,updated_at)
                       VALUES(?,?,?,?,?,1,?,?,?)""",
                    (session["company_id"], branch_id, code, name, location,
                     session["id"], now, now),
                ).lastrowid
                self.db.audit(
                    session["id"], "create", "warehouse", warehouse_id,
                    {"code": code, "name": name, "branch_id": branch_id},
                    company_id=session["company_id"],
                )
        except sqlite3.IntegrityError:
            return self.error_json(
                "Já existe um depósito com este código na empresa ativa", 409, "duplicate_warehouse",
            )
        return self.send_json({"ok": True, "id": warehouse_id}, 201)

    def inventory_movement_create(self, session):
        if not self.require_operation(session, "estoque", "move_stock"):
            return
        try:
            data = self.parse_json(max_bytes=64 * 1024)
            movement_type = str(data.get("movementType") or "").strip().upper()
            if movement_type not in INVENTORY_MOVEMENT_TYPES:
                raise ValueError("Tipo de movimento inválido")
            if movement_type in {"RESERVE", "RELEASE_RESERVATION", "TRANSFER_IN"}:
                raise ValueError("Use o fluxo específico de reserva ou transferência")
            quantity = self.inventory_micros(data.get("quantity"))
            warehouse_id = int(data.get("warehouseId"))
            product_id = int(data.get("productId"))
            lot_key = self.inventory_lot(data.get("lot"))
            origin_type, origin_id = self.inventory_origin(data)
            reference = str(data.get("reference") or "").strip()[:240] or None
            reason = str(data.get("reason") or "").strip()[:500] or None
            reservation_id = int(data.get("reservationId")) if data.get("reservationId") else None
            counterpart_id = int(data.get("counterpartWarehouseId")) if data.get("counterpartWarehouseId") else None
            unit_cost_cents = None
            if movement_type in INVENTORY_IN_TYPES:
                if not self.require_operation(session, "estoque", "view_values"):
                    return
                if data.get("unitCost") in (None, ""):
                    raise ValueError("Entradas manuais exigem o custo unitário")
                unit_cost_cents = self.money_cents(data.get("unitCost"), "Custo unitário")
            if movement_type.startswith("ADJUSTMENT_") and not reason:
                raise ValueError("Ajustes exigem uma justificativa")
            if (movement_type.startswith("ADJUSTMENT_")
                    and not self.require_operation(session, "estoque", "adjust_stock")):
                return
            if movement_type == "TRANSFER_OUT" and not counterpart_id:
                raise ValueError("Transferência exige o depósito de destino")
            if (movement_type == "TRANSFER_OUT"
                    and not self.require_operation(session, "estoque", "transfer_stock")):
                return
            if movement_type != "TRANSFER_OUT" and counterpart_id:
                raise ValueError("Depósito de destino só é aceito em transferências")
            if reservation_id and movement_type not in {"SALE_OUT", "SERVICE_ORDER_OUT"}:
                raise ValueError("Somente venda ou O.S. podem consumir uma reserva")
        except (ValueError, TypeError, OverflowError) as exc:
            return self.error_json(str(exc))
        company_id = session["company_id"]
        now = utc_now()
        try:
            with self.db.transaction(immediate=True):
                self.inventory_scope(company_id, warehouse_id, product_id)
                balance = self.inventory_balance(company_id, warehouse_id, product_id, lot_key, now)
                physical = int(balance["physical_quantity_micros"])
                reserved = int(balance["reserved_quantity_micros"])
                inventory_value = int(balance["inventory_value_cents"] or 0)
                physical_delta = 0
                reserved_delta = 0
                value_delta = 0
                if reservation_id:
                    reservation = self.db.connection().execute(
                        """SELECT * FROM inventory_reservations
                           WHERE id=? AND company_id=? AND status='ACTIVE'""",
                        (reservation_id, company_id),
                    ).fetchone()
                    if (not reservation or reservation["warehouse_id"] != warehouse_id or
                            reservation["product_record_id"] != product_id or
                            reservation["lot_key"] != lot_key):
                        raise ValueError("Reserva ativa não corresponde ao produto, lote e depósito")
                    if int(reservation["quantity_micros"]) != quantity:
                        raise ValueError("Consuma a quantidade integral da reserva nesta etapa")
                    if reserved < quantity or physical < quantity:
                        raise ValueError("Saldo reservado insuficiente para concluir a saída")
                    origin_type, origin_id = reservation["origin_type"], reservation["origin_id"]
                    reference = reference or reservation["reference"]
                    physical_delta = -quantity
                    reserved_delta = -quantity
                    value_delta = -self.inventory_proportional_value(
                        inventory_value, quantity, physical,
                    )
                    unit_cost_cents = self.inventory_average_cost(-value_delta, quantity)
                    self.db.connection().execute(
                        """UPDATE inventory_reservations
                           SET status='FULFILLED',released_by=?,updated_at=?
                           WHERE id=? AND company_id=? AND status='ACTIVE'""",
                        (session["id"], now, reservation_id, company_id),
                    )
                elif movement_type in INVENTORY_IN_TYPES:
                    physical_delta = quantity
                    value_delta = int(
                        (Decimal(quantity) * Decimal(unit_cost_cents or 0)
                         / INVENTORY_QUANTITY_SCALE).quantize(
                            Decimal("1"), rounding=ROUND_HALF_UP,
                        )
                    )
                elif movement_type in INVENTORY_OUT_TYPES:
                    if physical - reserved < quantity:
                        raise ValueError("Quantidade disponível insuficiente para esta saída")
                    physical_delta = -quantity
                    value_delta = -self.inventory_proportional_value(
                        inventory_value, quantity, physical,
                    )
                    unit_cost_cents = self.inventory_average_cost(-value_delta, quantity)
                new_inventory_value = inventory_value + value_delta
                self.db.connection().execute(
                    """UPDATE inventory_balances
                       SET physical_quantity_micros=physical_quantity_micros+?,
                           reserved_quantity_micros=reserved_quantity_micros+?,
                           inventory_value_cents=?,
                           revision=revision+1,updated_at=?
                       WHERE company_id=? AND warehouse_id=? AND product_record_id=? AND lot_key=?""",
                    (physical_delta, reserved_delta, new_inventory_value, now,
                     company_id, warehouse_id, product_id, lot_key),
                )
                movement_id = self.inventory_log_movement(
                    company_id=company_id, warehouse_id=warehouse_id, counterpart_id=counterpart_id,
                    product_id=product_id, lot_key=lot_key, movement_type=movement_type,
                    quantity_micros=quantity, physical_delta_micros=physical_delta,
                    reserved_delta_micros=reserved_delta, origin_type=origin_type,
                    origin_id=origin_id, reference=reference, reason=reason,
                    reservation_id=reservation_id, created_by=session["id"], created_at=now,
                    unit_cost_cents=unit_cost_cents, value_delta_cents=value_delta,
                    balance_value_cents=new_inventory_value,
                )
                paired_movement_id = None
                if movement_type == "TRANSFER_OUT":
                    if counterpart_id == warehouse_id:
                        raise ValueError("Origem e destino da transferência devem ser diferentes")
                    self.inventory_scope(company_id, counterpart_id, product_id)
                    destination = self.inventory_balance(
                        company_id, counterpart_id, product_id, lot_key, now,
                    )
                    destination_value = int(destination["inventory_value_cents"] or 0)
                    transferred_value = -value_delta
                    destination_new_value = destination_value + transferred_value
                    self.db.connection().execute(
                        """UPDATE inventory_balances
                           SET physical_quantity_micros=physical_quantity_micros+?,
                               inventory_value_cents=?,
                               revision=revision+1,updated_at=?
                           WHERE company_id=? AND warehouse_id=? AND product_record_id=? AND lot_key=?""",
                        (quantity, destination_new_value, now, company_id,
                         counterpart_id, product_id, lot_key),
                    )
                    paired_movement_id = self.inventory_log_movement(
                        company_id=company_id, warehouse_id=counterpart_id,
                        counterpart_id=warehouse_id, product_id=product_id, lot_key=lot_key,
                        movement_type="TRANSFER_IN", quantity_micros=quantity,
                        physical_delta_micros=quantity, reserved_delta_micros=0,
                        origin_type=origin_type, origin_id=origin_id, reference=reference,
                        reason=reason, reservation_id=None, created_by=session["id"], created_at=now,
                        unit_cost_cents=unit_cost_cents,
                        value_delta_cents=transferred_value,
                        balance_value_cents=destination_new_value,
                    )
                self.db.audit(
                    session["id"], "move", "inventory", movement_id,
                    {"movement_type": movement_type, "quantity_micros": quantity,
                     "warehouse_id": warehouse_id, "counterpart_warehouse_id": counterpart_id,
                     "product_record_id": product_id, "origin_type": origin_type,
                     "origin_id": origin_id, "paired_movement_id": paired_movement_id,
                     "unit_cost_cents": unit_cost_cents,
                     "value_delta_cents": value_delta,
                     "balance_value_cents": new_inventory_value},
                    company_id=company_id,
                )
        except (ValueError, sqlite3.IntegrityError) as exc:
            return self.error_json(str(exc), 409, "inventory_conflict")
        return self.send_json({"ok": True, "movementId": movement_id,
                               "pairedMovementId": paired_movement_id}, 201)

    def inventory_reservation_create(self, session):
        if not self.require_operation(session, "estoque", "reserve_stock"):
            return
        try:
            data = self.parse_json(max_bytes=64 * 1024)
            quantity = self.inventory_micros(data.get("quantity"))
            warehouse_id = int(data.get("warehouseId"))
            product_id = int(data.get("productId"))
            lot_key = self.inventory_lot(data.get("lot"))
            origin_type, origin_id = self.inventory_origin(data)
            reference = str(data.get("reference") or "").strip()[:240] or None
            expires_at = str(data.get("expiresAt") or "").strip() or None
            if expires_at:
                datetime.strptime(expires_at, "%Y-%m-%d")
        except (ValueError, TypeError, OverflowError) as exc:
            return self.error_json(str(exc))
        company_id = session["company_id"]
        now = utc_now()
        try:
            with self.db.transaction(immediate=True):
                self.inventory_scope(company_id, warehouse_id, product_id)
                balance = self.inventory_balance(company_id, warehouse_id, product_id, lot_key, now)
                available = (int(balance["physical_quantity_micros"]) -
                             int(balance["reserved_quantity_micros"]))
                if available < quantity:
                    raise ValueError("Quantidade disponível insuficiente para reservar")
                reservation_id = self.db.connection().execute(
                    """INSERT INTO inventory_reservations
                       (company_id,warehouse_id,product_record_id,lot_key,quantity_micros,status,
                        origin_type,origin_id,reference,expires_at,created_by,created_at,updated_at)
                       VALUES(?,?,?,?,?,'ACTIVE',?,?,?,?,?,?,?)""",
                    (company_id, warehouse_id, product_id, lot_key, quantity, origin_type,
                     origin_id, reference, expires_at, session["id"], now, now),
                ).lastrowid
                self.db.connection().execute(
                    """UPDATE inventory_balances
                       SET reserved_quantity_micros=reserved_quantity_micros+?,
                           revision=revision+1,updated_at=?
                       WHERE company_id=? AND warehouse_id=? AND product_record_id=? AND lot_key=?""",
                    (quantity, now, company_id, warehouse_id, product_id, lot_key),
                )
                movement_id = self.inventory_log_movement(
                    company_id=company_id, warehouse_id=warehouse_id, product_id=product_id,
                    lot_key=lot_key, movement_type="RESERVE", quantity_micros=quantity,
                    physical_delta_micros=0, reserved_delta_micros=quantity,
                    origin_type=origin_type, origin_id=origin_id, reference=reference,
                    reason=None, reservation_id=reservation_id, created_by=session["id"],
                    created_at=now,
                    balance_value_cents=int(balance["inventory_value_cents"] or 0),
                )
                self.db.audit(
                    session["id"], "reserve", "inventory", reservation_id,
                    {"movement_id": movement_id, "quantity_micros": quantity,
                     "warehouse_id": warehouse_id, "product_record_id": product_id,
                     "origin_type": origin_type, "origin_id": origin_id},
                    company_id=company_id,
                )
        except sqlite3.IntegrityError:
            return self.error_json(
                "Já existe uma reserva ativa para esta origem, produto, lote e depósito",
                409, "duplicate_reservation",
            )
        except ValueError as exc:
            return self.error_json(str(exc), 409, "inventory_conflict")
        return self.send_json({"ok": True, "id": reservation_id,
                               "movementId": movement_id}, 201)

    def inventory_reservation_release(self, path, session):
        if not self.require_operation(session, "estoque", "release_stock"):
            return
        pieces = path.split("/")
        if len(pieces) != 6 or not pieces[4].isdigit():
            return self.error_json("Reserva inválida", 404)
        reservation_id = int(pieces[4])
        company_id = session["company_id"]
        now = utc_now()
        try:
            with self.db.transaction(immediate=True):
                reservation = self.db.connection().execute(
                    """SELECT * FROM inventory_reservations
                       WHERE id=? AND company_id=? AND status='ACTIVE'""",
                    (reservation_id, company_id),
                ).fetchone()
                if not reservation:
                    raise ValueError("Reserva ativa não encontrada nesta empresa")
                balance = self.inventory_balance(
                    company_id, reservation["warehouse_id"], reservation["product_record_id"],
                    reservation["lot_key"], now,
                )
                quantity = int(reservation["quantity_micros"])
                if int(balance["reserved_quantity_micros"]) < quantity:
                    raise ValueError("Saldo reservado inconsistente; operação cancelada")
                self.db.connection().execute(
                    """UPDATE inventory_balances
                       SET reserved_quantity_micros=reserved_quantity_micros-?,
                           revision=revision+1,updated_at=?
                       WHERE company_id=? AND warehouse_id=? AND product_record_id=? AND lot_key=?""",
                    (quantity, now, company_id, reservation["warehouse_id"],
                     reservation["product_record_id"], reservation["lot_key"]),
                )
                self.db.connection().execute(
                    """UPDATE inventory_reservations
                       SET status='RELEASED',released_by=?,updated_at=?
                       WHERE id=? AND company_id=? AND status='ACTIVE'""",
                    (session["id"], now, reservation_id, company_id),
                )
                movement_id = self.inventory_log_movement(
                    company_id=company_id, warehouse_id=reservation["warehouse_id"],
                    product_id=reservation["product_record_id"], lot_key=reservation["lot_key"],
                    movement_type="RELEASE_RESERVATION", quantity_micros=quantity,
                    physical_delta_micros=0, reserved_delta_micros=-quantity,
                    origin_type=reservation["origin_type"], origin_id=reservation["origin_id"],
                    reference=reservation["reference"], reason="Liberação manual da reserva",
                    reservation_id=reservation_id, created_by=session["id"], created_at=now,
                    balance_value_cents=int(balance["inventory_value_cents"] or 0),
                )
                self.db.audit(
                    session["id"], "release", "inventory", reservation_id,
                    {"movement_id": movement_id, "quantity_micros": quantity},
                    company_id=company_id,
                )
        except ValueError as exc:
            return self.error_json(str(exc), 409, "inventory_conflict")
        return self.send_json({"ok": True, "movementId": movement_id})

    @staticmethod
    def money_cents(value, label="Valor"):
        text = str(value if value is not None else "").strip().replace(" ", "")
        if not text:
            return 0
        if "," in text:
            text = text.replace(".", "").replace(",", ".")
        try:
            number = Decimal(text)
        except (InvalidOperation, ValueError, TypeError):
            raise ValueError(f"{label}: informe um número válido") from None
        if not number.is_finite() or number < 0 or number > Decimal("10000000000000"):
            raise ValueError(f"{label}: valor fora do limite permitido")
        normalized = number.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        if normalized != number:
            raise ValueError(f"{label}: use no máximo duas casas decimais")
        return int(normalized * 100)

    def document_record(self, record_id, session, action="read"):
        record = self.db.connection().execute(
            """SELECT * FROM records WHERE id=? AND company_id=? AND deleted_at IS NULL""",
            (record_id, session["company_id"]),
        ).fetchone()
        if not record:
            self.error_json("Documento não encontrado", 404, "not_found")
            return None
        if record["module"] not in ITEM_DOCUMENT_MODULES:
            self.error_json("Este módulo não aceita itens estruturados", 409, "items_not_supported")
            return None
        allowed = self.allowed_modules(session, action)
        if record["module"] not in allowed:
            self.error_json(
                "Seu perfil não possui permissão para acessar os itens deste documento",
                403, "forbidden",
            )
            return None
        return record

    @staticmethod
    def document_item_json(row, show_values=True):
        item = dict(row)
        item["quantity"] = SIVSHandler.inventory_units(item.pop("quantity_micros"))
        unit_price = item.pop("unit_price_cents")
        discount = item.pop("discount_cents")
        total = item.pop("total_cents")
        item["unitPrice"] = unit_price / 100 if show_values else None
        item["discount"] = discount / 100 if show_values else None
        item["total"] = total / 100 if show_values else None
        item["valuesRestricted"] = not show_values
        item["itemKind"] = item.pop("item_kind")
        item["catalogRecordId"] = item.pop("catalog_record_id")
        item["warehouseId"] = item.pop("warehouse_id")
        item["warehouseName"] = item.pop("warehouse_name", None)
        item["lot"] = item.pop("lot_key")
        item["reservationId"] = item.pop("reservation_id")
        item["reservationStatus"] = item.pop("reservation_status", None)
        item["receiptMovementId"] = item.pop("receipt_movement_id", None)
        return item

    def document_totals(self, record_id, company_id):
        row = self.db.connection().execute(
            """SELECT COALESCE(SUM(total_cents + discount_cents),0) subtotal_cents,
                      COALESCE(SUM(discount_cents),0) discount_cents,
                      COALESCE(SUM(total_cents),0) total_cents,
                      COUNT(*) item_count
               FROM document_items WHERE record_id=? AND company_id=?""",
            (record_id, company_id),
        ).fetchone()
        return {
            "subtotalCents": int(row["subtotal_cents"] or 0),
            "discountCents": int(row["discount_cents"] or 0),
            "totalCents": int(row["total_cents"] or 0),
            "itemCount": int(row["item_count"] or 0),
        }

    def record_items_get(self, record_id, session):
        record = self.document_record(record_id, session)
        if not record:
            return
        company_id = session["company_id"]
        rows = self.db.connection().execute(
            """SELECT i.*,c.title catalog_title,w.name warehouse_name,
                      q.status reservation_status,
                      (SELECT m.id FROM inventory_movements m
                       WHERE m.company_id=i.company_id AND m.movement_type='PURCHASE_IN'
                         AND m.origin_type='PURCHASE_ORDER'
                         AND m.origin_id=CAST(i.record_id AS TEXT)||':'||CAST(i.id AS TEXT)
                       ORDER BY m.id LIMIT 1) receipt_movement_id
               FROM document_items i
               JOIN records c ON c.id=i.catalog_record_id AND c.company_id=i.company_id
               LEFT JOIN warehouses w ON w.id=i.warehouse_id AND w.company_id=i.company_id
               LEFT JOIN inventory_reservations q
                 ON q.id=i.reservation_id AND q.company_id=i.company_id
               WHERE i.record_id=? AND i.company_id=? ORDER BY i.sort_order,i.id""",
            (record_id, company_id),
        ).fetchall()
        readable = self.allowed_modules(session, "read")
        catalog = []
        catalog_modules = [module for module in ("produtos", "catalogo_servicos")
                           if module in readable]
        if catalog_modules:
            placeholders = ",".join("?" for _ in catalog_modules)
            catalog = [dict(row) for row in self.db.connection().execute(
                f"""SELECT id,module,title,status,payload FROM records
                    WHERE company_id=? AND module IN ({placeholders})
                      AND deleted_at IS NULL AND status NOT IN ('Cancelado','Cancelada','Obsoleto')
                    ORDER BY title COLLATE NOCASE LIMIT 3000""",
                (company_id, *catalog_modules),
            ).fetchall()]
            for option in catalog:
                payload = json.loads(option.pop("payload") or "{}")
                option["code"] = payload.get("codigo")
                option["defaultUnitPrice"] = payload.get("preco_venda") or 0
        warehouses = []
        if "estoque" in readable:
            warehouses = [dict(row) for row in self.db.connection().execute(
                """SELECT id,code,name FROM warehouses
                   WHERE company_id=? AND active=1 ORDER BY name""",
                (company_id,),
            ).fetchall()]
        totals = self.document_totals(record_id, company_id)
        show_values = "view_values" in self.allowed_operations(session, record["module"])
        item_list = [self.document_item_json(row, show_values) for row in rows]
        if not show_values:
            totals = {
                "subtotalCents": None, "discountCents": None,
                "totalCents": None, "itemCount": totals["itemCount"],
            }
            for option in catalog:
                option.pop("defaultUnitPrice", None)
        active_reservations = sum(
            1 for item in item_list if item["reservationStatus"] == "ACTIVE"
        )
        fulfilled_reservations = sum(
            1 for item in item_list if item["reservationStatus"] == "FULFILLED"
        )
        received_items = sum(1 for item in item_list if item["receiptMovementId"])
        document_actions = self.allowed_operations(session, record["module"])
        inventory_actions = self.allowed_operations(session, "estoque")
        can_manage = "manage_items" in document_actions
        can_reserve = (
            record["module"] in RESERVABLE_ITEM_MODULES
            and "reserve_stock" in document_actions
            and "reserve_stock" in inventory_actions
        )
        can_release = (
            record["module"] in RESERVABLE_ITEM_MODULES
            and "release_stock" in document_actions
            and "release_stock" in inventory_actions
        )
        reservable_statuses = {
            "vendas": {"Confirmado", "Separação"},
            "ordens_servico": {"Agendada", "Em execução", "Aguardando aprovação"},
        }
        fulfillable_statuses = {
            "vendas": {"Separação", "Faturado"},
            "ordens_servico": {"Em execução", "Aguardando aprovação"},
        }
        return self.send_json({
            "ok": True, "recordId": record_id, "module": record["module"],
            "recordRevision": record["revision"], "status": record["status"],
            "items": item_list, "totals": totals, "catalog": catalog,
            "warehouses": warehouses, "canManage": can_manage,
            "valuesVisible": show_values,
            "canReserve": can_reserve, "canRelease": can_release,
            "activeReservations": active_reservations,
            "fulfilledReservations": fulfilled_reservations,
            "receivedItems": received_items,
            "canReserveNow": can_reserve and record["status"] in reservable_statuses.get(record["module"], set()),
            "canFulfill": (
                "fulfill_stock" in document_actions and "move_stock" in inventory_actions
                and record["status"] in fulfillable_statuses.get(record["module"], set())
            ),
            "canReceive": (
                record["module"] == "pedidos_compra"
                and "receive_stock" in document_actions
                and "move_stock" in inventory_actions
                and record["status"] in {"Emitido", "Aguardando fornecedor", "Recebido parcial"}
            ),
        })

    def parse_document_item(self, data, company_id):
        item_kind = str(data.get("itemKind") or "").strip().upper()
        if item_kind not in {"PRODUCT", "SERVICE"}:
            raise ValueError("Escolha produto ou serviço")
        try:
            catalog_id = int(data.get("catalogRecordId"))
        except (ValueError, TypeError):
            raise ValueError("Selecione um item do catálogo") from None
        expected_module = "produtos" if item_kind == "PRODUCT" else "catalogo_servicos"
        catalog = self.db.connection().execute(
            """SELECT id,title FROM records
               WHERE id=? AND company_id=? AND module=? AND deleted_at IS NULL""",
            (catalog_id, company_id, expected_module),
        ).fetchone()
        if not catalog:
            raise ValueError("Item do catálogo não existe na empresa ativa")
        quantity = self.inventory_micros(data.get("quantity"), "Quantidade")
        unit_price = self.money_cents(data.get("unitPrice"), "Valor unitário")
        discount = self.money_cents(data.get("discount"), "Desconto")
        gross = int((Decimal(quantity) * Decimal(unit_price) / INVENTORY_QUANTITY_SCALE)
                    .quantize(Decimal("1"), rounding=ROUND_HALF_UP))
        if discount > gross:
            raise ValueError("O desconto não pode superar o valor bruto do item")
        description = str(data.get("description") or catalog["title"]).strip()[:500]
        if not description:
            raise ValueError("Informe a descrição do item")
        warehouse_id = None
        lot_key = ""
        if item_kind == "PRODUCT":
            if data.get("warehouseId") not in (None, ""):
                try:
                    warehouse_id = int(data.get("warehouseId"))
                except (ValueError, TypeError):
                    raise ValueError("Depósito inválido") from None
                self.inventory_scope(company_id, warehouse_id, catalog_id)
            lot_key = self.inventory_lot(data.get("lot"))
        elif data.get("warehouseId") not in (None, "") or str(data.get("lot") or "").strip():
            raise ValueError("Serviços não usam depósito ou lote")
        notes = str(data.get("notes") or "").strip()[:1000] or None
        return {
            "item_kind": item_kind, "catalog_id": catalog_id,
            "description": description, "quantity": quantity,
            "unit_price": unit_price, "discount": discount,
            "total": gross - discount, "warehouse_id": warehouse_id,
            "lot_key": lot_key, "notes": notes,
        }

    def touch_document_total(self, record_id, company_id, expected_revision, now):
        totals = self.document_totals(record_id, company_id)
        amount = totals["totalCents"] / 100 if totals["itemCount"] else None
        updated = self.db.connection().execute(
            """UPDATE records SET amount=?,revision=revision+1,updated_at=?
               WHERE id=? AND company_id=? AND revision=? AND deleted_at IS NULL""",
            (amount, now, record_id, company_id, expected_revision),
        )
        if updated.rowcount != 1:
            raise RuntimeError("Este documento foi alterado por outra pessoa. Recarregue antes de continuar.")
        self.db.connection().execute(
            """UPDATE approvals SET status='Expirada',decided_at=?,
               decision_comment='Itens alterados após a solicitação.'
               WHERE record_id=? AND company_id=? AND status='Pendente'""",
            (now, record_id, company_id),
        )
        return totals, expected_revision + 1

    def record_item_create(self, record_id, session):
        record = self.document_record(record_id, session, "write")
        if not record:
            return
        if not self.require_operation(session, record["module"], "manage_items"):
            return
        try:
            data = self.parse_json(max_bytes=64 * 1024)
            expected_revision = int(data.get("recordRevision"))
        except (ValueError, TypeError) as exc:
            return self.error_json(
                "A revisão atual do documento é obrigatória", 409, "revision_required",
            )
        company_id = session["company_id"]
        now = utc_now()
        try:
            with self.db.transaction(immediate=True):
                current = self.db.connection().execute(
                    """SELECT revision FROM records
                       WHERE id=? AND company_id=? AND deleted_at IS NULL""",
                    (record_id, company_id),
                ).fetchone()
                if not current or current["revision"] != expected_revision:
                    raise RuntimeError("Este documento foi alterado por outra pessoa. Recarregue antes de continuar.")
                values = self.parse_document_item(data, company_id)
                sort_order = self.db.scalar(
                    "SELECT COALESCE(MAX(sort_order),0)+10 FROM document_items WHERE record_id=? AND company_id=?",
                    (record_id, company_id),
                )
                item_id = self.db.connection().execute(
                    """INSERT INTO document_items
                       (company_id,record_id,item_kind,catalog_record_id,description,
                        quantity_micros,unit_price_cents,discount_cents,total_cents,
                        warehouse_id,lot_key,notes,sort_order,revision,created_by,created_at,updated_at)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,1,?,?,?)""",
                    (company_id, record_id, values["item_kind"], values["catalog_id"],
                     values["description"], values["quantity"], values["unit_price"],
                     values["discount"], values["total"], values["warehouse_id"],
                     values["lot_key"], values["notes"], sort_order, session["id"], now, now),
                ).lastrowid
                totals, revision = self.touch_document_total(
                    record_id, company_id, expected_revision, now,
                )
                self.db.audit(
                    session["id"], "create", "document_item", item_id,
                    {"record_id": record_id, "module": record["module"],
                     "catalog_record_id": values["catalog_id"],
                     "total_cents": values["total"]}, company_id=company_id,
                )
        except ValueError as exc:
            return self.error_json(str(exc))
        except (RuntimeError, sqlite3.IntegrityError) as exc:
            return self.error_json(str(exc), 409, "write_conflict")
        return self.send_json({"ok": True, "id": item_id, "totals": totals,
                               "recordRevision": revision}, 201)

    def record_item_write(self, method, record_id, item_id, session):
        record = self.document_record(record_id, session, "write")
        if not record:
            return
        if not self.require_operation(session, record["module"], "manage_items"):
            return
        try:
            data = self.parse_json(max_bytes=64 * 1024)
            expected_record_revision = int(data.get("recordRevision"))
            expected_item_revision = int(data.get("itemRevision"))
        except (ValueError, TypeError):
            return self.error_json(
                "As revisões do documento e do item são obrigatórias", 409, "revision_required",
            )
        company_id = session["company_id"]
        now = utc_now()
        try:
            with self.db.transaction(immediate=True):
                current = self.db.connection().execute(
                    """SELECT i.*,q.status reservation_status,
                              EXISTS(
                                SELECT 1 FROM inventory_movements m
                                WHERE m.company_id=i.company_id AND m.movement_type='PURCHASE_IN'
                                  AND m.origin_type='PURCHASE_ORDER'
                                  AND m.origin_id=CAST(i.record_id AS TEXT)||':'||CAST(i.id AS TEXT)
                              ) receipt_processed
                       FROM document_items i
                       LEFT JOIN inventory_reservations q ON q.id=i.reservation_id
                       WHERE i.id=? AND i.record_id=? AND i.company_id=?""",
                    (item_id, record_id, company_id),
                ).fetchone()
                if not current:
                    raise RuntimeError("Item não encontrado neste documento")
                current_record = self.db.connection().execute(
                    """SELECT revision FROM records
                       WHERE id=? AND company_id=? AND deleted_at IS NULL""",
                    (record_id, company_id),
                ).fetchone()
                if (current["revision"] != expected_item_revision
                        or not current_record
                        or current_record["revision"] != expected_record_revision):
                    raise RuntimeError("O documento ou item foi alterado. Recarregue antes de continuar.")
                if current["reservation_status"] == "ACTIVE":
                    raise RuntimeError("Libere a reserva de estoque antes de alterar ou excluir o item")
                if current["reservation_status"] == "FULFILLED":
                    raise RuntimeError("O item já foi baixado no estoque e não pode mais ser alterado")
                if current["receipt_processed"]:
                    raise RuntimeError("O item já foi recebido no estoque e não pode mais ser alterado")
                if method == "DELETE":
                    deleted = self.db.connection().execute(
                        """DELETE FROM document_items
                           WHERE id=? AND record_id=? AND company_id=? AND revision=?""",
                        (item_id, record_id, company_id, expected_item_revision),
                    )
                    if deleted.rowcount != 1:
                        raise RuntimeError("O item foi alterado por outra pessoa")
                    action = "delete"
                    audit_detail = {"record_id": record_id, "module": record["module"]}
                else:
                    values = self.parse_document_item(data, company_id)
                    updated = self.db.connection().execute(
                        """UPDATE document_items
                           SET item_kind=?,catalog_record_id=?,description=?,quantity_micros=?,
                               unit_price_cents=?,discount_cents=?,total_cents=?,warehouse_id=?,
                               lot_key=?,notes=?,revision=revision+1,updated_at=?
                           WHERE id=? AND record_id=? AND company_id=? AND revision=?""",
                        (values["item_kind"], values["catalog_id"], values["description"],
                         values["quantity"], values["unit_price"], values["discount"],
                         values["total"], values["warehouse_id"], values["lot_key"],
                         values["notes"], now, item_id, record_id, company_id,
                         expected_item_revision),
                    )
                    if updated.rowcount != 1:
                        raise RuntimeError("O item foi alterado por outra pessoa")
                    action = "update"
                    audit_detail = {"record_id": record_id, "module": record["module"],
                                    "catalog_record_id": values["catalog_id"],
                                    "total_cents": values["total"]}
                totals, revision = self.touch_document_total(
                    record_id, company_id, expected_record_revision, now,
                )
                self.db.audit(
                    session["id"], action, "document_item", item_id,
                    audit_detail, company_id=company_id,
                )
        except ValueError as exc:
            return self.error_json(str(exc))
        except (RuntimeError, sqlite3.IntegrityError) as exc:
            return self.error_json(str(exc), 409, "write_conflict")
        return self.send_json({"ok": True, "totals": totals, "recordRevision": revision})

    def record_items_reservation(self, record_id, action, session):
        record = self.document_record(record_id, session, "write")
        if not record:
            return
        if record["module"] not in RESERVABLE_ITEM_MODULES:
            return self.error_json(
                "Somente vendas e ordens de serviço reservam estoque",
                409, "reservation_not_supported",
            )
        operation = {
            "reserve-items": "reserve_stock",
            "release-items": "release_stock",
            "fulfill-items": "fulfill_stock",
        }[action]
        if (not self.require_operation(session, record["module"], operation)
                or not self.require_operation(
                    session, "estoque",
                    "release_stock" if action == "release-items" else
                    "reserve_stock" if action == "reserve-items" else "move_stock",
                )):
            return
        reservable_statuses = {
            "vendas": {"Confirmado", "Separação"},
            "ordens_servico": {"Agendada", "Em execução", "Aguardando aprovação"},
        }
        fulfillable_statuses = {
            "vendas": {"Separação", "Faturado"},
            "ordens_servico": {"Em execução", "Aguardando aprovação"},
        }
        if action == "reserve-items" and record["status"] not in reservable_statuses[record["module"]]:
            return self.error_json(
                "Confirme a venda ou agende a O.S. antes de reservar estoque",
                409, "document_status_not_reservable",
            )
        if action == "fulfill-items" and record["status"] not in fulfillable_statuses[record["module"]]:
            return self.error_json(
                "Coloque a venda em separação ou a O.S. em execução antes de baixar o estoque",
                409, "document_status_not_fulfillable",
            )
        company_id = session["company_id"]
        now = utc_now()
        movement_ids = []
        try:
            with self.db.transaction(immediate=True):
                items = self.db.connection().execute(
                    """SELECT i.*,q.status reservation_status
                       FROM document_items i
                       LEFT JOIN inventory_reservations q ON q.id=i.reservation_id
                       WHERE i.record_id=? AND i.company_id=? AND i.item_kind='PRODUCT'
                       ORDER BY i.sort_order,i.id""",
                    (record_id, company_id),
                ).fetchall()
                if not items:
                    raise ValueError("O documento não possui produtos para movimentar")
                changed = 0
                for item in items:
                    if action == "reserve-items":
                        if item["reservation_status"] == "ACTIVE":
                            continue
                        if item["reservation_status"] == "FULFILLED":
                            raise ValueError(
                                f'O estoque de “{item["description"]}” já foi baixado'
                            )
                        if not item["warehouse_id"]:
                            raise ValueError(
                                f'Defina o depósito do item “{item["description"]}” antes de reservar'
                            )
                        self.inventory_scope(
                            company_id, item["warehouse_id"], item["catalog_record_id"],
                        )
                        balance = self.inventory_balance(
                            company_id, item["warehouse_id"], item["catalog_record_id"],
                            item["lot_key"], now,
                        )
                        quantity = int(item["quantity_micros"])
                        available = (int(balance["physical_quantity_micros"])
                                     - int(balance["reserved_quantity_micros"]))
                        if available < quantity:
                            raise ValueError(
                                f'Saldo disponível insuficiente para “{item["description"]}”'
                            )
                        origin_type = "SALES_ORDER" if record["module"] == "vendas" else "SERVICE_ORDER"
                        origin_id = f"{record_id}:{item['id']}"
                        reservation_id = self.db.connection().execute(
                            """INSERT INTO inventory_reservations
                               (company_id,warehouse_id,product_record_id,lot_key,quantity_micros,
                                status,origin_type,origin_id,reference,created_by,created_at,updated_at)
                               VALUES(?,?,?,?,?,'ACTIVE',?,?,?,?,?,?)""",
                            (company_id, item["warehouse_id"], item["catalog_record_id"],
                             item["lot_key"], quantity, origin_type, origin_id,
                             record["title"], session["id"], now, now),
                        ).lastrowid
                        self.db.connection().execute(
                            """UPDATE inventory_balances
                               SET reserved_quantity_micros=reserved_quantity_micros+?,
                                   revision=revision+1,updated_at=?
                               WHERE company_id=? AND warehouse_id=?
                                 AND product_record_id=? AND lot_key=?""",
                            (quantity, now, company_id, item["warehouse_id"],
                             item["catalog_record_id"], item["lot_key"]),
                        )
                        movement_ids.append(self.inventory_log_movement(
                            company_id=company_id, warehouse_id=item["warehouse_id"],
                            product_id=item["catalog_record_id"], lot_key=item["lot_key"],
                            movement_type="RESERVE", quantity_micros=quantity,
                            physical_delta_micros=0, reserved_delta_micros=quantity,
                            origin_type=origin_type, origin_id=origin_id,
                            reference=record["title"], reason=None,
                            reservation_id=reservation_id, created_by=session["id"],
                            created_at=now,
                            balance_value_cents=int(balance["inventory_value_cents"] or 0),
                        ))
                        self.db.connection().execute(
                            """UPDATE document_items SET reservation_id=?,revision=revision+1,
                               updated_at=? WHERE id=? AND company_id=?""",
                            (reservation_id, now, item["id"], company_id),
                        )
                        changed += 1
                    elif action == "release-items":
                        if item["reservation_status"] != "ACTIVE" or not item["reservation_id"]:
                            continue
                        quantity = int(item["quantity_micros"])
                        balance = self.inventory_balance(
                            company_id, item["warehouse_id"], item["catalog_record_id"],
                            item["lot_key"], now,
                        )
                        if int(balance["reserved_quantity_micros"]) < quantity:
                            raise ValueError("Saldo reservado inconsistente; liberação cancelada")
                        self.db.connection().execute(
                            """UPDATE inventory_balances
                               SET reserved_quantity_micros=reserved_quantity_micros-?,
                                   revision=revision+1,updated_at=?
                               WHERE company_id=? AND warehouse_id=?
                                 AND product_record_id=? AND lot_key=?""",
                            (quantity, now, company_id, item["warehouse_id"],
                             item["catalog_record_id"], item["lot_key"]),
                        )
                        self.db.connection().execute(
                            """UPDATE inventory_reservations
                               SET status='RELEASED',released_by=?,updated_at=?
                               WHERE id=? AND company_id=? AND status='ACTIVE'""",
                            (session["id"], now, item["reservation_id"], company_id),
                        )
                        reservation = self.db.connection().execute(
                            """SELECT origin_type,origin_id,reference FROM inventory_reservations
                               WHERE id=? AND company_id=?""",
                            (item["reservation_id"], company_id),
                        ).fetchone()
                        movement_ids.append(self.inventory_log_movement(
                            company_id=company_id, warehouse_id=item["warehouse_id"],
                            product_id=item["catalog_record_id"], lot_key=item["lot_key"],
                            movement_type="RELEASE_RESERVATION", quantity_micros=quantity,
                            physical_delta_micros=0, reserved_delta_micros=-quantity,
                            origin_type=reservation["origin_type"], origin_id=reservation["origin_id"],
                            reference=reservation["reference"],
                            reason="Liberação pelo documento de origem",
                            reservation_id=item["reservation_id"], created_by=session["id"],
                            created_at=now,
                            balance_value_cents=int(balance["inventory_value_cents"] or 0),
                        ))
                        self.db.connection().execute(
                            """UPDATE document_items SET reservation_id=NULL,revision=revision+1,
                               updated_at=? WHERE id=? AND company_id=?""",
                            (now, item["id"], company_id),
                        )
                        changed += 1
                    else:
                        if item["reservation_status"] != "ACTIVE" or not item["reservation_id"]:
                            raise ValueError(
                                f'Reserve “{item["description"]}” antes de baixar o estoque'
                            )
                        quantity = int(item["quantity_micros"])
                        balance = self.inventory_balance(
                            company_id, item["warehouse_id"], item["catalog_record_id"],
                            item["lot_key"], now,
                        )
                        if (int(balance["reserved_quantity_micros"]) < quantity
                                or int(balance["physical_quantity_micros"]) < quantity):
                            raise ValueError("Saldo físico ou reservado inconsistente; baixa cancelada")
                        inventory_value = int(balance["inventory_value_cents"] or 0)
                        issue_value = self.inventory_proportional_value(
                            inventory_value, quantity,
                            int(balance["physical_quantity_micros"]),
                        )
                        new_inventory_value = inventory_value - issue_value
                        updated_balance = self.db.connection().execute(
                            """UPDATE inventory_balances
                               SET physical_quantity_micros=physical_quantity_micros-?,
                                   reserved_quantity_micros=reserved_quantity_micros-?,
                                   inventory_value_cents=?,
                                   revision=revision+1,updated_at=?
                               WHERE company_id=? AND warehouse_id=?
                                 AND product_record_id=? AND lot_key=?
                                 AND physical_quantity_micros>=?
                                 AND reserved_quantity_micros>=?""",
                            (quantity, quantity, new_inventory_value, now,
                             company_id, item["warehouse_id"],
                             item["catalog_record_id"], item["lot_key"], quantity, quantity),
                        )
                        fulfilled = self.db.connection().execute(
                            """UPDATE inventory_reservations
                               SET status='FULFILLED',released_by=?,updated_at=?
                               WHERE id=? AND company_id=? AND status='ACTIVE'""",
                            (session["id"], now, item["reservation_id"], company_id),
                        )
                        if updated_balance.rowcount != 1 or fulfilled.rowcount != 1:
                            raise ValueError("A reserva mudou durante a baixa; operação cancelada")
                        reservation = self.db.connection().execute(
                            """SELECT origin_type,origin_id,reference FROM inventory_reservations
                               WHERE id=? AND company_id=?""",
                            (item["reservation_id"], company_id),
                        ).fetchone()
                        movement_type = (
                            "SALE_OUT" if record["module"] == "vendas" else "SERVICE_ORDER_OUT"
                        )
                        movement_ids.append(self.inventory_log_movement(
                            company_id=company_id, warehouse_id=item["warehouse_id"],
                            product_id=item["catalog_record_id"], lot_key=item["lot_key"],
                            movement_type=movement_type, quantity_micros=quantity,
                            physical_delta_micros=-quantity, reserved_delta_micros=-quantity,
                            origin_type=reservation["origin_type"], origin_id=reservation["origin_id"],
                            reference=reservation["reference"],
                            reason="Baixa pelo documento de origem",
                            reservation_id=item["reservation_id"], created_by=session["id"],
                            created_at=now,
                            unit_cost_cents=self.inventory_average_cost(issue_value, quantity),
                            value_delta_cents=-issue_value,
                            balance_value_cents=new_inventory_value,
                        ))
                        self.db.connection().execute(
                            """UPDATE document_items SET revision=revision+1,updated_at=?
                               WHERE id=? AND company_id=?""",
                            (now, item["id"], company_id),
                        )
                        changed += 1
                if not changed:
                    messages = {
                        "reserve-items": "Todos os produtos já estão reservados",
                        "release-items": "Não há reservas ativas neste documento",
                        "fulfill-items": "Não há produtos reservados para baixar",
                    }
                    raise ValueError(messages[action])
                audit_actions = {
                    "reserve-items": "reserve", "release-items": "release",
                    "fulfill-items": "fulfill",
                }
                self.db.audit(
                    session["id"], audit_actions[action],
                    record["module"], record_id,
                    {"items": changed, "movement_ids": movement_ids}, company_id=company_id,
                )
        except (ValueError, sqlite3.IntegrityError) as exc:
            return self.error_json(str(exc), 409, "inventory_conflict")
        return self.send_json({"ok": True, "items": changed, "movementIds": movement_ids})

    def record_items_receive(self, record_id, session):
        record = self.document_record(record_id, session, "write")
        if not record:
            return
        if record["module"] != "pedidos_compra":
            return self.error_json(
                "Somente pedidos de compra geram recebimento de estoque",
                409, "receiving_not_supported",
            )
        if (not self.require_operation(session, "pedidos_compra", "receive_stock")
                or not self.require_operation(session, "estoque", "move_stock")):
            return
        if record["status"] not in {"Emitido", "Aguardando fornecedor", "Recebido parcial"}:
            return self.error_json(
                "Emita o pedido de compra antes de receber os produtos",
                409, "document_status_not_receivable",
            )
        company_id = session["company_id"]
        now = utc_now()
        movement_ids = []
        try:
            with self.db.transaction(immediate=True):
                current = self.db.connection().execute(
                    """SELECT id,title,status FROM records
                       WHERE id=? AND company_id=? AND module='pedidos_compra'
                         AND deleted_at IS NULL""",
                    (record_id, company_id),
                ).fetchone()
                if (not current
                        or current["status"] not in {"Emitido", "Aguardando fornecedor", "Recebido parcial"}):
                    raise ValueError("O pedido mudou de etapa durante o recebimento")
                items = self.db.connection().execute(
                    """SELECT i.*,
                              EXISTS(
                                SELECT 1 FROM inventory_movements m
                                WHERE m.company_id=i.company_id AND m.movement_type='PURCHASE_IN'
                                  AND m.origin_type='PURCHASE_ORDER'
                                  AND m.origin_id=CAST(i.record_id AS TEXT)||':'||CAST(i.id AS TEXT)
                              ) receipt_processed
                       FROM document_items i
                       WHERE i.record_id=? AND i.company_id=? AND i.item_kind='PRODUCT'
                       ORDER BY i.sort_order,i.id""",
                    (record_id, company_id),
                ).fetchall()
                if not items:
                    raise ValueError("O pedido não possui produtos para receber")
                for item in items:
                    if item["receipt_processed"]:
                        continue
                    if not item["warehouse_id"]:
                        raise ValueError(
                            f'Defina o depósito de “{item["description"]}” antes de receber'
                        )
                    self.inventory_scope(
                        company_id, item["warehouse_id"], item["catalog_record_id"],
                    )
                    quantity = int(item["quantity_micros"])
                    balance = self.inventory_balance(
                        company_id, item["warehouse_id"], item["catalog_record_id"],
                        item["lot_key"], now,
                    )
                    received_value = int(item["total_cents"] or 0)
                    new_inventory_value = (
                        int(balance["inventory_value_cents"] or 0) + received_value
                    )
                    updated = self.db.connection().execute(
                        """UPDATE inventory_balances
                           SET physical_quantity_micros=physical_quantity_micros+?,
                               inventory_value_cents=?,
                               revision=revision+1,updated_at=?
                           WHERE company_id=? AND warehouse_id=?
                             AND product_record_id=? AND lot_key=?""",
                        (quantity, new_inventory_value, now, company_id, item["warehouse_id"],
                         item["catalog_record_id"], item["lot_key"]),
                    )
                    if updated.rowcount != 1:
                        raise ValueError("O saldo mudou durante o recebimento; operação cancelada")
                    origin_id = f"{record_id}:{item['id']}"
                    movement_ids.append(self.inventory_log_movement(
                        company_id=company_id, warehouse_id=item["warehouse_id"],
                        product_id=item["catalog_record_id"], lot_key=item["lot_key"],
                        movement_type="PURCHASE_IN", quantity_micros=quantity,
                        physical_delta_micros=quantity, reserved_delta_micros=0,
                        origin_type="PURCHASE_ORDER", origin_id=origin_id,
                        reference=current["title"], reason="Recebimento do pedido de compra",
                        reservation_id=None, created_by=session["id"], created_at=now,
                        unit_cost_cents=self.inventory_average_cost(received_value, quantity),
                        value_delta_cents=received_value,
                        balance_value_cents=new_inventory_value,
                    ))
                    self.db.connection().execute(
                        """UPDATE document_items SET revision=revision+1,updated_at=?
                           WHERE id=? AND company_id=?""",
                        (now, item["id"], company_id),
                    )
                if not movement_ids:
                    raise ValueError("Todos os produtos deste pedido já foram recebidos")
                self.db.audit(
                    session["id"], "receive", "pedidos_compra", record_id,
                    {"items": len(movement_ids), "movement_ids": movement_ids},
                    company_id=company_id,
                )
        except (ValueError, sqlite3.IntegrityError) as exc:
            return self.error_json(str(exc), 409, "inventory_conflict")
        return self.send_json({
            "ok": True, "items": len(movement_ids), "movementIds": movement_ids,
        })

    def notifications_read(self, session):
        self.db.execute(
            """UPDATE notifications SET read_at=?
               WHERE company_id=? AND (user_id IS NULL OR user_id=?) AND read_at IS NULL""",
            (utc_now(), session["company_id"], session["id"]))
        return self.send_json({"ok": True})

    @staticmethod
    def record_amount_cents(value):
        if value in (None, ""):
            return 0
        try:
            amount = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        except (InvalidOperation, ValueError, TypeError):
            return 0
        return int(amount * 100)

    def management_overview(self, session):
        if not self.require_module_read(session, "controladoria"):
            return
        company_id = session["company_id"]
        db = self.db.connection()
        readable = self.allowed_modules(session, "read")
        operations = self.allowed_operations(session, "controladoria")

        def values_allowed(module):
            return (module in readable and
                    "view_values" in self.allowed_operations(session, module))

        visibility = {
            "billing": "view_billing" in operations and values_allowed("vendas"),
            "cashflow": "view_cashflow" in operations and (
                values_allowed("caixa") or values_allowed("financeiro")
            ),
            "inventoryValue": (
                "view_inventory_value" in operations and values_allowed("estoque")
            ),
            "overdue": "view_overdue" in operations and (
                values_allowed("contas_receber") or values_allowed("contas_pagar")
            ),
        }

        def amount_total(module, statuses=None, payload_type=None, due_before=None):
            if not values_allowed(module):
                return 0, 0
            sql = """SELECT amount FROM records
                     WHERE company_id=? AND module=? AND deleted_at IS NULL"""
            params = [company_id, module]
            if statuses:
                placeholders = ",".join("?" for _ in statuses)
                sql += f" AND status IN ({placeholders})"
                params.extend(statuses)
            if payload_type:
                field, value = payload_type
                sql += f" AND json_extract(payload,'$.{field}')=?"
                params.append(value)
            if due_before:
                sql += " AND due_date<?"
                params.append(due_before)
            rows = db.execute(sql, params).fetchall()
            return sum(self.record_amount_cents(row["amount"]) for row in rows), len(rows)

        billing_total, billing_count = amount_total(
            "vendas", ("Faturado", "Concluído"),
        )
        sales_open_total, sales_open_count = amount_total(
            "vendas", ("Confirmado", "Separação"),
        )
        receivable_open, receivable_count = amount_total(
            "contas_receber", ("Em aberto", "Parcial", "Vencido"),
        )
        payable_open, payable_count = amount_total(
            "contas_pagar", ("Em aberto", "Parcial", "Vencido"),
        )
        today = datetime.now(timezone.utc).date().isoformat()
        receivable_overdue, receivable_overdue_count = amount_total(
            "contas_receber", ("Em aberto", "Parcial", "Vencido"), due_before=today,
        )
        payable_overdue, payable_overdue_count = amount_total(
            "contas_pagar", ("Em aberto", "Parcial", "Vencido"), due_before=today,
        )
        cash_source = "caixa"
        cash_in, cash_in_count = amount_total(
            "caixa", payload_type=("tipo_movimento", "Entrada"),
        )
        cash_out, cash_out_count = amount_total(
            "caixa", payload_type=("tipo_movimento", "Saída"),
        )
        if values_allowed("financeiro") and not (cash_in_count or cash_out_count):
            cash_source = "financeiro"
            cash_in, cash_in_count = amount_total(
                "financeiro", payload_type=("tipo_lancamento", "Receita"),
            )
            cash_out, cash_out_count = amount_total(
                "financeiro", payload_type=("tipo_lancamento", "Despesa"),
            )

        inventory = db.execute(
            """SELECT COALESCE(SUM(inventory_value_cents),0) inventory_value,
                      SUM(CASE WHEN physical_quantity_micros>0
                                    AND inventory_value_cents=0 THEN 1 ELSE 0 END) unvalued
               FROM inventory_balances WHERE company_id=?""",
            (company_id,),
        ).fetchone()
        reserved_value = 0
        if visibility["inventoryValue"]:
            for row in db.execute(
                """SELECT physical_quantity_micros,reserved_quantity_micros,
                          inventory_value_cents FROM inventory_balances WHERE company_id=?""",
                (company_id,),
            ).fetchall():
                reserved_value += self.inventory_proportional_value(
                    int(row["inventory_value_cents"] or 0),
                    int(row["reserved_quantity_micros"] or 0),
                    int(row["physical_quantity_micros"] or 0),
                )
        inventory_value = int(inventory["inventory_value"] or 0)
        cost_of_sales = int(self.db.scalar(
            """SELECT COALESCE(-SUM(value_delta_cents),0) FROM inventory_movements
               WHERE company_id=? AND movement_type IN ('SALE_OUT','SERVICE_ORDER_OUT')""",
            (company_id,),
        ) or 0)

        month_keys = []
        cursor = datetime.now(timezone.utc).date().replace(day=1)
        for offset in range(5, -1, -1):
            year = cursor.year
            month = cursor.month - offset
            while month <= 0:
                year -= 1
                month += 12
            month_keys.append(f"{year:04d}-{month:02d}")
        series = []
        for month in month_keys:
            row = {"month": month, "billingCents": None, "cashInCents": None,
                   "cashOutCents": None}
            if visibility["billing"]:
                amounts = db.execute(
                    """SELECT amount FROM records WHERE company_id=? AND module='vendas'
                       AND deleted_at IS NULL AND status IN ('Faturado','Concluído')
                       AND substr(updated_at,1,7)=?""",
                    (company_id, month),
                ).fetchall()
                row["billingCents"] = sum(
                    self.record_amount_cents(item["amount"]) for item in amounts
                )
            if visibility["cashflow"]:
                if cash_source == "financeiro":
                    rows = db.execute(
                        """SELECT amount,json_extract(payload,'$.tipo_lancamento') movement
                           FROM records WHERE company_id=? AND module='financeiro'
                             AND deleted_at IS NULL AND substr(created_at,1,7)=?""",
                        (company_id, month),
                    ).fetchall()
                    incoming, outgoing = "Receita", "Despesa"
                else:
                    rows = db.execute(
                        """SELECT amount,json_extract(payload,'$.tipo_movimento') movement
                           FROM records WHERE company_id=? AND module='caixa'
                             AND deleted_at IS NULL AND substr(created_at,1,7)=?""",
                        (company_id, month),
                    ).fetchall()
                    incoming, outgoing = "Entrada", "Saída"
                row["cashInCents"] = sum(
                    self.record_amount_cents(item["amount"])
                    for item in rows if item["movement"] == incoming
                )
                row["cashOutCents"] = sum(
                    self.record_amount_cents(item["amount"])
                    for item in rows if item["movement"] == outgoing
                )
            series.append(row)

        return self.send_json({
            "ok": True,
            "visibility": visibility,
            "billing": {
                "totalCents": billing_total if visibility["billing"] else None,
                "count": billing_count if visibility["billing"] else None,
                "openOrdersCents": sales_open_total if visibility["billing"] else None,
                "openOrdersCount": sales_open_count if visibility["billing"] else None,
                "costOfSalesCents": cost_of_sales if (
                    visibility["billing"] and visibility["inventoryValue"]
                ) else None,
                "grossContributionCents": (
                    billing_total - cost_of_sales
                ) if (visibility["billing"] and visibility["inventoryValue"]) else None,
            },
            "cashflow": {
                "cashInCents": cash_in if visibility["cashflow"] else None,
                "cashOutCents": cash_out if visibility["cashflow"] else None,
                "balanceCents": (cash_in - cash_out) if visibility["cashflow"] else None,
                "receivableOpenCents": receivable_open if visibility["cashflow"] else None,
                "payableOpenCents": payable_open if visibility["cashflow"] else None,
                "receivableCount": receivable_count if visibility["cashflow"] else None,
                "payableCount": payable_count if visibility["cashflow"] else None,
            },
            "inventory": {
                "totalValueCents": inventory_value if visibility["inventoryValue"] else None,
                "reservedValueCents": reserved_value if visibility["inventoryValue"] else None,
                "availableValueCents": (
                    inventory_value - reserved_value
                ) if visibility["inventoryValue"] else None,
                "unvaluedBalances": int(inventory["unvalued"] or 0)
                if visibility["inventoryValue"] else None,
            },
            "overdue": {
                "receivableCents": receivable_overdue if visibility["overdue"] else None,
                "payableCents": payable_overdue if visibility["overdue"] else None,
                "receivableCount": receivable_overdue_count if visibility["overdue"] else None,
                "payableCount": payable_overdue_count if visibility["overdue"] else None,
            },
            "series": series,
            "asOf": utc_now(),
        })

    def dashboard(self, session):
        db = self.db.connection()
        company_id = session["company_id"]
        readable = sorted(self.allowed_modules(session, "read"))
        if not readable:
            return self.send_json({
                "ok": True, "counts": {}, "income": 0, "expense": 0, "alerts": [],
                "recent": [], "operationalTotal": 0, "pendingApprovals": 0,
                "unreadNotifications": 0, "financialVisible": False, "workItems": [],
            })
        placeholders = ",".join("?" for _ in readable)
        counts = {row["module"]: row["total"] for row in db.execute(
            f"""SELECT module,COUNT(*) total FROM records
               WHERE company_id=? AND deleted_at IS NULL AND module IN ({placeholders})
               GROUP BY module""", (company_id, *readable)).fetchall()}
        operational_total = self.db.scalar(
            f"""SELECT COUNT(*) FROM records WHERE company_id=? AND deleted_at IS NULL
               AND module IN ({placeholders})
               AND module NOT IN ('fontes','normas_tecnicas')
               AND COALESCE(json_extract(payload,'$.catalogo_seccol'),0)!=1""",
            (company_id, *readable),
        ) or 0
        financial_visible = any(
            module in readable and "view_values" in self.allowed_operations(session, module)
            for module in ("financeiro", "caixa")
        )
        financial = {"income": 0, "expense": 0}
        if ("financeiro" in readable
                and "view_values" in self.allowed_operations(session, "financeiro")
                and self.db.scalar(
            "SELECT COUNT(*) FROM records WHERE company_id=? AND module='financeiro' AND deleted_at IS NULL",
            (company_id,),
        )):
            financial = dict(db.execute(
                """SELECT
                   COALESCE(SUM(CASE WHEN json_extract(payload,'$.tipo_lancamento')='Receita'
                                     THEN ABS(COALESCE(amount,0)) ELSE 0 END),0) income,
                   COALESCE(SUM(CASE WHEN json_extract(payload,'$.tipo_lancamento')='Despesa'
                                     THEN ABS(COALESCE(amount,0)) ELSE 0 END),0) expense
                   FROM records WHERE company_id=? AND module='financeiro' AND deleted_at IS NULL""",
                (company_id,),
            ).fetchone())
        elif ("caixa" in readable
              and "view_values" in self.allowed_operations(session, "caixa")):
            financial = dict(db.execute(
                """SELECT
                   COALESCE(SUM(CASE WHEN json_extract(payload,'$.tipo_movimento')='Entrada'
                                     THEN ABS(COALESCE(amount,0)) ELSE 0 END),0) income,
                   COALESCE(SUM(CASE WHEN json_extract(payload,'$.tipo_movimento')='Saída'
                                     THEN ABS(COALESCE(amount,0)) ELSE 0 END),0) expense
                   FROM records WHERE company_id=? AND module='caixa' AND deleted_at IS NULL""",
                (company_id,),
            ).fetchone())
        alerts = [dict(row) for row in db.execute(
            f"""SELECT id,module,title,status,amount,due_date FROM records
               WHERE company_id=? AND deleted_at IS NULL AND due_date IS NOT NULL
               AND module IN ({placeholders})
               AND due_date <= date('now','+7 days')
               AND status NOT IN ('Concluído','Pago','Cancelado','Finalizado')
               ORDER BY due_date ASC LIMIT 8""", (company_id, *readable)
        ).fetchall()]
        for alert in alerts:
            if (alert["module"] in VALUE_SENSITIVE_MODULES and
                    "view_values" not in self.allowed_operations(session, alert["module"])):
                alert["amount"] = None
                alert["amountRestricted"] = True
        recent_rows = db.execute(
            f"""SELECT * FROM records WHERE company_id=? AND deleted_at IS NULL
               AND module IN ({placeholders})
               AND module NOT IN ('fontes','normas_tecnicas')
               AND COALESCE(json_extract(payload,'$.catalogo_seccol'),0)!=1
               ORDER BY updated_at DESC LIMIT 8""", (company_id, *readable)).fetchall()
        recent = self.records_json(recent_rows, session)
        approval_sql = f"""SELECT COUNT(*) FROM approvals a JOIN records r ON r.id=a.record_id
                            WHERE a.company_id=? AND a.status='Pendente'
                              AND r.module IN ({placeholders})"""
        approval_params = [company_id, *readable]
        if session["role"] not in {"admin", "manager", "approver"}:
            approval_sql += " AND (a.requested_by=? OR a.requested_to=?)"
            approval_params.extend([session["id"], session["id"]])
        pending_approvals = self.db.scalar(approval_sql, approval_params) or 0
        unread = self.db.scalar(
            """SELECT COUNT(*) FROM notifications WHERE company_id=? AND read_at IS NULL
               AND (user_id IS NULL OR user_id=?)""", (company_id, session["id"])) or 0
        approval_work_sql = f"""SELECT a.id approval_id,r.id record_id,r.module,r.title,
                                      a.approval_type,a.requested_at
                               FROM approvals a JOIN records r ON r.id=a.record_id
                               WHERE a.company_id=? AND a.status='Pendente' AND r.deleted_at IS NULL
                                 AND r.module IN ({placeholders})"""
        approval_work_params = [company_id, *readable]
        if session["role"] not in {"admin", "manager", "approver"}:
            approval_work_sql += " AND (a.requested_by=? OR a.requested_to=?)"
            approval_work_params.extend([session["id"], session["id"]])
        approval_work_sql += " ORDER BY a.requested_at ASC LIMIT 6"
        approval_work = db.execute(approval_work_sql, approval_work_params).fetchall()
        work_items = [{
            "kind": "approval", "priority": "high", "target": "aprovacoes",
            "recordId": row["record_id"], "module": row["module"], "title": row["title"],
            "meta": f'Aprovação: {row["approval_type"]}', "dueDate": None,
        } for row in approval_work]
        known_records = {item["recordId"] for item in work_items}
        today = datetime.now(timezone.utc).date().isoformat()
        for row in alerts:
            if row["id"] in known_records:
                continue
            overdue = str(row.get("due_date") or "") < today
            work_items.append({
                "kind": "overdue" if overdue else "deadline",
                "priority": "critical" if overdue else "normal",
                "target": row["module"], "recordId": row["id"], "module": row["module"],
                "title": row["title"],
                "meta": "Prazo vencido" if overdue else "Prazo próximo",
                "dueDate": row.get("due_date"),
            })
            known_records.add(row["id"])
        if len(work_items) < 10:
            own_rows = db.execute(
                f"""SELECT id,module,title,status,updated_at FROM records
                    WHERE company_id=? AND deleted_at IS NULL AND created_by=?
                      AND module IN ({placeholders})
                      AND status NOT IN ('Concluído','Pago','Cancelado','Finalizado','Obsoleto')
                      AND COALESCE(json_extract(payload,'$.catalogo_seccol'),0)!=1
                    ORDER BY updated_at DESC LIMIT 10""",
                (company_id, session["id"], *readable),
            ).fetchall()
            for row in own_rows:
                if row["id"] in known_records or len(work_items) >= 10:
                    continue
                work_items.append({
                    "kind": "followup", "priority": "low", "target": row["module"],
                    "recordId": row["id"], "module": row["module"], "title": row["title"],
                    "meta": f'Em acompanhamento · {row["status"]}', "dueDate": None,
                })
                known_records.add(row["id"])
        priority_order = {"critical": 0, "high": 1, "normal": 2, "low": 3}
        work_items.sort(key=lambda item: (priority_order[item["priority"]], item.get("dueDate") or "9999"))
        return self.send_json({"ok": True, "counts": counts, "income": financial["income"],
                               "expense": financial["expense"], "alerts": alerts, "recent": recent,
                               "operationalTotal": operational_total,
                               "pendingApprovals": pending_approvals, "unreadNotifications": unread,
                               "financialVisible": financial_visible, "workItems": work_items[:10]})

    def global_search(self, query, session):
        term = (query.get("q") or [""])[0].strip()
        if len(term) < 2:
            return self.send_json({"ok": True, "query": term, "items": []})
        if len(term) > 120:
            return self.error_json("A busca deve possuir no máximo 120 caracteres", 400, "invalid_search")
        readable = sorted(self.allowed_modules(session, "read"))
        if not readable:
            return self.send_json({"ok": True, "query": term, "items": []})
        escaped = term.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
        pattern = f"%{escaped}%"
        placeholders = ",".join("?" for _ in readable)
        rows = self.db.connection().execute(
            f"""SELECT id,module,title,status,due_date,updated_at
                FROM records WHERE company_id=? AND deleted_at IS NULL
                  AND module IN ({placeholders})
                  AND (title LIKE ? ESCAPE '\\' OR status LIKE ? ESCAPE '\\'
                       OR (module!='fontes' AND payload LIKE ? ESCAPE '\\'))
                ORDER BY CASE WHEN title LIKE ? ESCAPE '\\' THEN 0 ELSE 1 END,
                         updated_at DESC LIMIT 40""",
            (session["company_id"], *readable, pattern, pattern, pattern, pattern),
        ).fetchall()
        return self.send_json({
            "ok": True, "query": term,
            "items": [{
                "id": row["id"], "module": row["module"], "title": row["title"],
                "status": row["status"], "dueDate": row["due_date"], "updatedAt": row["updated_at"],
            } for row in rows],
        })

    @staticmethod
    def partner_lookup_json(url, headers=None):
        request = urllib.request.Request(
            url,
            headers={
                "Accept": "application/json",
                "User-Agent": f"SIVS/{VERSION}",
                **(headers or {}),
            },
        )
        with urllib.request.urlopen(request, timeout=PARTNER_LOOKUP_TIMEOUT) as response:
            return json.load(response)

    @staticmethod
    def partner_lookup_value(data, *path):
        current = data
        for key in path:
            if not isinstance(current, dict):
                return ""
            current = current.get(key)
        return str(current or "").strip()

    def partner_lookup(self, query, session):
        if not self.require_module_write(session, PARTY_MODULE):
            return
        cnpj = re.sub(r"\D", "", (query.get("cnpj") or [""])[0])
        cep = re.sub(r"\D", "", (query.get("cep") or [""])[0])
        if bool(cnpj) == bool(cep):
            return self.error_json("Informe exatamente um CNPJ ou um CEP para consulta")
        if cnpj:
            if not _valid_cnpj(cnpj):
                return self.error_json("CNPJ inválido")
            return self.partner_cnpj_lookup(cnpj, session)
        if len(cep) != 8:
            return self.error_json("CEP deve conter 8 dígitos")
        return self.partner_cep_lookup(cep, session)

    def partner_cnpj_lookup(self, cnpj, session):
        cache_key = ("cnpj", cnpj)
        cached = self.server.partner_lookup_cache_get(cache_key)  # type: ignore[attr-defined]
        if cached:
            return self.send_json({**cached, "cached": True})
        api_key = os.environ.get("CNPJA_API_KEY", "").strip()
        if not api_key:
            return self.send_json({
                "ok": True, "configured": False, "source": "CNPJá Comercial",
                "message": "Consulta de CNPJ não configurada. Preencha os dados manualmente.",
                "fields": {},
            })
        try:
            data = self.partner_lookup_json(
                f"https://api.cnpja.com/office/{cnpj}", {"Authorization": api_key}
            )
        except urllib.error.HTTPError as exc:
            if exc.code == 404:
                return self.error_json("CNPJ não encontrado na fonte cadastral", 404, "not_found")
            if exc.code in {401, 403}:
                return self.error_json("A credencial da consulta de CNPJ foi recusada", 503, "provider_unavailable")
            return self.error_json("A consulta de CNPJ está indisponível; preencha os dados manualmente.", 503, "provider_unavailable")
        except (urllib.error.URLError, TimeoutError, json.JSONDecodeError, OSError):
            return self.error_json("A consulta de CNPJ está indisponível; preencha os dados manualmente.", 503, "provider_unavailable")
        fields = {
            "razao_social": self.partner_lookup_value(data, "company", "name"),
            "nome_fantasia": self.partner_lookup_value(data, "alias"),
            "telefone": self.partner_lookup_value(data, "phones"),
            "email": self.partner_lookup_value(data, "emails"),
            "cep": re.sub(r"\D", "", self.partner_lookup_value(data, "address", "zip")),
            "logradouro": self.partner_lookup_value(data, "address", "street"),
            "bairro": self.partner_lookup_value(data, "address", "district"),
            "cidade": self.partner_lookup_value(data, "address", "city"),
            "uf": self.partner_lookup_value(data, "address", "state"),
        }
        # Telefones e e-mails podem vir como listas de objetos na API comercial.
        if isinstance(data.get("phones"), list):
            fields["telefone"] = next((
                f"{str(item.get('area') or '').strip()}{str(item.get('number') or '').strip()}"
                for item in data["phones"] if isinstance(item, dict) and item.get("number")
            ), "")
        if isinstance(data.get("emails"), list):
            fields["email"] = next((str(item.get("address") or "").strip() for item in data["emails"] if isinstance(item, dict) and item.get("address")), "")
        result = {"ok": True, "configured": True, "source": "CNPJá Comercial", "fields": {key: value for key, value in fields.items() if value}}
        self.server.partner_lookup_cache_put(cache_key, result)  # type: ignore[attr-defined]
        self.db.audit(session["id"], "partner_lookup", "cnpj", detail={"source": "cnpja"}, company_id=session["company_id"])
        return self.send_json({**result, "cached": False})

    def partner_cep_lookup(self, cep, session):
        cache_key = ("cep", cep)
        cached = self.server.partner_lookup_cache_get(cache_key)  # type: ignore[attr-defined]
        if cached:
            return self.send_json({**cached, "cached": True})
        try:
            data = self.partner_lookup_json(f"https://viacep.com.br/ws/{cep}/json/")
            if data.get("erro"):
                return self.error_json("CEP não encontrado", 404, "not_found")
        except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, json.JSONDecodeError, OSError):
            return self.error_json("A consulta de CEP está indisponível; preencha o endereço manualmente.", 503, "provider_unavailable")
        fields = {
            "logradouro": self.partner_lookup_value(data, "logradouro"),
            "bairro": self.partner_lookup_value(data, "bairro"),
            "cidade": self.partner_lookup_value(data, "localidade"),
            "uf": self.partner_lookup_value(data, "uf"),
        }
        result = {"ok": True, "configured": True, "source": "ViaCEP", "fields": {key: value for key, value in fields.items() if value}}
        self.server.partner_lookup_cache_put(cache_key, result)  # type: ignore[attr-defined]
        self.db.audit(session["id"], "partner_lookup", "cep", detail={"source": "viacep"}, company_id=session["company_id"])
        return self.send_json({**result, "cached": False})

    def records_get(self, path, query, session):
        company_id = session["company_id"]
        pieces = path.split("/")
        if len(pieces) == 4 and pieces[3].isdigit():
            row = self.db.connection().execute(
                "SELECT * FROM records WHERE id=? AND company_id=? AND deleted_at IS NULL",
                (int(pieces[3]), company_id)).fetchone()
            if not row:
                return self.error_json("Registro não encontrado", 404)
            if not self.require_module_read(session, row["module"]):
                return
            return self.send_json({"ok": True, "item": self.record_json(row, session)})
        module = (query.get("module") or [""])[0]
        if module not in MODULES:
            return self.error_json("Módulo inválido")
        if not self.require_module_read(session, module):
            return
        search = (query.get("q") or [""])[0].strip()
        status = (query.get("status") or [""])[0].strip()
        if module == PARTY_MODULE:
            readable_physical = [item for item in PARTY_PHYSICAL_MODULES
                                  if item in self.allowed_modules(session, "read")]
            if not readable_physical:
                return self.error_json("Seu perfil não possui permissão para consultar clientes ou fornecedores", 403, "forbidden")
            placeholders = ",".join("?" for _ in readable_physical)
            sql = f"SELECT * FROM records WHERE company_id=? AND module IN ({placeholders}) AND deleted_at IS NULL"
            params = [company_id, *readable_physical]
            if search:
                sql += " AND (title LIKE ? OR payload LIKE ?)"
                params += [f"%{search}%", f"%{search}%"]
            if status:
                sql += " AND status=?"
                params.append(status)
            sql += " ORDER BY updated_at DESC LIMIT 500"
            rows = self.db.connection().execute(sql, params).fetchall()
            return self.send_json({"ok": True, "items": self.records_json(rows, session)})
        sql = "SELECT * FROM records WHERE company_id=? AND module=? AND deleted_at IS NULL"
        params = [company_id, module]
        if search:
            sql += " AND (title LIKE ? OR payload LIKE ?)"
            params += [f"%{search}%", f"%{search}%"]
        if status:
            sql += " AND status=?"
            params.append(status)
        sql += " ORDER BY updated_at DESC LIMIT 500"
        rows = self.db.connection().execute(sql, params).fetchall()
        return self.send_json({"ok": True, "items": self.records_json(rows, session)})

    def assistant_query(self, session):
        """Responde perguntas usando apenas um contexto SQL filtrado no servidor."""
        try:
            data = self.parse_json(max_bytes=64 * 1024)
        except ValueError as exc:
            return self.error_json(str(exc))
        question = str(data.get("question") or "").strip()
        if len(question) < 3:
            return self.error_json("Escreva uma pergunta para o assistente")
        if len(question) > 800:
            return self.error_json("A pergunta deve possuir no máximo 800 caracteres")
        plan = self.assistant_plan(question, session)
        context = self.assistant_context(plan, session)
        ai_enabled = bool(os.environ.get("OPENROUTER_API_KEY"))
        model_used = "deterministic-context"
        notice = None
        if ai_enabled:
            try:
                result, model_used = self.openrouter_assistant(question, plan, context)
            except (OSError, ValueError, json.JSONDecodeError, urllib.error.URLError) as exc:
                result = self.assistant_fallback(question, plan, context)
                notice = "A IA não respondeu; exibindo o resultado seguro do SIVS."
                model_used = "deterministic-fallback"
                print(f"[AVISO ASSISTENTE] {type(exc).__name__}: {exc}")
        else:
            result = self.assistant_fallback(question, plan, context)
            notice = "Configure OPENROUTER_API_KEY para ativar a análise generativa."
        detail = {
            "question": question[:800], "intent": plan["intent"],
            "modules": plan["modules"], "source_count": len(context),
            "source_ids": [item["id"] for item in context[:100]],
            "model": model_used, "ai_enabled": ai_enabled,
            "response": str(result.get("answer", ""))[:4000],
        }
        self.db.audit(session["id"], "assistant_query", "assistant", detail=detail,
                      company_id=session["company_id"])
        return self.send_json({
            "ok": True, "answer": result.get("answer", "Não encontrei dados suficientes."),
            "confidence": result.get("confidence", "media"),
            "suggestions": result.get("suggestions", []),
            "sources": [{"id": item["id"], "module": item["module"], "title": item["title"]}
                        for item in context],
            "intent": plan["intent"], "model": model_used,
            "aiEnabled": ai_enabled, "notice": notice,
        })

    def assistant_plan(self, question, session):
        normalized = self.normalized_text(question)
        readable = self.allowed_modules(session, "read")
        today = datetime.now(timezone.utc).date()
        plan = {"intent": "search", "modules": [], "term": "", "start": None,
                "end": None, "threshold": None, "status_exclude": []}
        if "licit" in normalized and ("aderencia" in normalized or "70" in normalized or "pont" in normalized):
            plan.update(intent="tender_score", modules=["editais"], threshold=70)
            match = re.search(r"(\d{1,3})\s*%", normalized)
            if match:
                plan["threshold"] = min(100, max(0, int(match.group(1))))
        elif "proposta" in normalized and any(word in normalized for word in ("vence", "venc", "prazo")):
            plan.update(intent="proposal_deadline", modules=["propostas"], start=today.isoformat(),
                        end=(today + timedelta(days=7)).isoformat())
        elif "certific" in normalized and any(word in normalized for word in ("cliente", "venc", "valid")):
            plan.update(intent="certificate_deadline", modules=["certificados"], start=today.isoformat(),
                        end=(today + timedelta(days=30)).isoformat())
            plan["term"] = self.assistant_extract_term(normalized, ("cliente", "certificado"))
        elif ("ordem de servico" in normalized or "o.s" in normalized or "os atras" in normalized) and "atras" in normalized:
            plan.update(intent="overdue_work", modules=["ordens_servico"], end=today.isoformat(),
                        status_exclude=["Concluída", "Concluído", "Cancelada", "Cancelado"])
        elif "equipamento" in normalized and any(word in normalized for word in ("calibr", "valid")):
            plan.update(intent="equipment_calibration", modules=["equipamentos"], end=(today + timedelta(days=30)).isoformat())
            plan["term"] = self.assistant_extract_term(normalized, ("equipamento", "calibracao"))
        elif "historico" in normalized and "cliente" in normalized:
            plan["intent"] = "customer_history"
            plan["modules"] = sorted(readable - {"fontes", "normas_tecnicas"})
            plan["term"] = self.assistant_extract_term(normalized, ("historico", "cliente"))
        elif any(word in normalized for word in ("redija", "rascunho", "email", "e-mail", "acompanhamento")):
            plan["intent"] = "commercial_draft"
            plan["modules"] = [module for module in ("clientes", "crm", "propostas", "produtos", "catalogo_servicos")
                                if module in readable]
            plan["term"] = self.assistant_extract_term(normalized, ("redija", "rascunho", "email", "e-mail", "acompanhamento"))
        elif "proximo passo" in normalized or "próximo passo" in question.lower():
            plan["intent"] = "commercial_next_step"
            plan["modules"] = [module for module in ("crm", "propostas", "clientes") if module in readable]
        else:
            plan["modules"] = sorted(readable - {"fontes"})
            plan["term"] = question
        plan["modules"] = [module for module in plan["modules"] if module in readable or module == "editais"]
        return plan

    @staticmethod
    def assistant_extract_term(normalized, stop_words):
        text = normalized
        for marker in stop_words:
            text = text.replace(marker, " ")
        text = re.sub(r"\b(quais|qual|estao|estão|proximos|proximas|do|da|de|dos|das|o|a|os|as|e|com)\b", " ", text)
        return " ".join(part for part in text.split() if len(part) > 2)[:100]

    def assistant_context(self, plan, session):
        company_id = session["company_id"]
        items = []
        modules = plan["modules"]
        if plan["intent"] == "tender_score" and "editais" in modules:
            if not self.require_module_read(session, "editais"):
                return []
            threshold = plan["threshold"] or 70
            rows = self.db.connection().execute(
                """SELECT id,title,object_text,agency,uf,deadline,relevance_score,status,source_url
                   FROM tender_results WHERE company_id=? AND relevance_score>=?
                   ORDER BY relevance_score DESC,deadline LIMIT 30""", (company_id, threshold)).fetchall()
            return [{"id": f"tender:{row['id']}", "module": "editais", "title": row["title"],
                     "status": row["status"], "due_date": row["deadline"],
                     "fields": {"aderencia": row["relevance_score"], "objeto": row["object_text"][:500],
                                "orgao": row["agency"], "uf": row["uf"]},
                     "source_url": row["source_url"]} for row in rows]
        if not modules:
            return []
        readable = [module for module in modules if module in self.allowed_modules(session, "read")]
        if not readable:
            return []
        placeholders = ",".join("?" for _ in readable)
        sql = f"SELECT id,module,title,status,amount,due_date,payload,updated_at FROM records WHERE company_id=? AND deleted_at IS NULL AND module IN ({placeholders})"
        params = [company_id, *readable]
        if plan["intent"] == "proposal_deadline" or plan["intent"] == "certificate_deadline":
            sql += " AND COALESCE(due_date,json_extract(payload,'$.validade'),json_extract(payload,'$.proxima_calibracao')) BETWEEN ? AND ?"
            params.extend([plan["start"], plan["end"]])
        elif plan["intent"] == "overdue_work":
            sql += " AND due_date IS NOT NULL AND due_date < ?"
            params.append(plan["end"])
        elif plan["intent"] == "equipment_calibration":
            sql += " AND COALESCE(json_extract(payload,'$.proxima_calibracao'),due_date) <= ?"
            params.append(plan["end"])
        if plan["term"] and plan["intent"] in {"certificate_deadline", "equipment_calibration", "customer_history", "commercial_draft"}:
            pattern = f"%{plan['term']}%"
            sql += " AND (title LIKE ? OR payload LIKE ?)"
            params.extend([pattern, pattern])
        if plan["status_exclude"]:
            excluded = ",".join("?" for _ in plan["status_exclude"])
            sql += f" AND status NOT IN ({excluded})"
            params.extend(plan["status_exclude"])
        sql += " ORDER BY COALESCE(due_date,updated_at) LIMIT 40"
        rows = self.db.connection().execute(sql, params).fetchall()
        items = []
        for row in rows:
            item = self.assistant_record_context(row)
            if "view_values" not in self.allowed_operations(session, row["module"]):
                item["amount"] = None
            items.append(item)
        return items

    @staticmethod
    def assistant_record_context(row):
        try:
            payload = json.loads(row["payload"] or "{}")
        except (ValueError, TypeError):
            payload = {}
        safe_keys = ("cliente", "razao_social", "validade", "proxima_calibracao", "etapa",
                     "proximo_passo", "equipamento", "numero", "tipo", "observacoes", "notes")
        fields = {key: str(payload[key])[:300] for key in safe_keys if payload.get(key) not in (None, "")}
        return {"id": row["id"], "module": row["module"], "title": row["title"],
                "status": row["status"], "due_date": row["due_date"], "amount": row["amount"],
                "updated_at": row["updated_at"], "fields": fields}

    def assistant_fallback(self, question, plan, context):
        if not context:
            return {"answer": "Não encontrei registros autorizados para essa pergunta.",
                    "confidence": "alta", "suggestions": ["Tente outro termo ou verifique suas permissões."]}
        labels = {"proposal_deadline": "propostas com prazo nesta semana",
                  "certificate_deadline": "certificados próximos do vencimento",
                  "tender_score": f"licitações com aderência de pelo menos {plan['threshold'] or 70}%",
                  "overdue_work": "ordens de serviço atrasadas",
                  "equipment_calibration": "equipamentos com calibração vencida ou próxima",
                  "customer_history": "registros encontrados no histórico do cliente",
                  "commercial_draft": "registros comerciais para preparar um rascunho",
                  "commercial_next_step": "registros comerciais para sugerir o próximo passo"}
        lines = [f"Encontrei {len(context)} registro(s) em {labels.get(plan['intent'], 'sua busca')}:"]
        for item in context[:12]:
            detail = item.get("due_date") or item.get("status") or ""
            lines.append(f"• {item['title']} — {detail}")
        return {"answer": "\n".join(lines), "confidence": "alta", "suggestions": []}

    def openrouter_assistant(self, question, plan, context):
        key = os.environ.get("OPENROUTER_API_KEY", "").strip()
        if not key:
            raise ValueError("OPENROUTER_API_KEY ausente")
        model = os.environ.get("OPENROUTER_ASSISTANT_MODEL", "openai/gpt-5.4-mini")
        schema = {"type": "object", "additionalProperties": False, "properties": {
            "answer": {"type": "string"}, "confidence": {"type": "string", "enum": ["alta", "media", "baixa"]},
            "suggestions": {"type": "array", "items": {"type": "string"}},
        }, "required": ["answer", "confidence", "suggestions"]}
        body = {"model": model,
                "messages": [{"role": "system", "content":
                    "Você é o assistente interno do SIVS. Responda somente com base no CONTEXTO autorizado. "
                    "Não invente valores, prazos, preços ou permissões. Se faltar informação, diga isso. "
                    "Sugestões de CRM/propostas devem ser rascunhos e nunca alterar condições comerciais."},
                           {"role": "user", "content": f"PERGUNTA:\n{question}\n\nCONTEXTO:\n{json_dumps(context)[:30000]}"}],
                "response_format": {"type": "json_schema", "json_schema": {"name": "sivs_assistant", "strict": True, "schema": schema}},
                "temperature": 0.1, "max_tokens": 900}
        request = urllib.request.Request("https://openrouter.ai/api/v1/chat/completions",
                                         data=json_dumps(body).encode("utf-8"),
                                         headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json",
                                                  "HTTP-Referer": "https://sivs-seccol.local", "X-Title": "SIVS SECCOL"}, method="POST")
        with urllib.request.urlopen(request, timeout=45) as response:
            data = json.load(response)
        content = data["choices"][0]["message"]["content"]
        result = json.loads(content) if isinstance(content, str) else content
        return result, data.get("model") or model

    def record_json(self, row, session=None):
        if row is None:
            return None
        return self.records_json([row], session)[0]

    def records_json(self, rows, session=None):
        """Serializa registros em lote sem consultas adicionais por linha."""
        raw_items = [dict(row) for row in rows]
        if not raw_items:
            return []
        hydrated = {}
        grouped = collections.defaultdict(list)
        for raw in raw_items:
            grouped[raw.get("company_id")].append(raw)
        connection = self.db.connection()
        for company_id, company_items in grouped.items():
            record_ids = [int(item["id"]) for item in company_items]
            placeholders = ",".join("?" for _ in record_ids)
            params = [*record_ids, company_id]
            subject_ids = sorted({int(item["subject_id"]) for item in company_items
                                  if item.get("subject_id")})
            subjects_by_id = {}
            if subject_ids:
                subject_placeholders = ",".join("?" for _ in subject_ids)
                subject_rows = connection.execute(
                    f"""SELECT id,name,status FROM subjects
                        WHERE id IN ({subject_placeholders}) AND company_id=?""",
                    (*subject_ids, company_id),
                ).fetchall()
                subjects_by_id = {row["id"]: row for row in subject_rows}
            relations_by_record = collections.defaultdict(list)
            for relation in connection.execute(
                f"""SELECT rr.from_record_id,rr.to_record_id,rr.relationship_type,
                            r.module,r.title
                     FROM record_relationships rr JOIN records r ON r.id=rr.to_record_id
                     WHERE rr.from_record_id IN ({placeholders})
                       AND r.company_id=? AND r.deleted_at IS NULL ORDER BY rr.id""",
                params,
            ).fetchall():
                relations_by_record[relation["from_record_id"]].append(relation)
            record_subjects = collections.defaultdict(list)
            for subject_row in connection.execute(
                f"""SELECT rs.record_id,s.id,s.name,rs.relationship_type,rs.is_primary
                     FROM record_subjects rs JOIN subjects s ON s.id=rs.subject_id
                     WHERE rs.record_id IN ({placeholders}) AND s.company_id=?
                     ORDER BY rs.is_primary DESC,s.name""",
                params,
            ).fetchall():
                record_subjects[subject_row["record_id"]].append(subject_row)
            attachments_by_record = collections.defaultdict(list)
            for attachment in connection.execute(
                f"""SELECT record_id,id,filename,mime_type,size,category,version,sha256,
                            license_confirmed,created_at
                     FROM attachments WHERE record_id IN ({placeholders}) AND company_id=?
                     ORDER BY id DESC""",
                params,
            ).fetchall():
                attachments_by_record[attachment["record_id"]].append(attachment)
            approvals_by_record = collections.defaultdict(list)
            for approval in connection.execute(
                f"""SELECT a.*,u0.name requested_by_name,u1.name requested_to_name,
                            u2.name decided_by_name
                     FROM approvals a LEFT JOIN users u0 ON u0.id=a.requested_by
                     LEFT JOIN users u1 ON u1.id=a.requested_to
                     LEFT JOIN users u2 ON u2.id=a.decided_by
                     WHERE a.record_id IN ({placeholders}) AND a.company_id=? ORDER BY a.id DESC""",
                params,
            ).fetchall():
                approvals_by_record[approval["record_id"]].append(approval)
            for raw in company_items:
                item = dict(raw)
                item["payload"] = json.loads(item["payload"] or "{}")
                subject = subjects_by_id.get(item.get("subject_id"))
                relations = relations_by_record[item["id"]]
                if subject:
                    item["payload"]["assunto"] = subject["name"]
                    item["subject"] = dict(subject)
                related_by_id = {
                    int(relation["to_record_id"]): relation for relation in relations
                }
                for field in RECORD_REFERENCE_RULES:
                    try:
                        target = related_by_id.get(
                            int(item["payload"].get(f"{field}_id") or 0)
                        )
                    except (ValueError, TypeError):
                        target = None
                    if target:
                        item["payload"][field] = target["title"]
                item["payload"]["relacionamentos"] = [
                    {"record": f'{relation["module"]}:{relation["to_record_id"]}',
                     "type": relation["relationship_type"], "label": relation["title"]}
                    for relation in relations
                ]
                item["subjects"] = [
                    {key: row[key] for key in row.keys() if key != "record_id"}
                    for row in record_subjects[item["id"]]
                ]
                item["attachments"] = [
                    {key: row[key] for key in row.keys() if key != "record_id"}
                    for row in attachments_by_record[item["id"]]
                ]
                item["approvals"] = [dict(row) for row in approvals_by_record[item["id"]]]
                hydrated[item["id"]] = item
        result = [hydrated[item["id"]] for item in raw_items]
        if session is not None:
            for item in result:
                module = item.get("module")
                if (module in VALUE_SENSITIVE_MODULES
                        and "view_values" not in self.allowed_operations(session, module)):
                    item["amount"] = None
                    item["amountRestricted"] = True
                    for key in SENSITIVE_PAYLOAD_FIELDS.get(module, set()):
                        item["payload"].pop(key, None)
        return result

    @staticmethod
    def _validate_json_shape(value, path="payload", depth=0):
        if depth > 10:
            raise ValueError(f"{path}: estrutura excede 10 níveis")
        if isinstance(value, float) and not math.isfinite(value):
            raise ValueError(f"{path}: número não finito")
        if isinstance(value, str) and len(value) > 20_000:
            raise ValueError(f"{path}: texto excede 20.000 caracteres")
        if isinstance(value, dict):
            if len(value) > 300:
                raise ValueError(f"{path}: campos em excesso")
            for key, child in value.items():
                if not isinstance(key, str) or len(key) > 120:
                    raise ValueError(f"{path}: nome de campo inválido")
                SIVSHandler._validate_json_shape(child, f"{path}.{key}", depth + 1)
        elif isinstance(value, list):
            if len(value) > 1_000:
                raise ValueError(f"{path}: itens em excesso")
            for index, child in enumerate(value):
                SIVSHandler._validate_json_shape(child, f"{path}[{index}]", depth + 1)

    def validate_record_payload(self, module, payload):
        self._validate_json_shape(payload)
        try:
            encoded = json_dumps(payload).encode("utf-8")
        except (ValueError, TypeError) as exc:
            raise ValueError(f"Detalhes incompatíveis com JSON: {exc}") from None
        if len(encoded) > MAX_RECORD_PAYLOAD:
            raise ValueError("Detalhes do registro excedem 1 MB")

        subject = str(payload.get("assunto") or "").strip()
        if len(subject) < 3:
            raise ValueError("Assunto principal é obrigatório e deve possuir ao menos 3 caracteres")
        if len(subject) > 180:
            raise ValueError("Assunto principal excede 180 caracteres")

        missing = [key for key in REQUIRED_PAYLOAD_FIELDS.get(module, ()) if _blank(payload.get(key))]
        if missing:
            labels = ", ".join(key.replace("_", " ") for key in missing[:8])
            suffix = "…" if len(missing) > 8 else ""
            raise ValueError(f"Campos obrigatórios ausentes: {labels}{suffix}")

        for key in DATE_FIELDS.get(module, set()):
            value = payload.get(key)
            if _blank(value):
                continue
            try:
                datetime.strptime(str(value), "%Y-%m-%d")
            except ValueError:
                raise ValueError(f"{key.replace('_', ' ').title()}: data inválida; use AAAA-MM-DD") from None
        for key in DATETIME_FIELDS.get(module, set()):
            value = payload.get(key)
            if _blank(value):
                continue
            try:
                datetime.fromisoformat(str(value).replace("Z", "+00:00"))
            except ValueError:
                raise ValueError(f"{key.replace('_', ' ').title()}: data e hora inválidas") from None
        for key in TIME_FIELDS.get(module, set()):
            value = payload.get(key)
            if _blank(value):
                continue
            try:
                datetime.strptime(str(value), "%H:%M")
            except ValueError:
                raise ValueError(f"{key.replace('_', ' ').title()}: hora inválida") from None
        for key in NUMBER_FIELDS.get(module, set()):
            value = payload.get(key)
            if _blank(value):
                continue
            try:
                numeric = float(value)
            except (ValueError, TypeError):
                raise ValueError(f"{key.replace('_', ' ').title()}: número inválido") from None
            if not math.isfinite(numeric) or abs(numeric) > 1_000_000_000_000_000:
                raise ValueError(f"{key.replace('_', ' ').title()}: número fora do limite")
            if key in {"probabilidade"} and not 0 <= numeric <= 100:
                raise ValueError("Probabilidade deve ficar entre 0 e 100")
            if key in {"tempo_minutos", "quilometragem", "proxima_km", "preco_venda", "quantidade", "horas"} and numeric < 0:
                raise ValueError(f"{key.replace('_', ' ').title()} não pode ser negativo")

        for key in EMAIL_FIELDS.get(module, set()):
            value = str(payload.get(key) or "").strip()
            if value and not re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", value):
                raise ValueError(f"{key.replace('_', ' ').title()}: e-mail inválido")
        for key in URL_FIELDS.get(module, set()):
            value = str(payload.get(key) or "").strip()
            if not value:
                continue
            parsed = urlparse(value)
            if parsed.scheme not in {"http", "https"} or not parsed.netloc:
                raise ValueError(f"{key.replace('_', ' ').title()}: URL deve usar HTTP ou HTTPS")

        if module in {"clientes", "fornecedores"}:
            _validate_document(str(payload.get("documento") or ""))
        elif module == "colaboradores":
            if not _valid_cpf(str(payload.get("cpf") or "")):
                raise ValueError("CPF inválido")
        elif module == "concorrentes":
            if not _valid_cnpj(str(payload.get("cnpj") or "")):
                raise ValueError("CNPJ inválido")

        if module == "fiscal" and str(payload.get("tipo_nota") or "") == "NF-e":
            key = re.sub(r"\D", "", str(payload.get("chave") or ""))
            if len(key) != 44:
                raise ValueError("Chave de NF-e deve possuir 44 dígitos")

        date_pairs = {
            "contratos": ("inicio", "fim"),
            "treinamentos": ("data", "validade"),
            "calibracoes": ("data_calibracao", "proxima_calibracao"),
            "ordens_servico": ("inicio", "fim"),
        }
        if module in date_pairs:
            start_key, end_key = date_pairs[module]
            start, end = payload.get(start_key), payload.get(end_key)
            if start and end:
                try:
                    if datetime.fromisoformat(str(end)) < datetime.fromisoformat(str(start)):
                        raise ValueError(f"{end_key.replace('_', ' ').title()} não pode ser anterior a {start_key.replace('_', ' ')}")
                except ValueError as exc:
                    if "não pode" in str(exc):
                        raise

        relationships = payload.get("relacionamentos") or []
        if not isinstance(relationships, list) or len(relationships) > 50:
            raise ValueError("Relacionamentos devem ser uma lista de até 50 itens")
        for relation in relationships:
            if not isinstance(relation, dict):
                raise ValueError("Relacionamento inválido")
            reference = str(relation.get("record") or relation.get("registro") or "")
            match = re.fullmatch(r"([a-z_]+):(\d+)", reference)
            if not match or match.group(1) not in MODULES or int(match.group(2)) <= 0:
                raise ValueError("Referência de relacionamento inválida")
            relation_type = str(relation.get("type") or relation.get("tipo") or "Relacionado a").strip()
            if not relation_type or len(relation_type) > 80:
                raise ValueError("Tipo de relacionamento inválido")

    def normalized_record(self, data, existing_status=None):
        if not isinstance(data, dict):
            raise ValueError("Registro deve ser um objeto")
        module = str(data.get("module", "")).strip()
        title = str(data.get("title", "")).strip()
        status = str(data.get("status") or "").strip()
        amount = data.get("amount")
        due_date = str(data.get("due_date") or "").strip() or None
        payload = data.get("payload") or {}
        if module == PARTY_MODULE:
            party_type = str(payload.get("tipo_cadastro") or "").strip()
            party_digits = re.sub(r"\D", "", str(payload.get("documento") or ""))
            if not party_type and len(party_digits) in {11, 14}:
                party_type = "C" if len(party_digits) == 11 else "F"
            party_type = {
                "Cliente": "C", "Cliente (C)": "C",
                "Fornecedor": "F", "Fornecedor (F)": "F",
                "Cliente e fornecedor": "A", "Cliente e fornecedor (A)": "A",
                # Compatibilidade com rascunhos e clientes antigos que
                # enviavam a opção curta exibida antes da padronização.
                "C e F": "A", "A": "A",
            }.get(party_type, party_type)
            if party_type not in {"C", "F", "A"}:
                raise ValueError("Escolha Cliente, Fornecedor ou Cliente e fornecedor")
            payload["tipo_cadastro"] = party_type
            # O cadastro unificado não deve falhar por um campo operacional
            # que ainda não foi preenchido. Fornecedor nasce como pendente
            # para posterior avaliação, preservando a validação do módulo
            # físico e evitando um erro silencioso no primeiro cadastro.
            if party_type == "F" and not str(payload.get("avaliacao") or "").strip():
                payload["avaliacao"] = "Pendente"
            # A permanece um único cadastro canônico; a aba unificada o
            # apresenta nos dois contextos sem duplicar a pessoa no banco.
            module = "fornecedores" if party_type == "F" else "clientes"
        if module in {"clientes", "fornecedores"}:
            digits = re.sub(r"\D", "", str(payload.get("documento") or ""))
            if len(digits) in {11, 14}:
                derived_person = "Pessoa física" if len(digits) == 11 else "Pessoa jurídica"
                declared_person = str(payload.get("tipo_pessoa") or "").strip()
                if declared_person and declared_person != derived_person:
                    raise ValueError("O tipo de pessoa deve corresponder ao documento: CPF = Pessoa física; CNPJ = Pessoa jurídica")
                payload["tipo_pessoa"] = derived_person
                payload["documento"] = digits
        if module == "contas_pagar":
            partner_type = str(payload.get("tipo_parte") or "Fornecedor (F)").strip()
            if partner_type not in {"Fornecedor (F)", "Cliente e fornecedor (A)"}:
                raise ValueError("Contas a pagar devem estar vinculadas a Fornecedor (F) ou Cliente e fornecedor (A)")
            payload["tipo_parte"] = partner_type
        elif module == "contas_receber":
            partner_type = str(payload.get("tipo_parte") or "Cliente (C)").strip()
            if partner_type not in {"Cliente (C)", "Cliente e fornecedor (A)"}:
                raise ValueError("Contas a receber devem estar vinculadas a Cliente (C) ou Cliente e fornecedor (A)")
            payload["tipo_parte"] = partner_type
        if module not in MODULES or not title:
            raise ValueError("Módulo e título são obrigatórios")
        if not status:
            status = MODULE_INITIAL_STATUSES.get(module, "Ativo")
        if len(title) > 240 or any(ord(char) < 32 and char not in "\t\n" for char in title):
            raise ValueError("Título inválido ou superior a 240 caracteres")
        allowed_statuses = MODULE_STATUSES.get(module, DEFAULT_STATUSES)
        if status not in allowed_statuses and status != existing_status:
            raise ValueError("Status inválido para este módulo")
        transitions = MODULE_STATUS_TRANSITIONS.get(module)
        if transitions and existing_status is None and status != MODULE_INITIAL_STATUSES[module]:
            raise ValueError(
                f"Novos registros deste fluxo devem iniciar em {MODULE_INITIAL_STATUSES[module]}"
            )
        if transitions and existing_status is not None and status != existing_status:
            allowed_next = transitions.get(existing_status, {MODULE_INITIAL_STATUSES[module]})
            if status not in allowed_next:
                raise ValueError(f"Transição de {existing_status} para {status} não é permitida")
        if amount in ("", None):
            amount = None
        else:
            try:
                amount = float(amount)
            except (ValueError, TypeError):
                raise ValueError("Valor financeiro inválido") from None
            if not math.isfinite(amount) or abs(amount) > 1_000_000_000_000_000:
                raise ValueError("Valor financeiro fora do limite permitido")
        if due_date:
            try:
                datetime.strptime(due_date, "%Y-%m-%d")
            except ValueError:
                raise ValueError("Prazo inválido; use AAAA-MM-DD") from None
        if not isinstance(payload, dict):
            raise ValueError("Detalhes inválidos")
        self.validate_record_payload(module, payload)
        return module, title[:240], status[:80], amount, due_date, json_dumps(payload)

    def resolve_record_references(self, values, session):
        """Valida IDs de cadastros mestres e materializa seus vínculos internos."""
        payload = json.loads(values[5])
        relationships = list(payload.get("relacionamentos") or [])
        readable = self.allowed_modules(session, "read")
        company_id = session["company_id"]
        party_aliases = {
            "Cliente": "C", "Cliente (C)": "C", "C": "C",
            "Fornecedor": "F", "Fornecedor (F)": "F", "F": "F",
            "Cliente e fornecedor": "A", "Cliente e fornecedor (A)": "A",
            "C e F": "A", "A": "A",
        }
        for field, rule in RECORD_REFERENCE_RULES.items():
            id_key = f"{field}_id"
            raw_id = payload.get(id_key)
            if raw_id in (None, ""):
                continue
            try:
                target_id = int(raw_id)
            except (ValueError, TypeError):
                raise ValueError(f"{field.replace('_', ' ').title()}: selecione um cadastro válido") from None
            if target_id <= 0:
                raise ValueError(f"{field.replace('_', ' ').title()}: selecione um cadastro válido")
            target = self.db.connection().execute(
                """SELECT id,module,title,payload FROM records
                   WHERE id=? AND company_id=? AND deleted_at IS NULL""",
                (target_id, company_id),
            ).fetchone()
            if (not target or target["module"] not in rule["modules"] or
                    target["module"] not in readable):
                raise ValueError(
                    f"{field.replace('_', ' ').title()}: o cadastro selecionado não existe "
                    "nesta empresa ou não está autorizado"
                )
            expected_role = rule.get("party_role")
            if expected_role:
                target_payload = json.loads(target["payload"] or "{}")
                fallback = "F" if target["module"] == "fornecedores" else "C"
                target_role = party_aliases.get(
                    str(target_payload.get("tipo_cadastro") or fallback).strip(), fallback
                )
                allowed_roles = {"C", "A"} if expected_role == "C" else (
                    {"F", "A"} if expected_role == "F" else {"C", "F", "A"}
                )
                if target_role not in allowed_roles:
                    label = "cliente" if expected_role == "C" else "fornecedor"
                    raise ValueError(f"{field.replace('_', ' ').title()}: selecione um {label} compatível")
            payload[id_key] = target_id
            payload[field] = target["title"]
            reference = f"{target['module']}:{target_id}"
            relationship_type = rule["relation"]
            relationships = [
                item for item in relationships
                if not (isinstance(item, dict) and
                        str(item.get("type") or item.get("tipo") or "") == relationship_type)
            ]
            relationships.append({"record": reference, "type": relationship_type})
        if len(relationships) > 50:
            raise ValueError("Relacionamentos devem ser uma lista de até 50 itens")
        payload["relacionamentos"] = relationships
        self.validate_record_payload(values[0], payload)
        return (*values[:5], json_dumps(payload))

    def validate_operational_partner(self, values, session):
        """Impede documentos operacionais com parceiro textual ou bloqueado."""
        module = values[0]
        payload = json.loads(values[5])
        customer_modules = {"propostas", "vendas", "contas_receber", "ordens_servico"}
        supplier_modules = {"pedidos_compra", "contas_pagar"}
        if module not in customer_modules | supplier_modules:
            return values
        field = "cliente" if module in customer_modules else "fornecedor"
        raw_id = payload.get(f"{field}_id")
        if not raw_id:
            raise ValueError(
                f"{field.title()}: selecione um cadastro validado da empresa; texto livre não é aceito neste fluxo"
            )
        partner = self.db.connection().execute(
            """SELECT id,title,payload FROM records
               WHERE id=? AND company_id=? AND module IN ('clientes','fornecedores')
                 AND deleted_at IS NULL""",
            (int(raw_id), session["company_id"]),
        ).fetchone()
        if not partner:
            raise ValueError(f"{field.title()}: cadastro não encontrado na empresa ativa")
        partner_payload = json.loads(partner["payload"] or "{}")
        if module in customer_modules and partner_payload.get("bloqueado"):
            raise ValueError(f"Cliente bloqueado: revise “{partner['title']}” antes de continuar")
        if module in {"vendas", "contas_receber"} and not partner_payload.get("aprovado_faturamento"):
            raise ValueError(
                f"Cliente não aprovado para faturamento: revise “{partner['title']}”"
            )
        if module in supplier_modules:
            if partner_payload.get("avaliacao") == "Reprovado":
                raise ValueError(f"Fornecedor reprovado: “{partner['title']}”")
            if not partner_payload.get("aprovado_compras"):
                raise ValueError(
                    f"Fornecedor não aprovado para compras: revise “{partner['title']}”"
                )
        return values

    def assign_party_code(self, values, company_id, party_type):
        """Gera identificacao curta e sequencial dentro da empresa (C/F/A-0001)."""
        if party_type not in {"C", "F", "A"}:
            return values
        payload = json.loads(values[5])
        if payload.get("codigo_cadastro"):
            return values
        pattern = f"{party_type}-%"
        count = self.db.scalar(
            """SELECT COUNT(*) FROM records
               WHERE company_id=? AND module IN ('clientes','fornecedores')
                 AND json_extract(payload,'$.codigo_cadastro') LIKE ?""",
            (company_id, pattern),
        )
        payload["codigo_cadastro"] = f"{party_type}-{int(count or 0) + 1:04d}"
        return (*values[:5], json_dumps(payload))

    def duplicate_party_document(self, company_id, module, payload, exclude_id=None):
        if module not in {"clientes", "fornecedores"}:
            return None
        digits = re.sub(r"\D", "", str(payload.get("documento") or ""))
        if len(digits) not in {11, 14}:
            return None
        query = """
            SELECT id,title,module FROM records
            WHERE company_id=? AND deleted_at IS NULL
              AND module IN ('clientes','fornecedores')
              AND replace(replace(replace(replace(
                    CAST(json_extract(payload,'$.documento') AS TEXT),
                    '.',''),'-',''),'/',''),' ','')=?
        """
        params = [company_id, digits]
        if exclude_id is not None:
            query += " AND id<>?"
            params.append(exclude_id)
        return self.db.connection().execute(query + " LIMIT 1", params).fetchone()

    def duplicate_party_response(self, payload, duplicate):
        digits = re.sub(r"\D", "", str(payload.get("documento") or ""))
        document_type = "CPF" if len(digits) == 11 else "CNPJ"
        return self.error_json(
            f"Este {document_type} já está cadastrado em “{duplicate['title']}”. Abra o cadastro existente.",
            409, "duplicate_party_document",
        )

    def validate_unique_business_key(self, company_id, module, payload, exclude_id=None):
        field = BUSINESS_UNIQUE_FIELDS.get(module)
        if not field:
            return
        value = str(payload.get(field) or "").strip()
        if not value:
            return
        sql = """SELECT id,title FROM records
                 WHERE company_id=? AND module=? AND deleted_at IS NULL
                   AND lower(trim(CAST(json_extract(payload,?) AS TEXT)))=lower(?)"""
        params = [company_id, module, f"$.{field}", value]
        if exclude_id is not None:
            sql += " AND id<>?"
            params.append(exclude_id)
        duplicate = self.db.connection().execute(sql + " LIMIT 1", params).fetchone()
        if duplicate:
            label = field.replace("_", " ").title()
            raise BusinessKeyConflict(
                f"{label} “{value}” já pertence a “{duplicate['title']}” nesta empresa"
            )

    def records_write(self, method, path, session):
        try:
            data = self.parse_json() if method != "DELETE" else {}
        except ValueError as exc:
            return self.error_json(str(exc))
        pieces = path.split("/")
        record_id = int(pieces[3]) if len(pieces) == 4 and pieces[3].isdigit() else None
        if method == "POST" and path == "/api/records":
            if str(data.get("module") or "").strip() == "estoque":
                return self.error_json(
                    "Estoque só pode ser alterado por movimentação auditável.",
                    409, "inventory_ledger_required",
                )
            try:
                values = self.normalized_record(data)
                values = self.resolve_record_references(values, session)
                values = self.validate_operational_partner(values, session)
                self.db.validate_normative_base(values[0], json.loads(values[5]), session["company_id"])
            except (ValueError, TypeError) as exc:
                return self.error_json(str(exc))
            normalized_payload = json.loads(values[5])
            requested_module = str(data.get("module") or "").strip()
            if requested_module == PARTY_MODULE:
                party_type = str(normalized_payload.get("tipo_cadastro") or "").strip()
                required_modules = (
                    PARTY_PHYSICAL_MODULES if party_type == "A"
                    else ("fornecedores",) if party_type == "F"
                    else ("clientes",)
                )
                for required_module in required_modules:
                    if not self.require_operation(session, required_module, "create"):
                        return
            elif not self.require_operation(session, values[0], "create"):
                return
            duplicate = self.duplicate_party_document(
                session["company_id"], values[0], normalized_payload
            )
            if duplicate:
                return self.duplicate_party_response(normalized_payload, duplicate)
            now = utc_now()
            try:
                with self.db.transaction(immediate=True):
                    self.validate_unique_business_key(
                        session["company_id"], values[0], normalized_payload,
                    )
                    if str(data.get("module") or "") == PARTY_MODULE:
                        values = self.assign_party_code(
                            values, session["company_id"],
                            str(json.loads(values[5]).get("tipo_cadastro") or "").strip(),
                        )
                    cursor = self.db.execute(
                        """INSERT INTO records
                           (module,title,status,amount,due_date,payload,created_by,created_at,updated_at,
                            company_id,revision)
                           VALUES(?,?,?,?,?,?,?,?,?,?,1)""",
                        (*values, session["id"], now, now, session["company_id"]))
                    record_id = cursor.lastrowid
                    self.db.sync_relationships(
                        record_id, json.loads(values[5]), session["id"], session["company_id"]
                    )
                    self.db.audit(
                        session["id"], "create", values[0], record_id,
                        {"title": values[1], "revision": 1}, company_id=session["company_id"],
                    )
            except BusinessKeyConflict as exc:
                return self.error_json(str(exc), 409, "duplicate_business_key")
            except (ValueError, sqlite3.Error) as exc:
                return self.error_json(str(exc))
            row = self.db.connection().execute(
                "SELECT * FROM records WHERE id=? AND company_id=?", (record_id, session["company_id"])
            ).fetchone()
            return self.send_json({"ok": True, "item": self.record_json(row)}, 201)
        if not record_id:
            return self.error_json("Registro inválido", 404)
        existing = self.db.connection().execute(
            "SELECT * FROM records WHERE id=? AND company_id=? AND deleted_at IS NULL",
            (record_id, session["company_id"])).fetchone()
        if not existing:
            return self.error_json("Registro não encontrado", 404)
        requested_action = "delete" if method == "DELETE" else "update"
        if not self.require_operation(session, existing["module"], requested_action):
            return
        if method == "PUT" and existing["module"] == "estoque":
            return self.error_json(
                "Movimentos legados de estoque não são editáveis; use o ledger de movimentações.",
                409, "inventory_ledger_required",
            )
        if method == "DELETE":
            if existing["module"] in ITEM_DOCUMENT_MODULES:
                active_reservations = self.db.scalar(
                    """SELECT COUNT(*) FROM document_items i
                       JOIN inventory_reservations q ON q.id=i.reservation_id
                       WHERE i.record_id=? AND i.company_id=? AND q.status='ACTIVE'""",
                    (record_id, session["company_id"]),
                )
                if active_reservations:
                    return self.error_json(
                        "Libere as reservas de estoque antes de excluir o documento.",
                        409, "active_inventory_reservations",
                    )
            if existing["module"] in {"produtos", "catalogo_servicos"}:
                used_as_item = self.db.scalar(
                    """SELECT COUNT(*) FROM document_items i JOIN records r ON r.id=i.record_id
                       WHERE i.catalog_record_id=? AND i.company_id=? AND r.deleted_at IS NULL""",
                    (record_id, session["company_id"]),
                )
                if used_as_item:
                    return self.error_json(
                        "Este item pertence a documento(s) ativo(s). Inative-o em vez de excluir.",
                        409, "catalog_item_in_use",
                    )
            if existing["module"] == "normas_tecnicas":
                referenced = self.db.scalar(
                    """SELECT COUNT(*) FROM record_relationships rr
                       JOIN records r ON r.id=rr.from_record_id
                       WHERE rr.to_record_id=? AND r.company_id=? AND r.deleted_at IS NULL
                       AND r.module IN ('certificados','laudos_tecnicos','estudos_tecnicos')""",
                    (record_id, session["company_id"]),
                )
                if referenced:
                    return self.error_json(
                        "Esta norma fundamenta documento(s) ativo(s) e não pode ser excluída.",
                        409, "norm_in_use",
                    )
            now = utc_now()
            with self.db.transaction(immediate=True):
                current = self.db.connection().execute(
                    "SELECT * FROM records WHERE id=? AND company_id=? AND deleted_at IS NULL",
                    (record_id, session["company_id"]),
                ).fetchone()
                if not current:
                    return self.error_json("O registro já foi alterado ou excluído", 409, "write_conflict")
                self.save_record_version(current, session["id"])
                self.db.execute(
                    """UPDATE records SET deleted_at=?,updated_at=?,revision=revision+1
                       WHERE id=? AND company_id=?""",
                    (now, now, record_id, session["company_id"]),
                )
                self.db.execute(
                    """UPDATE approvals SET status='Expirada',decided_at=?,
                       decision_comment='Registro excluído após a solicitação.'
                       WHERE record_id=? AND company_id=? AND status='Pendente'""",
                    (now, record_id, session["company_id"]),
                )
                self.db.audit(
                    session["id"], "delete", current["module"], record_id,
                    {"title": current["title"], "revision": current["revision"] + 1},
                    company_id=session["company_id"],
                )
            return self.send_json({"ok": True})
        if method == "PUT":
            try:
                expected_revision = int(data.get("revision"))
                if expected_revision < 1:
                    raise ValueError
            except (ValueError, TypeError):
                return self.error_json(
                    "A revisão do registro é obrigatória para salvar alterações",
                    409, "revision_required",
                )
            try:
                values = self.normalized_record(data, existing_status=existing["status"])
                values = self.resolve_record_references(values, session)
                values = self.validate_operational_partner(values, session)
                self.db.validate_normative_base(values[0], json.loads(values[5]), session["company_id"])
            except (ValueError, TypeError) as exc:
                return self.error_json(str(exc))
            if values[0] != existing["module"]:
                return self.error_json("O módulo de um registro existente não pode ser alterado")
            if values[2] != existing["status"]:
                if ("transition" in MODULE_ACTIONS.get(values[0], ())
                        and not self.require_operation(session, values[0], "transition")):
                    return
                if (values[0] == "vendas" and values[2] == "Faturado"
                        and not self.require_operation(session, values[0], "bill_sales")):
                    return
                if (values[0] in {"contas_pagar", "contas_receber"}
                        and values[2] in {"Pago", "Recebido"}
                        and not self.require_operation(
                            session, values[0], "settle_financial"
                        )):
                    return
                if (values[0] in {"contas_pagar", "contas_receber"}
                        and values[2] == "Cancelado"
                        and not self.require_operation(
                            session, values[0], "cancel_financial"
                        )):
                    return
            normalized_payload = json.loads(values[5])
            if values[0] in {"clientes", "fornecedores"}:
                prior_payload = json.loads(existing["payload"] or "{}")
                controlled_fields = {
                    "aprovado_faturamento", "aprovado_compras", "bloqueado", "avaliacao",
                }
                if (any(prior_payload.get(key) != normalized_payload.get(key)
                        for key in controlled_fields)
                        and not self.require_operation(
                            session, values[0], "partner_control"
                        )):
                    return
            duplicate = self.duplicate_party_document(
                session["company_id"], values[0], normalized_payload, record_id
            )
            if duplicate:
                return self.duplicate_party_response(normalized_payload, duplicate)
            now = utc_now()
            try:
                with self.db.transaction(immediate=True):
                    current = self.db.connection().execute(
                        "SELECT * FROM records WHERE id=? AND company_id=? AND deleted_at IS NULL",
                        (record_id, session["company_id"]),
                    ).fetchone()
                    if not current or current["revision"] != expected_revision:
                        return self.error_json(
                            "Este registro foi alterado por outra pessoa. Recarregue antes de salvar.",
                            409, "write_conflict",
                        )
                    if (current["module"] in RESERVABLE_ITEM_MODULES
                            and values[2] != current["status"]
                            and values[2] in {"Concluído", "Concluída", "Cancelado", "Cancelada"}):
                        active_reservations = self.db.scalar(
                            """SELECT COUNT(*) FROM document_items i
                               JOIN inventory_reservations q ON q.id=i.reservation_id
                               WHERE i.record_id=? AND i.company_id=? AND q.status='ACTIVE'""",
                            (record_id, session["company_id"]),
                        )
                        if active_reservations:
                            raise InventoryWorkflowConflict(
                                "Baixe ou libere todas as reservas antes de concluir ou cancelar o documento"
                            )
                    if current["module"] == "pedidos_compra" and values[2] != current["status"]:
                        received_products = self.db.scalar(
                            """SELECT COUNT(*) FROM document_items i
                               WHERE i.record_id=? AND i.company_id=? AND i.item_kind='PRODUCT'
                                 AND EXISTS(
                                   SELECT 1 FROM inventory_movements m
                                   WHERE m.company_id=i.company_id
                                     AND m.movement_type='PURCHASE_IN'
                                     AND m.origin_type='PURCHASE_ORDER'
                                     AND m.origin_id=CAST(i.record_id AS TEXT)||':'||CAST(i.id AS TEXT)
                                 )""",
                            (record_id, session["company_id"]),
                        )
                        if values[2] == "Cancelado" and received_products:
                            raise InventoryWorkflowConflict(
                                "Este pedido já possui entrada de estoque e não pode ser cancelado"
                            )
                        if values[2] == "Recebido":
                            product_count = self.db.scalar(
                                """SELECT COUNT(*) FROM document_items
                                   WHERE record_id=? AND company_id=? AND item_kind='PRODUCT'""",
                                (record_id, session["company_id"]),
                            )
                            if received_products < product_count:
                                raise InventoryWorkflowConflict(
                                    "Receba todos os produtos no estoque antes de concluir o pedido"
                                )
                    self.validate_unique_business_key(
                        session["company_id"], values[0], normalized_payload, record_id,
                    )
                    self.save_record_version(current, session["id"])
                    cursor = self.db.execute(
                        """UPDATE records
                           SET title=?,status=?,amount=?,due_date=?,payload=?,updated_at=?,revision=revision+1
                           WHERE id=? AND company_id=? AND revision=?""",
                        (values[1], values[2], values[3], values[4], values[5], now,
                         record_id, session["company_id"], expected_revision),
                    )
                    if cursor.rowcount != 1:
                        raise sqlite3.IntegrityError("conflito de gravação")
                    self.db.sync_relationships(
                        record_id, json.loads(values[5]), session["id"], session["company_id"]
                    )
                    self.db.execute(
                        """UPDATE approvals SET status='Expirada',decided_at=?,
                           decision_comment='Registro alterado após a solicitação.'
                           WHERE record_id=? AND company_id=? AND status='Pendente'""",
                        (now, record_id, session["company_id"]),
                    )
                    self.db.audit(
                        session["id"], "update", values[0], record_id,
                        {"title": values[1], "from_revision": expected_revision,
                         "to_revision": expected_revision + 1},
                        company_id=session["company_id"],
                    )
            except InventoryWorkflowConflict as exc:
                return self.error_json(str(exc), 409, "active_inventory_reservations")
            except BusinessKeyConflict as exc:
                return self.error_json(str(exc), 409, "duplicate_business_key")
            except ValueError as exc:
                return self.error_json(str(exc))
            except sqlite3.IntegrityError:
                return self.error_json(
                    "Este registro foi alterado por outra pessoa. Recarregue antes de salvar.",
                    409, "write_conflict",
                )
            row = self.db.connection().execute(
                "SELECT * FROM records WHERE id=? AND company_id=?", (record_id, session["company_id"])
            ).fetchone()
            return self.send_json({"ok": True, "item": self.record_json(row)})
        return self.error_json("Método não permitido", 405)

    def save_record_version(self, row, user_id):
        snapshot = self.record_json(row)
        self.db.execute(
            """INSERT INTO record_versions
               (record_id,snapshot,changed_by,created_at,company_id) VALUES(?,?,?,?,?)""",
            (row["id"], json_dumps(snapshot), user_id, utc_now(), row["company_id"]),
        )

    def trash_record_blockers(self, record_id, company_id):
        blockers = []
        references = self.db.connection().execute(
            """SELECT r.id,r.title,r.module FROM record_relationships rr
               JOIN records r ON r.id=rr.from_record_id
               WHERE rr.to_record_id=? AND r.company_id=? AND r.deleted_at IS NULL
               ORDER BY r.updated_at DESC LIMIT 5""",
            (record_id, company_id),
        ).fetchall()
        blockers.extend(
            f"{MODULES.get(row['module'], row['module'])}: {row['title']}" for row in references
        )
        tender = self.db.connection().execute(
            """SELECT id,title FROM tender_results
               WHERE company_id=? AND converted_record_id=? LIMIT 1""",
            (company_id, record_id),
        ).fetchone()
        if tender:
            blockers.append(f"Licitação convertida: {tender['title']}")
        return blockers

    def trash_purge(self, path, session):
        if not self.require_admin(session):
            return
        try:
            data = self.parse_json(max_bytes=8 * 1024)
        except ValueError as exc:
            return self.error_json(str(exc))
        pieces = path.split("/")
        record_id = int(pieces[3]) if len(pieces) == 4 and pieces[3].isdigit() else None
        if path != "/api/trash" and record_id is None:
            return self.error_json("Registro inválido", 404)
        expected = "EXCLUIR" if record_id else "ESVAZIAR"
        if str(data.get("confirmation") or "").strip().upper() != expected:
            return self.error_json(f"Digite {expected} para confirmar a exclusão definitiva")
        company_id = session["company_id"]
        with self.db.transaction(immediate=True):
            if record_id:
                rows = self.db.connection().execute(
                    """SELECT id,module,title FROM records
                       WHERE id=? AND company_id=? AND deleted_at IS NOT NULL""",
                    (record_id, company_id),
                ).fetchall()
            else:
                rows = self.db.connection().execute(
                    """SELECT id,module,title FROM records
                       WHERE company_id=? AND deleted_at IS NOT NULL ORDER BY deleted_at,id""",
                    (company_id,),
                ).fetchall()
            if record_id and not rows:
                return self.error_json("Registro excluído não encontrado", 404)
            blocked = []
            purgeable = []
            for row in rows:
                reasons = self.trash_record_blockers(row["id"], company_id)
                if reasons:
                    blocked.append({"id": row["id"], "title": row["title"], "reasons": reasons})
                else:
                    purgeable.append(row)
            if record_id and blocked:
                return self.send_json({
                    "ok": False, "error": "record_referenced",
                    "message": "Este registro ainda é usado por cadastro(s) ativo(s) e não pode ser apagado definitivamente.",
                    "blockedItems": blocked,
                }, 409)
            purge_ids = [row["id"] for row in purgeable]
            deleted_count = 0
            for offset in range(0, len(purge_ids), 400):
                batch = purge_ids[offset:offset + 400]
                placeholders = ",".join("?" for _ in batch)
                self.db.execute(
                    f"DELETE FROM record_versions WHERE company_id=? AND record_id IN ({placeholders})",
                    (company_id, *batch),
                )
                deleted = self.db.execute(
                    f"""DELETE FROM records WHERE company_id=? AND deleted_at IS NOT NULL
                        AND id IN ({placeholders})""",
                    (company_id, *batch),
                )
                deleted_count += deleted.rowcount
                if deleted.rowcount != len(batch):
                    raise sqlite3.IntegrityError("A lixeira foi alterada durante a exclusão")
            if deleted_count != len(purge_ids):
                raise sqlite3.IntegrityError("Nem todos os registros da lixeira foram excluídos")
            detail = {
                "purged_count": len(purge_ids), "purged_ids": purge_ids[:200],
                "blocked_count": len(blocked), "blocked_ids": [item["id"] for item in blocked[:200]],
            }
            self.db.audit(
                session["id"], "purge", "trash", str(record_id or "all"), detail,
                company_id=company_id,
            )
        return self.send_json({
            "ok": True, "purged": len(purge_ids), "blocked": len(blocked),
            "blockedItems": blocked,
        })

    def record_restore(self, path, session):
        pieces = path.split("/")
        if len(pieces) != 4 or not pieces[3].isdigit():
            return self.error_json("Registro inválido", 404)
        record_id = int(pieces[3])
        row = self.db.connection().execute(
            "SELECT * FROM records WHERE id=? AND company_id=? AND deleted_at IS NOT NULL",
            (record_id, session["company_id"])
        ).fetchone()
        if not row:
            return self.error_json("Registro excluído não encontrado", 404)
        if not self.require_operation(session, row["module"], "restore"):
            return
        now = utc_now()
        with self.db.transaction(immediate=True):
            self.save_record_version(row, session["id"])
            updated = self.db.execute(
                """UPDATE records SET deleted_at=NULL,updated_at=?,revision=revision+1
                   WHERE id=? AND company_id=? AND deleted_at IS NOT NULL""",
                (now, record_id, session["company_id"]),
            )
            if updated.rowcount != 1:
                return self.error_json(
                    "O registro já foi restaurado por outra pessoa", 409, "write_conflict"
                )
            self.db.audit(
                session["id"], "restore", row["module"], record_id,
                {"title": row["title"], "revision": row["revision"] + 1},
                company_id=session["company_id"],
            )
        return self.send_json({"ok": True})

    def user_create(self, session):
        if not self.require_admin(session):
            return
        try:
            data = self.parse_json()
        except ValueError as exc:
            return self.error_json(str(exc))
        name = str(data.get("name", "")).strip()
        email = str(data.get("email", "")).strip().lower()
        password = str(data.get("password", ""))
        role = str(data.get("role", "operator"))
        if role not in ROLE_MODULES:
            return self.error_json("Perfil inválido")
        permission_spec = {}
        if ("effectivePermissions" in data or "effectiveCapabilities" in data
                or "effectiveActions" in data):
            try:
                permission_spec = self.effective_permission_spec(
                    role, data.get("effectivePermissions", {}),
                    data.get("effectiveCapabilities"), data.get("effectiveActions"),
                )
            except ValueError as exc:
                return self.error_json(str(exc))
        if (len(name) < 2 or not re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", email)
                or len(password) < 10):
            return self.error_json("Informe nome, e-mail válido e senha com pelo menos 10 caracteres")
        now = utc_now()
        try:
            with self.db.transaction(immediate=True):
                existing = self.db.connection().execute(
                    "SELECT id FROM users WHERE email=?", (email,)
                ).fetchone()
                if existing:
                    user_id = existing["id"]
                    membership = self.db.connection().execute(
                        "SELECT 1 FROM company_memberships WHERE company_id=? AND user_id=?",
                        (session["company_id"], user_id),
                    ).fetchone()
                    if membership:
                        return self.error_json(
                            "Este e-mail já possui acesso nesta empresa. Use Redefinir senha se necessário.",
                            409, "duplicate_membership",
                        )
                else:
                    cursor = self.db.execute(
                        """INSERT INTO users(name,email,password_hash,role,created_at,updated_at)
                           VALUES(?,?,?,?,?,?)""",
                        (name, email, password_hash(password), role, now, now),
                    )
                    user_id = cursor.lastrowid
                self.db.execute(
                    """INSERT INTO company_memberships
                       (company_id,user_id,role,permissions,active,created_at,updated_at)
                       VALUES(?,?,?,?,1,?,?)""",
                    (session["company_id"], user_id, role,
                     json_dumps(permission_spec), now, now),
                )
                self.db.audit(
                    session["id"], "create", "user", user_id,
                    {"email": email, "role": role, "permissions": permission_spec},
                    company_id=session["company_id"],
                )
        except sqlite3.IntegrityError:
            return self.error_json("Não foi possível vincular o usuário à empresa", 409, "duplicate_membership")
        return self.send_json({"ok": True, "id": user_id, "existingAccount": bool(existing)}, 201)

    def user_password_reset(self, path, session):
        """Redefinição administrativa, sem expor a senha nem o hash."""
        if not self.require_admin(session):
            return
        pieces = path.split("/")
        if len(pieces) != 5 or not pieces[3].isdigit() or pieces[4] != "password":
            return self.error_json("Usuário inválido", 404)
        try:
            data = self.parse_json()
        except ValueError as exc:
            return self.error_json(str(exc))
        password = str(data.get("password", ""))
        if len(password) < 10:
            return self.error_json("A nova senha deve possuir ao menos 10 caracteres")
        user_id = int(pieces[3])
        membership = self.db.connection().execute(
            "SELECT 1 FROM company_memberships WHERE company_id=? AND user_id=?",
            (session["company_id"], user_id),
        ).fetchone()
        if not membership:
            return self.error_json("Usuário não encontrado", 404)
        now = utc_now()
        with self.db.transaction(immediate=True):
            self.db.execute(
                "UPDATE users SET password_hash=?,active=1,updated_at=? WHERE id=?",
                (password_hash(password), now, user_id),
            )
            # Tokens anteriores deixam de valer assim que a senha muda.
            self.db.execute("DELETE FROM sessions WHERE user_id=?", (user_id,))
            self.db.audit(
                session["id"], "password_reset", "user", user_id,
                {"by_admin": True}, company_id=session["company_id"],
            )
        return self.send_json({"ok": True})

    def user_update(self, path, session):
        if not self.require_admin(session):
            return
        pieces = path.split("/")
        if len(pieces) != 4 or not pieces[3].isdigit():
            return self.error_json("Usuário inválido", 404)
        user_id = int(pieces[3])
        try:
            data = self.parse_json()
        except ValueError as exc:
            return self.error_json(str(exc))
        role = str(data.get("role", "operator"))
        active = 1 if data.get("active", True) else 0
        if role not in ROLE_MODULES:
            return self.error_json("Perfil inválido")
        if user_id == session["id"] and not active:
            return self.error_json("O administrador não pode desativar a própria conta")
        if user_id == session["id"] and role != "admin":
            return self.error_json("O administrador não pode remover o próprio perfil administrativo")
        target = self.db.connection().execute(
            "SELECT role,active,permissions FROM company_memberships WHERE company_id=? AND user_id=?",
            (session["company_id"], user_id)).fetchone()
        if not target:
            return self.error_json("Usuário não encontrado", 404)
        if target["role"] == "admin" and target["active"] and (not active or role != "admin"):
            active_admins = self.db.scalar(
                """SELECT COUNT(*) FROM company_memberships
                   WHERE company_id=? AND role='admin' AND active=1""", (session["company_id"],))
            if active_admins <= 1:
                return self.error_json("O sistema deve manter ao menos um administrador ativo")
        permission_spec = self.permission_spec(target)
        if ("effectivePermissions" in data or "effectiveCapabilities" in data
                or "effectiveActions" in data):
            try:
                permission_spec = self.effective_permission_spec(
                    role, data.get("effectivePermissions", {}),
                    data.get("effectiveCapabilities"), data.get("effectiveActions"),
                )
            except ValueError as exc:
                return self.error_json(str(exc))
        elif role != target["role"]:
            # Uma troca explícita de perfil aplica exatamente a nova matriz-base.
            # Exceções antigas pertencem ao perfil anterior e não devem ganhar um
            # significado diferente só porque a base de comparação mudou.
            permission_spec = {}
        with self.db.transaction(immediate=True):
            self.db.execute(
                """UPDATE company_memberships SET role=?,active=?,permissions=?,updated_at=?
                   WHERE company_id=? AND user_id=?""",
                (role, active, json_dumps(permission_spec), utc_now(),
                 session["company_id"], user_id))
            if not active:
                self.db.execute(
                    "DELETE FROM sessions WHERE user_id=? AND company_id=?",
                    (user_id, session["company_id"]),
                )
            self.db.audit(
                session["id"], "update", "user", user_id,
                {"role": role, "active": bool(active), "permissions": permission_spec},
                company_id=session["company_id"],
            )
        return self.send_json({"ok": True})

    @staticmethod
    def normalized_text(value):
        text = unicodedata.normalize("NFD", str(value or "").lower())
        return "".join(char for char in text if unicodedata.category(char) != "Mn")

    TENDER_GENERIC_WORDS = {
        "a", "ao", "aos", "as", "com", "contratacao", "da", "das", "de", "do", "dos",
        "e", "em", "equipamento", "equipamentos", "fornecimento", "manutencao", "na", "nas",
        "no", "nos", "o", "os", "para", "por", "produto", "produtos", "realizacao", "servico",
        "servicos", "sistema", "sistemas", "tecnica", "tecnico",
    }

    @classmethod
    def tender_significant_tokens(cls, value):
        return {
            token for token in re.findall(r"[a-z0-9]+", cls.normalized_text(value))
            if len(token) >= 3 and token not in cls.TENDER_GENERIC_WORDS
        }

    def tender_portfolio(self, company_id):
        """Lê somente o catálogo ativo da empresa corrente, sem cruzar tenants."""
        rows = self.db.connection().execute(
            """SELECT id,module,title,payload FROM records
               WHERE company_id=? AND module IN ('produtos','catalogo_servicos')
                 AND deleted_at IS NULL AND lower(status) NOT IN ('inativo','cancelado','descartado')
               ORDER BY module,id""",
            (company_id,),
        ).fetchall()
        portfolio = []
        for row in rows:
            try:
                payload = json.loads(row["payload"] or "{}")
            except (TypeError, json.JSONDecodeError):
                payload = {}
            title = str(row["title"] or payload.get("title") or "").strip()
            descriptive = " ".join(str(payload.get(key) or "") for key in (
                "descricao", "familia", "categoria", "tipo_item", "tipo_servico",
            ))
            tokens = self.tender_significant_tokens(f"{title} {descriptive}")
            title_tokens = self.tender_significant_tokens(title)
            if tokens:
                portfolio.append({
                    "id": row["id"], "module": row["module"], "title": title,
                    "tokens": tokens, "title_tokens": title_tokens,
                })
        return portfolio

    def tender_portfolio_matches(self, candidate_text, matched_terms, portfolio):
        """Exige evidência técnica do catálogo; palavras genéricas isoladas não bastam."""
        evidence_text = " ".join([
            str(candidate_text or ""),
            " ".join(str(term or "") for term in (matched_terms or [])),
        ])
        evidence_tokens = self.tender_significant_tokens(evidence_text)
        matches = []
        for record in portfolio:
            shared = record["tokens"] & evidence_tokens
            title_shared = record["title_tokens"] & evidence_tokens
            distinctive = shared & {"hepa", "ulpa", "pao", "uvc", "csb", "hvac", "iso14644"}
            title_required = min(2, len(record["title_tokens"]))
            strong = bool(distinctive) or (
                len(shared) >= 2 and len(title_shared) >= title_required
            )
            if strong:
                matches.append({
                    "id": record["id"], "module": record["module"], "title": record["title"],
                    "evidence": sorted(shared)[:8],
                })
        return matches[:8]

    def tender_result_portfolio_data(self, row, portfolio, official_items=None):
        matched_terms = json.loads(row["matched_terms"] or "[]")
        item_text = " ".join(
            str(item.get("descricao") or "") for item in (official_items or []) if isinstance(item, dict)
        )
        matches = self.tender_portfolio_matches(
            f"{row['object_text'] or ''} {item_text}", matched_terms, portfolio,
        )
        return matched_terms, matches

    @classmethod
    def tender_text_queries(cls, keywords, offset=0):
        """Seleciona um lote rotativo sem fingir que todos os termos cabem na cota do PNCP."""
        candidates = keywords
        selected = []
        seen = set()
        for item in candidates:
            value = str(item).strip().strip('"')
            key = cls.normalized_text(value)
            if len(key) < 3 or key in seen:
                continue
            selected.append(value)
            seen.add(key)
        if len(selected) <= PNCP_TEXT_QUERIES_PER_SEARCH:
            return selected
        start = max(0, int(offset or 0)) % len(selected)
        ordered = selected[start:] + selected[:start]
        return ordered[:PNCP_TEXT_QUERIES_PER_SEARCH]

    @classmethod
    def normalize_tender_keywords(cls, raw_keywords, limit=80):
        """Normaliza, limita e remove duplicatas sem perder a grafia informada."""
        if isinstance(raw_keywords, str):
            candidates = re.split(r"[\n,;\t]+", raw_keywords)
        elif isinstance(raw_keywords, list):
            candidates = raw_keywords
        else:
            raise ValueError("Palavras-chave inválidas")
        keywords = []
        seen = set()
        for raw in candidates:
            value = str(raw or "").strip().strip("'\"")[:180]
            normalized = cls.normalized_text(value).strip()
            if len(normalized) < 3 or normalized in seen:
                continue
            seen.add(normalized)
            keywords.append(value)
            if len(keywords) >= limit:
                break
        if not keywords:
            raise ValueError("Informe ao menos uma palavra-chave")
        return keywords

    @classmethod
    def tender_spreadsheet_keywords(cls, filename, content):
        """Extrai termos de CSV/XLSX de forma limitada e independente da ordem das colunas."""
        extension = Path(filename or "").suffix.lower()
        if extension not in {".csv", ".txt", ".xlsx"}:
            raise ValueError("Use uma planilha XLSX ou CSV. Arquivos XLS antigos devem ser salvos como XLSX.")
        if not content or len(content) > 2 * 1024 * 1024:
            raise ValueError("A planilha deve possuir entre 1 byte e 2 MB")
        sheet_name = Path(filename).name
        rows = []
        if extension == ".xlsx":
            try:
                with zipfile.ZipFile(io.BytesIO(content)) as archive:
                    if len(archive.infolist()) > 2000 or sum(item.file_size for item in archive.infolist()) > 20 * 1024 * 1024:
                        raise ValueError("A planilha XLSX expandida excede o limite seguro")
                from openpyxl import load_workbook
            except ImportError:
                raise ValueError("O leitor XLSX não está instalado no servidor") from None
            except zipfile.BadZipFile:
                raise ValueError("Arquivo XLSX inválido") from None
            try:
                workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                worksheet = workbook.active
                sheet_name = worksheet.title
                for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                    if row_index > 5000:
                        break
                    rows.append(["" if value is None else str(value).strip() for value in row[:20]])
                workbook.close()
            except (OSError, ValueError, TypeError, zipfile.BadZipFile) as exc:
                raise ValueError(f"Não foi possível ler a planilha XLSX: {exc}") from None
        else:
            try:
                text = content.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = content.decode("cp1252")
            sample = text[:4096]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
            except csv.Error:
                dialect = csv.excel
                dialect.delimiter = ";"
            rows = [list(row[:20]) for index, row in enumerate(csv.reader(io.StringIO(text), dialect)) if index < 5000]

        rows = [[str(cell or "").strip() for cell in row] for row in rows if any(str(cell or "").strip() for cell in row)]
        if not rows:
            raise ValueError("A planilha não contém dados")
        aliases = {
            "keyword": {"palavra chave", "palavras chave", "keyword", "keywords", "termo", "termos", "expressao", "expressões", "expressao de busca"},
            "category": {"categoria", "grupo", "tema", "familia", "família"},
            "active": {"ativa", "ativo", "status", "usar", "incluir"},
        }
        normalized_header = [re.sub(r"[^a-z0-9]+", " ", cls.normalized_text(cell)).strip() for cell in rows[0]]
        indexes = {}
        for kind, names in aliases.items():
            indexes[kind] = next((index for index, value in enumerate(normalized_header) if value in {cls.normalized_text(name) for name in names}), None)
        has_header = indexes["keyword"] is not None
        keyword_index = indexes["keyword"] if has_header else 0
        entries = []
        ignored = 0
        duplicate_count = 0
        seen = set()
        inactive_values = {"0", "nao", "não", "n", "false", "inativa", "inativo", "pausada", "pausado"}
        for source_row, row in enumerate(rows[1:] if has_header else rows, start=2 if has_header else 1):
            if keyword_index >= len(row):
                ignored += 1
                continue
            keyword = str(row[keyword_index] or "").strip().strip("'\"")[:180]
            active_index = indexes.get("active")
            active = row[active_index] if active_index is not None and active_index < len(row) else "sim"
            normalized = cls.normalized_text(keyword).strip()
            if len(normalized) < 3 or cls.normalized_text(active).strip() in inactive_values:
                ignored += 1
                continue
            if normalized in seen:
                duplicate_count += 1
                continue
            seen.add(normalized)
            category_index = indexes.get("category")
            category = row[category_index][:80] if category_index is not None and category_index < len(row) else ""
            significant = cls.tender_significant_tokens(keyword)
            entries.append({
                "keyword": keyword, "category": category, "row": source_row,
                "specificity": "específica" if len(significant) >= 2 or any(len(token) <= 5 for token in significant) else "revisar",
            })
            if len(entries) >= 80:
                ignored += max(0, len(rows) - source_row)
                break
        if not entries:
            raise ValueError("Nenhuma palavra-chave ativa foi identificada na planilha")
        return {
            "keywords": [entry["keyword"] for entry in entries], "entries": entries,
            "duplicates": duplicate_count, "ignored": ignored, "sheet": sheet_name,
            "headerDetected": has_header,
        }

    def tender_keywords_import(self, session):
        try:
            data = self.parse_json(max_bytes=3 * 1024 * 1024)
            filename = Path(str(data.get("filename") or "planilha.csv")).name
            content = base64.b64decode(str(data.get("content") or ""), validate=True)
            result = self.tender_spreadsheet_keywords(filename, content)
        except (ValueError, binascii.Error) as exc:
            return self.error_json(str(exc))
        self.db.audit(
            session["id"], "parse", "tender_keywords", detail={
                "filename": filename[:180], "terms": len(result["keywords"]),
                "duplicates": result["duplicates"], "ignored": result["ignored"],
            }, company_id=session["company_id"],
        )
        return self.send_json({"ok": True, **result})

    @staticmethod
    def normalize_pncp_search_item(item):
        item_url = str(item.get("item_url") or "").strip()
        if item_url.startswith("/"):
            item_url = "https://pncp.gov.br/app" + item_url
        return {
            "objetoCompra": item.get("description") or item.get("title") or "Contratação PNCP",
            "informacaoComplementar": "",
            "numeroControlePNCP": item.get("numero_controle_pncp"),
            "orgaoEntidadeRazaoSocial": item.get("orgao_nome"),
            "unidadeOrgaoUfSigla": item.get("uf"),
            "unidadeOrgaoMunicipioNome": item.get("municipio_nome"),
            "modalidadeNome": item.get("modalidade_licitacao_nome") or item.get("tipo_nome"),
            "valorTotalEstimado": item.get("valor_global"),
            "dataPublicacaoPncp": item.get("data_publicacao_pncp") or item.get("createdAt"),
            "dataEncerramentoProposta": item.get("data_fim_vigencia"),
            "linkSistemaOrigem": item_url or None,
            "_portal_search": item,
        }

    def competitor_insights(self, session):
        """Entrega somente um benchmark agregado de preços para a tela competitiva."""
        if not self.require_module_read(session, "concorrentes"):
            return
        if "editais" not in self.allowed_modules(session, "read"):
            return self.send_json({"ok": True, "available": False, "reason": "editais_forbidden", "count": 0, "average": None, "latest": []})
        if "view_values" not in self.allowed_operations(session, "editais"):
            return self.send_json({"ok": True, "available": False, "reason": "values_forbidden", "count": 0, "average": None, "latest": []})
        rows = self.db.connection().execute(
            """SELECT id,title,object_text,agency,uf,modality,estimated_value,deadline,published_at,status,source_url
               FROM tender_results
               WHERE company_id=? AND estimated_value IS NOT NULL AND estimated_value>0
               ORDER BY COALESCE(published_at,updated_at) DESC, id DESC LIMIT 30""",
            (session["company_id"],),
        ).fetchall()
        values = [float(row["estimated_value"]) for row in rows]
        latest = [{
            "id": row["id"], "title": row["title"], "object": row["object_text"],
            "agency": row["agency"], "uf": row["uf"], "modality": row["modality"],
            "value": row["estimated_value"], "deadline": row["deadline"],
            "publishedAt": row["published_at"], "status": row["status"],
            "sourceUrl": row["source_url"],
        } for row in rows[:10]]
        return self.send_json({
            "ok": True, "available": bool(values), "count": len(values),
            "average": round(sum(values) / len(values), 2) if values else None,
            "latest": latest,
        })

    def tender_results_get(self, query, session):
        status = (query.get("status") or [""])[0].strip()
        search = (query.get("q") or [""])[0].strip()
        sql = "SELECT * FROM tender_results WHERE company_id=?"
        params = [session["company_id"]]
        if status:
            sql += " AND status=?"
            params.append(status)
        if search:
            sql += " AND (title LIKE ? OR object_text LIKE ? OR agency LIKE ?)"
            params.extend([f"%{search}%"] * 3)
        sql += " ORDER BY CASE status WHEN 'Novo' THEN 0 WHEN 'Analisar' THEN 1 ELSE 2 END, relevance_score DESC, deadline ASC LIMIT 1000"
        rows = self.db.connection().execute(sql, params).fetchall()
        portfolio = self.tender_portfolio(session["company_id"])
        details = self.db.connection().execute(
            "SELECT tender_result_id,items_json FROM tender_details WHERE company_id=?",
            (session["company_id"],),
        ).fetchall()
        official_items = {}
        for detail in details:
            try:
                official_items[detail["tender_result_id"]] = json.loads(detail["items_json"] or "[]")
            except (TypeError, json.JSONDecodeError):
                official_items[detail["tender_result_id"]] = []
        items = []
        show_values = "view_values" in self.allowed_operations(session, "editais")
        for row in rows:
            item = dict(row)
            matched_terms, matches = self.tender_result_portfolio_data(
                row, portfolio, official_items.get(row["id"]),
            )
            item["matched_terms"] = matched_terms
            item["portfolio_matches"] = matches
            item["strict_match"] = bool(matches)
            if not show_values:
                item["estimated_value"] = None
                item["values_restricted"] = True
            items.append(item)
        quality_row = self.db.connection().execute(
            """SELECT COUNT(*) evaluated,
                      SUM(CASE WHEN relevance_feedback='relevant' THEN 1 ELSE 0 END) relevant,
                      SUM(CASE WHEN relevance_feedback='irrelevant' THEN 1 ELSE 0 END) irrelevant
               FROM tender_results WHERE company_id=? AND relevance_feedback IS NOT NULL""",
            (session["company_id"],),
        ).fetchone()
        evaluated = int(quality_row["evaluated"] or 0)
        relevant = int(quality_row["relevant"] or 0)
        return self.send_json({
            "ok": True, "items": items, "valuesVisible": show_values,
            "portfolioCount": len(portfolio),
            "strictCount": sum(1 for item in items if item["strict_match"]),
            "quality": {
                "evaluated": evaluated, "relevant": relevant,
                "irrelevant": int(quality_row["irrelevant"] or 0),
                "precisionPercent": round(100 * relevant / evaluated, 1) if evaluated else None,
                "minimumSampleReached": evaluated >= 30,
            },
        })

    @staticmethod
    def pncp_purchase_parts(external_id):
        """Extrai CNPJ, ano e sequencial do numero de controle do PNCP."""
        try:
            left, year = str(external_id).split("/", 1)
            cnpj, _marker, sequence = left.split("-", 2)
            if not (cnpj.isdigit() and len(cnpj) == 14 and year.isdigit() and sequence.isdigit()):
                raise ValueError
            return cnpj, int(year), int(sequence)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def tender_value_from_official_data(detail, items):
        """Nunca infere valor quando o PNCP informa orçamento sigiloso."""
        secret = int(detail.get("orcamentoSigilosoCodigo") or 0) in {1, 2, 3}
        if secret or any(bool(item.get("orcamentoSigiloso")) for item in items):
            return None, "sigiloso"
        declared = detail.get("valorTotalEstimado")
        if isinstance(declared, (int, float)) and declared > 0:
            return float(declared), "valor_total_pncp"
        totals = [item.get("valorTotal") for item in items]
        if totals and all(isinstance(value, (int, float)) and value >= 0 for value in totals):
            total = sum(totals)
            if total > 0:
                return float(total), "soma_itens_pncp"
        return None, "nao_publicado"

    def tender_result_get(self, path, session):
        pieces = path.split("/")
        if len(pieces) == 5 and pieces[4].isdigit():
            result_id = int(pieces[4])
            row = self.db.connection().execute(
                "SELECT * FROM tender_results WHERE id=? AND company_id=?",
                (result_id, session["company_id"]),
            ).fetchone()
            if not row:
                return self.error_json("Oportunidade não encontrada", 404)
            detail = self.db.connection().execute(
                "SELECT * FROM tender_details WHERE tender_result_id=? AND company_id=?",
                (result_id, session["company_id"]),
            ).fetchone()
            payload = dict(row)
            payload["matched_terms"] = json.loads(payload["matched_terms"] or "[]")
            payload["official"] = ({
                "data": json.loads(detail["official_data"] or "{}"),
                "items": json.loads(detail["items_json"] or "[]"),
                "documents": json.loads(detail["documents_json"] or "[]"),
                "valueSource": detail["value_source"],
                "analysis": json.loads(detail["analysis_json"] or "{}"),
                "refreshedAt": detail["refreshed_at"],
                "refreshError": detail["refresh_error"],
            } if detail else None)
            show_values = "view_values" in self.allowed_operations(session, "editais")
            if not show_values:
                payload["estimated_value"] = None
                payload["official"] = self.redact_nested_values(payload["official"])
                payload["values_restricted"] = True
            return self.send_json({"ok": True, "item": payload})
        if len(pieces) == 7 and pieces[4].isdigit() and pieces[5] == "documentos" and pieces[6].isdigit():
            return self.tender_document_download(int(pieces[4]), int(pieces[6]), session)
        return self.error_json("Oportunidade ou documento inválido", 404)

    def tender_result_refresh(self, path, session):
        pieces = path.split("/")
        if len(pieces) != 6 or not pieces[4].isdigit() or pieces[5] != "refresh":
            return self.error_json("Oportunidade inválida", 404)
        result_id = int(pieces[4])
        row = self.db.connection().execute(
            "SELECT * FROM tender_results WHERE id=? AND company_id=?",
            (result_id, session["company_id"]),
        ).fetchone()
        if not row:
            return self.error_json("Oportunidade não encontrada", 404)
        parts = self.pncp_purchase_parts(row["external_id"])
        if not parts:
            return self.error_json("Este resultado não possui identificador PNCP utilizável", 422)
        cnpj, year, sequence = parts
        detail_url = f"https://pncp.gov.br/api/consulta/v1/orgaos/{cnpj}/compras/{year}/{sequence}"
        api_base = f"https://pncp.gov.br/api/pncp/v1/orgaos/{cnpj}/compras/{year}/{sequence}"
        try:
            official_data = self.fetch_tender_json(detail_url, timeout=18, attempts=2)
            items = self.fetch_tender_json(api_base + "/itens", timeout=18, attempts=2)
            documents = self.fetch_tender_json(api_base + "/arquivos", timeout=18, attempts=2)
        except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, json.JSONDecodeError, ConnectionError) as exc:
            return self.error_json(f"PNCP não respondeu para este edital: {exc}", 502, "pncp_unavailable")
        if not isinstance(items, list):
            items = []
        if not isinstance(documents, list):
            documents = []
        value, value_source = self.tender_value_from_official_data(official_data, items)
        now = utc_now()
        with self.db.transaction(immediate=True):
            self.db.execute(
                """INSERT INTO tender_details
                   (tender_result_id,company_id,official_data,items_json,documents_json,value_source,refreshed_at,refresh_error)
                   VALUES(?,?,?,?,?,?,?,NULL)
                   ON CONFLICT(tender_result_id) DO UPDATE SET official_data=excluded.official_data,
                     items_json=excluded.items_json,documents_json=excluded.documents_json,
                     value_source=excluded.value_source,refreshed_at=excluded.refreshed_at,refresh_error=NULL""",
                (result_id, session["company_id"], json_dumps(official_data), json_dumps(items),
                 json_dumps(documents), value_source, now),
            )
            self.db.execute(
                "UPDATE tender_results SET estimated_value=?,updated_at=? WHERE id=? AND company_id=?",
                (value, now, result_id, session["company_id"]),
            )
            self.db.audit(session["id"], "refresh", "tender_result", result_id,
                          {"source": "PNCP", "value_source": value_source, "documents": len(documents)},
                          company_id=session["company_id"])
        show_values = "view_values" in self.allowed_operations(session, "editais")
        return self.send_json({"ok": True, "value": value if show_values else None,
                               "valueSource": value_source if show_values else None,
                               "documents": len(documents), "items": len(items), "refreshedAt": now})

    @classmethod
    def redact_nested_values(cls, value):
        """Remove campos monetários da resposta sem alterar o dado oficial persistido."""
        if isinstance(value, list):
            return [cls.redact_nested_values(item) for item in value]
        if not isinstance(value, dict):
            return value
        restricted_terms = ("valor", "preco", "price", "amount", "budget", "orcamento")
        return {
            key: cls.redact_nested_values(item)
            for key, item in value.items()
            if not any(term in cls.normalized_text(key) for term in restricted_terms)
        }

    @staticmethod
    def tender_document_bytes(document):
        """Obtém somente arquivo HTTPS do PNCP, sem seguir redirecionamentos."""
        url = str(document.get("url") or document.get("uri") or "")
        parsed = urllib.parse.urlparse(url)
        if parsed.scheme != "https" or parsed.hostname not in {"pncp.gov.br", "www.pncp.gov.br"}:
            raise ValueError("O documento não possui URL oficial PNCP válida")
        class NoRedirect(urllib.request.HTTPRedirectHandler):
            def redirect_request(self, req, fp, code, msg, headers, newurl):
                return None
        request = urllib.request.Request(url, headers={"User-Agent": f"SIVS/{VERSION}"})
        with urllib.request.build_opener(NoRedirect).open(request, timeout=30) as response:
            length = int(response.headers.get("Content-Length") or 0)
            if length > MAX_TENDER_DOCUMENT:
                raise ValueError("Documento oficial excede o limite de 20 MB")
            body = response.read(MAX_TENDER_DOCUMENT + 1)
            if len(body) > MAX_TENDER_DOCUMENT:
                raise ValueError("Documento oficial excede o limite de 20 MB")
            return body, response.headers.get_content_type() or "application/octet-stream"

    def tender_document_download(self, result_id, document_index, session):
        detail = self.db.connection().execute(
            "SELECT documents_json FROM tender_details WHERE tender_result_id=? AND company_id=?",
            (result_id, session["company_id"]),
        ).fetchone()
        if not detail:
            return self.error_json("Atualize os dados oficiais antes de abrir documentos", 404)
        documents = json.loads(detail["documents_json"] or "[]")
        if document_index < 0 or document_index >= len(documents) or not isinstance(documents[document_index], dict):
            return self.error_json("Documento não encontrado", 404)
        document = documents[document_index]
        try:
            body, mime_type = self.tender_document_bytes(document)
        except ValueError as exc:
            return self.error_json(str(exc), 422)
        except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError) as exc:
            return self.error_json(f"Não foi possível obter o documento no PNCP: {exc}", 502, "pncp_document_unavailable")
        filename = re.sub(r"[^A-Za-z0-9._ -]", "_", str(document.get("titulo") or "edital-pncp"))[:180]
        if not filename:
            filename = "edital-pncp"
        self.db.audit(session["id"], "download", "tender_document", result_id,
                      {"document": filename, "source": "PNCP"}, company_id=session["company_id"])
        self._response_started = True
        self.send_response(200)
        self.send_header("Content-Type", mime_type)
        self.send_header("Content-Disposition", f'inline; filename="{filename}"')
        self.send_header("Content-Length", str(len(body)))
        self.send_header("X-Content-Type-Options", "nosniff")
        self.security_headers("SAMEORIGIN")
        self.end_headers()
        self.wfile.write(body)

    @staticmethod
    def tender_pdf_text(body, document_name):
        """Extrai texto com limites para que o edital não exceda o contexto da IA."""
        reader = PdfReader(io.BytesIO(body), strict=False)
        pages = []
        for page_number, page in enumerate(reader.pages[:120], start=1):
            try:
                text = (page.extract_text() or "").strip()
            except Exception:
                text = ""
            try:
                has_images = len(page.images) > 0
            except Exception:
                has_images = False
            if text or has_images:
                pages.append({"document": document_name, "page": page_number, "text": text[:8000], "hasImages": has_images})
            if sum(len(item["text"]) for item in pages) >= 50_000:
                break
        return pages

    @staticmethod
    def tender_pages_markdown(pages, max_chars=90_000):
        """Formata os trechos em Markdown; páginas com imagem apontam para o PDF em vez de texto perdido."""
        page_blocks = []
        current_document = None
        for item in pages:
            parts = []
            if item["document"] != current_document:
                current_document = item["document"]
                parts.append(f"# {current_document}")
            parts.append(f"## Página {item['page']}")
            if item["text"]:
                parts.append(item["text"])
            if item.get("hasImages"):
                parts.append(
                    "[Página com imagem ou tabela em formato de imagem, não convertida para texto — "
                    "não descreva o conteúdo visual; registre como pendência e recomende consultar o PDF original nesta página.]"
                )
            elif not item["text"]:
                parts.append("[Página sem texto extraível.]")
            page_blocks.append("\n\n".join(parts))
        # Junta blocos inteiros até o limite, para nunca cortar um marcador de página/imagem no meio.
        included, used = [], 0
        for block in page_blocks:
            added = len(block) + (2 if included else 0)
            if included and used + added > max_chars:
                break
            included.append(block)
            used += added
        return "\n\n".join(included)

    def openrouter_tender_analysis(self, tender, pages):
        key = os.environ.get("OPENROUTER_API_KEY", "").strip()
        if not key:
            raise ValueError("OPENROUTER_API_KEY ausente")
        model = os.environ.get("OPENROUTER_TENDER_MODEL") or os.environ.get(
            "OPENROUTER_ASSISTANT_MODEL", "openai/gpt-5.4-mini"
        )
        source_text = self.tender_pages_markdown(pages)
        prompt = (
            "Analise exclusivamente os trechos de documentos oficiais do PNCP abaixo, formatados em Markdown "
            "com um título por documento (#) e uma seção por página (##). "
            "Não invente fatos, não conclua conformidade jurídica e não afirme direito de impugnar sem apontar "
            "a cláusula e a página. Quando uma página estiver marcada como imagem não convertida, não descreva "
            "o conteúdo visual: registre a limitação em riscos_pendencias e recomende consultar o PDF original "
            "nessa página. Produza JSON com as chaves: resumo (string), prazos (lista de objetos com evento, "
            "data, citacao), habilitacao (lista), requisitos_tecnicos (lista), obrigacoes_contratadas (lista), "
            "criterios_julgamento (lista), participacao (objeto com situacao em 'apta_para_revisao', 'pendencias' ou "
            "'nao_verificada', itens e justificativa), riscos_pendencias (lista), recomendacao (string), "
            "minuta_esclarecimento (string), minuta_impugnacao (string), citacoes (lista de objetos com documento, pagina, achado). "
            "As minutas são rascunhos para revisão jurídica, devem conter [PREENCHER] onde faltar dado e nunca alegar fato não presente. "
            "Limite cada lista a 8 itens e as citações a 12, com achados de até 240 caracteres. Cada achado deve ter citação; "
            "quando o trecho não bastar, escreva 'não identificado nos trechos lidos'.\n\n"
            f"CONTRATAÇÃO: {json_dumps(tender)}\n\nTRECHOS:\n{source_text}"
        )
        body = {
            "model": model,
            "messages": [
                {"role": "system", "content": "Você é analista de editais públicos. Responda somente JSON válido."},
                {"role": "user", "content": prompt},
            ],
            "response_format": {"type": "json_object"},
            "temperature": 0.1,
            "max_tokens": 4000,
        }
        request = urllib.request.Request(
            "https://openrouter.ai/api/v1/chat/completions", data=json_dumps(body).encode("utf-8"),
            headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json",
                     "HTTP-Referer": "https://sivs-seccol.local", "X-Title": "SIVS SECCOL"}, method="POST",
        )
        with urllib.request.urlopen(request, timeout=75) as response:
            payload = json.load(response)
        content = payload["choices"][0]["message"]["content"]
        analysis = json.loads(content) if isinstance(content, str) else content
        if not isinstance(analysis, dict):
            raise ValueError("A IA retornou uma análise inválida")
        # Alguns provedores preservam a palavra inglesa apesar da instrução em
        # português; normalizamos a chave antes de persistir o contrato da UI.
        if "riscos_pendencias" not in analysis and "risks_pendencias" in analysis:
            analysis["riscos_pendencias"] = analysis.pop("risks_pendencias")
        return analysis, payload.get("model") or model

    def tender_result_analyze(self, path, session):
        pieces = path.split("/")
        if len(pieces) != 6 or not pieces[4].isdigit() or pieces[5] != "analyze":
            return self.error_json("Oportunidade inválida", 404)
        result_id = int(pieces[4])
        row = self.db.connection().execute(
            "SELECT * FROM tender_results WHERE id=? AND company_id=?", (result_id, session["company_id"])
        ).fetchone()
        detail = self.db.connection().execute(
            "SELECT documents_json FROM tender_details WHERE tender_result_id=? AND company_id=?",
            (result_id, session["company_id"]),
        ).fetchone()
        if not row or not detail:
            return self.error_json("Atualize os dados oficiais do PNCP antes de solicitar a leitura", 409)
        documents = json.loads(detail["documents_json"] or "[]")
        pages, skipped = [], []
        priority = {"edital": 0, "aviso": 1, "termo de referência": 2, "projeto básico": 3}
        documents = sorted(
            documents, key=lambda document: next(
                (rank for term, rank in priority.items() if term in str(document.get("tipoDocumentoNome") or document.get("titulo") or "").lower()),
                9,
            )
        )
        for document in documents[:8]:
            name = str(document.get("titulo") or document.get("tipoDocumentoNome") or "Documento PNCP")
            try:
                body, mime_type = self.tender_document_bytes(document)
                if mime_type != "application/pdf" and not name.lower().endswith(".pdf"):
                    skipped.append(f"{name}: formato não textual")
                    continue
                extracted = self.tender_pdf_text(body, name)
                if extracted:
                    pages.extend(extracted)
                else:
                    skipped.append(f"{name}: PDF sem texto extraível (requer OCR)")
            except (OSError, ValueError, PyPdfError, urllib.error.URLError, urllib.error.HTTPError, TimeoutError) as exc:
                skipped.append(f"{name}: não foi possível ler ({type(exc).__name__})")
            if sum(len(item["text"]) for item in pages) >= 70_000:
                break
        if not pages:
            return self.error_json("Nenhum texto de edital pôde ser extraído. Os PDFs podem exigir OCR.", 422)
        tender = {key: row[key] for key in ("external_id", "object_text", "agency", "modality", "deadline")}
        try:
            analysis, model = self.openrouter_tender_analysis(tender, pages)
        except (OSError, ValueError, KeyError, json.JSONDecodeError, urllib.error.URLError, urllib.error.HTTPError) as exc:
            return self.error_json(f"A IA não concluiu a leitura do edital: {exc}", 502, "ai_analysis_unavailable")
        image_pages = [{"document": item["document"], "page": item["page"]} for item in pages if item.get("hasImages")]
        stored = {"status": "completed", "generatedAt": utc_now(), "model": model,
                  "documentsRead": sorted({item["document"] for item in pages}), "pagesRead": len(pages),
                  "skipped": skipped, "imagePages": image_pages, "result": analysis}
        with self.db.transaction(immediate=True):
            self.db.execute(
                "UPDATE tender_details SET analysis_json=? WHERE tender_result_id=? AND company_id=?",
                (json_dumps(stored), result_id, session["company_id"]),
            )
            self.db.audit(session["id"], "analyze", "tender_result", result_id,
                          {"model": model, "pages": len(pages), "documents": len(stored["documentsRead"])},
                          company_id=session["company_id"])
        return self.send_json({"ok": True, "analysis": stored})

    def tender_search(self, session):
        try:
            data = self.parse_json()
        except ValueError as exc:
            return self.error_json(str(exc))
        try:
            with self.db.transaction(immediate=True):
                active = self.db.connection().execute(
                    """SELECT id,status,progress,stage FROM tender_jobs
                       WHERE company_id=? AND status IN ('queued','running')
                       ORDER BY id DESC LIMIT 1""",
                    (session["company_id"],),
                ).fetchone()
                if active:
                    return self.send_json({
                        "ok": True, "jobId": active["id"], "status": active["status"],
                        "progress": active["progress"], "stage": active["stage"],
                        "alreadyRunning": True,
                    }, 202)
                cursor = self.db.execute(
                    """INSERT INTO tender_jobs
                       (company_id,status,request_json,progress,stage,created_by,created_at)
                       VALUES(?,'queued',?,0,'Pesquisa enfileirada',?,?)""",
                    (session["company_id"], json_dumps(data), session["id"], utc_now()),
                )
                job_id = cursor.lastrowid
        except sqlite3.IntegrityError:
            active = self.db.connection().execute(
                """SELECT id,status,progress,stage FROM tender_jobs
                   WHERE company_id=? AND status IN ('queued','running') ORDER BY id DESC LIMIT 1""",
                (session["company_id"],),
            ).fetchone()
            if active:
                return self.send_json({"ok": True, "jobId": active["id"],
                                       "status": active["status"], "alreadyRunning": True}, 202)
            return self.error_json("Não foi possível enfileirar a pesquisa", 409, "job_conflict")
        worker = threading.Thread(
            target=self.server.run_tender_job,
            args=(self, job_id, dict(session), data),
            name=f"sivs-tender-{job_id}", daemon=True,
        )
        worker.start()
        return self.send_json({"ok": True, "jobId": job_id, "status": "queued"}, 202)

    def tender_job_get(self, job_id, session):
        row = self.db.connection().execute(
            """SELECT id,status,progress,stage,result_json,error_detail,
                      created_at,started_at,heartbeat_at,finished_at
               FROM tender_jobs WHERE id=? AND company_id=?""",
            (job_id, session["company_id"]),
        ).fetchone()
        if not row:
            return self.error_json("Pesquisa não encontrada", 404)
        item = dict(row)
        try:
            item["result"] = json_loads_strict(item.pop("result_json")) if item["result_json"] else None
        except (ValueError, json.JSONDecodeError):
            item["result"] = None
        return self.send_json({"ok": True, "job": item})

    def _run_tender_job(self, job_id, session, data):
        now = utc_now()
        self.db.execute(
            """UPDATE tender_jobs SET status='running',progress=2,stage='Conectando às fontes oficiais',
               started_at=?,heartbeat_at=? WHERE id=? AND status='queued'""",
            (now, now, job_id),
        )

        def progress(percent, stage):
            self.db.execute(
                """UPDATE tender_jobs SET progress=?,stage=?,heartbeat_at=?
                   WHERE id=? AND status='running'""",
                (max(0, min(int(percent), 99)), str(stage)[:240], utc_now(), job_id),
            )

        try:
            result = self.execute_tender_search(session, data, progress)
            finished = utc_now()
            self.db.execute(
                """UPDATE tender_jobs SET status='completed',progress=100,stage='Pesquisa concluída',
                   result_json=?,heartbeat_at=?,finished_at=? WHERE id=?""",
                (json_dumps(result), finished, finished, job_id),
            )
            job = self.db.connection().execute(
                "SELECT schedule_id FROM tender_jobs WHERE id=?", (job_id,)
            ).fetchone()
            if job and job["schedule_id"]:
                self.db.execute(
                    "UPDATE search_schedules SET last_run_at=?,updated_at=? WHERE id=?",
                    (finished, finished, job["schedule_id"]),
                )
        except Exception as exc:
            reference = secrets.token_hex(6)
            print(f"[ERRO PESQUISA {reference}] job={job_id}")
            traceback.print_exc()
            finished = utc_now()
            self.db.execute(
                """UPDATE tender_jobs SET status='failed',stage='Pesquisa não concluída',
                   error_detail=?,heartbeat_at=?,finished_at=? WHERE id=?""",
                (f"Falha interna; referência {reference}", finished, finished, job_id),
            )

    def execute_tender_search(self, session, data, progress=None):
        progress = progress or (lambda _percent, _stage: None)
        progress(5, "Validando parâmetros da pesquisa")
        raw_keywords = data.get("keywords") or DEFAULT_TENDER_KEYWORDS
        company_id = session["company_id"]
        try:
            keywords = self.normalize_tender_keywords(raw_keywords)
        except ValueError as exc:
            return self.error_json(str(exc))
        uf = str(data.get("uf", "")).strip().upper()[:2]
        try:
            days = max(1, min(int(data.get("days", 7)), 30))
            modalities = data.get("modalities") or [4, 5, 6, 7, 8, 9, 12]
            modalities = [int(value) for value in modalities
                          if int(value) in {4, 5, 6, 7, 8, 9, 12}]
        except (ValueError, TypeError):
            return self.error_json("Período ou modalidade inválida")
        if not modalities:
            modalities = [4, 5, 6, 7, 8, 9, 12]
        # O PNCP publica timestamps em UTC; usar a mesma referência evita
        # descartar uma publicação feita após meia-noite UTC e ainda no dia
        # anterior no fuso local da empresa.
        end = datetime.now(timezone.utc).date()
        start = end - timedelta(days=days)
        found = 0
        inserted = 0
        errors = []
        successful_pages = 0
        normalized_keywords = [(keyword, self.normalized_text(keyword)) for keyword in keywords]
        normalized_context = [(term, self.normalized_text(term)) for term in SECCOL_CONTEXT_TERMS]
        portfolio = self.tender_portfolio(company_id)
        source_status = {"pncp": "indisponível", "comprasgov": "não acionado"}
        sources_used = ["pncp"]
        planned_pages = 0
        completed_jobs = 0

        def store_item(item, retrieved_via, matched_override=None):
            nonlocal found, inserted
            haystack = self.normalized_text(f"{item.get('objetoCompra','')} {item.get('informacaoComplementar','')}")
            matched = list(matched_override or [
                original for original, normalized in normalized_keywords
                if normalized and normalized in haystack
            ])
            if not matched:
                return
            object_text = str(item.get("objetoCompra") or "").strip()
            portfolio_matches = self.tender_portfolio_matches(object_text, matched, portfolio)
            if not portfolio_matches:
                return
            context_hits = [original for original, normalized in normalized_context if normalized in haystack]
            external_id = str(item.get("numeroControlePNCP") or "").strip()
            if not external_id:
                return
            found += 1
            agency_data = item.get("orgaoEntidade") or {}
            unit_data = item.get("unidadeOrgao") or {}
            agency = (agency_data.get("razaoSocial") or item.get("orgaoEntidadeRazaoSocial") or
                      item.get("nomeOrgao") or "Órgão não informado")
            item_uf = unit_data.get("ufSigla") or item.get("unidadeOrgaoUfSigla") or uf or None
            municipality = unit_data.get("municipioNome") or item.get("unidadeOrgaoMunicipioNome")
            modality_name = item.get("modalidadeNome") or "Contratação"
            title = f"{modality_name} — {agency}"
            source_url = item.get("linkSistemaOrigem") or self.pncp_public_url(external_id)
            deadline = item.get("dataEncerramentoProposta") or item.get("dataEncerramentoPropostaPncp")
            raw_item = dict(item)
            raw_item["_recuperado_via"] = retrieved_via
            raw_item["_portfolio_matches"] = portfolio_matches
            raw_item["_strict_match"] = True
            now = utc_now()
            stored_source_key = "pncp" if company_id == 1 else f"{company_id}:pncp"
            cursor = self.db.execute(
                """INSERT OR IGNORE INTO tender_results
                   (source_key,external_id,title,object_text,agency,uf,municipality,modality,estimated_value,
                    published_at,deadline,source_url,matched_terms,relevance_score,status,raw_json,created_at,updated_at,
                    company_id)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (stored_source_key, external_id, title[:500], object_text, agency, item_uf, municipality, modality_name,
                 item.get("valorTotalEstimado"), item.get("dataPublicacaoPncp"), deadline, source_url,
                 json_dumps(matched), min(
                     100,
                     (55 if matched_override else 40) + len(matched) * 10 +
                     min(3, len(portfolio_matches)) * 8 + min(4, len(context_hits)) * 3,
                 ), "Novo",
                 json_dumps(raw_item), now, now, company_id),
            )
            inserted += cursor.rowcount

        def fetch_json(job, timeout=14):
            _modality, _page, url = job
            return self.fetch_tender_json(url, timeout=timeout)

        # O indice textual usado pelo proprio portal encontra termos tambem nos itens e anexos.
        # A API cronologica, mantida abaixo como contingencia, tem centenas de paginas por
        # modalidade e fazia o SIVS examinar apenas uma fracao arbitraria do periodo.
        # O PNCP limita chamadas em sequência. Cada execução usa oito termos
        # e avança para o próximo lote, cobrindo todo o vocabulário salvo em
        # execuções sucessivas, sem deixar silenciosamente os demais termos de fora.
        keyword_signature = json_dumps(keywords)
        previous_searches = self.db.scalar(
            "SELECT COUNT(*) FROM tender_searches WHERE company_id=? AND keywords=?",
            (company_id, keyword_signature),
        ) or 0
        text_queries = self.tender_text_queries(
            keywords, offset=previous_searches * PNCP_TEXT_QUERIES_PER_SEARCH,
        )
        portal_responses = 0
        portal_candidates = {}
        planned_pages += len(text_queries)
        progress(10, "Pesquisando termos técnicos no índice oficial do PNCP")
        for query_index, search_term in enumerate(text_queries, start=1):
            params = {
                "q": f'"{search_term}"',
                "tipos_documento": "edital",
                "pagina": 1,
                "tam_pagina": PNCP_TEXT_RESULTS_PER_QUERY,
                "status": "recebendo_proposta",
                "ordenacao": "-data",
            }
            url = "https://pncp.gov.br/api/search/?" + urllib.parse.urlencode(params)
            try:
                payload = self.fetch_tender_json(url, timeout=10, attempts=2)
                portal_responses += 1
                successful_pages += 1
                for raw_item in payload.get("items", []):
                    if raw_item.get("cancelado"):
                        continue
                    if uf and str(raw_item.get("uf") or "").upper() != uf:
                        continue
                    published = str(
                        raw_item.get("data_publicacao_pncp") or raw_item.get("createdAt") or ""
                    )[:10]
                    if published and not (start.isoformat() <= published <= end.isoformat()):
                        continue
                    external_id = str(raw_item.get("numero_controle_pncp") or "").strip()
                    if not external_id:
                        continue
                    candidate = portal_candidates.setdefault(external_id, {
                        "item": self.normalize_pncp_search_item(raw_item),
                        "matched": [],
                    })
                    if search_term not in candidate["matched"]:
                        candidate["matched"].append(search_term)
            except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError,
                    json.JSONDecodeError, ConnectionError) as exc:
                errors.append(f"PNCP busca textual '{search_term}': {exc}")
            completed_jobs += 1
            progress(
                10 + int(35 * query_index / max(1, len(text_queries))),
                f"PNCP textual: {query_index}/{len(text_queries)} termo(s) consultado(s)",
            )
            if query_index < len(text_queries):
                time.sleep(0.2)
        for candidate in portal_candidates.values():
            store_item(candidate["item"], "PNCP — busca textual", candidate["matched"])

        # Primeiro testa uma página de cada modalidade. Só amplia a paginação se a fonte responder.
        first_jobs = []
        publication_modalities = [] if portal_responses else modalities
        for modality in publication_modalities:
            params = {"dataInicial": start.strftime("%Y%m%d"), "dataFinal": end.strftime("%Y%m%d"),
                      "codigoModalidadeContratacao": modality, "pagina": 1, "tamanhoPagina": 50}
            if uf:
                params["uf"] = uf
            first_jobs.append((modality, 1, "https://pncp.gov.br/api/consulta/v1/contratacoes/publicacao?" +
                               urllib.parse.urlencode(params)))
        planned_pages += len(first_jobs)
        progress(10, "Consultando primeiras páginas do PNCP")
        pncp_payloads = []
        # O PNCP limita rajadas. A execução paralela anterior provocava HTTP
        # 429 em parte das páginas e, mesmo assim, apresentava sucesso total.
        for job in first_jobs:
            try:
                payload = fetch_json(job)
                pncp_payloads.append((job, payload))
                successful_pages += 1
                for item in payload.get("data", []):
                    store_item(item, "PNCP")
            except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, json.JSONDecodeError) as exc:
                errors.append(f"PNCP modalidade {job[0]}, página 1: {exc}")
            completed_jobs += 1
            progress(10 + int(35 * completed_jobs / max(1, planned_pages)),
                     f"PNCP: {completed_jobs}/{planned_pages} consulta(s) processadas")

        if pncp_payloads or portal_responses:
            followups = []
            for (modality, _page, _url), payload in pncp_payloads:
                remaining = min(3 if modality in {4, 6, 8} else 1, int(payload.get("paginasRestantes", 0) or 0))
                for page in range(2, remaining + 2):
                    params = {"dataInicial": start.strftime("%Y%m%d"), "dataFinal": end.strftime("%Y%m%d"),
                              "codigoModalidadeContratacao": modality, "pagina": page, "tamanhoPagina": 50}
                    if uf:
                        params["uf"] = uf
                    followups.append((modality, page, "https://pncp.gov.br/api/consulta/v1/contratacoes/publicacao?" +
                                      urllib.parse.urlencode(params)))
            # O serviço recusa rajadas a partir de aproximadamente dez
            # chamadas. Reservamos uma chamada da cota e priorizamos as
            # primeiras páginas de todas as modalidades antes de aprofundar.
            remaining_budget = max(0, PNCP_MAX_REQUESTS_PER_SEARCH - len(first_jobs))
            followups = followups[:remaining_budget]
            planned_pages += len(followups)
            progress(48, "Ampliando a paginação das modalidades aderentes")
            for job in followups:
                try:
                    payload = fetch_json(job)
                    successful_pages += 1
                    for item in payload.get("data", []):
                        store_item(item, "PNCP")
                except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, json.JSONDecodeError) as exc:
                    errors.append(f"PNCP modalidade {job[0]}, página {job[1]}: {exc}")
                completed_jobs += 1
                progress(45 + int(35 * completed_jobs / max(1, planned_pages)),
                         f"PNCP: {completed_jobs}/{planned_pages} consulta(s) processadas")
            source_status["pncp"] = "concluído" if successful_pages == planned_pages else "parcial"
        else:
            # Fallback oficial: API de Dados Abertos do Compras.gov.br.
            sources_used.append("comprasgov")
            source_status["comprasgov"] = "consultando"
            modality_map = {4: 3, 6: 5, 8: 6, 9: 7}
            fallback_jobs = []
            for pncp_modality in modalities:
                compras_modality = modality_map.get(pncp_modality)
                if not compras_modality:
                    continue
                params = {"pagina": 1, "tamanhoPagina": 100,
                          "dataPublicacaoPncpInicial": start.isoformat(), "dataPublicacaoPncpFinal": end.isoformat(),
                          "codigoModalidade": compras_modality}
                if uf:
                    params["unidadeOrgaoUfSigla"] = uf
                url = ("https://dadosabertos.compras.gov.br/modulo-contratacoes/"
                       "1_consultarContratacoes_PNCP_14133?" + urllib.parse.urlencode(params))
                fallback_jobs.append((pncp_modality, 1, url))
            planned_pages += len(fallback_jobs)
            progress(48, "PNCP indisponível; acionando contingência Compras.gov.br")
            fallback_success = 0
            for job in fallback_jobs:
                try:
                    payload = fetch_json(job, 20)
                    fallback_success += 1
                    successful_pages += 1
                    for item in payload.get("resultado", []):
                        store_item(item, "Compras.gov.br")
                except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, json.JSONDecodeError) as exc:
                    errors.append(f"Compras.gov modalidade {job[0]}: {exc}")
                completed_jobs += 1
                progress(50 + int(30 * completed_jobs / max(1, planned_pages)),
                         f"Compras.gov.br: {completed_jobs}/{planned_pages} consulta(s) processadas")
            source_status["comprasgov"] = (
                "concluído" if fallback_success == len(fallback_jobs)
                else "parcial" if fallback_success else "indisponível"
            )
        progress(86, "Consolidando, deduplicando e registrando resultados")
        self.db.execute(
            """INSERT INTO tender_searches
               (keywords,uf,days,sources_searched,found_count,new_count,error_detail,created_by,created_at,company_id)
               VALUES(?,?,?,?,?,?,?,?,?,?)""",
            (keyword_signature, uf or None, days, json_dumps(sources_used), found, inserted,
             "\n".join(errors) if errors else None, session["id"], utc_now(), company_id),
        )
        execution_time = utc_now()
        for source_key in ("pncp", "comprasgov"):
            source_state = source_status[source_key]
            succeeded = source_state == "concluído"
            self.db.execute(
                """UPDATE records SET
                     payload=json_set(payload,'$.ultima_execucao',?,'$.ultimo_sucesso',
                                      CASE WHEN ? THEN ? ELSE json_extract(payload,'$.ultimo_sucesso') END,
                                      '$.ultimo_estado',?),updated_at=?
                   WHERE company_id=? AND module='fontes'
                     AND json_extract(payload,'$.source_key')=?""",
                (execution_time, 1 if succeeded else 0, execution_time, source_state,
                 execution_time, company_id, source_key))
        self.db.audit(session["id"], "search", "tenders",
                      detail={"found": found, "new": inserted, "uf": uf, "days": days}, company_id=company_id)
        if successful_pages == 0:
            message = "PNCP e Compras.gov.br não responderam. Nenhum resultado foi descartado; tente novamente."
        elif source_status["pncp"] == "indisponível":
            message = (f"PNCP indisponível; fallback oficial do Compras.gov.br concluído: "
                       f"{found} oportunidade(s) aderente(s), {inserted} nova(s).")
        elif source_status["pncp"] == "parcial":
            message = (f"Pesquisa parcial no PNCP: {successful_pages}/{planned_pages} consulta(s) responderam, "
                       f"com {found} oportunidade(s) aderente(s) e {inserted} nova(s). "
                       "Tente novamente para completar.")
        elif portal_responses:
            message = (f"Pesquisa textual concluída no PNCP: {found} oportunidade(s) aderente(s), "
                       f"{inserted} nova(s), em {len(text_queries)} termo(s) técnico(s).")
        else:
            message = f"Pesquisa concluída: {found} oportunidade(s) aderente(s), {inserted} nova(s)."
        progress(97, "Atualizando fontes, histórico e trilha de auditoria")
        return {"ok": True, "found": found, "new": inserted, "errors": errors,
                "pagesChecked": successful_pages, "pagesPlanned": planned_pages,
                "sourceStatus": source_status, "queriesUsed": text_queries,
                "keywordTotal": len(keywords), "queryCount": len(text_queries),
                "coveragePercent": round(100 * len(text_queries) / max(1, len(keywords))),
                "message": message}

    @staticmethod
    def fetch_tender_json(url, timeout=14, attempts=4):
        """Consulta uma fonte oficial e repete somente falhas transitórias."""
        transient_statuses = {429, 500, 502, 503, 504}
        for attempt in range(attempts):
            request = urllib.request.Request(
                url,
                headers={"Accept": "application/json", "User-Agent": f"SIVS/{VERSION}"},
            )
            try:
                with urllib.request.urlopen(request, timeout=timeout) as response:
                    if response.status == 204:
                        return {}
                    return json.load(response)
            except urllib.error.HTTPError as exc:
                if exc.code not in transient_statuses or attempt + 1 >= attempts:
                    raise
                retry_after = exc.headers.get("Retry-After") if exc.headers else None
                try:
                    delay = float(retry_after)
                except (TypeError, ValueError):
                    delay = 1.5 * (2 ** attempt)
                time.sleep(max(0.5, min(delay, 12)))
            except TimeoutError:
                # Um segundo timeout já indica indisponibilidade daquela
                # página; quatro tentativas podiam reter o job por um minuto.
                if attempt >= 1 or attempt + 1 >= attempts:
                    raise
                time.sleep(0.75)
            except (urllib.error.URLError, ConnectionError):
                if attempt + 1 >= attempts:
                    raise
                time.sleep(min(1.5 * (2 ** attempt), 12))
        raise RuntimeError("Fonte oficial não respondeu")

    @staticmethod
    def pncp_public_url(external_id):
        try:
            left, year = external_id.split("/")
            cnpj, marker, sequence = left.split("-")
            return f"https://pncp.gov.br/app/editais/{cnpj}/{year}/{int(sequence)}"
        except (ValueError, TypeError):
            return "https://pncp.gov.br/app/editais"

    def tender_result_update(self, path, session):
        pieces = path.split("/")
        if len(pieces) != 5 or not pieces[4].isdigit():
            return self.error_json("Oportunidade inválida", 404)
        try:
            data = self.parse_json()
        except ValueError as exc:
            return self.error_json(str(exc))
        status = str(data.get("status") or "").strip()
        feedback = str(data.get("relevanceFeedback") or "").strip()
        reason = str(data.get("feedbackReason") or "").strip()[:500]
        if status and status not in {"Novo", "Analisar", "Aprovado", "Descartado", "Convertido"}:
            return self.error_json("Situação inválida")
        if feedback and feedback not in {"relevant", "irrelevant"}:
            return self.error_json("Avaliação de aderência inválida")
        if not status and not feedback:
            return self.error_json("Informe a situação ou a avaliação de aderência")
        result_id = int(pieces[4])
        now = utc_now()
        if status and feedback:
            cursor = self.db.execute(
                """UPDATE tender_results SET status=?,relevance_feedback=?,feedback_reason=?,
                          feedback_at=?,feedback_by=?,updated_at=? WHERE id=? AND company_id=?""",
                (status, feedback, reason or None, now, session["id"], now,
                 result_id, session["company_id"]),
            )
        elif feedback:
            cursor = self.db.execute(
                """UPDATE tender_results SET relevance_feedback=?,feedback_reason=?,
                          feedback_at=?,feedback_by=?,updated_at=? WHERE id=? AND company_id=?""",
                (feedback, reason or None, now, session["id"], now,
                 result_id, session["company_id"]),
            )
        else:
            cursor = self.db.execute(
                "UPDATE tender_results SET status=?,updated_at=? WHERE id=? AND company_id=?",
                (status, now, result_id, session["company_id"]),
            )
        if not cursor.rowcount:
            return self.error_json("Oportunidade não encontrada", 404)
        self.db.audit(session["id"], "triage", "tender_result", result_id,
                      {"status": status or None, "relevance": feedback or None},
                      company_id=session["company_id"])
        return self.send_json({"ok": True})

    def tender_convert(self, path, session):
        pieces = path.split("/")
        if len(pieces) != 5 or not pieces[4].isdigit():
            return self.error_json("Oportunidade inválida", 404)
        result_id = int(pieces[4])
        row = self.db.connection().execute(
            "SELECT * FROM tender_results WHERE id=? AND company_id=?",
            (result_id, session["company_id"])).fetchone()
        if not row:
            return self.error_json("Oportunidade não encontrada", 404)
        if row["converted_record_id"]:
            return self.send_json({"ok": True, "recordId": row["converted_record_id"], "alreadyConverted": True})
        now = utc_now()
        opening_date = str(row["deadline"] or row["published_at"] or datetime.now().date().isoformat())[:10]
        record_payload = {
            "orgao": row["agency"], "edital": row["external_id"], "portal": row["source_url"],
            "modalidade": row["modality"], "data_abertura": opening_date,
            "fonte_resultado_id": row["id"], "notes": row["object_text"], "etapa": "Captação",
            "assunto": row["title"], "relacionamentos": []
        }
        try:
            values = self.normalized_record({
                "module": "licitacoes", "title": row["title"], "status": "Captação",
                "amount": row["estimated_value"], "due_date": opening_date,
                "payload": record_payload,
            })
        except ValueError as exc:
            return self.error_json(f"A oportunidade não possui dados suficientes para conversão: {exc}")
        with self.db.transaction(immediate=True):
            current = self.db.connection().execute(
                "SELECT converted_record_id FROM tender_results WHERE id=? AND company_id=?",
                (result_id, session["company_id"]),
            ).fetchone()
            if current and current["converted_record_id"]:
                return self.send_json({"ok": True, "recordId": current["converted_record_id"],
                                       "alreadyConverted": True})
            cursor = self.db.execute(
                """INSERT INTO records
                   (module,title,status,amount,due_date,payload,created_by,created_at,updated_at,
                    company_id,revision)
                   VALUES(?,?,?,?,?,?,?,?,?,?,1)""",
                (*values, session["id"], now, now, session["company_id"]),
            )
            record_id = cursor.lastrowid
            self.db.sync_relationships(
                record_id, record_payload, session["id"], session["company_id"]
            )
            updated = self.db.execute(
                """UPDATE tender_results SET status='Convertido',converted_record_id=?,
                          relevance_feedback='relevant',feedback_at=?,feedback_by=?,updated_at=?
                   WHERE id=? AND company_id=? AND converted_record_id IS NULL""",
                (record_id, now, session["id"], now, result_id, session["company_id"]),
            )
            if updated.rowcount != 1:
                raise sqlite3.IntegrityError("conversão concorrente")
            self.db.audit(
                session["id"], "convert", "tender_result", result_id,
                {"record_id": record_id}, company_id=session["company_id"],
            )
        return self.send_json({"ok": True, "recordId": record_id})

    def search_schedule_save(self, session):
        try:
            data = self.parse_json()
        except ValueError as exc:
            return self.error_json(str(exc))
        name = str(data.get("name") or "Monitor diário de editais").strip()[:180]
        try:
            keywords = self.normalize_tender_keywords(data.get("keywords") or DEFAULT_TENDER_KEYWORDS)
        except ValueError as exc:
            return self.error_json(str(exc))
        frequency = str(data.get("frequency") or "daily")
        if frequency not in {"manual", "daily", "weekly"}:
            return self.error_json("Frequência inválida")
        try:
            days = max(1, min(int(data.get("days") or 7), 30))
        except (ValueError, TypeError):
            return self.error_json("Período inválido")
        now = utc_now()
        next_run = None
        if frequency == "daily":
            next_run = (datetime.now(timezone.utc) + timedelta(days=1)).isoformat(timespec="seconds")
        elif frequency == "weekly":
            next_run = (datetime.now(timezone.utc) + timedelta(days=7)).isoformat(timespec="seconds")
        cursor = self.db.execute(
            """INSERT INTO search_schedules
               (company_id,name,keywords,uf,days,frequency,active,next_run_at,created_by,created_at,updated_at)
               VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
            (session["company_id"], name, json_dumps(keywords),
             str(data.get("uf") or "").upper()[:2] or None,
             days, frequency,
             1 if data.get("active", True) else 0, next_run, session["id"], now, now))
        self.db.audit(session["id"], "create", "search_schedule", cursor.lastrowid,
                      {"frequency": frequency}, company_id=session["company_id"])
        return self.send_json({"ok": True, "id": cursor.lastrowid}, 201)

    def attachment_upload(self, path, session):
        parts = path.split("/")
        if len(parts) != 5 or not parts[3].isdigit():
            return self.error_json("Registro inválido", 404)
        record_id = int(parts[3])
        record = self.db.connection().execute(
            "SELECT module FROM records WHERE id=? AND company_id=? AND deleted_at IS NULL",
            (record_id, session["company_id"])).fetchone()
        if not record:
            return self.error_json("Registro não encontrado", 404)
        if not self.require_operation(session, record["module"], "manage_attachments"):
            return
        try:
            data = self.parse_json()
            encoded = str(data.get("content") or "")
            if encoded.startswith("data:"):
                encoded = encoded.split(",", 1)[-1]
            content = base64.b64decode(encoded, validate=True)
        except (ValueError, TypeError, binascii.Error) as exc:
            return self.error_json(f"Arquivo inválido: {exc}")
        if not content or len(content) > MAX_ATTACHMENT:
            return self.error_json("O arquivo deve possuir até 10 MB")
        filename = Path(str(data.get("filename") or "arquivo.bin")).name[:240]
        category = str(data.get("category") or "Evidência").strip()[:100]
        license_confirmed = bool(data.get("license_confirmed"))
        if category == "Cópia normativa licenciada" and not license_confirmed:
            return self.error_json(
                "Confirme que a empresa possui licença para anexar a íntegra da norma.",
                409, "license_confirmation_required",
            )
        try:
            mime_type = self.detect_attachment_mime(content, filename)
        except ValueError as exc:
            return self.error_json(str(exc))
        declared = str(data.get("mime_type") or "").split(";", 1)[0].strip().lower()
        compatible = {
            "application/zip": {
                "application/zip",
                "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            },
            "text/plain": {"text/plain", "text/csv", "application/json", "application/xml", "text/xml"},
        }
        if declared and declared != "application/octet-stream" and declared != mime_type:
            if declared not in compatible.get(mime_type, set()):
                return self.error_json("O tipo declarado não corresponde ao conteúdo do arquivo")
        digest = hashlib.sha256(content).hexdigest()
        now = utc_now()
        with self.db.transaction(immediate=True):
            cursor = self.db.execute(
                """INSERT INTO attachments
                   (company_id,record_id,filename,mime_type,content,size,category,version,
                    uploaded_by,created_at,sha256,license_confirmed)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
                (session["company_id"], record_id, filename, mime_type, content, len(content),
                 category, str(data.get("version") or "")[:40] or None,
                 session["id"], now, digest, 1 if license_confirmed else 0),
            )
            self.db.audit(
                session["id"], "upload", "attachment", cursor.lastrowid,
                {"record_id": record_id, "filename": filename, "sha256": digest,
                 "size": len(content), "category": category},
                company_id=session["company_id"],
            )
        return self.send_json({"ok": True, "id": cursor.lastrowid, "filename": filename}, 201)

    @staticmethod
    def detect_attachment_mime(content, filename):
        extension = Path(filename).suffix.lower()
        if content.startswith(b"%PDF-"):
            return "application/pdf"
        if content.startswith(b"\x89PNG\r\n\x1a\n"):
            return "image/png"
        if content.startswith(b"\xff\xd8\xff"):
            return "image/jpeg"
        if content.startswith(b"PK\x03\x04") and extension in {".zip", ".docx", ".xlsx"}:
            return "application/zip"
        if content.startswith((b"MZ", b"\x7fELF")):
            raise ValueError("Arquivos executáveis não são permitidos")
        if extension in {".xml", ".txt", ".csv", ".json"}:
            try:
                text = content.decode("utf-8-sig")
            except UnicodeDecodeError:
                raise ValueError("O arquivo textual deve usar codificação UTF-8") from None
            if "\x00" in text:
                raise ValueError("Conteúdo binário incompatível com arquivo textual")
            if extension == ".json":
                try:
                    json_loads_strict(text)
                except (ValueError, json.JSONDecodeError) as exc:
                    raise ValueError(f"JSON anexado é inválido: {exc}") from None
                return "application/json"
            if extension == ".xml":
                if "<!DOCTYPE" in text.upper() or "<!ENTITY" in text.upper():
                    raise ValueError("XML com DTD ou entidade externa não é permitido")
                try:
                    ET.fromstring(text)
                except ET.ParseError as exc:
                    raise ValueError(f"XML anexado é inválido: {exc}") from None
                return "application/xml"
            return "text/plain"
        raise ValueError(
            "Formato não permitido. Use PDF, PNG, JPEG, ZIP/DOCX/XLSX ou texto XML/JSON/CSV/TXT."
        )

    def attachment_download(self, attachment_id, session):
        row = self.db.connection().execute(
            """SELECT a.*,r.module,r.deleted_at FROM attachments a
               JOIN records r ON r.id=a.record_id
               WHERE a.id=? AND a.company_id=?""",
            (attachment_id, session["company_id"])).fetchone()
        if not row or row["deleted_at"]:
            return self.error_json("Arquivo não encontrado", 404)
        if not self.require_module_read(session, row["module"]):
            return
        body = row["content"]
        safe_name = str(row["filename"]).replace('"', "").replace("\r", "").replace("\n", "")
        self.db.audit(
            session["id"], "download", "attachment", attachment_id,
            {"record_id": row["record_id"], "filename": safe_name, "sha256": row["sha256"]},
            company_id=session["company_id"],
        )
        self._response_started = True
        self.send_response(200)
        self.send_header("Content-Type", row["mime_type"] or "application/octet-stream")
        self.send_header("Content-Disposition", f'attachment; filename="{safe_name}"')
        self.send_header("Content-Length", str(len(body)))
        self.send_header("X-Content-SHA256", row["sha256"] or hashlib.sha256(body).hexdigest())
        self.security_headers()
        self.end_headers()
        self.wfile.write(body)

    def membership_can_decide(self, company_id, user_id, module):
        membership = self.db.connection().execute(
            """SELECT u.id,cm.role,cm.permissions,cm.company_id
               FROM company_memberships cm JOIN users u ON u.id=cm.user_id
               WHERE cm.company_id=? AND cm.user_id=? AND cm.active=1 AND u.active=1""",
            (company_id, user_id),
        ).fetchone()
        return bool(
            membership
            and self.capabilities(membership)["approvals"]
            and (
                membership["role"] in {"admin", "manager", "approver"}
                or "decide_approval" in self.allowed_operations(membership, module)
            )
        )

    def approval_create(self, path, session):
        parts = path.split("/")
        if len(parts) != 5 or not parts[3].isdigit():
            return self.error_json("Registro inválido", 404)
        record_id = int(parts[3])
        record = self.db.connection().execute(
            """SELECT module,title,revision FROM records
               WHERE id=? AND company_id=? AND deleted_at IS NULL""",
            (record_id, session["company_id"])).fetchone()
        if not record:
            return self.error_json("Registro não encontrado", 404)
        if not self.require_operation(session, record["module"], "request_approval"):
            return
        try:
            data = self.parse_json()
            requested_to = int(data.get("requested_to")) if data.get("requested_to") else None
        except (ValueError, TypeError):
            return self.error_json("Aprovador inválido")
        approval_type = str(data.get("approval_type") or "Aprovação").strip()[:100]
        request_comment = str(data.get("comment") or "").strip()[:1000] or None
        if not approval_type:
            return self.error_json("Informe o tipo de aprovação")
        if requested_to == session["id"]:
            return self.error_json(
                "Solicitante e aprovador devem ser pessoas diferentes", 409, "segregation_required"
            )
        if requested_to:
            if not self.membership_can_decide(
                    session["company_id"], requested_to, record["module"]):
                return self.error_json("Selecione um aprovador ativo com perfil habilitado")
        else:
            memberships = self.db.connection().execute(
                """SELECT user_id,role FROM company_memberships
                   WHERE company_id=? AND user_id<>? AND active=1
                   ORDER BY CASE role WHEN 'approver' THEN 0 WHEN 'manager' THEN 1
                                      WHEN 'admin' THEN 2 ELSE 3 END,user_id""",
                (session["company_id"], session["id"]),
            ).fetchall()
            requested_to = next((
                membership["user_id"] for membership in memberships
                if self.membership_can_decide(
                    session["company_id"], membership["user_id"], record["module"]
                )
            ), None)
        if not requested_to:
            return self.error_json(
                "Cadastre outro usuário como aprovador, gestor ou administrador antes de solicitar.",
                409, "approver_unavailable",
            )
        now = utc_now()
        try:
            with self.db.transaction(immediate=True):
                current = self.db.connection().execute(
                    """SELECT title,revision FROM records
                       WHERE id=? AND company_id=? AND deleted_at IS NULL""",
                    (record_id, session["company_id"]),
                ).fetchone()
                if not current:
                    return self.error_json("Registro não encontrado", 404)
                cursor = self.db.execute(
                    """INSERT INTO approvals
                       (company_id,record_id,approval_type,status,requested_to,requested_by,
                        record_revision,comment,request_comment,requested_at)
                       VALUES(?,?,?,'Pendente',?,?,?,?,?,?)""",
                    (session["company_id"], record_id, approval_type, requested_to,
                     session["id"], current["revision"], request_comment, request_comment, now),
                )
                approval_id = cursor.lastrowid
                self.db.execute(
                    """INSERT INTO notifications
                       (company_id,user_id,title,message,record_id,level,created_at)
                       VALUES(?,?,?,?,?,'warning',?)""",
                    (session["company_id"], requested_to, "Aprovação pendente",
                     f'{current["title"]} aguarda sua análise.', record_id, now),
                )
                self.db.audit(
                    session["id"], "request", "approval", approval_id,
                    {"record_id": record_id, "record_revision": current["revision"],
                     "requested_to": requested_to}, company_id=session["company_id"],
                )
        except sqlite3.IntegrityError:
            return self.error_json(
                "Já existe uma solicitação pendente deste tipo para o registro.",
                409, "approval_already_pending",
            )
        return self.send_json({"ok": True, "id": approval_id, "requestedTo": requested_to}, 201)

    def approval_can_decide(self, session, approval):
        """Espelha a autorização da decisão para a UI sem enfraquecer o POST."""
        if approval.get("status") != "Pendente":
            return False
        if approval.get("requested_by") == session["id"]:
            return False
        assigned_scope = (
            approval.get("requested_to") == session["id"]
            and (
                session["role"] in {"admin", "manager", "approver"}
                or "decide_approval" in self.allowed_operations(
                    session, approval.get("module")
                )
            )
            and self.capabilities(session)["approvals"]
        )
        if (not assigned_scope and
                "decide_approval" not in self.allowed_operations(session, approval.get("module"))):
            return False
        return (
            approval.get("requested_to") == session["id"]
            or session["role"] in {"admin", "manager"}
        )

    def approval_decide(self, path, session):
        parts = path.split("/")
        if len(parts) != 4 or not parts[3].isdigit():
            return self.error_json("Aprovação inválida", 404)
        approval_id = int(parts[3])
        approval = self.db.connection().execute(
            """SELECT a.*,r.module,r.title,r.revision current_revision,r.deleted_at
               FROM approvals a JOIN records r ON r.id=a.record_id
               WHERE a.id=? AND a.company_id=? AND a.status='Pendente'""",
            (approval_id, session["company_id"])).fetchone()
        if not approval:
            return self.error_json("Aprovação pendente não encontrada", 404)
        assigned_scope = (
            approval["requested_to"] == session["id"]
            and (
                session["role"] in {"admin", "manager", "approver"}
                or "decide_approval" in self.allowed_operations(
                    session, approval["module"]
                )
            )
            and self.capabilities(session)["approvals"]
        )
        if (not assigned_scope and
                not self.require_operation(session, approval["module"], "decide_approval")):
            return
        if approval["requested_to"] != session["id"]:
            if not self.require_module_read(session, approval["module"]):
                return
        if approval["requested_by"] == session["id"]:
            return self.error_json(
                "Quem solicitou não pode decidir a própria aprovação.",
                409, "segregation_required",
            )
        if approval["requested_to"] != session["id"] and session["role"] not in {"admin", "manager"}:
            return self.error_json("A aprovação pertence a outro responsável", 403, "forbidden")
        try:
            data = self.parse_json()
        except ValueError as exc:
            return self.error_json(str(exc))
        status = str(data.get("status") or "")
        if status not in {"Aprovado", "Rejeitado"}:
            return self.error_json("Decisão inválida")
        decision_comment = str(data.get("comment") or "").strip()[:1000] or None
        if status == "Rejeitado" and not decision_comment:
            return self.error_json("Informe o motivo da rejeição")
        now = utc_now()
        stale = bool(
            approval["deleted_at"] or approval["record_revision"] != approval["current_revision"]
        )
        with self.db.transaction(immediate=True):
            if stale:
                self.db.execute(
                    """UPDATE approvals SET status='Expirada',decided_at=?,
                       decision_comment='O registro mudou após a solicitação.'
                       WHERE id=? AND company_id=? AND status='Pendente'""",
                    (now, approval_id, session["company_id"]),
                )
                self.db.audit(
                    session["id"], "expire", "approval", approval_id,
                    {"requested_revision": approval["record_revision"],
                     "current_revision": approval["current_revision"]},
                    company_id=session["company_id"],
                )
            else:
                updated = self.db.execute(
                    """UPDATE approvals
                       SET status=?,decided_by=?,decision_comment=?,decided_at=?
                       WHERE id=? AND company_id=? AND status='Pendente'""",
                    (status, session["id"], decision_comment, now,
                     approval_id, session["company_id"]),
                )
                if updated.rowcount != 1:
                    return self.error_json(
                        "Esta aprovação já foi decidida por outra pessoa.", 409, "approval_conflict"
                    )
                self.db.execute(
                    """INSERT INTO notifications
                       (company_id,user_id,title,message,record_id,level,created_at)
                       VALUES(?,?,?,?,?,?,?)""",
                    (session["company_id"], approval["requested_by"],
                     f"Solicitação {status.lower()}",
                     f'{approval["title"]}: {status}.', approval["record_id"],
                     "success" if status == "Aprovado" else "warning", now),
                )
                self.db.audit(
                    session["id"], "decide", "approval", approval_id,
                    {"status": status, "record_revision": approval["record_revision"]},
                    company_id=session["company_id"],
                )
        if stale:
            return self.error_json(
                "A solicitação expirou porque o registro foi alterado. Gere uma nova aprovação.",
                409, "approval_stale",
            )
        return self.send_json({"ok": True, "status": status})

    def technical_report_context(self, record_id, session, require_final=False):
        row = self.db.connection().execute(
            """SELECT * FROM records
               WHERE id=? AND company_id=? AND deleted_at IS NULL
                 AND module IN ('certificados','laudos_tecnicos','estudos_tecnicos')""",
            (record_id, session["company_id"]),
        ).fetchone()
        if not row:
            raise LookupError("Documento técnico não encontrado")
        if not self.require_module_read(session, row["module"]):
            return None
        norms = self.db.connection().execute(
            """SELECT n.id,n.title,n.status,n.payload,
                      SUM(CASE WHEN a.category='Cópia normativa licenciada'
                               AND a.license_confirmed=1 THEN 1 ELSE 0 END) licensed_copies,
                      COUNT(a.id) attachment_count
               FROM record_relationships rr JOIN records n ON n.id=rr.to_record_id
               LEFT JOIN attachments a ON a.record_id=n.id AND a.company_id=n.company_id
               WHERE rr.from_record_id=? AND n.company_id=? AND n.module='normas_tecnicas'
                 AND n.deleted_at IS NULL
               GROUP BY n.id ORDER BY n.title""",
            (record_id, session["company_id"]),
        ).fetchall()
        if not norms:
            raise ValueError("O documento não possui base normativa vinculada")
        parsed_norms = []
        for norm in norms:
            item = dict(norm)
            item["payload"] = json_loads_strict(item["payload"] or "{}")
            parsed_norms.append(item)
        if require_final:
            invalid = [norm["title"] for norm in parsed_norms
                       if norm["status"] in {"Obsoleta", "Cancelada"}]
            missing_licensed = [norm["title"] for norm in parsed_norms
                                if "Comercial" in str(norm["payload"].get("licenciamento") or "")
                                and not norm["licensed_copies"]]
            if invalid:
                raise ValueError(
                    "Base normativa obsoleta/cancelada: " + ", ".join(invalid[:5])
                )
            if missing_licensed:
                raise ValueError(
                    "Anexe e confirme a cópia licenciada antes da emissão final: " +
                    ", ".join(missing_licensed[:5])
                )
            approved = self.db.connection().execute(
                """SELECT id,decided_by,decided_at FROM approvals
                   WHERE company_id=? AND record_id=? AND status='Aprovado'
                     AND record_revision=? ORDER BY decided_at DESC,id DESC LIMIT 1""",
                (session["company_id"], record_id, row["revision"]),
            ).fetchone()
            if not approved:
                raise ValueError(
                    "A emissão final exige aprovação válida para a revisão atual do documento"
                )
        else:
            approved = None
        company = self.db.connection().execute(
            "SELECT name,cnpj,phone,email,address FROM companies WHERE id=?",
            (session["company_id"],),
        ).fetchone()
        return row, parsed_norms, approved, dict(company or {})

    @staticmethod
    def build_technical_report_pdf(record, norms, company, final=False, approval=None):
        try:
            from reportlab.lib import colors
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.platypus import (
                KeepTogether, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
            )
        except ImportError as exc:
            raise RuntimeError(
                "O gerador PDF não está instalado. Execute: pip install reportlab"
            ) from exc

        payload = json_loads_strict(record["payload"] or "{}")
        buffer = io.BytesIO()
        document = SimpleDocTemplate(
            buffer, pagesize=A4, rightMargin=18 * mm, leftMargin=18 * mm,
            topMargin=20 * mm, bottomMargin=18 * mm,
            title=record["title"], author=company.get("name") or "SECCOL",
        )
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(
            name="SIVSTitle", parent=styles["Title"], fontName="Helvetica-Bold",
            fontSize=17, leading=21, textColor=colors.HexColor("#171717"), spaceAfter=6,
        ))
        styles.add(ParagraphStyle(
            name="SIVSSection", parent=styles["Heading2"], fontName="Helvetica-Bold",
            fontSize=10, leading=13, textColor=colors.HexColor("#A84718"),
            spaceBefore=10, spaceAfter=6, uppercase=True,
        ))
        styles.add(ParagraphStyle(
            name="SIVSSmall", parent=styles["BodyText"], fontSize=7.5, leading=10,
            textColor=colors.HexColor("#666666"),
        ))
        styles.add(ParagraphStyle(
            name="SIVSRight", parent=styles["SIVSSmall"], alignment=TA_RIGHT,
        ))
        styles.add(ParagraphStyle(
            name="SIVSCenter", parent=styles["BodyText"], alignment=TA_CENTER,
            fontSize=8, leading=11,
        ))

        module_titles = {
            "certificados": "CERTIFICADO TÉCNICO",
            "laudos_tecnicos": "LAUDO TÉCNICO",
            "estudos_tecnicos": "ESTUDO TÉCNICO",
        }
        code = payload.get("numero") or f"SIVS-{record['id']}"
        story = [
            Table([
                [Paragraph(f"<b>{html.escape(company.get('name') or 'SECCOL')}</b><br/>"
                           f"<font size='7'>{html.escape(company.get('cnpj') or 'CNPJ não informado')}</font>",
                           styles["BodyText"]),
                 Paragraph(f"<b>{module_titles[record['module']]}</b><br/>"
                           f"{html.escape(str(code))} · revisão {record['revision']}", styles["SIVSRight"])],
            ], colWidths=[105 * mm, 51 * mm], style=TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LINEBELOW", (0, 0), (-1, -1), 1.2, colors.HexColor("#C85D23")),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ])),
            Spacer(1, 8),
            Paragraph(html.escape(record["title"]), styles["SIVSTitle"]),
            Paragraph(
                "EMISSÃO CONTROLADA" if final else "PRÉVIA — NÃO ASSINAR NEM UTILIZAR COMO DOCUMENTO FINAL",
                ParagraphStyle("State", parent=styles["SIVSCenter"], fontName="Helvetica-Bold",
                               textColor=colors.HexColor("#167A74" if final else "#B42318"),
                               backColor=colors.HexColor("#E8F5F2" if final else "#FDECEC"),
                               borderPadding=6, spaceAfter=10),
            ),
        ]

        core_fields = {
            "Cliente": payload.get("cliente") or payload.get("contato"),
            "Ordem de serviço": payload.get("os"),
            "Local avaliado": payload.get("local_avaliado"),
            "Equipamento": payload.get("equipamento"),
            "Objeto": payload.get("objeto"),
            "Responsável técnico": payload.get("responsavel_tecnico") or payload.get("aprovador"),
            "Data de emissão": payload.get("data_emissao"),
            "Assunto controlado": payload.get("assunto"),
        }
        rows = [[Paragraph(f"<b>{html.escape(label)}</b>", styles["SIVSSmall"]),
                 Paragraph(html.escape(str(value)), styles["BodyText"])]
                for label, value in core_fields.items() if value]
        story.extend([
            Paragraph("IDENTIFICAÇÃO E RASTREABILIDADE", styles["SIVSSection"]),
            Table(rows, colWidths=[45 * mm, 111 * mm], style=TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#DDDDDD")),
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F5F6F7")),
                ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ])),
        ])

        narrative_fields = [
            ("OBJETO E PREMISSAS", payload.get("premissas") or payload.get("notes")),
            ("MÉTODO / METODOLOGIA", payload.get("metodo") or payload.get("metodologia")),
            ("REGRA DE DECISÃO", payload.get("regra_decisao")),
            ("CONCLUSÃO TÉCNICA", payload.get("conclusao")),
            ("RECOMENDAÇÕES", payload.get("recomendacoes")),
        ]
        for label, value in narrative_fields:
            if value:
                paragraphs = [Paragraph(html.escape(part), styles["BodyText"])
                              for part in str(value).splitlines() if part.strip()]
                story.append(KeepTogether([Paragraph(label, styles["SIVSSection"]), *paragraphs]))

        story.extend([Paragraph("BASE NORMATIVA CONTROLADA", styles["SIVSSection"])])
        norm_rows = [[Paragraph("Referência / edição", styles["SIVSSmall"]),
                      Paragraph("Aplicabilidade registrada", styles["SIVSSmall"])]]
        for norm in norms:
            norm_payload = norm["payload"]
            norm_rows.append([
                Paragraph(f"<b>{html.escape(norm['title'])}</b><br/>"
                          f"{html.escape(str(norm_payload.get('edicao') or 'Edição não informada'))} · "
                          f"{html.escape(norm['status'])}", styles["SIVSSmall"]),
                Paragraph(html.escape(str(norm_payload.get("aplicabilidade_seccol") or
                                          norm_payload.get("escopo_resumido") or "—")), styles["SIVSSmall"]),
            ])
        story.extend([
            Table(norm_rows, repeatRows=1, colWidths=[62 * mm, 94 * mm], style=TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D8D8D8")),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#171717")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ])),
            Spacer(1, 8),
            Paragraph(
                "As fichas e citações acima não substituem a íntegra licenciada. A execução e a assinatura "
                "devem observar edição, emendas, escopo contratado, método aprovado, rastreabilidade e "
                "competência do responsável técnico.", styles["SIVSSmall"],
            ),
            Paragraph("RESPONSABILIDADE E APROVAÇÃO", styles["SIVSSection"]),
            Paragraph(
                f"Responsável técnico: {html.escape(str(payload.get('responsavel_tecnico') or payload.get('aprovador') or 'A definir'))}<br/>"
                f"Aprovação eletrônica SIVS: {html.escape(str(approval['id'])) if approval else 'não aplicável à prévia'}<br/>"
                f"Revisão de dados: {record['revision']} · registro SIVS #{record['id']}", styles["BodyText"],
            ),
        ])

        def page_footer(canvas, doc):
            canvas.saveState()
            canvas.setStrokeColor(colors.HexColor("#DDDDDD"))
            canvas.line(18 * mm, 13 * mm, 192 * mm, 13 * mm)
            canvas.setFont("Helvetica", 7)
            canvas.setFillColor(colors.HexColor("#666666"))
            canvas.drawString(18 * mm, 8.5 * mm, f"SIVS 2.2 · registro #{record['id']} · revisão {record['revision']}")
            canvas.drawRightString(192 * mm, 8.5 * mm, f"Página {doc.page}")
            if not final:
                canvas.setFont("Helvetica-Bold", 34)
                canvas.setFillColor(colors.Color(0.75, 0.1, 0.05, alpha=0.09))
                canvas.translate(105 * mm, 145 * mm)
                canvas.rotate(35)
                canvas.drawCentredString(0, 0, "PRÉVIA — NÃO CONTROLADA")
            canvas.restoreState()

        document.build(story, onFirstPage=page_footer, onLaterPages=page_footer)
        return buffer.getvalue()

    def technical_report_preview(self, record_id, session):
        try:
            context = self.technical_report_context(record_id, session, require_final=False)
            if context is None:
                return
            record, norms, approval, company = context
            body = self.build_technical_report_pdf(record, norms, company, final=False, approval=approval)
        except LookupError as exc:
            return self.error_json(str(exc), 404)
        except (ValueError, RuntimeError, json.JSONDecodeError) as exc:
            return self.error_json(str(exc))
        self.db.audit(
            session["id"], "preview", record["module"], record_id,
            {"revision": record["revision"], "norms": [norm["id"] for norm in norms]},
            company_id=session["company_id"],
        )
        filename = f"previa-{record['module']}-{record_id}-r{record['revision']}.pdf"
        return self.send_pdf(body, filename)

    def technical_report_issue(self, record_id, session):
        try:
            context = self.technical_report_context(record_id, session, require_final=True)
            if context is None:
                return
            record, norms, approval, company = context
            if not self.require_operation(session, record["module"], "issue_report"):
                return
            body = self.build_technical_report_pdf(record, norms, company, final=True, approval=approval)
        except LookupError as exc:
            return self.error_json(str(exc), 404)
        except (ValueError, RuntimeError, json.JSONDecodeError) as exc:
            return self.error_json(str(exc), 409, "issuance_blocked")
        digest = hashlib.sha256(body).hexdigest()
        version = f"r{record['revision']}"
        filename = f"{record['module']}-{record_id}-{version}.pdf"
        now = utc_now()
        with self.db.transaction(immediate=True):
            latest_record = self.db.connection().execute(
                """SELECT revision FROM records
                   WHERE id=? AND company_id=? AND deleted_at IS NULL""",
                (record_id, session["company_id"]),
            ).fetchone()
            latest_approval = self.db.connection().execute(
                """SELECT id FROM approvals
                   WHERE company_id=? AND record_id=? AND status='Aprovado'
                     AND record_revision=? ORDER BY decided_at DESC,id DESC LIMIT 1""",
                (session["company_id"], record_id, record["revision"]),
            ).fetchone()
            invalid_norms = self.db.scalar(
                """SELECT COUNT(*) FROM record_relationships rr
                   JOIN records n ON n.id=rr.to_record_id
                   WHERE rr.from_record_id=? AND n.company_id=? AND n.module='normas_tecnicas'
                     AND (n.deleted_at IS NOT NULL OR n.status IN ('Obsoleta','Cancelada'))""",
                (record_id, session["company_id"]),
            )
            missing_licensed = self.db.scalar(
                """SELECT COUNT(*) FROM record_relationships rr
                   JOIN records n ON n.id=rr.to_record_id
                   WHERE rr.from_record_id=? AND n.company_id=? AND n.module='normas_tecnicas'
                     AND json_extract(n.payload,'$.licenciamento') LIKE '%Comercial%'
                     AND NOT EXISTS (
                       SELECT 1 FROM attachments a
                       WHERE a.record_id=n.id AND a.company_id=n.company_id
                         AND a.category='Cópia normativa licenciada' AND a.license_confirmed=1
                     )""",
                (record_id, session["company_id"]),
            )
            if (not latest_record or latest_record["revision"] != record["revision"]
                    or not latest_approval or latest_approval["id"] != approval["id"]
                    or invalid_norms or missing_licensed):
                return self.error_json(
                    "O documento, a aprovação ou a base normativa mudou durante a emissão. Gere o PDF novamente.",
                    409, "issuance_context_changed",
                )
            existing = self.db.connection().execute(
                """SELECT id FROM attachments WHERE company_id=? AND record_id=?
                   AND category='Documento técnico emitido' AND version=? AND sha256=?""",
                (session["company_id"], record_id, version, digest),
            ).fetchone()
            if existing:
                attachment_id = existing["id"]
            else:
                cursor = self.db.execute(
                    """INSERT INTO attachments
                       (company_id,record_id,filename,mime_type,content,size,category,version,
                        uploaded_by,created_at,sha256,license_confirmed)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,1)""",
                    (session["company_id"], record_id, filename, "application/pdf", body, len(body),
                     "Documento técnico emitido", version, session["id"], now, digest),
                )
                attachment_id = cursor.lastrowid
            final_status = "Publicado" if record["module"] == "certificados" else "Emitido"
            self.db.execute(
                "UPDATE records SET status=?,updated_at=? WHERE id=? AND company_id=?",
                (final_status, now, record_id, session["company_id"]),
            )
            self.db.audit(
                session["id"], "issue", record["module"], record_id,
                {"revision": record["revision"], "attachment_id": attachment_id,
                 "sha256": digest, "approval_id": approval["id"],
                 "norms": [norm["id"] for norm in norms]},
                company_id=session["company_id"],
            )
        return self.send_json({
            "ok": True, "attachmentId": attachment_id,
            "downloadUrl": f"/api/attachments/{attachment_id}", "sha256": digest,
            "revision": record["revision"], "status": final_status,
        }, 201 if not existing else 200)

    def send_pdf(self, body, filename):
        self._response_started = True
        self.send_response(200)
        self.send_header("Content-Type", "application/pdf")
        self.send_header("Content-Disposition", f'inline; filename="{filename}"')
        self.send_header("Content-Length", str(len(body)))
        self.send_header("X-Content-SHA256", hashlib.sha256(body).hexdigest())
        self.security_headers()
        self.end_headers()
        self.wfile.write(body)

    @staticmethod
    def _xml_local(element, name):
        if element is None:
            return None
        for child in element.iter():
            if child.tag.rsplit("}", 1)[-1] == name:
                return child
        return None

    @classmethod
    def _xml_text(cls, element, name, default=""):
        child = cls._xml_local(element, name)
        return (child.text or default).strip() if child is not None else default

    def xml_import(self, session):
        if not self.require_operation(session, "importacoes_xml", "import_xml"):
            return
        try:
            data = self.parse_json()
        except ValueError as exc:
            return self.error_json(str(exc))
        xml_text = str(data.get("xml") or "")
        if not xml_text or len(xml_text.encode("utf-8")) > 4 * 1024 * 1024:
            return self.error_json("Selecione um XML NF-e válido de até 4 MB")
        if "<!DOCTYPE" in xml_text.upper() or "<!ENTITY" in xml_text.upper():
            return self.error_json("XML com DTD ou entidade externa não é aceito por segurança")
        try:
            root = ET.fromstring(xml_text)
        except ET.ParseError as exc:
            return self.error_json(f"XML inválido: {exc}")
        inf = self._xml_local(root, "infNFe")
        if inf is None:
            return self.error_json("O arquivo não contém uma NF-e reconhecível")
        ide = self._xml_local(inf, "ide")
        emit = self._xml_local(inf, "emit")
        dest = self._xml_local(inf, "dest")
        chave = str(inf.attrib.get("Id") or "").replace("NFe", "")
        numero = self._xml_text(ide, "nNF")
        emit_cnpj = self._xml_text(emit, "CNPJ") or self._xml_text(emit, "CPF")
        emit_name = self._xml_text(emit, "xNome") or "Fornecedor não identificado"
        destination_document = self._xml_text(dest, "CNPJ") or self._xml_text(dest, "CPF")
        if not re.fullmatch(r"\d{44}", chave):
            return self.error_json("A chave de acesso da NF-e deve possuir 44 dígitos")
        try:
            _validate_document(emit_cnpj, "CPF/CNPJ do emitente")
        except ValueError as exc:
            return self.error_json(str(exc))
        company = self.db.connection().execute(
            "SELECT name,cnpj FROM companies WHERE id=? AND active=1",
            (session["company_id"],),
        ).fetchone()
        company_document = re.sub(r"\D", "", str(company["cnpj"] or "")) if company else ""
        if not company_document:
            return self.error_json(
                "Cadastre o CNPJ da empresa ativa em Configurações antes de importar uma NF-e.",
                409, "company_document_required",
            )
        try:
            _validate_document(destination_document, "CPF/CNPJ do destinatário")
        except ValueError as exc:
            return self.error_json(str(exc))
        if re.sub(r"\D", "", destination_document) != company_document:
            return self.error_json(
                "O destinatário da NF-e não corresponde ao CNPJ da empresa ativa.",
                409, "invoice_recipient_mismatch",
            )
        existing = self.db.connection().execute(
            """SELECT id FROM records WHERE company_id=? AND module='importacoes_xml'
               AND json_extract(payload,'$.chave')=? AND deleted_at IS NULL""",
            (session["company_id"], chave)).fetchone() if chave else None
        if existing:
            return self.error_json("Esta NF-e já foi importada", 409, "duplicate_invoice")
        items = []
        for det in [node for node in inf.iter() if node.tag.rsplit("}", 1)[-1] == "det"]:
            prod = self._xml_local(det, "prod")
            items.append({
                "numero_item": det.attrib.get("nItem"), "codigo": self._xml_text(prod, "cProd"),
                "descricao": self._xml_text(prod, "xProd"), "ncm": self._xml_text(prod, "NCM"),
                "cfop": self._xml_text(prod, "CFOP"), "unidade": self._xml_text(prod, "uCom"),
                "quantidade": self._xml_text(prod, "qCom"), "valor_unitario": self._xml_text(prod, "vUnCom"),
                "valor_total": self._xml_text(prod, "vProd")
            })
        parcels = []
        for dup in [node for node in inf.iter() if node.tag.rsplit("}", 1)[-1] == "dup"]:
            parcels.append({"numero": self._xml_text(dup, "nDup"), "vencimento": self._xml_text(dup, "dVenc"),
                            "valor": self._xml_text(dup, "vDup")})
        total = self._xml_text(self._xml_local(inf, "ICMSTot"), "vNF")
        subject_name = str(data.get("assunto") or f"NF-e {numero or chave[-8:]} — {emit_name}")[:180]
        now = utc_now()
        payload = {
            "assunto": subject_name, "chave": chave, "numero": numero,
            "natureza_operacao": self._xml_text(ide, "natOp"),
            "data_emissao": (self._xml_text(ide, "dhEmi") or self._xml_text(ide, "dEmi"))[:10],
            "fornecedor": emit_name, "fornecedor_documento": emit_cnpj,
            "destinatario": self._xml_text(dest, "xNome"),
            "destinatario_documento": destination_document,
            "itens": items, "parcelas": parcels, "valor_total": total,
            "status_importacao": "Importada",
            "assinatura_xml_presente": self._xml_local(root, "Signature") is not None,
            "relacionamentos": []
        }
        try:
            amount = float(total or 0)
        except ValueError:
            amount = None
        if amount is not None and not math.isfinite(amount):
            return self.error_json("Valor total da NF-e inválido")
        self.db.begin_manual_transaction(immediate=True)
        concurrent = self.db.connection().execute(
            """SELECT id FROM records WHERE company_id=? AND module='importacoes_xml'
               AND json_extract(payload,'$.chave')=? AND deleted_at IS NULL""",
            (session["company_id"], chave),
        ).fetchone()
        if concurrent:
            self.db.finish_manual_transaction(commit=False)
            return self.error_json("Esta NF-e já foi importada", 409, "duplicate_invoice")
        cursor = self.db.execute(
            """INSERT INTO records
               (module,title,status,amount,payload,created_by,created_at,updated_at,company_id)
               VALUES('importacoes_xml',?,'Importada',?,?,?,?,?,?)""",
            (f"NF-e {numero or chave[-8:]} — {emit_name}", amount, json_dumps(payload),
             session["id"], now, now, session["company_id"]))
        import_id = cursor.lastrowid
        self.db.sync_relationships(import_id, payload, session["id"], session["company_id"])
        xml_bytes = xml_text.encode("utf-8")
        attachment = self.db.execute(
            """INSERT INTO attachments
               (company_id,record_id,filename,mime_type,content,size,category,version,uploaded_by,
                created_at,sha256,license_confirmed)
               VALUES(?,?,?,'application/xml',?,?,?,'1',?,?,?,0)""",
            (session["company_id"], import_id, Path(str(data.get("filename") or f"nfe-{numero}.xml")).name,
             xml_bytes, len(xml_bytes), "XML NF-e", session["id"], now,
             hashlib.sha256(xml_bytes).hexdigest()))
        self.db.execute("UPDATE records SET payload=json_set(payload,'$.xml_attachment_id',?) WHERE id=?",
                        (attachment.lastrowid, import_id))

        supplier = self.db.connection().execute(
            """SELECT id FROM records WHERE company_id=? AND module='fornecedores'
               AND json_extract(payload,'$.documento')=? AND deleted_at IS NULL""",
            (session["company_id"], emit_cnpj)).fetchone() if emit_cnpj else None
        if not supplier:
            supplier_payload = {"assunto": subject_name, "documento": emit_cnpj, "razao_social": emit_name,
                                "tipo_pessoa": "Pessoa jurídica" if len(emit_cnpj) == 14 else "Pessoa física",
                                "avaliacao": "Pendente",
                                "relacionamentos": [{"record": f"importacoes_xml:{import_id}", "type": "Originado de"}]}
            supplier_cursor = self.db.execute(
                """INSERT INTO records(module,title,status,payload,created_by,created_at,updated_at,company_id)
                   VALUES('fornecedores',?,'Ativo',?,?,?,?,?)""",
                (emit_name, json_dumps(supplier_payload), session["id"], now, now, session["company_id"]))
            self.db.sync_relationships(supplier_cursor.lastrowid, supplier_payload, session["id"], session["company_id"])
        product_links = []
        created_products = 0
        for item in items:
            product = None
            if item.get("codigo"):
                product = self.db.connection().execute(
                    """SELECT id FROM records WHERE company_id=? AND module='produtos'
                       AND json_extract(payload,'$.codigo')=? AND deleted_at IS NULL""",
                    (session["company_id"], item["codigo"])).fetchone()
            if product:
                product_id = product["id"]
            else:
                product_payload = {
                    "assunto": subject_name, "codigo": item.get("codigo"),
                    "descricao": item.get("descricao"), "ncm": item.get("ncm"),
                    "cfop": item.get("cfop"), "unidade": item.get("unidade"),
                    "familia": "Importado de NF-e", "tipo_item": "Produto adquirido",
                    "preco_venda": 0,
                    "origem": "Importação XML NF-e",
                    "relacionamentos": [
                        {"record": f"importacoes_xml:{import_id}", "type": "Originado de"}
                    ]
                }
                product_cursor = self.db.execute(
                    """INSERT INTO records(module,title,status,payload,created_by,created_at,updated_at,company_id)
                       VALUES('produtos',?,'Ativo',?,?,?,?,?)""",
                    (item.get("descricao") or item.get("codigo") or "Produto da NF-e",
                     json_dumps(product_payload), session["id"], now, now, session["company_id"]))
                product_id = product_cursor.lastrowid
                self.db.sync_relationships(
                    product_id, product_payload, session["id"], session["company_id"])
                created_products += 1
            item["produto_id"] = product_id
            product_links.append({"record": f"produtos:{product_id}", "type": "Contém produto"})
        if product_links:
            payload["relacionamentos"] = product_links
            payload["itens"] = items
            self.db.execute("UPDATE records SET payload=?,updated_at=? WHERE id=?",
                            (json_dumps(payload), utc_now(), import_id))
            self.db.sync_relationships(
                import_id, payload, session["id"], session["company_id"])
        for parcel in parcels:
            try:
                parcel_value = -abs(float(parcel["valor"] or 0))
            except ValueError:
                parcel_value = None
            payable_payload = {
                "assunto": subject_name, "fornecedor": emit_name, "documento": numero,
                "parcela": parcel["numero"], "categoria": "Compras",
                "centro_custo": "A classificar", "origem": "Importação XML NF-e",
                "relacionamentos": [{"record": f"importacoes_xml:{import_id}", "type": "Originado de"}]
            }
            payable_cursor = self.db.execute(
                """INSERT INTO records
                   (module,title,status,amount,due_date,payload,created_by,created_at,updated_at,company_id)
                   VALUES('contas_pagar',?,'Em aberto',?,?,?,?,?,?,?,?)""",
                (f"NF-e {numero} — parcela {parcel['numero'] or len(parcels)}", parcel_value,
                 parcel["vencimento"] or None, json_dumps(payable_payload), session["id"], now, now,
                 session["company_id"]))
            self.db.sync_relationships(payable_cursor.lastrowid, payable_payload, session["id"], session["company_id"])
        self.db.audit(session["id"], "import", "nfe_xml", import_id,
                      {"chave": chave, "items": len(items), "parcels": len(parcels),
                       "created_products": created_products},
                      company_id=session["company_id"])
        self.db.finish_manual_transaction(commit=True)
        return self.send_json({"ok": True, "recordId": import_id, "items": len(items),
                               "parcels": len(parcels), "supplier": emit_name,
                               "createdProducts": created_products}, 201)

    @staticmethod
    def fiscal_master_key():
        raw = str(os.environ.get("SIVS_FISCAL_MASTER_KEY") or "").strip()
        if not raw:
            raise ValueError(
                "Configure SIVS_FISCAL_MASTER_KEY com uma chave Base64 de 32 bytes"
            )
        try:
            key = base64.b64decode(raw, validate=True)
        except (ValueError, binascii.Error):
            key = b""
        if len(key) != 32:
            raise ValueError(
                "SIVS_FISCAL_MASTER_KEY deve decodificar exatamente 32 bytes"
            )
        return key

    @staticmethod
    def validate_sefaz_url(value, label):
        text = str(value or "").strip()
        parsed = urlparse(text)
        hostname = str(parsed.hostname or "").lower().rstrip(".")
        if (parsed.scheme != "https" or not hostname.endswith(".gov.br") or
                parsed.username or parsed.password or parsed.fragment or
                (parsed.port not in (None, 443))):
            raise ValueError(f"{label} deve usar HTTPS em domínio oficial gov.br")
        if parsed.query and parsed.query.lower() != "wsdl":
            raise ValueError(f"{label} contém parâmetros não permitidos")
        clean_path = parsed.path or "/"
        return urllib.parse.urlunparse(("https", hostname, clean_path, "", "", ""))

    @staticmethod
    def fiscal_certificate_public(row):
        if not row:
            return None
        return {
            "id": row["id"], "branchId": row["branch_id"],
            "type": row["certificate_type"], "subject": row["subject_name"],
            "fingerprintSha256": row["fingerprint_sha256"],
            "serialNumber": row["serial_number"], "issuer": row["issuer_name"],
            "keyAlgorithm": row["key_algorithm"], "validFrom": row["valid_from"],
            "validTo": row["valid_to"], "status": row["status"],
            "lastUsedAt": row["last_used_at"], "createdAt": row["created_at"],
        }

    def fiscal_readiness(self, session):
        if not self.require_module_read(session, "fiscal"):
            return
        company_id = session["company_id"]
        db = self.db.connection()
        company = db.execute(
            """SELECT id,name,legal_name,cnpj,state_registration,municipal_registration,
                      uf,municipality_code,tax_regime,address
               FROM companies WHERE id=?""", (company_id,),
        ).fetchone()
        branch = db.execute(
            """SELECT id,code,name,cnpj,state_registration,municipal_registration,
                      uf,municipality_code,address
               FROM branches WHERE company_id=? AND active=1
               ORDER BY is_headquarters DESC,id LIMIT 1""", (company_id,),
        ).fetchone()
        configurations = db.execute(
            """SELECT id,branch_id,environment,uf,state_code,service_version,
                      status_service_url,source_url,source_verified_at,enabled,
                      last_status_code,last_status_reason,last_checked_at
               FROM sefaz_configurations WHERE company_id=?
               ORDER BY environment DESC,id DESC""", (company_id,),
        ).fetchall()
        certificate = db.execute(
            """SELECT * FROM fiscal_certificates
               WHERE company_id=? AND status='ACTIVE'
               ORDER BY id DESC LIMIT 1""", (company_id,),
        ).fetchone()
        schema = db.execute(
            """SELECT id,document_type,version,environment,schema_reference,valid_from,valid_to
               FROM fiscal_schema_versions
               WHERE active=1 AND upper(document_type) IN ('NFE','NF-E','55')
               ORDER BY id DESC LIMIT 1"""
        ).fetchone()
        rule_count = int(db.execute(
            "SELECT COUNT(*) FROM tax_rules WHERE company_id=? AND active=1",
            (company_id,),
        ).fetchone()[0])
        profile_count = int(db.execute(
            "SELECT COUNT(*) FROM company_fiscal_profiles WHERE company_id=? AND active=1",
            (company_id,),
        ).fetchone()[0])
        product_profile_count = int(db.execute(
            "SELECT COUNT(*) FROM product_fiscal_profiles WHERE company_id=? AND active=1",
            (company_id,),
        ).fetchone()[0])
        now = datetime.now(timezone.utc)
        certificate_valid = False
        if certificate and certificate["valid_to"]:
            try:
                certificate_valid = datetime.fromisoformat(
                    str(certificate["valid_to"]).replace("Z", "+00:00")
                ) > now
            except ValueError:
                certificate_valid = False
        master_key_configured = False
        try:
            self.fiscal_master_key()
            master_key_configured = True
        except ValueError:
            pass
        company_data = dict(company) if company else {}
        branch_data = dict(branch) if branch else {}
        effective_uf = str(branch_data.get("uf") or company_data.get("uf") or "").upper()
        effective_cnpj = str(branch_data.get("cnpj") or company_data.get("cnpj") or "")
        effective_ie = str(branch_data.get("state_registration") or
                           company_data.get("state_registration") or "").strip()
        effective_municipality = str(branch_data.get("municipality_code") or
                                     company_data.get("municipality_code") or "")
        homologation = next(
            (row for row in configurations if row["environment"] == "HOMOLOGATION" and row["enabled"]),
            None,
        )
        checks = [
            {"key": "cnpj", "label": "CNPJ válido", "ready": _valid_cnpj(effective_cnpj)},
            {"key": "stateRegistration", "label": "Inscrição estadual", "ready": bool(effective_ie)},
            {"key": "uf", "label": "UF e código autorizador", "ready": effective_uf in UF_CODES},
            {"key": "municipality", "label": "Código IBGE do município", "ready": bool(re.fullmatch(r"\d{7}", effective_municipality))},
            {"key": "taxRegime", "label": "Regime tributário", "ready": bool(company_data.get("tax_regime"))},
            {"key": "homologation", "label": "Endpoint de homologação habilitado", "ready": homologation is not None},
            {"key": "masterKey", "label": "Chave do cofre fiscal", "ready": master_key_configured},
            {"key": "certificate", "label": "Certificado A1 válido", "ready": certificate_valid},
            {"key": "schema", "label": "Schema NF-e oficial versionado", "ready": schema is not None},
            {"key": "taxRules", "label": "Perfil e regras fiscais revisados", "ready": bool(profile_count and rule_count)},
        ]
        can_status = bool(
            homologation and certificate_valid and master_key_configured and
            "check_sefaz_status" in self.allowed_operations(session, "fiscal")
        )
        return self.send_json({
            "ok": True,
            "company": company_data,
            "branch": branch_data,
            "configurations": [dict(row) for row in configurations],
            "certificate": self.fiscal_certificate_public(certificate),
            "schema": dict(schema) if schema else None,
            "counts": {"companyProfiles": profile_count, "taxRules": rule_count,
                       "productProfiles": product_profile_count},
            "checks": checks,
            "readyCount": sum(1 for item in checks if item["ready"]),
            "totalChecks": len(checks),
            "canCheckStatus": can_status,
            "canIssue": False,
            "issueBlockReason": (
                "Emissão permanece bloqueada até schemas oficiais, regras determinísticas e "
                "cenários fiscais da empresa serem homologados."
            ),
            "productionAllowed": os.environ.get("SIVS_ALLOW_SEFAZ_PRODUCTION") == "1",
            "officialReferences": {
                "services": SEFAZ_OFFICIAL_REFERENCE,
                "schemas": SEFAZ_SCHEMA_REFERENCE,
                "verifiedAt": "2026-08-18",
            },
        })

    def fiscal_configuration_update(self, session):
        if not self.require_operation(session, "fiscal", "manage_fiscal_config"):
            return
        try:
            data = self.parse_json(max_bytes=64 * 1024)
        except ValueError as exc:
            return self.error_json(str(exc))
        company_id = session["company_id"]
        try:
            branch_id = int(data.get("branchId") or 0)
        except (TypeError, ValueError):
            branch_id = 0
        branch = self.db.connection().execute(
            "SELECT id FROM branches WHERE id=? AND company_id=? AND active=1",
            (branch_id, company_id),
        ).fetchone()
        if not branch:
            return self.error_json("Unidade fiscal inválida")
        uf = str(data.get("uf") or "").strip().upper()
        if uf not in UF_CODES:
            return self.error_json("Selecione uma UF brasileira válida")
        municipality_code = re.sub(r"\D", "", str(data.get("municipalityCode") or ""))
        if not re.fullmatch(r"\d{7}", municipality_code):
            return self.error_json("Código IBGE do município deve possuir 7 dígitos")
        state_registration = str(data.get("stateRegistration") or "").strip()[:30]
        if not state_registration:
            return self.error_json("Informe a inscrição estadual")
        tax_regime = str(data.get("taxRegime") or "").strip().upper()
        if tax_regime not in {"SIMPLES_NACIONAL", "SIMPLES_EXCESSO", "REGIME_NORMAL"}:
            return self.error_json("Regime tributário inválido")
        environment = str(data.get("environment") or "HOMOLOGATION").strip().upper()
        if environment not in {"HOMOLOGATION", "PRODUCTION"}:
            return self.error_json("Ambiente fiscal inválido")
        enabled = bool(data.get("enabled", True))
        if (environment == "PRODUCTION" and enabled and
                os.environ.get("SIVS_ALLOW_SEFAZ_PRODUCTION") != "1"):
            return self.error_json(
                "Produção SEFAZ está bloqueada. Homologue primeiro e habilite "
                "SIVS_ALLOW_SEFAZ_PRODUCTION=1 conscientemente.",
                409, "sefaz_production_locked",
            )
        endpoints = data.get("endpoints") if isinstance(data.get("endpoints"), dict) else {}
        if uf == "GO" and data.get("useOfficialPreset", True):
            endpoints = SEFAZ_GO_ENDPOINTS[environment]
        try:
            normalized = {
                key: self.validate_sefaz_url(endpoints.get(key), f"Endpoint {key}")
                for key in ("status", "authorization", "authorization_return",
                            "protocol", "events", "invalidation")
            }
        except ValueError as exc:
            return self.error_json(str(exc))
        now = utc_now()
        cnpj = str(data.get("cnpj") or "").strip()
        if not _valid_cnpj(cnpj):
            return self.error_json("Informe o CNPJ válido da unidade fiscal")
        legal_name = str(data.get("legalName") or "").strip()[:200]
        if not legal_name:
            return self.error_json("Informe a razão social")
        with self.db.transaction(immediate=True):
            self.db.execute(
                """UPDATE companies SET legal_name=?,cnpj=?,state_registration=?,
                          municipal_registration=?,uf=?,municipality_code=?,tax_regime=?,updated_at=?
                   WHERE id=?""",
                (legal_name, cnpj, state_registration,
                 str(data.get("municipalRegistration") or "").strip()[:30] or None,
                 uf, municipality_code, tax_regime, now, company_id),
            )
            self.db.execute(
                """UPDATE branches SET cnpj=?,state_registration=?,municipal_registration=?,
                          uf=?,municipality_code=?,updated_at=?
                   WHERE id=? AND company_id=?""",
                (cnpj, state_registration,
                 str(data.get("municipalRegistration") or "").strip()[:30] or None,
                 uf, municipality_code, now, branch_id, company_id),
            )
            self.db.execute(
                """INSERT INTO sefaz_configurations
                   (company_id,branch_id,environment,uf,state_code,service_version,
                    status_service_url,authorization_service_url,authorization_return_url,
                    protocol_service_url,event_service_url,invalidation_service_url,
                    source_url,source_verified_at,enabled,created_by,updated_by,created_at,updated_at)
                   VALUES(?,?,?,?,?,'4.00',?,?,?,?,?,?,?,?,?,?,?,?,?)
                   ON CONFLICT(company_id,branch_id,environment) DO UPDATE SET
                     uf=excluded.uf,state_code=excluded.state_code,
                     service_version=excluded.service_version,
                     status_service_url=excluded.status_service_url,
                     authorization_service_url=excluded.authorization_service_url,
                     authorization_return_url=excluded.authorization_return_url,
                     protocol_service_url=excluded.protocol_service_url,
                     event_service_url=excluded.event_service_url,
                     invalidation_service_url=excluded.invalidation_service_url,
                     source_url=excluded.source_url,source_verified_at=excluded.source_verified_at,
                     enabled=excluded.enabled,updated_by=excluded.updated_by,updated_at=excluded.updated_at""",
                (company_id, branch_id, environment, uf, UF_CODES[uf],
                 normalized["status"], normalized["authorization"],
                 normalized["authorization_return"], normalized["protocol"],
                 normalized["events"], normalized["invalidation"],
                 SEFAZ_OFFICIAL_REFERENCE, "2026-08-18", 1 if enabled else 0,
                 session["id"], session["id"], now, now),
            )
            self.db.audit(
                session["id"], "configure", "sefaz", branch_id,
                {"environment": environment, "uf": uf, "enabled": enabled,
                 "endpoint_host": urlparse(normalized["status"]).hostname,
                 "source": SEFAZ_OFFICIAL_REFERENCE},
                company_id=company_id,
            )
        return self.fiscal_readiness(session)

    def fiscal_certificate_upload(self, session):
        if not self.require_operation(session, "fiscal", "manage_fiscal_certificate"):
            return
        try:
            data = self.parse_json(max_bytes=MAX_FISCAL_CERTIFICATE * 2)
            branch_id = int(data.get("branchId") or 0)
            raw = base64.b64decode(str(data.get("contentBase64") or ""), validate=True)
        except (ValueError, TypeError, binascii.Error) as exc:
            return self.error_json(f"Certificado A1 inválido: {exc}")
        if not raw or len(raw) > MAX_FISCAL_CERTIFICATE:
            return self.error_json("Certificado A1 deve possuir no máximo 2 MB")
        branch = self.db.connection().execute(
            "SELECT id FROM branches WHERE id=? AND company_id=? AND active=1",
            (branch_id, session["company_id"]),
        ).fetchone()
        if not branch:
            return self.error_json("Unidade fiscal inválida")
        password = str(data.get("password") or "")
        if len(password) > 512:
            return self.error_json("Senha do certificado inválida")
        try:
            from cryptography.hazmat.primitives import hashes, serialization
            from cryptography.hazmat.primitives.ciphers.aead import AESGCM
            from cryptography.hazmat.primitives.serialization import pkcs12
            private_key, certificate, chain = pkcs12.load_key_and_certificates(
                raw, password.encode("utf-8") if password else None,
            )
            if private_key is None or certificate is None:
                raise ValueError("o arquivo não contém chave privada e certificado")
            valid_from = getattr(certificate, "not_valid_before_utc", None)
            valid_to = getattr(certificate, "not_valid_after_utc", None)
            if valid_from is None:
                valid_from = certificate.not_valid_before.replace(tzinfo=timezone.utc)
                valid_to = certificate.not_valid_after.replace(tzinfo=timezone.utc)
            if valid_to <= datetime.now(timezone.utc):
                raise ValueError("o certificado está vencido")
            subject = certificate.subject.rfc4514_string()
            issuer = certificate.issuer.rfc4514_string()
            fingerprint = certificate.fingerprint(hashes.SHA256()).hex()
            bundle = {
                "privateKeyPem": base64.b64encode(private_key.private_bytes(
                    serialization.Encoding.PEM,
                    serialization.PrivateFormat.PKCS8,
                    serialization.NoEncryption(),
                )).decode("ascii"),
                "certificatePem": base64.b64encode(certificate.public_bytes(
                    serialization.Encoding.PEM,
                )).decode("ascii"),
                "chainPem": [base64.b64encode(item.public_bytes(
                    serialization.Encoding.PEM,
                )).decode("ascii") for item in (chain or ())],
            }
            plaintext = json_dumps(bundle).encode("utf-8")
            nonce = secrets.token_bytes(12)
            aad = f"SIVS-A1-1:{session['company_id']}:{branch_id}:{fingerprint}".encode("ascii")
            encrypted = b"SIVSA11" + nonce + AESGCM(self.fiscal_master_key()).encrypt(
                nonce, plaintext, aad,
            )
        except ImportError:
            return self.error_json(
                "O componente cryptography é necessário para o certificado A1",
                503, "crypto_unavailable",
            )
        except (ValueError, TypeError) as exc:
            return self.error_json(f"Não foi possível abrir o certificado A1: {exc}")
        now = utc_now()
        key_algorithm = type(private_key).__name__.replace("PrivateKey", "")
        with self.db.transaction(immediate=True):
            self.db.execute(
                "UPDATE fiscal_certificates SET status='INACTIVE',updated_at=? WHERE company_id=? AND branch_id=?",
                (now, session["company_id"], branch_id),
            )
            self.db.execute(
                """INSERT INTO fiscal_certificates
                   (company_id,branch_id,certificate_type,subject_name,fingerprint_sha256,
                    encrypted_content,valid_from,valid_to,status,created_by,created_at,updated_at,
                    serial_number,issuer_name,key_algorithm)
                   VALUES(?,?,'A1',?,?,?,?,?,'ACTIVE',?,?,?,?,?,?)
                   ON CONFLICT(company_id,fingerprint_sha256) DO UPDATE SET
                     branch_id=excluded.branch_id,certificate_type='A1',
                     subject_name=excluded.subject_name,encrypted_content=excluded.encrypted_content,
                     valid_from=excluded.valid_from,valid_to=excluded.valid_to,status='ACTIVE',
                     updated_at=excluded.updated_at,serial_number=excluded.serial_number,
                     issuer_name=excluded.issuer_name,key_algorithm=excluded.key_algorithm""",
                (session["company_id"], branch_id, subject, fingerprint, encrypted,
                 valid_from.isoformat(timespec="seconds"), valid_to.isoformat(timespec="seconds"),
                 session["id"], now, now, format(certificate.serial_number, "x"),
                 issuer, key_algorithm),
            )
            certificate_id = self.db.scalar(
                "SELECT id FROM fiscal_certificates WHERE company_id=? AND fingerprint_sha256=?",
                (session["company_id"], fingerprint),
            )
            self.db.audit(
                session["id"], "activate", "fiscal_certificate", certificate_id,
                {"branch_id": branch_id, "fingerprint_sha256": fingerprint,
                 "valid_to": valid_to.isoformat(timespec="seconds")},
                company_id=session["company_id"],
            )
        # Remove referências de senha e conteúdo o quanto antes; nenhum dos dois
        # é persistido, auditado ou devolvido ao navegador.
        password = ""
        raw = b""
        return self.send_json({
            "ok": True,
            "certificate": self.fiscal_certificate_public(self.db.connection().execute(
                "SELECT * FROM fiscal_certificates WHERE id=? AND company_id=?",
                (certificate_id, session["company_id"]),
            ).fetchone()),
        }, 201)

    def fiscal_certificate_delete(self, path, session):
        if not self.require_operation(session, "fiscal", "manage_fiscal_certificate"):
            return
        certificate_id = int(path.rsplit("/", 1)[-1])
        row = self.db.connection().execute(
            "SELECT id,branch_id,fingerprint_sha256 FROM fiscal_certificates WHERE id=? AND company_id=?",
            (certificate_id, session["company_id"]),
        ).fetchone()
        if not row:
            return self.error_json("Certificado não encontrado", 404)
        with self.db.transaction(immediate=True):
            self.db.execute(
                "DELETE FROM fiscal_certificates WHERE id=? AND company_id=?",
                (certificate_id, session["company_id"]),
            )
            self.db.audit(
                session["id"], "delete", "fiscal_certificate", certificate_id,
                {"branch_id": row["branch_id"],
                 "fingerprint_sha256": row["fingerprint_sha256"]},
                company_id=session["company_id"],
            )
        return self.send_json({"ok": True})

    def fiscal_certificate_bundle(self, certificate):
        from cryptography.hazmat.primitives.ciphers.aead import AESGCM
        encrypted = bytes(certificate["encrypted_content"] or b"")
        if not encrypted.startswith(b"SIVSA11") or len(encrypted) < 20:
            raise ValueError("formato criptográfico do certificado não reconhecido")
        nonce = encrypted[7:19]
        aad = (
            f"SIVS-A1-1:{certificate['company_id']}:{certificate['branch_id']}:"
            f"{certificate['fingerprint_sha256']}"
        ).encode("ascii")
        plaintext = AESGCM(self.fiscal_master_key()).decrypt(
            nonce, encrypted[19:], aad,
        )
        return json.loads(plaintext.decode("utf-8"))

    @staticmethod
    def sefaz_status_transport(endpoint, context, state_code, environment):
        soap_namespace = "http://www.w3.org/2003/05/soap-envelope"
        wsdl_namespace = "http://www.portalfiscal.inf.br/nfe/wsdl/NFeStatusServico4"
        nfe_namespace = "http://www.portalfiscal.inf.br/nfe"
        ET.register_namespace("soap12", soap_namespace)
        envelope = ET.Element(f"{{{soap_namespace}}}Envelope")
        body = ET.SubElement(envelope, f"{{{soap_namespace}}}Body")
        message = ET.SubElement(body, f"{{{wsdl_namespace}}}nfeDadosMsg")
        request = ET.SubElement(message, f"{{{nfe_namespace}}}consStatServ", {"versao": "4.00"})
        ET.SubElement(request, f"{{{nfe_namespace}}}tpAmb").text = "2" if environment == "HOMOLOGATION" else "1"
        ET.SubElement(request, f"{{{nfe_namespace}}}cUF").text = state_code
        ET.SubElement(request, f"{{{nfe_namespace}}}xServ").text = "STATUS"
        payload = ET.tostring(envelope, encoding="utf-8", xml_declaration=True)
        parsed = urlparse(endpoint)
        connection = http.client.HTTPSConnection(
            parsed.hostname, parsed.port or 443, context=context, timeout=20,
        )
        try:
            connection.request(
                "POST", parsed.path or "/", body=payload,
                headers={
                    "Content-Type": (
                        "application/soap+xml; charset=utf-8; "
                        'action="http://www.portalfiscal.inf.br/nfe/wsdl/'
                        'NFeStatusServico4/nfeStatusServicoNF"'
                    ),
                    "Accept": "application/soap+xml, application/xml",
                    "User-Agent": "SIVS-SECCOL/2.2",
                },
            )
            response = connection.getresponse()
            content = response.read(2 * 1024 * 1024 + 1)
            if len(content) > 2 * 1024 * 1024:
                raise ValueError("resposta da SEFAZ excedeu 2 MB")
            if response.status < 200 or response.status >= 300:
                raise ValueError(f"SEFAZ respondeu HTTP {response.status}")
        finally:
            connection.close()
        try:
            root = ET.fromstring(content)
        except ET.ParseError as exc:
            raise ValueError(f"resposta XML inválida da SEFAZ: {exc}") from None
        values = {}
        for key in ("tpAmb", "verAplic", "cStat", "xMotivo", "cUF", "dhRecbto", "tMed"):
            element = next((item for item in root.iter() if item.tag.rsplit("}", 1)[-1] == key), None)
            values[key] = str(element.text or "").strip() if element is not None else None
        if not values["cStat"] or not values["xMotivo"]:
            raise ValueError("resposta da SEFAZ sem status reconhecível")
        return values

    def fiscal_sefaz_status(self, session):
        if not self.require_operation(session, "fiscal", "check_sefaz_status"):
            return
        try:
            data = self.parse_json(max_bytes=8 * 1024)
        except ValueError as exc:
            return self.error_json(str(exc))
        environment = str(data.get("environment") or "HOMOLOGATION").upper()
        if environment not in {"HOMOLOGATION", "PRODUCTION"}:
            return self.error_json("Ambiente fiscal inválido")
        if (environment == "PRODUCTION" and
                os.environ.get("SIVS_ALLOW_SEFAZ_PRODUCTION") != "1"):
            return self.error_json(
                "Consulta em produção está bloqueada até a homologação ser concluída",
                409, "sefaz_production_locked",
            )
        try:
            branch_id = int(data.get("branchId") or 0)
        except (TypeError, ValueError):
            branch_id = 0
        config = self.db.connection().execute(
            """SELECT * FROM sefaz_configurations
               WHERE company_id=? AND branch_id=? AND environment=? AND enabled=1""",
            (session["company_id"], branch_id, environment),
        ).fetchone()
        if not config:
            return self.error_json("Configuração SEFAZ habilitada não encontrada", 409, "sefaz_not_configured")
        certificate = self.db.connection().execute(
            """SELECT * FROM fiscal_certificates
               WHERE company_id=? AND branch_id=? AND status='ACTIVE'
               ORDER BY id DESC LIMIT 1""",
            (session["company_id"], branch_id),
        ).fetchone()
        if not certificate:
            return self.error_json("Certificado A1 ativo não encontrado", 409, "fiscal_certificate_missing")
        try:
            certificate_expires_at = datetime.fromisoformat(
                str(certificate["valid_to"]).replace("Z", "+00:00")
            )
            if certificate_expires_at.tzinfo is None:
                certificate_expires_at = certificate_expires_at.replace(tzinfo=timezone.utc)
        except (TypeError, ValueError):
            return self.error_json(
                "A validade do certificado A1 não pôde ser verificada",
                409, "fiscal_certificate_invalid_validity",
            )
        if certificate_expires_at <= datetime.now(timezone.utc):
            return self.error_json(
                "O certificado A1 está vencido",
                409, "fiscal_certificate_expired",
            )
        try:
            bundle = self.fiscal_certificate_bundle(certificate)
            with tempfile.TemporaryDirectory(prefix="sivs-sefaz-") as directory:
                folder = Path(directory)
                key_path = folder / "client-key.pem"
                certificate_path = folder / "client-chain.pem"
                key_path.write_bytes(base64.b64decode(bundle["privateKeyPem"], validate=True))
                certificate_path.write_bytes(
                    base64.b64decode(bundle["certificatePem"], validate=True) +
                    b"".join(base64.b64decode(item, validate=True)
                             for item in bundle.get("chainPem", []))
                )
                with contextlib.suppress(OSError):
                    os.chmod(key_path, 0o600)
                    os.chmod(certificate_path, 0o600)
                context = ssl.create_default_context()
                context.minimum_version = ssl.TLSVersion.TLSv1_2
                context.load_cert_chain(certificate_path, key_path)
                result = self.sefaz_status_transport(
                    config["status_service_url"], context,
                    config["state_code"], environment,
                )
        except ImportError:
            return self.error_json("Componente de criptografia indisponível", 503, "crypto_unavailable")
        except (ValueError, OSError, ssl.SSLError, binascii.Error, KeyError) as exc:
            self.db.system_event(
                "warning", "integration", "sefaz_status_failed",
                "Falha na consulta de status da SEFAZ",
                company_id=session["company_id"], user_id=session["id"],
                detail={"environment": environment, "uf": config["uf"],
                        "reason": str(exc)[:500]},
            )
            return self.error_json(
                f"Não foi possível consultar a SEFAZ: {exc}", 502, "sefaz_unavailable",
            )
        now = utc_now()
        with self.db.transaction(immediate=True):
            self.db.execute(
                """UPDATE sefaz_configurations
                   SET last_status_code=?,last_status_reason=?,last_checked_at=?,updated_at=?
                   WHERE id=? AND company_id=?""",
                (result["cStat"], result["xMotivo"], now, now,
                 config["id"], session["company_id"]),
            )
            self.db.execute(
                "UPDATE fiscal_certificates SET last_used_at=?,updated_at=? WHERE id=? AND company_id=?",
                (now, now, certificate["id"], session["company_id"]),
            )
            self.db.audit(
                session["id"], "status", "sefaz", config["id"],
                {"environment": environment, "uf": config["uf"],
                 "status_code": result["cStat"], "reason": result["xMotivo"]},
                company_id=session["company_id"],
            )
        return self.send_json({
            "ok": True, "operational": result["cStat"] == "107",
            "environment": environment, "uf": config["uf"],
            "statusCode": result["cStat"], "reason": result["xMotivo"],
            "applicationVersion": result["verAplic"], "receivedAt": result["dhRecbto"],
            "averageTime": result["tMed"], "checkedAt": now,
        })

    @staticmethod
    def accounting_csv(headers, rows):
        output = io.StringIO(newline="")
        writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
        writer.writerow(headers)
        for row in rows:
            safe = []
            for value in row:
                if value is None:
                    safe.append("")
                    continue
                text = str(value)
                if text.startswith(("=", "+", "-", "@")) and not re.fullmatch(r"-?\d+(?:[.,]\d+)?", text):
                    text = "'" + text
                safe.append(text)
            writer.writerow(safe)
        return ("\ufeff" + output.getvalue()).encode("utf-8")

    def accounting_export(self, query, session):
        if (not self.require_module_export(session, "fiscal") or
                not self.require_operation(session, "fiscal", "export_accounting") or
                not self.require_operation(session, "fiscal", "view_values")):
            return
        period = str((query.get("period") or [""])[0]).strip()
        match = re.fullmatch(r"(20\d{2})-(0[1-9]|1[0-2])", period)
        if not match:
            return self.error_json("Período deve usar o formato AAAA-MM")
        year, month = int(match.group(1)), int(match.group(2))
        start = datetime(year, month, 1, tzinfo=timezone.utc)
        end = datetime(year + (month == 12), 1 if month == 12 else month + 1, 1,
                       tzinfo=timezone.utc)
        start_iso, end_iso = start.isoformat(), end.isoformat()
        start_date, end_date = start.date().isoformat(), end.date().isoformat()
        company_id = session["company_id"]
        db = self.db.connection()
        company = db.execute(
            """SELECT id,name,legal_name,cnpj,state_registration,municipal_registration,
                      uf,municipality_code,tax_regime,address,email,phone
               FROM companies WHERE id=?""", (company_id,),
        ).fetchone()
        records = db.execute(
            """SELECT id,module,title,status,amount,due_date,payload,created_at,updated_at,revision
               FROM records WHERE company_id=? AND deleted_at IS NULL
                 AND module IN ('fiscal','importacoes_xml','vendas','pedidos_compra',
                                'contas_pagar','contas_receber','financeiro','caixa')
                 AND ((created_at>=? AND created_at<?) OR (updated_at>=? AND updated_at<?)
                      OR (due_date>=? AND due_date<?)
                      OR (json_extract(payload,'$.data_emissao')>=?
                          AND json_extract(payload,'$.data_emissao')<?)
                      OR (json_extract(payload,'$.data_pagamento')>=?
                          AND json_extract(payload,'$.data_pagamento')<?)
                      OR (json_extract(payload,'$.data_recebimento')>=?
                          AND json_extract(payload,'$.data_recebimento')<?))
               ORDER BY module,id""",
            (company_id, start_iso, end_iso, start_iso, end_iso, start_date, end_date,
             start_date, end_date, start_date, end_date, start_date, end_date),
        ).fetchall()
        movements = db.execute(
            """SELECT m.id,m.movement_type,m.quantity_micros,m.unit_cost_cents,
                      m.value_delta_cents,m.balance_value_cents,m.lot_key,m.origin_type,
                      m.origin_id,m.reference,m.reason,m.created_at,
                      w.code warehouse_code,w.name warehouse_name,
                      r.id product_id,r.title product_name,json_extract(r.payload,'$.codigo') product_code
               FROM inventory_movements m
               JOIN warehouses w ON w.id=m.warehouse_id
               JOIN records r ON r.id=m.product_record_id
               WHERE m.company_id=? AND m.created_at>=? AND m.created_at<? ORDER BY m.id""",
            (company_id, start_iso, end_iso),
        ).fetchall()
        items = db.execute(
            """SELECT di.*,r.module record_module,r.title record_title
               FROM document_items di JOIN records r ON r.id=di.record_id
               WHERE di.company_id=? AND r.updated_at>=? AND r.updated_at<?
               ORDER BY di.record_id,di.sort_order,di.id""",
            (company_id, start_iso, end_iso),
        ).fetchall()
        files = {}
        files["cadastros/empresa.json"] = json.dumps(
            dict(company) if company else {}, ensure_ascii=False, indent=2,
        ).encode("utf-8")
        files["cadastros/unidades.json"] = json.dumps([
            dict(row) for row in db.execute(
                """SELECT id,code,name,cnpj,state_registration,municipal_registration,
                          uf,municipality_code,address,active,is_headquarters
                   FROM branches WHERE company_id=? ORDER BY is_headquarters DESC,id""",
                (company_id,),
            ).fetchall()
        ], ensure_ascii=False, indent=2).encode("utf-8")
        record_payloads = {
            row["id"]: json.loads(row["payload"] or "{}") for row in records
        }
        files["lancamentos/registros.csv"] = self.accounting_csv(
            ["id", "modulo", "titulo", "situacao", "valor_centavos", "vencimento",
             "criado_em", "atualizado_em", "revisao", "dados_json"],
            [(row["id"], row["module"], row["title"], row["status"],
              self.record_amount_cents(row["amount"]), row["due_date"], row["created_at"],
              row["updated_at"], row["revision"], row["payload"]) for row in records],
        )
        financial_modules = {"contas_pagar", "contas_receber", "financeiro", "caixa"}
        files["financeiro/lancamentos.csv"] = self.accounting_csv(
            ["modulo", "id", "titulo", "situacao", "valor_centavos", "vencimento",
             "tipo_lancamento", "documento", "parcela", "categoria", "centro_custo",
             "conta", "parte", "data_pagamento", "data_recebimento", "atualizado_em"],
            [(row["module"], row["id"], row["title"], row["status"],
              self.record_amount_cents(row["amount"]), row["due_date"],
              record_payloads[row["id"]].get("tipo_lancamento") or
              record_payloads[row["id"]].get("tipo_movimento"),
              record_payloads[row["id"]].get("documento"),
              record_payloads[row["id"]].get("parcela"),
              record_payloads[row["id"]].get("categoria"),
              record_payloads[row["id"]].get("centro_custo"),
              record_payloads[row["id"]].get("conta"),
              record_payloads[row["id"]].get("cliente") or
              record_payloads[row["id"]].get("fornecedor"),
              record_payloads[row["id"]].get("data_pagamento"),
              record_payloads[row["id"]].get("data_recebimento"), row["updated_at"])
             for row in records if row["module"] in financial_modules],
        )
        fiscal_modules = {"fiscal", "importacoes_xml", "vendas", "pedidos_compra"}
        files["fiscal/documentos.csv"] = self.accounting_csv(
            ["modulo", "id", "titulo", "tipo_documento", "numero", "serie", "chave",
             "parte", "cfop", "finalidade", "valor_centavos", "data_emissao",
             "situacao", "atualizado_em"],
            [(row["module"], row["id"], row["title"],
              record_payloads[row["id"]].get("tipo_nota") or row["module"],
              record_payloads[row["id"]].get("numero") or
              record_payloads[row["id"]].get("documento"),
              record_payloads[row["id"]].get("serie"),
              record_payloads[row["id"]].get("chave"),
              record_payloads[row["id"]].get("destinatario") or
              record_payloads[row["id"]].get("fornecedor") or
              record_payloads[row["id"]].get("cliente"),
              record_payloads[row["id"]].get("cfop"),
              record_payloads[row["id"]].get("finalidade"),
              self.record_amount_cents(row["amount"]),
              record_payloads[row["id"]].get("data_emissao"),
              row["status"], row["updated_at"])
             for row in records if row["module"] in fiscal_modules],
        )
        files["lancamentos/itens_documentos.csv"] = self.accounting_csv(
            ["id", "documento_id", "modulo", "documento", "linha", "tipo_item",
             "cadastro_item_id", "descricao", "quantidade_micros",
             "valor_unitario_centavos", "desconto_centavos", "total_centavos",
             "deposito_id", "lote", "reserva_id", "observacoes"],
            [(row["id"], row["record_id"], row["record_module"], row["record_title"],
              row["sort_order"], row["item_kind"], row["catalog_record_id"],
              row["description"], row["quantity_micros"], row["unit_price_cents"],
              row["discount_cents"], row["total_cents"], row["warehouse_id"],
              row["lot_key"], row["reservation_id"], row["notes"]) for row in items],
        )
        files["estoque/movimentos.csv"] = self.accounting_csv(
            ["id", "data", "tipo", "produto_id", "codigo_produto", "produto",
             "deposito_codigo", "deposito", "lote", "quantidade_micros",
             "custo_unitario_centavos", "variacao_valor_centavos", "saldo_valor_centavos",
             "tipo_origem", "id_origem", "referencia", "justificativa"],
            [(row["id"], row["created_at"], row["movement_type"], row["product_id"],
              row["product_code"], row["product_name"], row["warehouse_code"],
              row["warehouse_name"], row["lot_key"], row["quantity_micros"],
              row["unit_cost_cents"], row["value_delta_cents"], row["balance_value_cents"],
              row["origin_type"], row["origin_id"], row["reference"], row["reason"])
             for row in movements],
        )
        xml_count = 0
        for row in db.execute(
            """SELECT x.id,x.document_role,x.content FROM xml_documents x
               WHERE x.company_id=? AND x.created_at>=? AND x.created_at<? ORDER BY x.id""",
            (company_id, start_iso, end_iso),
        ).fetchall():
            files[f"xml/fiscal-{row['id']}-{re.sub(r'[^A-Za-z0-9_-]', '_', row['document_role'])}.xml"] = bytes(row["content"])
            xml_count += 1
        for row in db.execute(
            """SELECT a.id,a.filename,a.content FROM attachments a
               JOIN records r ON r.id=a.record_id
               WHERE a.company_id=? AND r.module='importacoes_xml'
                 AND a.created_at>=? AND a.created_at<?
                 AND (lower(a.filename) LIKE '%.xml' OR a.mime_type IN ('application/xml','text/xml'))
               ORDER BY a.id""",
            (company_id, start_iso, end_iso),
        ).fetchall():
            filename = re.sub(r"[^A-Za-z0-9._-]", "_", Path(row["filename"]).name)
            files[f"xml/entrada-{row['id']}-{filename}"] = bytes(row["content"])
            xml_count += 1
        files["LEIA-ME.txt"] = (
            "PACOTE CONTÁBIL SIVS\r\n\r\n"
            f"Empresa: {company['legal_name'] or company['name'] if company else 'Não informada'}\r\n"
            f"CNPJ: {company['cnpj'] if company else 'Não informado'}\r\n"
            f"Competência: {period}\r\n\r\n"
            "Use manifest.json para conferir contagens e SHA-256 de cada arquivo. "
            "Os valores estão em centavos e as quantidades de estoque em micros.\r\n"
            "Este pacote apoia a escrituração e não substitui SPED, livros fiscais, "
            "conciliação ou validação do contador.\r\n"
        ).encode("utf-8-sig")
        file_manifest = [
            {"path": name, "bytes": len(content),
             "sha256": hashlib.sha256(content).hexdigest()}
            for name, content in sorted(files.items())
        ]
        totals = {
            "records": len(records), "documentItems": len(items),
            "inventoryMovements": len(movements), "xmlDocuments": xml_count,
            "recordAmountCents": sum(self.record_amount_cents(row["amount"]) for row in records),
            "inventoryValueDeltaCents": sum(int(row["value_delta_cents"] or 0) for row in movements),
        }
        manifest = {
            "format": "SIVS-ACCOUNTING-1", "version": VERSION,
            "generatedAt": utc_now(), "period": period,
            "company": {"id": company_id, "name": company["name"] if company else None,
                        "cnpj": company["cnpj"] if company else None},
            "basis": (
                "Registros criados, atualizados, vencidos, emitidos ou liquidados no período; "
                "movimentos e XML ocorridos no mesmo intervalo"
            ),
            "totals": totals, "files": file_manifest,
            "notice": "Pacote de apoio à escrituração; não substitui SPED, livros ou validação do contador.",
        }
        files["manifest.json"] = json.dumps(
            manifest, ensure_ascii=False, indent=2, allow_nan=False,
        ).encode("utf-8")
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED, compresslevel=9) as archive:
            for name, content in sorted(files.items()):
                archive.writestr(name, content)
        body = buffer.getvalue()
        checksum = hashlib.sha256(body).hexdigest()
        now = utc_now()
        with self.db.transaction(immediate=True):
            export_id = self.db.execute(
                """INSERT INTO accounting_exports
                   (company_id,period,format_version,sha256,file_size,totals_json,generated_by,created_at)
                   VALUES(?,?,'SIVS-ACCOUNTING-1',?,?,?,?,?)""",
                (company_id, period, checksum, len(body), json_dumps(totals), session["id"], now),
            ).lastrowid
            self.db.audit(
                session["id"], "export", "accounting", export_id,
                {"period": period, "sha256": checksum, **totals},
                company_id=company_id,
            )
        filename = f"sivs-contabilidade-{period}-{re.sub(r'\D', '', str(company['cnpj'] or 'empresa'))}.zip"
        self._response_started = True
        self.send_response(200)
        self.send_header("Content-Type", "application/zip")
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(body)))
        self.send_header("X-Content-SHA256", checksum)
        self.send_header("X-SIVS-Format", "SIVS-ACCOUNTING-1")
        self.security_headers()
        self.end_headers()
        self.wfile.write(body)

    def fiscal_action(self, path, session):
        parts = path.split("/")
        if len(parts) != 5 or not parts[3].isdigit():
            return self.error_json("Documento fiscal inválido", 404)
        record_id, action = int(parts[3]), parts[4]
        known_actions = {"registrar", "cce", "cancelar", "inutilizar", "reenviar", "email"}
        if action not in known_actions:
            return self.error_json("Ação fiscal inválida")
        record = self.db.connection().execute(
            "SELECT id FROM records WHERE id=? AND company_id=? AND module='fiscal' AND deleted_at IS NULL",
            (record_id, session["company_id"])).fetchone()
        if not record:
            return self.error_json("Documento fiscal não encontrado", 404)
        if action == "registrar" and not self.require_operation(
                session, "fiscal", "register_fiscal"):
            return
        try:
            data = self.parse_json()
        except ValueError as exc:
            return self.error_json(str(exc))
        # Emissão, eventos e transmissão fiscal ainda não foram implementados.
        # Não simulamos SEFAZ nem dependemos de um ERP/conector externo.
        if action != "registrar":
            return self.error_json(
                "A emissão fiscal própria ainda não está habilitada. Esta ação será implementada "
                "somente com os schemas e manuais oficiais vigentes da SEFAZ.",
                501, "fiscal_engine_not_implemented")
        status = "Registrado localmente"
        cursor = self.db.execute(
            """INSERT INTO fiscal_events
               (company_id,record_id,event_type,status,protocol,response_detail,created_by,created_at)
               VALUES(?,?,?,?,?,?,?,?)""",
            (session["company_id"], record_id, action, status,
             str(data.get("protocol") or "")[:120] or None,
             str(data.get("detail") or "")[:4000] or None, session["id"], utc_now()))
        self.db.audit(session["id"], "event", "fiscal", record_id,
                      {"event": action, "event_id": cursor.lastrowid}, company_id=session["company_id"])
        return self.send_json({"ok": True, "eventId": cursor.lastrowid, "status": status}, 201)

    def subject_action(self, path, session):
        parts = path.split("/")
        if len(parts) != 5 or not parts[3].isdigit():
            return self.error_json("Assunto inválido", 404)
        subject_id, action = int(parts[3]), parts[4]
        subject = self.db.connection().execute(
            "SELECT * FROM subjects WHERE id=? AND company_id=?", (subject_id, session["company_id"])).fetchone()
        if not subject:
            return self.error_json("Assunto não encontrado", 404)
        try:
            data = self.parse_json()
        except ValueError as exc:
            return self.error_json(str(exc))
        if action == "rename":
            name = str(data.get("name") or "").strip()[:180]
            if not name:
                return self.error_json("Informe o novo nome")
            key = f'{session["company_id"]}:{self.db.normalize_subject(name)}'
        elif action == "merge":
            try:
                target_id = int(data.get("target_id"))
            except (ValueError, TypeError):
                return self.error_json("Informe o assunto de destino")
            if target_id == subject_id:
                return self.error_json("Selecione outro assunto para unificar")
            target = self.db.connection().execute(
                "SELECT id FROM subjects WHERE id=? AND company_id=?", (target_id, session["company_id"])).fetchone()
            if not target:
                return self.error_json("Assunto de destino não encontrado")
        elif action != "archive":
            return self.error_json("Ação de assunto inválida")
        try:
            with self.db.transaction(immediate=True):
                if action == "archive":
                    status = "Arquivado" if data.get("archived", True) else "Ativo"
                    self.db.execute(
                        "UPDATE subjects SET status=?,updated_at=? WHERE id=?",
                        (status, utc_now(), subject_id),
                    )
                elif action == "rename":
                    self.db.execute(
                        "UPDATE subjects SET name=?,normalized_name=?,updated_at=? WHERE id=?",
                        (name, key, utc_now(), subject_id),
                    )
                else:
                    db = self.db.connection()
                    rows = db.execute(
                        "SELECT * FROM record_subjects WHERE subject_id=?", (subject_id,)
                    ).fetchall()
                    for row in rows:
                        db.execute(
                            """INSERT OR IGNORE INTO record_subjects
                               (record_id,subject_id,relationship_type,is_primary,created_by,created_at)
                               VALUES(?,?,?,?,?,?)""",
                            (row["record_id"], target_id, row["relationship_type"], row["is_primary"],
                             session["id"], utc_now()),
                        )
                    db.execute(
                        "UPDATE records SET subject_id=? WHERE subject_id=? AND company_id=?",
                        (target_id, subject_id, session["company_id"]),
                    )
                    db.execute("DELETE FROM record_subjects WHERE subject_id=?", (subject_id,))
                    db.execute(
                        "UPDATE subjects SET status='Unificado',updated_at=? WHERE id=?",
                        (utc_now(), subject_id),
                    )
                self.db.audit(
                    session["id"], action, "subject", subject_id, data,
                    company_id=session["company_id"],
                )
        except sqlite3.IntegrityError:
            return self.error_json("Já existe um assunto com este nome", 409, "duplicate_subject")
        return self.send_json({"ok": True})

    def settings_update(self, session):
        try:
            data = self.parse_json()
        except ValueError as exc:
            return self.error_json(str(exc))
        now = utc_now()
        company = data.get("company")
        if isinstance(company, dict):
            name = str(company.get("name") or "").strip()
            if not name:
                return self.error_json("Informe o nome da empresa")
            cnpj = str(company.get("cnpj") or "").strip() or None
            email = str(company.get("email") or "").strip().lower() or None
            if cnpj and not _valid_cnpj(cnpj):
                return self.error_json("CNPJ da empresa inválido")
            if email and not re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", email):
                return self.error_json("E-mail da empresa inválido")
        with self.db.transaction(immediate=True):
            if isinstance(company, dict):
                self.db.execute(
                    """UPDATE companies SET name=?,cnpj=?,phone=?,email=?,address=?,updated_at=?
                       WHERE id=?""",
                    (name, cnpj, str(company.get("phone") or "").strip() or None, email,
                     str(company.get("address") or "").strip() or None,
                     now, session["company_id"]),
                )
            for key, value in data.items():
                if key not in {"preferences", "email", "banking", "certweb"}:
                    continue
                self.db.execute(
                    """INSERT OR REPLACE INTO company_settings(company_id,key,value,updated_at)
                       VALUES(?,?,?,?)""", (session["company_id"], key, json_dumps(value), now))
            self.db.audit(session["id"], "update", "settings", company_id=session["company_id"])
        return self.send_json({"ok": True})

    def export_data(self, query, session):
        module = (query.get("module") or [""])[0]
        if module and module not in MODULES:
            return self.error_json("Módulo inválido")
        if module:
            if not self.require_module_export(session, module):
                return
            if (module in VALUE_SENSITIVE_MODULES
                    and not self.require_operation(session, module, "view_values")):
                return
        elif not self.capabilities(session)["full_backup"]:
            return self.error_json(
                "A exportação consolidada exige perfil de administrador", 403, "forbidden"
            )
        sql = "SELECT * FROM records WHERE company_id=? AND deleted_at IS NULL"
        params = [session["company_id"]]
        if module == PARTY_MODULE:
            physical = [item for item in PARTY_PHYSICAL_MODULES
                         if item in self.allowed_modules(session, "export")]
            if not physical:
                return self.error_json("Seu perfil não possui permissão para exportar clientes ou fornecedores", 403, "forbidden")
            placeholders = ",".join("?" for _ in physical)
            sql += f" AND module IN ({placeholders})"
            params.extend(physical)
        elif module:
            sql += " AND module=?"
            params.append(module)
        sql += " ORDER BY module,id"
        rows = self.db.connection().execute(sql, params).fetchall()
        payload = {"format": "SIVS-3", "version": VERSION, "exported_at": utc_now(),
                   "scope": module or "business-data", "records": [self.record_json(r) for r in rows]}
        if not module:
            company = self.db.connection().execute("SELECT * FROM companies WHERE id=?",
                                                   (session["company_id"],)).fetchone()
            payload["company"] = dict(company) if company else {}
            settings = self.db.connection().execute(
                "SELECT key,value,updated_at FROM company_settings WHERE company_id=?",
                (session["company_id"],)).fetchall()
            payload["settings"] = [{"key": row["key"], "value": json.loads(row["value"]),
                                    "updated_at": row["updated_at"]} for row in settings]
            payload["tender_results"] = [dict(row) for row in self.db.connection().execute(
                "SELECT * FROM tender_results WHERE company_id=? ORDER BY id",
                (session["company_id"],)).fetchall()]
            payload["tender_searches"] = [dict(row) for row in self.db.connection().execute(
                "SELECT * FROM tender_searches WHERE company_id=? ORDER BY id",
                (session["company_id"],)).fetchall()]
            payload["attachments"] = [
                {**{key: row[key] for key in row.keys() if key != "content"},
                 "content": base64.b64encode(row["content"]).decode("ascii")}
                for row in self.db.connection().execute(
                    "SELECT * FROM attachments WHERE company_id=? ORDER BY id",
                    (session["company_id"],)).fetchall()
            ]
            payload["approvals"] = [dict(row) for row in self.db.connection().execute(
                "SELECT * FROM approvals WHERE company_id=? ORDER BY id",
                (session["company_id"],)).fetchall()]
            payload["fiscal_events"] = [dict(row) for row in self.db.connection().execute(
                "SELECT * FROM fiscal_events WHERE company_id=? ORDER BY id",
                (session["company_id"],)).fetchall()]
            payload["search_schedules"] = [dict(row) for row in self.db.connection().execute(
                "SELECT * FROM search_schedules WHERE company_id=? ORDER BY id",
                (session["company_id"],)).fetchall()]
        self.db.audit(
            session["id"], "export", module or "business-data",
            detail={"records": len(rows)}, company_id=session["company_id"],
        )
        body = json.dumps(
            payload, ensure_ascii=False, indent=2, allow_nan=False
        ).encode("utf-8")
        self._response_started = True
        self.send_response(200)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header(
            "Content-Disposition",
            f'attachment; filename="sivs-dados-{module or "consolidados"}-{datetime.now():%Y%m%d}.json"',
        )
        self.send_header("Content-Length", str(len(body)))
        self.security_headers()
        self.end_headers()
        self.wfile.write(body)

    def database_backup(self, session):
        """Gera cópia SQLite íntegra e a cifra com AES-256-GCM."""
        try:
            data = self.parse_json(max_bytes=8 * 1024)
        except ValueError as exc:
            return self.error_json(str(exc))
        passphrase = str(data.get("passphrase") or "")
        if len(passphrase) < 12 or len(passphrase) > 256:
            return self.error_json(
                "A senha do backup deve possuir entre 12 e 256 caracteres"
            )
        inaccessible = self.db.scalar(
            """SELECT COUNT(*) FROM companies c WHERE c.active=1 AND NOT EXISTS (
                 SELECT 1 FROM company_memberships cm
                 WHERE cm.company_id=c.id AND cm.user_id=? AND cm.active=1 AND cm.role='admin'
               )""",
            (session["id"],),
        )
        if inaccessible:
            return self.error_json(
                "O backup integral contém todas as empresas e exige administração em cada uma delas.",
                403, "forbidden",
            )
        try:
            from cryptography.hazmat.primitives import hashes
            from cryptography.hazmat.primitives.ciphers.aead import AESGCM
            from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
        except ImportError:
            return self.error_json(
                "O componente de criptografia não está instalado. Execute: pip install cryptography",
                503, "crypto_unavailable",
            )

        self.db.audit(
            session["id"], "backup", "database",
            detail={"format": "SIVS-BACKUP-2", "encrypted": True},
            company_id=session["company_id"],
        )
        temporary = tempfile.NamedTemporaryFile(prefix="sivs-backup-", suffix=".sqlite3", delete=False)
        temporary_path = Path(temporary.name)
        temporary.close()
        try:
            destination = sqlite3.connect(temporary_path)
            try:
                self.db.connection().backup(destination)
                # Sessões são deliberadamente invalidadas no artefato para impedir replay após restauração.
                destination.execute("DELETE FROM sessions")
                destination.commit()
                integrity = destination.execute("PRAGMA integrity_check").fetchone()[0]
            finally:
                destination.close()
            if integrity != "ok":
                raise sqlite3.DatabaseError(f"verificação de integridade: {integrity}")
            plaintext = temporary_path.read_bytes()
        except (OSError, sqlite3.Error) as exc:
            return self.error_json(f"Não foi possível gerar o backup íntegro: {exc}", 500, "backup_failed")
        finally:
            try:
                temporary_path.unlink(missing_ok=True)
            except OSError:
                pass

        magic = b"SIVSBKP2"
        salt = secrets.token_bytes(16)
        nonce = secrets.token_bytes(12)
        iterations = 600_000
        key = PBKDF2HMAC(
            algorithm=hashes.SHA256(), length=32, salt=salt, iterations=iterations
        ).derive(passphrase.encode("utf-8"))
        header = magic + iterations.to_bytes(4, "big") + salt + nonce
        encrypted = header + AESGCM(key).encrypt(nonce, plaintext, header)
        checksum = hashlib.sha256(encrypted).hexdigest()
        filename = f"sivs-backup-{datetime.now():%Y%m%d-%H%M%S}.sivsbackup"
        self._response_started = True
        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.sivs.backup")
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(encrypted)))
        self.send_header("X-SIVS-Format", "SIVS-BACKUP-2")
        self.send_header("X-Content-SHA256", checksum)
        self.security_headers()
        self.end_headers()
        self.wfile.write(encrypted)

    def import_data(self, session):
        try:
            data = self.parse_json(max_bytes=MAX_IMPORT_BODY)
        except ValueError as exc:
            return self.error_json(str(exc))
        records = data.get("records", [])
        if not isinstance(records, list) or len(records) > 10_000:
            return self.error_json("Arquivo de importação inválido ou excessivamente grande")
        count = 0
        now = utc_now()
        transaction_context = self.db.transaction(immediate=True)
        try:
            db = transaction_context.__enter__()
            id_map = {}
            staged = []
            for record in records:
                values = self.normalized_record(record)
                if values[0] == "estoque":
                    raise ValueError(
                        "Registros genéricos de estoque não podem ser importados; use o ledger dedicado"
                    )
                source_key = (record.get("payload") or {}).get("source_key") if isinstance(record, dict) else None
                existing = db.execute(
                    """SELECT id FROM records WHERE company_id=? AND module='fontes'
                       AND json_extract(payload,'$.source_key')=?""",
                    (session["company_id"], source_key)).fetchone() if values[0] == "fontes" and source_key else None
                if existing:
                    record_id = existing["id"]
                else:
                    cursor = db.execute(
                        """INSERT INTO records
                           (module,title,status,amount,due_date,payload,created_by,created_at,updated_at,company_id)
                           VALUES(?,?,?,?,?,?,?,?,?,?)""",
                        (*values, session["id"], now, now, session["company_id"]))
                    record_id = cursor.lastrowid
                    count += 1
                if isinstance(record, dict) and record.get("id") is not None:
                    id_map[int(record["id"])] = record_id
                staged.append((record_id, json.loads(values[5])))
            for record_id, payload in staged:
                rewritten = []
                relations = payload.get("relacionamentos") if isinstance(payload.get("relacionamentos"), list) else []
                primary = payload.get("registro_relacionado")
                if primary:
                    relations = relations + [{"record": primary, "type": payload.get("tipo_relacao") or "Relacionado a"}]
                for relation in relations:
                    if not isinstance(relation, dict):
                        continue
                    reference = str(relation.get("record") or relation.get("registro") or "")
                    try:
                        old_target = int(reference.rsplit(":", 1)[-1])
                    except ValueError:
                        continue
                    new_target = id_map.get(old_target)
                    if new_target:
                        target = db.execute(
                            "SELECT module FROM records WHERE id=? AND company_id=?",
                            (new_target, session["company_id"]),
                        ).fetchone()
                        if target:
                            rewritten.append({
                                "record": f'{target["module"]}:{new_target}',
                                "type": relation.get("type") or relation.get("tipo") or "Relacionado a",
                            })
                payload["registro_relacionado"] = ""
                payload["relacionamentos"] = rewritten
                module_row = db.execute("SELECT module FROM records WHERE id=?", (record_id,)).fetchone()
                self.db.validate_normative_base(
                    module_row["module"], payload, session["company_id"],
                )
                db.execute("UPDATE records SET payload=? WHERE id=?", (json_dumps(payload), record_id))
                self.db.sync_relationships(record_id, payload, session["id"], session["company_id"])
            if isinstance(data.get("company"), dict):
                company = data["company"]
                db.execute(
                    """UPDATE companies SET name=?,cnpj=?,phone=?,email=?,address=?,updated_at=? WHERE id=?""",
                    (str(company.get("name") or "SECCOL")[:200], company.get("cnpj"), company.get("phone"),
                     company.get("email"), company.get("address"), now, session["company_id"]))
            if isinstance(data.get("settings"), list):
                for setting in data["settings"][:100]:
                    if isinstance(setting, dict) and setting.get("key") in {
                        "preferences", "email", "banking", "certweb"
                    }:
                        db.execute(
                            """INSERT OR REPLACE INTO company_settings(company_id,key,value,updated_at)
                               VALUES(?,?,?,?)""",
                            (session["company_id"], setting["key"], json_dumps(setting.get("value")), now))
            for result in data.get("tender_results", [])[:10000] if isinstance(data.get("tender_results"), list) else []:
                if not isinstance(result, dict):
                    continue
                converted = id_map.get(result.get("converted_record_id")) if result.get("converted_record_id") else None
                source_key = result.get("source_key") or "pncp"
                if session["company_id"] != 1 and not str(source_key).startswith(f'{session["company_id"]}:'):
                    source_key = f'{session["company_id"]}:{source_key}'
                db.execute(
                    """INSERT OR IGNORE INTO tender_results
                       (source_key,external_id,title,object_text,agency,uf,municipality,modality,estimated_value,
                        published_at,deadline,source_url,matched_terms,relevance_score,status,raw_json,
                        converted_record_id,relevance_feedback,feedback_reason,feedback_at,
                        created_at,updated_at,company_id)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (source_key, result.get("external_id"), result.get("title"),
                     result.get("object_text"), result.get("agency"), result.get("uf"), result.get("municipality"),
                     result.get("modality"), result.get("estimated_value"), result.get("published_at"),
                     result.get("deadline"), result.get("source_url"), result.get("matched_terms") or "[]",
                     result.get("relevance_score") or 0, result.get("status") or "Novo",
                     result.get("raw_json") or "{}", converted, result.get("relevance_feedback"),
                     result.get("feedback_reason"), result.get("feedback_at"), result.get("created_at") or now,
                     result.get("updated_at") or now, session["company_id"]))
            for search in data.get("tender_searches", [])[:10000] if isinstance(data.get("tender_searches"), list) else []:
                if not isinstance(search, dict):
                    continue
                db.execute(
                    """INSERT INTO tender_searches
                       (keywords,uf,days,sources_searched,found_count,new_count,error_detail,created_by,created_at,company_id)
                       VALUES(?,?,?,?,?,?,?,?,?,?)""",
                    (search.get("keywords") or "[]", search.get("uf"), search.get("days") or 7,
                     search.get("sources_searched") or "[]", search.get("found_count") or 0,
                     search.get("new_count") or 0, search.get("error_detail"), session["id"],
                     search.get("created_at") or now, session["company_id"]))
            for attachment in data.get("attachments", [])[:10000] if isinstance(data.get("attachments"), list) else []:
                if not isinstance(attachment, dict):
                    continue
                mapped_record = id_map.get(attachment.get("record_id"))
                if not mapped_record or not attachment.get("content"):
                    continue
                try:
                    content = base64.b64decode(attachment["content"], validate=True)
                except (ValueError, TypeError, binascii.Error):
                    continue
                if len(content) > MAX_ATTACHMENT:
                    continue
                db.execute(
                    """INSERT INTO attachments
                       (company_id,record_id,filename,mime_type,content,size,category,version,
                        uploaded_by,created_at,sha256,license_confirmed)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (session["company_id"], mapped_record, Path(str(attachment.get("filename") or "arquivo.bin")).name,
                     attachment.get("mime_type") or "application/octet-stream", content, len(content),
                     attachment.get("category"), attachment.get("version"), session["id"],
                     attachment.get("created_at") or now, hashlib.sha256(content).hexdigest(),
                     1 if attachment.get("license_confirmed") else 0))
            for approval in data.get("approvals", [])[:10000] if isinstance(data.get("approvals"), list) else []:
                if not isinstance(approval, dict):
                    continue
                mapped_record = id_map.get(approval.get("record_id"))
                if not mapped_record:
                    continue
                db.execute(
                    """INSERT INTO approvals
                       (company_id,record_id,approval_type,status,comment,requested_at,decided_at)
                       VALUES(?,?,?,?,?,?,?)""",
                    (session["company_id"], mapped_record,
                     approval.get("approval_type") or "Aprovação",
                     approval.get("status") or "Pendente", approval.get("comment"),
                     approval.get("requested_at") or now, approval.get("decided_at")))
            for event in data.get("fiscal_events", [])[:10000] if isinstance(data.get("fiscal_events"), list) else []:
                if not isinstance(event, dict):
                    continue
                mapped_record = id_map.get(event.get("record_id"))
                if not mapped_record:
                    continue
                db.execute(
                    """INSERT INTO fiscal_events
                       (company_id,record_id,event_type,status,protocol,response_detail,created_by,created_at)
                       VALUES(?,?,?,?,?,?,?,?)""",
                    (session["company_id"], mapped_record,
                     event.get("event_type") or "registrar", event.get("status") or "Importado",
                     event.get("protocol"), event.get("response_detail"), session["id"],
                     event.get("created_at") or now))
            for schedule in data.get("search_schedules", [])[:1000] if isinstance(data.get("search_schedules"), list) else []:
                if not isinstance(schedule, dict):
                    continue
                db.execute(
                    """INSERT INTO search_schedules
                       (company_id,name,keywords,uf,days,frequency,active,last_run_at,next_run_at,
                        created_by,created_at,updated_at)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (session["company_id"], schedule.get("name") or "Monitor importado",
                     schedule.get("keywords") or "[]", schedule.get("uf"), schedule.get("days") or 7,
                     schedule.get("frequency") or "manual", 1 if schedule.get("active", True) else 0,
                     schedule.get("last_run_at"), schedule.get("next_run_at"), session["id"],
                     schedule.get("created_at") or now, schedule.get("updated_at") or now))
            transaction_context.__exit__(None, None, None)
        except Exception as exc:
            transaction_context.__exit__(type(exc), exc, exc.__traceback__)
            if isinstance(exc, (ValueError, TypeError, KeyError, sqlite3.Error)):
                return self.error_json(f"Importação cancelada: {exc}")
            return self.error_json(
                "Importação cancelada porque o arquivo contém dados incompatíveis",
                400, "import_failed",
            )
        self.db.audit(session["id"], "import", "records", detail={"count": count},
                      company_id=session["company_id"])
        return self.send_json({"ok": True, "imported": count})

    def client_error_report(self, session):
        if not self.allow_request("client-error", 20, 5 * 60):
            return
        try:
            data = self.parse_json(max_bytes=64 * 1024)
        except ValueError as exc:
            return self.error_json(str(exc))
        message = Database._log_text(data.get("message"), 500)
        if not message:
            return self.error_json("Informe a mensagem do erro")
        detail = {
            "source": Database._log_text(data.get("source"), 300),
            "line": int(data.get("line") or 0) if str(data.get("line") or "0").isdigit() else 0,
            "column": int(data.get("column") or 0) if str(data.get("column") or "0").isdigit() else 0,
            "stack": str(data.get("stack") or "")[:4000],
        }
        self.db.system_event(
            "error", "client", "javascript_error", message,
            company_id=session["company_id"], user_id=session["id"], detail=detail,
            request_id=self._request_id, path=Database._log_text(data.get("page"), 300),
            method="CLIENT", client_ip=self.client_ip(),
            user_agent=self.headers.get("User-Agent", ""),
        )
        return self.send_json({"ok": True}, 202)

    @staticmethod
    def _epoch_iso(value):
        if not value:
            return None
        return datetime.fromtimestamp(int(value), timezone.utc).isoformat(timespec="seconds")

    def control_center_get(self, session, query):
        if not self.capabilities(session)["control_center"]:
            return self.error_json(
                "O Centro de Controle exige perfil de administrador", 403, "forbidden"
            )
        company_id = session["company_id"]
        now_epoch = int(time.time())
        active_after = now_epoch - SESSION_ACTIVE_SECONDS
        sessions = self.db.connection().execute(
            """SELECT s.public_id,s.created_at,s.last_activity_at,s.expires_at,
                      s.ip_address,s.user_agent,u.id user_id,u.name,u.email,cm.role
               FROM sessions s
               JOIN users u ON u.id=s.user_id AND u.active=1
               JOIN company_memberships cm
                 ON cm.user_id=u.id AND cm.company_id=s.company_id AND cm.active=1
               WHERE s.company_id=? AND s.expires_at>=?
               ORDER BY s.last_activity_at DESC LIMIT 100""",
            (company_id, now_epoch),
        ).fetchall()
        session_items = [{
            "id": row["public_id"], "userId": row["user_id"], "name": row["name"],
            "email": row["email"], "role": row["role"], "createdAt": row["created_at"],
            "lastActivityAt": self._epoch_iso(row["last_activity_at"]),
            "expiresAt": self._epoch_iso(row["expires_at"]),
            "ipAddress": row["ip_address"] or "Não identificado",
            "userAgent": row["user_agent"] or "Não identificado",
            "activeNow": bool(row["last_activity_at"] and row["last_activity_at"] >= active_after),
            "current": row["public_id"] == session["public_id"],
        } for row in sessions]

        audit_rows = self.db.connection().execute(
            """SELECT a.id,a.action,a.entity_type,a.entity_id,a.detail,a.created_at,
                      u.id user_id,u.name user_name
               FROM audit_log a LEFT JOIN users u ON u.id=a.user_id
               WHERE a.company_id=? ORDER BY a.id DESC LIMIT 100""",
            (company_id,),
        ).fetchall()
        changes = [dict(row) for row in audit_rows]
        event_rows = self.db.connection().execute(
            """SELECT e.id,e.severity,e.category,e.event_type,e.message,e.detail,
                      e.request_id,e.path,e.method,e.client_ip,e.user_agent,e.resolved_at,
                      e.created_at,u.name user_name
               FROM system_events e LEFT JOIN users u ON u.id=e.user_id
               WHERE e.company_id=? ORDER BY e.id DESC LIMIT 100""",
            (company_id,),
        ).fetchall()
        events = []
        for row in event_rows:
            item = dict(row)
            try:
                item["detail"] = json_loads_strict(item["detail"]) if item["detail"] else None
            except (ValueError, TypeError, json.JSONDecodeError):
                item["detail"] = None
            events.append(item)

        users = self.db.connection().execute(
            """SELECT COUNT(*) total,
                      SUM(CASE WHEN cm.active=1 AND u.active=1 THEN 1 ELSE 0 END) active
               FROM company_memberships cm JOIN users u ON u.id=cm.user_id
               WHERE cm.company_id=?""",
            (company_id,),
        ).fetchone()
        job_rows = self.db.connection().execute(
            """SELECT status,COUNT(*) total FROM tender_jobs
               WHERE company_id=? GROUP BY status""",
            (company_id,),
        ).fetchall()
        jobs = {row["status"]: row["total"] for row in job_rows}
        last_backup = self.db.scalar(
            """SELECT created_at FROM audit_log
               WHERE company_id=? AND action='backup' ORDER BY id DESC LIMIT 1""",
            (company_id,),
        )
        open_errors = self.db.scalar(
            """SELECT COUNT(*) FROM system_events
               WHERE company_id=? AND resolved_at IS NULL AND severity='error'""",
            (company_id,),
        ) or 0
        db_path = self.db.path.resolve()
        disk = shutil.disk_usage(db_path.parent)
        mount_required = os.environ.get("SIVS_REQUIRE_PERSISTENT_DB") == "1"
        storage_verified = database_directory_is_mount(db_path) if mount_required else None
        health = {
            "version": VERSION,
            "uptimeSeconds": int(time.time() - self.server.started_at),  # type: ignore[attr-defined]
            "databaseBytes": db_path.stat().st_size if db_path.exists() else 0,
            "walBytes": Path(str(db_path) + "-wal").stat().st_size
                if Path(str(db_path) + "-wal").exists() else 0,
            "diskFreeBytes": disk.free,
            "diskTotalBytes": disk.total,
            "persistentStorageRequired": mount_required,
            "persistentStorageVerified": storage_verified,
            "schedulerRunning": self.server._scheduler.is_alive(),  # type: ignore[attr-defined]
            "lastBackupAt": last_backup,
            "aiConfigured": bool(os.environ.get("OPENROUTER_API_KEY")),
            "cnpjLookupConfigured": bool(os.environ.get("CNPJA_API_KEY")),
        }
        active_users = len({item["userId"] for item in session_items if item["activeNow"]})
        return self.send_json({
            "ok": True,
            "generatedAt": utc_now(),
            "summary": {
                "activeUsers": active_users,
                "activeSessions": sum(1 for item in session_items if item["activeNow"]),
                "validSessions": len(session_items),
                "usersTotal": int(users["total"] or 0),
                "usersEnabled": int(users["active"] or 0),
                "openErrors": int(open_errors),
            },
            "health": health,
            "requests": self.server.telemetry_snapshot(),  # type: ignore[attr-defined]
            "jobs": jobs,
            "sessions": session_items,
            "changes": changes,
            "events": events,
        })

    def control_center_session_delete(self, path, session):
        if not self.capabilities(session)["control_center"]:
            return self.error_json("Operação exclusiva de administrador", 403, "forbidden")
        public_id = path.rsplit("/", 1)[-1]
        if not re.fullmatch(r"[a-f0-9]{24}", public_id):
            return self.error_json("Sessão inválida")
        row = self.db.connection().execute(
            """SELECT s.public_id,s.user_id,u.name FROM sessions s JOIN users u ON u.id=s.user_id
               WHERE s.public_id=? AND s.company_id=?""",
            (public_id, session["company_id"]),
        ).fetchone()
        if not row:
            return self.error_json("Sessão não encontrada", 404, "not_found")
        if public_id == session["public_id"]:
            return self.error_json(
                "Use Encerrar sessão para sair deste dispositivo", 409, "current_session"
            )
        with self.db.transaction(immediate=True):
            self.db.execute("DELETE FROM sessions WHERE public_id=?", (public_id,))
            self.db.audit(
                session["id"], "terminate", "session", public_id,
                {"user_id": row["user_id"], "user_name": row["name"]},
                company_id=session["company_id"],
            )
        return self.send_json({"ok": True})

    def control_center_event_resolve(self, path, session):
        if not self.capabilities(session)["control_center"]:
            return self.error_json("Operação exclusiva de administrador", 403, "forbidden")
        parts = path.split("/")
        if len(parts) != 6 or not parts[4].isdigit():
            return self.error_json("Evento inválido")
        event_id = int(parts[4])
        with self.db.transaction(immediate=True):
            updated = self.db.execute(
                """UPDATE system_events SET resolved_at=?,resolved_by=?
                   WHERE id=? AND company_id=? AND resolved_at IS NULL""",
                (utc_now(), session["id"], event_id, session["company_id"]),
            )
            if updated.rowcount != 1:
                return self.error_json("Evento não encontrado ou já resolvido", 404, "not_found")
            self.db.audit(
                session["id"], "resolve", "system_event", event_id,
                company_id=session["company_id"],
            )
        return self.send_json({"ok": True})

    def static_get(self, path):
        if path == "/":
            path = "/index.html"
        requested = (STATIC_DIR / path.lstrip("/")).resolve()
        if STATIC_DIR.resolve() not in requested.parents and requested != STATIC_DIR.resolve():
            return self.send_error(403)
        if not requested.exists() or not requested.is_file():
            if Path(path).suffix:
                return self.error_json("Arquivo não encontrado", 404, "not_found")
            requested = STATIC_DIR / "index.html"
        body = requested.read_bytes()
        content_type = mimetypes.guess_type(requested.name)[0] or "application/octet-stream"
        self._response_started = True
        self.send_response(200)
        self.send_header("Content-Type", content_type + ("; charset=utf-8" if content_type.startswith("text/") else ""))
        self.send_header("Content-Length", str(len(body)))
        volatile_asset = requested.name in {"index.html", "service-worker.js", "manifest.json"} or requested.suffix in {".js", ".css"}
        self.send_header(
            "Cache-Control",
            "no-cache, no-store, must-revalidate" if volatile_asset else "public, max-age=3600",
        )
        if volatile_asset:
            self.send_header("Pragma", "no-cache")
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("X-Frame-Options", "DENY")
        self.send_header("Referrer-Policy", "same-origin")
        self.send_header("Content-Security-Policy", "default-src 'self'; img-src 'self' data: https:; style-src 'self' 'unsafe-inline'; script-src 'self'; connect-src 'self'")
        self.end_headers()
        self.wfile.write(body)


class SIVSServer(ThreadingHTTPServer):
    daemon_threads = True

    def __init__(self, address, handler, db):
        super().__init__(address, handler)
        self.db = db
        self.started_at = time.time()
        self._telemetry_lock = threading.Lock()
        self._request_samples = collections.deque(maxlen=5000)
        self._request_total = 0
        self._rate_lock = threading.Lock()
        self._rate_buckets = {}
        self._partner_lookup_lock = threading.Lock()
        self._partner_lookup_cache = {}
        self._stop_workers = threading.Event()
        self._scheduler = threading.Thread(
            target=self._scheduler_loop, name="sivs-scheduler", daemon=True
        )
        self._scheduler.start()

    def record_request(self, method, path, status, duration_ms):
        normalized = re.sub(r"/\d+(?=/|$)", "/:id", str(path))[:240]
        with self._telemetry_lock:
            self._request_total += 1
            self._request_samples.append({
                "at": time.time(), "method": str(method)[:12], "path": normalized,
                "status": int(status), "durationMs": round(float(duration_ms), 2),
            })

    def telemetry_snapshot(self):
        cutoff = time.time() - 15 * 60
        with self._telemetry_lock:
            samples = [dict(item) for item in self._request_samples if item["at"] >= cutoff]
            total = self._request_total
        durations = sorted(item["durationMs"] for item in samples)
        p95 = durations[math.ceil(len(durations) * 0.95) - 1] if durations else 0
        slowest = sorted(samples, key=lambda item: item["durationMs"], reverse=True)[:8]
        return {
            "sinceStart": total,
            "last15Minutes": len(samples),
            "clientErrors": sum(1 for item in samples if 400 <= item["status"] < 500),
            "serverErrors": sum(1 for item in samples if item["status"] >= 500),
            "averageMs": round(sum(durations) / len(durations), 2) if durations else 0,
            "p95Ms": p95,
            "slowest": slowest,
        }

    def rate_limit(self, bucket, client, limit, window_seconds):
        now = time.monotonic()
        key = (str(bucket), str(client))
        with self._rate_lock:
            recent = [stamp for stamp in self._rate_buckets.get(key, []) if now - stamp < window_seconds]
            if len(recent) >= limit:
                retry_after = max(1, int(window_seconds - (now - recent[0])) + 1)
                self._rate_buckets[key] = recent
                return False, retry_after
            recent.append(now)
            self._rate_buckets[key] = recent
            if len(self._rate_buckets) > 10_000:
                self._rate_buckets = {key: recent}
            return True, 0

    def partner_lookup_cache_get(self, key):
        now = time.monotonic()
        with self._partner_lookup_lock:
            item = self._partner_lookup_cache.get(key)
            if not item or now - item[0] >= PARTNER_LOOKUP_CACHE_SECONDS:
                self._partner_lookup_cache.pop(key, None)
                return None
            return dict(item[1])

    def partner_lookup_cache_put(self, key, value):
        with self._partner_lookup_lock:
            self._partner_lookup_cache[key] = (time.monotonic(), dict(value))
            if len(self._partner_lookup_cache) > 2_000:
                oldest = min(self._partner_lookup_cache, key=lambda item: self._partner_lookup_cache[item][0])
                self._partner_lookup_cache.pop(oldest, None)

    def process_request_thread(self, request, client_address):
        try:
            super().process_request_thread(request, client_address)
        finally:
            self.db.close_thread_connection()

    def run_tender_job(self, runner, job_id, session, request_data):
        try:
            runner._run_tender_job(job_id, session, request_data)
        finally:
            self.db.close_thread_connection()

    def _scheduler_loop(self):
        while not self._stop_workers.is_set():
            try:
                self._release_expired_inventory_reservations()
                self._enqueue_due_tender_schedules()
            except Exception:
                print("[ERRO AGENDADOR] Não foi possível processar os planos de pesquisa")
                traceback.print_exc()
            self._stop_workers.wait(30)
        self.db.close_thread_connection()

    def _release_expired_inventory_reservations(self):
        """Libera reservas vencidas sem editar ou apagar o histórico do ledger."""
        today = datetime.now().astimezone().date().isoformat()
        now = utc_now()
        released = 0
        has_expired = self.db.connection().execute(
            """SELECT 1 FROM inventory_reservations
               WHERE status='ACTIVE' AND expires_at IS NOT NULL AND expires_at<? LIMIT 1""",
            (today,),
        ).fetchone()
        if not has_expired:
            return 0
        with self.db.transaction(immediate=True):
            reservations = self.db.connection().execute(
                """SELECT * FROM inventory_reservations
                   WHERE status='ACTIVE' AND expires_at IS NOT NULL AND expires_at<?
                   ORDER BY expires_at,id LIMIT 200""",
                (today,),
            ).fetchall()
            for reservation in reservations:
                quantity = int(reservation["quantity_micros"])
                balance = self.db.connection().execute(
                    """UPDATE inventory_balances
                       SET reserved_quantity_micros=reserved_quantity_micros-?,
                           revision=revision+1,updated_at=?
                       WHERE company_id=? AND warehouse_id=? AND product_record_id=?
                         AND lot_key=? AND reserved_quantity_micros>=?""",
                    (quantity, now, reservation["company_id"], reservation["warehouse_id"],
                     reservation["product_record_id"], reservation["lot_key"], quantity),
                )
                if balance.rowcount != 1:
                    raise sqlite3.IntegrityError(
                        f"reserva vencida {reservation['id']} possui saldo inconsistente"
                    )
                updated = self.db.connection().execute(
                    """UPDATE inventory_reservations
                       SET status='RELEASED',released_by=NULL,updated_at=?
                       WHERE id=? AND company_id=? AND status='ACTIVE'""",
                    (now, reservation["id"], reservation["company_id"]),
                )
                if updated.rowcount != 1:
                    raise sqlite3.IntegrityError(
                        f"reserva vencida {reservation['id']} mudou durante a liberação"
                    )
                movement_id = self.db.connection().execute(
                    """INSERT INTO inventory_movements
                       (company_id,warehouse_id,counterpart_warehouse_id,product_record_id,
                        lot_key,movement_type,quantity_micros,physical_delta_micros,
                        reserved_delta_micros,origin_type,origin_id,reference,reason,
                        reservation_id,created_by,created_at)
                       VALUES(?,?,NULL,?,?,'RELEASE_RESERVATION',?,0,?,?,?,?,?,?,NULL,?)""",
                    (reservation["company_id"], reservation["warehouse_id"],
                     reservation["product_record_id"], reservation["lot_key"], quantity,
                     -quantity, reservation["origin_type"], reservation["origin_id"],
                     reservation["reference"], "Expiração automática da reserva",
                     reservation["id"], now),
                ).lastrowid
                self.db.audit(
                    None, "expire", "inventory", reservation["id"],
                    {"movement_id": movement_id, "expires_at": reservation["expires_at"],
                     "quantity_micros": quantity},
                    company_id=reservation["company_id"],
                )
                released += 1
        return released

    def _enqueue_due_tender_schedules(self):
        now = utc_now()
        queued = []
        with self.db.transaction(immediate=True):
            schedules = self.db.connection().execute(
                """SELECT * FROM search_schedules
                   WHERE active=1 AND frequency IN ('daily','weekly')
                     AND next_run_at IS NOT NULL AND next_run_at<=?
                   ORDER BY next_run_at,id LIMIT 20""",
                (now,),
            ).fetchall()
            for schedule in schedules:
                active = self.db.connection().execute(
                    """SELECT 1 FROM tender_jobs
                       WHERE company_id=? AND status IN ('queued','running') LIMIT 1""",
                    (schedule["company_id"],),
                ).fetchone()
                delta = timedelta(days=1 if schedule["frequency"] == "daily" else 7)
                next_run = (datetime.now(timezone.utc) + delta).isoformat(timespec="seconds")
                if active:
                    self.db.execute(
                        "UPDATE search_schedules SET next_run_at=?,updated_at=? WHERE id=?",
                        (next_run, now, schedule["id"]),
                    )
                    continue
                request_data = {
                    "keywords": json_loads_strict(schedule["keywords"] or "[]"),
                    "uf": schedule["uf"] or "",
                    "days": schedule["days"],
                }
                cursor = self.db.execute(
                    """INSERT INTO tender_jobs
                       (company_id,schedule_id,status,request_json,progress,stage,created_by,created_at)
                       VALUES(?,?,'queued',?,0,'Pesquisa agendada enfileirada',?,?)""",
                    (schedule["company_id"], schedule["id"], json_dumps(request_data),
                     schedule["created_by"], now),
                )
                self.db.execute(
                    "UPDATE search_schedules SET next_run_at=?,updated_at=? WHERE id=?",
                    (next_run, now, schedule["id"]),
                )
                queued.append((cursor.lastrowid, schedule["company_id"],
                               schedule["created_by"], request_data))
        for job_id, company_id, user_id, request_data in queued:
            runner = object.__new__(SIVSHandler)
            runner.server = self
            threading.Thread(
                target=self.run_tender_job,
                args=(runner, job_id, {"id": user_id, "company_id": company_id}, request_data),
                name=f"sivs-scheduled-tender-{job_id}", daemon=True,
            ).start()

    def server_close(self):
        self._stop_workers.set()
        if self._scheduler.is_alive() and threading.current_thread() is not self._scheduler:
            self._scheduler.join(timeout=2)
        super().server_close()
        self.db.close_thread_connection()


def main():
    parser = argparse.ArgumentParser(description="Servidor local do SIVS")
    parser.add_argument("--host", default=os.environ.get("SIVS_HOST", "127.0.0.1"))
    parser.add_argument(
        "--port",
        type=int,
        default=int(os.environ.get("PORT") or os.environ.get("SIVS_PORT", "8844")),
    )
    parser.add_argument("--db", type=Path, default=Path(os.environ.get("SIVS_DB", DEFAULT_DB)))
    parser.add_argument(
        "--allow-insecure-network", action="store_true",
        default=os.environ.get("SIVS_ALLOW_INSECURE_NETWORK") == "1",
        help="permite HTTP em interface não local; use apenas em rede isolada e temporariamente",
    )
    args = parser.parse_args()
    local_hosts = {"127.0.0.1", "localhost", "::1"}
    proxy_secure = (
        os.environ.get("SIVS_TRUST_PROXY") == "1" and
        os.environ.get("SIVS_SECURE_COOKIE") == "1"
    )
    if args.host not in local_hosts and not (args.allow_insecure_network or proxy_secure):
        parser.error(
            "acesso em rede exige proxy HTTPS. Mantenha --host 127.0.0.1 ou, assumindo o risco, "
            "use --allow-insecure-network"
        )
    if args.host not in local_hosts and args.allow_insecure_network and not proxy_secure:
        print("AVISO CRÍTICO: HTTP em rede foi liberado explicitamente; credenciais não terão proteção TLS.")
    persistent_storage = require_persistent_database(args.db)
    persistent_state = None
    prestart_backup = None
    if persistent_storage:
        persistent_state = validate_persistent_database_state(args.db)
        if not persistent_state["bootstrap"]:
            prestart_backup = create_prestart_database_backup(args.db)
    db = Database(args.db)
    server = SIVSServer((args.host, args.port), SIVSHandler, db)
    print(f"SIVS disponível em http://{args.host}:{args.port}")
    print(f"Banco de dados: {args.db.resolve()}")
    if persistent_storage:
        print("Persistencia do banco: volume montado e verificado")
        if persistent_state and persistent_state["bootstrap"]:
            print("AVISO: bootstrap vazio autorizado temporariamente; remova SIVS_ALLOW_EMPTY_DB_INITIALIZATION")
        elif prestart_backup:
            print(f"Snapshot pre-start verificado: {prestart_backup}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nSIVS encerrado.")
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
